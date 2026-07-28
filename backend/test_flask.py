import json
import time
import logging
import threading
import requests
import urllib3
import pandas as pd
import os
import glob
import re
import random
import smtplib
import urllib.request
import urllib.error
import urllib.parse
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from collections import defaultdict
from io import BytesIO
from datetime import datetime, timedelta

from flask import Flask, jsonify, request, Response, stream_with_context, send_file, g
from flask_cors import CORS
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

from auth import LDAPAuth, create_jwt, decode_jwt, jwt_required

# ======================================================
# In-memory credential cache (stores LDAP passwords for Jita calls)
# Passwords are cached on login and expire after the same TTL as JWT.
# ======================================================
_credential_cache = {}  # {username: {"password": str, "expires_at": float}}
_credential_cache_lock = threading.Lock()
CREDENTIAL_TTL_SECONDS = int(os.environ.get("JWT_EXPIRY_HOURS", "24")) * 3600


def _store_user_credentials(username, password):
    """Cache user credentials on successful login."""
    with _credential_cache_lock:
        _credential_cache[username] = {
            "password": password,
            "expires_at": time.time() + CREDENTIAL_TTL_SECONDS,
        }


def _get_user_credentials(username):
    """Retrieve cached credentials. Returns (username, password) tuple or None."""
    with _credential_cache_lock:
        entry = _credential_cache.get(username)
        if entry and entry["expires_at"] > time.time():
            return (username, entry["password"])
        _credential_cache.pop(username, None)
        return None


# ======================================================
# Flask App
# ======================================================
app = Flask(__name__)
CORS(app, supports_credentials=True)

# ======================================================
# Disable SSL warnings
# ======================================================
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ======================================================
# Logging
# ======================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)
logger = logging.getLogger(__name__)

# ======================================================
# Constants
# ======================================================
JITA_BASE = "https://jita.eng.nutanix.com/api/v2"
# QMS coupon service (used by JITA's Provider page "Validate" button for global pool).
QMS_BASE_URL = os.getenv("QMS_BASE_URL", "https://qms-api.nucloud.ntnxdpro.com").rstrip("/")
# JITA manage UI: /manage/job_profiles/<id> often 404s. Open the list and pre-fill name search (filter bar syntax).
JITA_WEB_ORIGIN = os.getenv("JITA_WEB_ORIGIN", "https://jita.eng.nutanix.com").rstrip("/")
# Query key JITA reads on page load (override if your build uses e.g. "filter" or "q").
JITA_WEB_MANAGE_SEARCH_PARAM = os.getenv("JITA_WEB_MANAGE_SEARCH_PARAM", "search").strip().strip("?&=") or "search"
JITA_WEB_JOB_PROFILE_URL = os.getenv(
    "JITA_WEB_JOB_PROFILE_URL",
    "{origin}/manage/job_profiles?" + JITA_WEB_MANAGE_SEARCH_PARAM + "={search_query}",
)
JITA_WEB_TEST_SET_URL = os.getenv(
    "JITA_WEB_TEST_SET_URL",
    "{origin}/manage/test_sets?" + JITA_WEB_MANAGE_SEARCH_PARAM + "={search_query}",
)
TRIAGE_GENIE_BASE = "http://triage-genie.eng.nutanix.com/api"
LOGIN_URL = os.getenv("TRIAGE_GENIE_LOGIN_URL", "http://triage-genie.eng.nutanix.com/login")
TRIAGE_GENIE_USERNAME = os.getenv("TRIAGE_GENIE_USERNAME", "")
TRIAGE_GENIE_PASSWORD = os.getenv("TRIAGE_GENIE_PASSWORD", "")
PHX_BASE = "https://jita-phx1-webserver-2.eng.nutanix.com/api/v2"
TCMS_BASE = "https://tcms.eng.nutanix.com/api-readonly/v1"
TCMS_SUMMARY_BASE = "https://tcms.eng.nutanix.com/api/v1"
TCMS_WRITE_BASE = "https://tcms.eng.nutanix.com/api/v1"
TCMS_TESTDB_BASE = "https://quality-pipeline.eng.nutanix.com/testdb/api/v1"

# TCMS auth (base64-encoded defaults; override with env vars in production)
TCMS_USER = os.getenv("TCMS_USER", "agave_bot")
TCMS_PASSWORD = os.getenv("TCMS_PASSWORD", "admin")

TESTCASE_MGMT_DATA_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data"
)


def _jita_browser_entity_url(template, entity_id=None, entity_name=None):
    """Build JITA manage UI URL from env template.

    Placeholders:
      {origin}         — JITA_WEB_ORIGIN (no trailing slash)
      {id}             — entity id (if template uses a direct record path)
      {search_query}   — urllib-quoted ``showAll:true name:<entity_name>`` (manage list filter bar)
    Query key for list+search defaults: ``JITA_WEB_MANAGE_SEARCH_PARAM`` (default ``search``), or set
    ``JITA_WEB_JOB_PROFILE_URL`` / ``JITA_WEB_TEST_SET_URL`` to the exact pattern your JITA build expects.
    """
    if not template:
        return None
    try:
        out = str(template).replace("{origin}", JITA_WEB_ORIGIN)
    except Exception:
        return None
    eid = str(entity_id).strip() if entity_id else ""
    if "{id}" in out:
        if not eid:
            return None
        out = out.replace("{id}", eid)
    if "{search_query}" in out:
        if not entity_name or not str(entity_name).strip():
            return None
        raw_q = f"showAll:true name:{str(entity_name).strip()}"
        out = out.replace("{search_query}", urllib.parse.quote(raw_q, safe=""))
    out = out.strip()
    if out.startswith("http://") or out.startswith("https://"):
        return out
    if out.startswith("/"):
        return f"{JITA_WEB_ORIGIN}{out}"
    return None


# Tag-to-team configuration for TCMS QI lookups.
# Each key is a tag pattern; value holds the TCMS team name and fallback branch.
# "default" is used when no specific tag match is found.
TEAM_CONFIG = {
    "cdp_master_full_reg": {"team": "CDP", "default_branch": "master"},
    "default":             {"team": "CDP", "default_branch": "master"},
}

# Maps full branch names (as shown in the Run Summary table) to the short
# milestone names expected by the TCMS API.  "master" stays as-is.
BRANCH_SHORT_NAME_MAP = {
    "master": "master",
    "ganges-7.6-stable": "7.6",
    "ganges-7.5-stable": "7.5",
    "ganges-7.5.1-stable": "7.5.1",
}

# AI Endpoint for failure summary
AI_BASE = "https://hkn12.ai.nutanix.com/enterpriseai/v1"
AI_API_KEY = "ddb2b793-1004-49a1-b005-4ddf4c2ade8c"

# SSL context for AI endpoint (skip TLS verify)
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

MAX_FAILED_TESTS = 50
MAX_WORKERS = 5

HEADERS = {
    "Authorization": "Bearer TOKEN",  # move to env later
    "Content-Type": "application/json",
    "Accept": "application/json"
}

# In-memory storage for manual tasks
# Structure: {tag: {branch: [task_ids]}}
manual_tasks_store = {}

# Run Plan storage file
RUN_PLAN_STORAGE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "run_plans.json")

def load_run_plans():
    """Load run plans from JSON file, backfilling missing fields on older entries."""
    try:
        if os.path.exists(RUN_PLAN_STORAGE):
            with open(RUN_PLAN_STORAGE, 'r') as f:
                data = json.load(f)
            dirty = False
            for rp in data.get("run_plans", []):
                if "schedule_triggered" not in rp:
                    rp["schedule_triggered"] = False
                    dirty = True
                if "branch" not in rp:
                    rp["branch"] = ""
                    dirty = True
                if "service_account" not in rp:
                    rp["service_account"] = ""
                    dirty = True
            if dirty:
                save_run_plans(data)
            return data
        return {"run_plans": [], "history": []}
    except Exception as e:
        logger.error(f"Error loading run plans: {e}")
        return {"run_plans": [], "history": []}

def save_run_plans(data):
    """Save run plans to JSON file"""
    try:
        with open(RUN_PLAN_STORAGE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving run plans: {e}")
        raise

# ======================================================
# Run Plan Scheduler — checks every 30 min for due scheduled runs
# Uses service account (JITA_SVC_AUTH) so no user session is needed.
# ======================================================
_scheduler_lock = threading.Lock()
SCHEDULER_INTERVAL_SECONDS = int(os.environ.get("SCHEDULER_INTERVAL_SECONDS", "1800"))


def _trigger_scheduled_run_plan(run_plan, data):
    """Trigger a single run plan via the service account and record history."""
    rp_id = run_plan["id"]
    rp_name = run_plan.get("name", rp_id)
    job_profile_ids = [
        jp for jp in run_plan.get("job_profiles", [])
        if jp and isinstance(jp, str) and jp.strip()
    ]
    if not job_profile_ids:
        logger.warning(f"[scheduler] Run plan '{rp_name}' has no valid job profiles — skipping")
        return

    svc_name = run_plan.get("service_account", "")
    trigger_auth = RUN_PLAN_SERVICE_ACCOUNTS.get(svc_name, JITA_SVC_AUTH) if svc_name else JITA_SVC_AUTH
    logger.info(f"[scheduler] Triggering '{rp_name}' ({len(job_profile_ids)} JP(s)) via {'svc:' + svc_name if svc_name else 'default svc account'}")
    task_ids = []
    failed_jobs = []

    for jp_id in job_profile_ids:
        try:
            url = f"{JITA_BASE}/job_profiles/{jp_id}/trigger"
            resp = requests.post(
                url, json={}, headers={"Content-Type": "application/json"},
                auth=trigger_auth, verify=False, timeout=60
            )
            if resp.status_code == 200:
                res_data = resp.json()
                if res_data.get("success") and "task_ids" in res_data:
                    ids = [
                        item["$oid"] if isinstance(item, dict) and "$oid" in item else item
                        for item in res_data["task_ids"]
                    ]
                    task_ids.extend(ids)
                else:
                    failed_jobs.append({
                        "job_id": jp_id, "success": False,
                        "error": res_data.get("message", "Trigger returned failure")
                    })
            else:
                failed_jobs.append({
                    "job_id": jp_id, "success": False,
                    "error": f"HTTP {resp.status_code}: {resp.text[:200]}"
                })
        except Exception as exc:
            failed_jobs.append({"job_id": jp_id, "success": False, "error": str(exc)})

    now_iso = datetime.now().isoformat()
    run_plan["last_triggered"] = now_iso
    run_plan["schedule_triggered"] = True

    if "history" not in data:
        data["history"] = []
    data["history"].append({
        "id": str(int(time.time() * 1000)),
        "run_plan_id": rp_id,
        "triggered_at": now_iso,
        "triggered_by": "scheduler (service account)",
        "task_ids": task_ids,
        "failed_jobs": failed_jobs,
        "status": "success" if not failed_jobs else "partial"
    })
    logger.info(
        f"[scheduler] '{rp_name}' done — {len(task_ids)} task(s), {len(failed_jobs)} failure(s)"
    )


def _run_plan_scheduler_loop():
    """Background loop: every SCHEDULER_INTERVAL_SECONDS, check for due scheduled run plans."""
    while True:
        try:
            with _scheduler_lock:
                data = load_run_plans()
                now = datetime.now()
                triggered_any = False
                for rp in data.get("run_plans", []):
                    sched = rp.get("schedule_date")
                    if not sched:
                        continue
                    if rp.get("schedule_triggered"):
                        continue
                    try:
                        sched_dt = datetime.fromisoformat(sched)
                    except (ValueError, TypeError):
                        continue
                    if sched_dt <= now:
                        _trigger_scheduled_run_plan(rp, data)
                        triggered_any = True
                if triggered_any:
                    save_run_plans(data)
        except Exception as exc:
            logger.error(f"[scheduler] Unhandled error: {exc}", exc_info=True)
        time.sleep(SCHEDULER_INTERVAL_SECONDS)

# Triage Genie jobs storage file
TRIAGE_GENIE_STORAGE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "triage_genie_jobs.json")

def load_triage_genie_jobs():
    """Load Triage Genie jobs from JSON file"""
    try:
        if os.path.exists(TRIAGE_GENIE_STORAGE):
            with open(TRIAGE_GENIE_STORAGE, 'r') as f:
                return json.load(f)
        return {"jobs": []}
    except Exception as e:
        logger.error(f"Error loading triage genie jobs: {e}")
        return {"jobs": []}

def save_triage_genie_jobs(data):
    """Save Triage Genie jobs to JSON file"""
    try:
        with open(TRIAGE_GENIE_STORAGE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving triage genie jobs: {e}")
        raise

# Regression Dashboard Configuration storage file
REGRESSION_CONFIG_STORAGE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "regression_config.json")

def load_regression_config():
    """Load regression dashboard configuration from JSON file. Migrates legacy schema."""
    try:
        if os.path.exists(REGRESSION_CONFIG_STORAGE):
            with open(REGRESSION_CONFIG_STORAGE, 'r') as f:
                config = json.load(f)
            # Migration: add default_tag, added_tags if missing
            added = config.get("added_tags", [])
            if not isinstance(added, list):
                added = []
            if "default_tag" not in config:
                existing_tag = config.get("tag", "").strip()
                if existing_tag:
                    config["default_tag"] = existing_tag
                    if existing_tag not in added:
                        added = list(added) + [existing_tag]
                else:
                    config["default_tag"] = None
            config["added_tags"] = added
            return config
        return {
            "input_mode": "tag",
            "tag": "cdp_master_full_reg",
            "default_tag": "cdp_master_full_reg",
            "added_tags": ["cdp_master_full_reg"],
            "task_ids": []
        }
    except Exception as e:
        logger.error(f"Error loading regression config: {e}")
        return {
            "input_mode": "tag",
            "tag": "cdp_master_full_reg",
            "default_tag": "cdp_master_full_reg",
            "added_tags": ["cdp_master_full_reg"],
            "task_ids": []
        }

def save_regression_config(data):
    """Save regression dashboard configuration to JSON file"""
    try:
        with open(REGRESSION_CONFIG_STORAGE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving regression config: {e}")
        raise

# Triage Accuracy Analyzer data storage
TRIAGE_ACCURACY_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
TRIAGE_ACCURACY_TASKIDS_FILE = "triage_accuracy_data_taskids.json"

def _sanitize_tag_for_filename(tag):
    """Replace unsafe chars for filenames; collapse multiple underscores."""
    if not tag or not isinstance(tag, str):
        return "unknown"
    s = tag.strip()
    for c in r'|/:*?"<>\\':
        s = s.replace(c, "_")
    s = re.sub(r"_+", "_", s).strip("_")
    return s if s else "unknown"

def _triage_accuracy_path(tag=None):
    """Get path for triage accuracy JSON. tag=None means task_ids mode."""
    os.makedirs(TRIAGE_ACCURACY_DATA_DIR, exist_ok=True)
    if tag:
        sanitized = _sanitize_tag_for_filename(tag)
        return os.path.join(TRIAGE_ACCURACY_DATA_DIR, f"triage_accuracy_data_{sanitized}.json")
    return os.path.join(TRIAGE_ACCURACY_DATA_DIR, TRIAGE_ACCURACY_TASKIDS_FILE)

def load_triage_accuracy_data(tag=None):
    """Load triage accuracy data from JSON file. tag=None for task_ids mode."""
    try:
        path = _triage_accuracy_path(tag)
        if os.path.exists(path):
            with open(path, 'r') as f:
                return json.load(f)
        # Migration: copy legacy triage_accuracy_data.json to per-tag file if tag matches
        if tag:
            legacy_path = os.path.join(TRIAGE_ACCURACY_DATA_DIR, "triage_accuracy_data.json")
            if os.path.exists(legacy_path):
                with open(legacy_path, 'r') as f:
                    data = json.load(f)
                cached_tag = (data.get("tag") or "").strip()
                if cached_tag == tag:
                    save_triage_accuracy_data(data, tag)
                    return data
        return None
    except Exception as e:
        logger.error(f"Error loading triage accuracy data: {e}")
        return None

def save_triage_accuracy_data(data, tag=None):
    """Save triage accuracy data to JSON file. tag=None for task_ids mode."""
    try:
        path = _triage_accuracy_path(tag)
        os.makedirs(TRIAGE_ACCURACY_DATA_DIR, exist_ok=True)
        with open(path, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving triage accuracy data: {e}")
        raise

def invalidate_triage_accuracy_cache(tag=None):
    """Delete triage accuracy cache. tag=None invalidates only task_ids file; pass tag for per-tag file."""
    try:
        path = _triage_accuracy_path(tag)
        if os.path.exists(path):
            os.remove(path)
            logger.info(f"Invalidated triage accuracy cache: {path}")
    except Exception as e:
        logger.warning(f"Could not invalidate triage accuracy cache: {e}")

# --------------- Failed Analysis Saved Tags storage ---------------
FAILED_ANALYSIS_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
FAILED_ANALYSIS_TAGS_FILE = os.path.join(FAILED_ANALYSIS_DATA_DIR, "failed_analysis_saved_tags.json")

def _failed_analysis_results_path(tag):
    sanitized = _sanitize_tag_for_filename(tag)
    return os.path.join(FAILED_ANALYSIS_DATA_DIR, f"failed_analysis_{sanitized}.json")

def load_failed_analysis_tags():
    try:
        if os.path.exists(FAILED_ANALYSIS_TAGS_FILE):
            with open(FAILED_ANALYSIS_TAGS_FILE, 'r') as f:
                data = json.load(f)
                if isinstance(data, dict) and "tags" in data:
                    return data
        return {"tags": []}
    except Exception as e:
        logger.error(f"Error loading failed analysis tags: {e}")
        return {"tags": []}

def save_failed_analysis_tags(data):
    try:
        os.makedirs(FAILED_ANALYSIS_DATA_DIR, exist_ok=True)
        with open(FAILED_ANALYSIS_TAGS_FILE, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving failed analysis tags: {e}")
        raise

def load_failed_analysis_results(tag):
    try:
        path = _failed_analysis_results_path(tag)
        if os.path.exists(path):
            with open(path, 'r') as f:
                return json.load(f)
        return None
    except Exception as e:
        logger.error(f"Error loading failed analysis results for tag '{tag}': {e}")
        return None

def save_failed_analysis_results(tag, data):
    try:
        os.makedirs(FAILED_ANALYSIS_DATA_DIR, exist_ok=True)
        path = _failed_analysis_results_path(tag)
        with open(path, 'w') as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        logger.error(f"Error saving failed analysis results for tag '{tag}': {e}")
        raise

def delete_failed_analysis_results(tag):
    try:
        path = _failed_analysis_results_path(tag)
        if os.path.exists(path):
            os.remove(path)
            logger.info(f"Deleted failed analysis cache: {path}")
    except Exception as e:
        logger.warning(f"Could not delete failed analysis cache for tag '{tag}': {e}")

# Load regression owners mapping from CSV
CSV_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "regression_owners.csv")
owner_mapping = {}

def load_owner_mapping():
    """Load test prefix to owner mapping from CSV"""
    global owner_mapping
    try:
        if os.path.exists(CSV_PATH):
            df = pd.read_csv(CSV_PATH, header=0)
            # CSV format: "Test Area,Regression Owner"
            for _, row in df.iterrows():
                test_prefix = str(row.iloc[0]).strip()
                owner = str(row.iloc[1]).strip() if len(row) > 1 else "Unknown"
                if test_prefix and owner and test_prefix != "Test Area":  # Skip header row
                    owner_mapping[test_prefix] = owner
            logger.info(f"Loaded {len(owner_mapping)} owner mappings from CSV")
        else:
            logger.warning(f"CSV file not found at {CSV_PATH}")
    except Exception as e:
        logger.error(f"Error loading owner mapping: {e}")

# Load owner mapping on startup
load_owner_mapping()

# ======================================================
# Session (reused)
# ======================================================
session = requests.Session()
session.headers.update(HEADERS)
session.verify = False

# ======================================================
# Helpers
# ======================================================
def should_process_task(status):
    """Process only non-successful runs"""
    return status not in ("Succeeded", "Pending")


def fetch_test_result(testcase_id):
    resp = session.get(
        f"{PHX_BASE}/agave_test_results/{testcase_id}",
        timeout=30
    )
    resp.raise_for_status()
    return resp.json().get("data", {})


def process_failed_tests(task_id, agave_task):
    failed_tests = []

    agave_results = agave_task.get("AgaveTestResults", [])
    if not agave_results:
        return failed_tests

    failed_ids = [
        tr["$oid"]
        for tr in agave_results
        if tr.get("status") in ("Failed", "Warning")
    ][:MAX_FAILED_TESTS]

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [
            executor.submit(fetch_test_result, tid)
            for tid in failed_ids
        ]

        for future in as_completed(futures):
            try:
                failure = future.result()
                failed_tests.append({
                    "testcase_id": failure.get("_id", {}).get("$oid"),
                    "test_name": failure.get("test", {}).get("name"),
                    "status": failure.get("status"),
                    "jira_tickets": failure.get("jira_tickets", []),
                    "exception_summary": failure.get("exception_summary"),
                    "log_url": failure.get("test_log_url")
                })
            except Exception as e:
                logger.error(f"[ERROR] Failed testcase fetch: {e}")

    return failed_tests


# ======================================================
# API-1: Fetch Regression Tasks
# ======================================================
def fetch_regression_tasks(tag=None, task_ids=None):
    """
    Fetch regression tasks either by tag or by task IDs
    
    Args:
        tag: Tag name to filter tasks
        task_ids: List of task IDs to fetch
    
    Returns:
        List of task data
    """
    if task_ids:
        # Fetch tasks by task IDs
        raw_query = {
            "_id": {
                "$in": [{"$oid": tid} for tid in task_ids]
            }
        }
    elif tag:
        # Fetch tasks by tag (original behavior)
        raw_query = {
            "$or": [
                {"created_by": {
                    "$in": ["shilpa.sattigeri", "sudharshan.musali"]
                    }
                },
                {"user_groups": {"$in": ["cdp_reg_jarvis"]}}
            ],
            "tester_tags": {"$in": [tag]},
            "system_under_test.component": "main"
        }
    else:
        raise ValueError("Either tag or task_ids must be provided")

    params = {
        "limit": 2000,
        "start": 0,
        "sort": "-_id",
        "only": (
            "label,branch,status,created_by,test_result_count,"
            "created_at,end_time"
        ),
        "raw_query": json.dumps(raw_query)
    }

    try:
        resp = session.get(
            f"{JITA_BASE}/tasks",
            params=params,
            timeout=30
        )
        resp.raise_for_status()
        return resp.json().get("data", [])
    except requests.exceptions.ConnectionError as e:
        logger.error(f"Connection error fetching regression tasks: {e}")
        raise ConnectionError(f"Failed to connect to JITA API. Please check your network connection and ensure 'jita.eng.nutanix.com' is accessible.")
    except requests.exceptions.Timeout as e:
        logger.error(f"Timeout error fetching regression tasks: {e}")
        raise TimeoutError(f"Request to JITA API timed out. Please try again.")
    except requests.exceptions.RequestException as e:
        logger.error(f"Request error fetching regression tasks: {e}")
        raise Exception(f"Error fetching regression tasks: {str(e)}")


# ======================================================
# API-2: Fetch Agave Task
# ======================================================
def fetch_agave_task(task_id):
    resp = session.get(
        f"{JITA_BASE}/agave_tasks/{task_id}",
        timeout=30
    )
    resp.raise_for_status()
    return resp.json().get("data", {})

# ======================================================
# Flask Endpoint
# ======================================================
def fetch_test_results_batch_with_pagination(task_ids, limit=2000, timeout=120):
    """
    Fetch test results for multiple tasks in batch.
    Always uses merge=True to get merged results across all tasks.
    Fetches all results in a single request with limit=2000.
    
    Args:
        task_ids: List of task IDs to fetch results for
        limit: Number of results to fetch (default: 2000)
        timeout: Request timeout in seconds (default: 120)
    
    Returns:
        List of merged test results
    
    Raises:
        requests.exceptions.Timeout: If request times out
        requests.exceptions.RequestException: For other request errors
    """
    if not task_ids:
        return []
    
    # Increase timeout for large task sets (more than 50 tasks)
    if len(task_ids) > 50:
        timeout = max(timeout, 180)  # At least 3 minutes for large sets
    
    logger.info(f"Fetching test results for {len(task_ids)} tasks (timeout: {timeout}s, limit: {limit})")
    
    # Correct payload structure for merged test results
    # Verified format: raw_query at top level with agave_task_id query, merge at top level
    payload = {
        "raw_query": {
            "agave_task_id": {
                "$in": [{"$oid": tid} for tid in task_ids]
            }
        },
        "only": (
            "_id,test,status,agave_task_id,jira_tickets,triaged_by,exception_summary,"
            "test_log_url,comments"
        ),
        "start": 0,
        "limit": limit,
        "sort": "agave_task_id,status",
        "merge": True  # Must be at top level to get merged results
    }
    
    # Log payload for verification
    #logger.info(f"[agave_test_results API] Payload: {json.dumps(payload, indent=2)}")
    #logger.info(f"[agave_test_results API] merge parameter: {payload.get('merge')}")
    logger.info(f"[agave_test_results API] Number of task_ids in query: {len(task_ids)}")
    
    try:
        resp = session.post(
            f"{JITA_BASE}/reports/agave_test_results",
            json=payload,
            timeout=timeout
        )
        
        resp.raise_for_status()
        response_data = resp.json()
        results = response_data.get("data", [])
        total = response_data.get("total", 0)
        
        # Log response info
        logger.info(f"[agave_test_results API] Response - Total: {total}, Returned: {len(results)}, Merge enabled: {payload.get('merge')}")
        if results:
            # Log sample test names to verify merge (first 3 unique test names)
            sample_tests = []
            seen_sample = set()
            for r in results:
                test_name = r.get("test", {}).get("name", "")
                if test_name and test_name not in seen_sample and len(sample_tests) < 3:
                    sample_tests.append(test_name)
                    seen_sample.add(test_name)
            logger.info(f"[agave_test_results API] Sample test names (first 3 unique): {sample_tests}")
        
        logger.info(f"Fetched {len(results)} merged test results from {len(task_ids)} tasks")
        return results
        
    except requests.exceptions.Timeout as e:
        logger.error(f"Timeout fetching test results: {e}")
        raise requests.exceptions.Timeout(
            f"Request timed out after {timeout}s while fetching test results. "
            f"This may be due to a large number of tasks ({len(task_ids)}). "
            f"Try reducing the number of tasks or increasing the timeout."
        ) from e
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching test results: {e}")
        raise

# ======================================================
# Auth Routes (no @jwt_required)
# ======================================================
@app.route("/mcp/regression/auth/login", methods=["POST"])
def auth_login():
    data = request.get_json(silent=True) or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"error": "Username and password are required"}), 400

    ldap_auth = LDAPAuth()
    user_info = ldap_auth.authenticate(username, password)
    if user_info is None:
        return jsonify({"error": "Invalid username or password"}), 401

    # Cache credentials for downstream Jita calls (trigger, triage, etc.)
    _store_user_credentials(user_info["username"], password)

    token = create_jwt(
        user_info["username"],
        user_info.get("displayName", ""),
        user_info.get("email", ""),
    )
    return jsonify({"token": token, "user": user_info})


@app.route("/mcp/regression/auth/me", methods=["GET"])
@jwt_required
def auth_me():
    return jsonify({"user": g.current_user})


@app.route("/mcp/regression/auth/logout", methods=["POST"])
def auth_logout():
    return jsonify({"success": True})


# ======================================================
# Protected Routes
# ======================================================
@app.route("/mcp/regression/home", methods=["GET"])
@jwt_required
def regression_home():
    start = time.time()

    tag = request.args.get("tag")
    task_ids_param = request.args.get("task_ids")  # Comma-separated task IDs
    
    # Parse task_ids if provided
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]
    
    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400

    if tag:
        logger.info(f"[START] Regression Home | tag={tag}")
    else:
        logger.info(f"[START] Regression Home | task_ids={len(task_ids)} tasks")
        logger.info(f"[DEBUG] Requested task IDs: {task_ids[:5]}..." if len(task_ids) > 5 else f"[DEBUG] Requested task IDs: {task_ids}")

    # Store original requested task IDs for comparison
    requested_task_ids = task_ids.copy() if task_ids else None

    try:
        tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
    except TimeoutError as e:
        logger.error(f"Regression home: JITA task list timeout: {e}")
        return jsonify({
            "error": str(e),
            "type": "jita_timeout",
            "tag": tag,
            "generated_at": datetime.utcnow().isoformat(),
            "total_runs": 0,
            "runs": [],
            "branch_start_dates": {},
            "oldest_start_date": None,
        }), 504
    except ConnectionError as e:
        logger.error(f"Regression home: JITA connection error: {e}")
        return jsonify({
            "error": str(e),
            "type": "jita_connection_error",
            "tag": tag,
            "generated_at": datetime.utcnow().isoformat(),
            "total_runs": 0,
            "runs": [],
            "branch_start_dates": {},
            "oldest_start_date": None,
        }), 503
    
    # Log which task IDs were found vs requested
    if requested_task_ids and not tag:
        found_task_ids = [task["_id"]["$oid"] for task in tasks]
        missing_task_ids = set(requested_task_ids) - set(found_task_ids)
        logger.info(f"[DEBUG] Found {len(found_task_ids)}/{len(requested_task_ids)} tasks")
        if missing_task_ids:
            logger.warning(f"[DEBUG] Missing task IDs ({len(missing_task_ids)}): {list(missing_task_ids)[:5]}..." if len(missing_task_ids) > 5 else f"[DEBUG] Missing task IDs: {list(missing_task_ids)}")
        
        # Log branch distribution from raw JITA data
        if tasks:
            branch_distribution = {}
            for task in tasks:
                raw_branch = task.get("branch", "None")
                branch_distribution[raw_branch] = branch_distribution.get(raw_branch, 0) + 1
            logger.info(f"[DEBUG] Raw branch distribution from JITA: {branch_distribution}")
    
    # Collect all task IDs from found tasks
    task_ids = [task["_id"]["$oid"] for task in tasks]
    
    # Fetch test results using agave_test_results API for accurate counts
    logger.info(f"Fetching test results for {len(task_ids)} tasks using agave_test_results API")
    test_results = []
    if task_ids:
        try:
            test_results = fetch_test_results_batch_with_pagination(task_ids)
            logger.info(f"Fetched {len(test_results)} test results")
        except Exception as e:
            logger.warning(f"Failed to fetch test results from agave_test_results API: {e}. Falling back to test_result_count.")
            test_results = []
    
    # Track whether we successfully fetched merged results
    merged_results_available = len(test_results) > 0
    
    # Group test results by task_id
    test_results_by_task = defaultdict(list)
    for test_result in test_results:
        agave_task_id = test_result.get("agave_task_id")
        if agave_task_id:
            # Handle both string and $oid format
            if isinstance(agave_task_id, dict) and "$oid" in agave_task_id:
                task_id = agave_task_id["$oid"]
            else:
                task_id = str(agave_task_id)
            test_results_by_task[task_id].append(test_result)
    
    runs = []
    
    # Track created_at times for finding oldest start date
    created_at_times = []

    for task in tasks:
        task_id = task["_id"]["$oid"]
        status = task.get("status")
        
        # Collect created_at time for oldest date calculation
        created_at = task.get("created_at")
        if created_at:
            created_at_times.append(created_at)
        
        # Count test statuses from agave_test_results if available
        test_counts = {
            "total": 0,
            "Succeeded": 0,
            "Failed": 0,
            "Pending": 0,
            "Warning": 0,
            "Running": 0,
            "Skipped": 0,
            "Killed": 0
        }
        
        if task_id in test_results_by_task:
            # Count statuses from actual test results
            for test_result in test_results_by_task[task_id]:
                test_status = test_result.get("status", "")
                test_counts["total"] += 1
                
                # Normalize status names (handle case-insensitive matching)
                status_lower = test_status.lower() if test_status else ""
                
                if status_lower == "succeeded" or status_lower == "success":
                    test_counts["Succeeded"] += 1
                elif status_lower == "failed" or status_lower == "failure":
                    test_counts["Failed"] += 1
                elif status_lower == "pending" or status_lower == "waiting":
                    test_counts["Pending"] += 1
                elif status_lower == "warning" or status_lower == "warn":
                    test_counts["Warning"] += 1
                elif status_lower == "running" or status_lower == "executing" or status_lower == "in_progress":
                    test_counts["Running"] += 1
                elif status_lower == "skipped" or status_lower == "skip":
                    test_counts["Skipped"] += 1
                elif status_lower == "killed" or status_lower == "terminated" or status_lower == "cancelled":
                    test_counts["Killed"] += 1
                else:
                    # For unknown statuses, try to infer from common patterns
                    # But don't default to pending - log it for debugging
                    logger.debug(f"Unknown test status: {test_status} for task {task_id}")
                    # Map unknown statuses based on common patterns
                    if any(x in status_lower for x in ["pending", "waiting", "queued"]):
                        test_counts["Pending"] += 1
                    elif any(x in status_lower for x in ["running", "executing", "in_progress", "active"]):
                        test_counts["Running"] += 1
                    elif any(x in status_lower for x in ["skipped", "skip"]):
                        test_counts["Skipped"] += 1
                    elif any(x in status_lower for x in ["killed", "terminated", "cancelled", "aborted"]):
                        test_counts["Killed"] += 1
                    else:
                        # Default to pending only if truly unknown
                        test_counts["Pending"] += 1
        elif not merged_results_available:
            # Fallback to test_result_count ONLY when the agave_test_results API
            # call failed entirely. When merge=True results were fetched successfully,
            # tasks not in the map have their tests already counted via other tasks'
            # merged results — using the stale test_result_count would double-count.
            tc = task.get("test_result_count", {})
            test_counts = {
                "total": tc.get("Total", 0),
                "Succeeded": tc.get("Succeeded", 0),
                "Failed": tc.get("Failed", 0),
                "Pending": tc.get("Pending", 0),
                "Warning": tc.get("Warning", 0),
                "Running": tc.get("Running", 0),
                "Skipped": tc.get("Skipped", 0),
                "Killed": tc.get("Killed", 0)
            }

        # Get branch and normalize it
        original_branch = task.get("branch")
        label = task.get("label", "")
        branch = None
        
        if original_branch:
            # Branch exists, normalize it
            branch = original_branch.strip()
            # Normalize master branch variations (case-insensitive)
            branch_lower = branch.lower()
            if branch_lower in ["master", "main"]:
                branch = "master"
            elif branch_lower in ["ganges-7.5-stable", "ganges_7.5_stable"]:
                branch = "ganges-7.5-stable"
        else:
            # Branch is missing (None, empty string, etc.)
            # Try to infer from label
            label_lower = label.lower()
            
            # Check for master branch indicators in label
            master_keywords = ["master", "main", "cdp_master", "master_full", "master_reg"]
            if any(keyword in label_lower for keyword in master_keywords):
                branch = "master"
                logger.info(f"[BRANCH_INFER] Task {task_id}: No branch field, inferred 'master' from label: '{label}'")
            # Check for other known branch patterns
            elif "ganges" in label_lower or "7.5" in label_lower:
                branch = "ganges-7.5-stable"
                logger.info(f"[BRANCH_INFER] Task {task_id}: No branch field, inferred 'ganges-7.5-stable' from label: '{label}'")
            else:
                branch = "unknown"
                logger.warning(f"[BRANCH_MISSING] Task {task_id}: No branch field and cannot infer from label: '{label}'")
        
        # Log master branch detection for debugging
        if branch == "master":
            logger.info(f"[MASTER_BRANCH] Task {task_id}: branch='{original_branch}' -> normalized to 'master', label='{label}'")
        
        # Get created_at time for this task
        created_at = task.get("created_at")
        
        run = {
            "task_id": task_id,
            "label": label,
            "branch": branch,
            "status": status,
            "created_by": task.get("created_by"),
            "created_at": created_at,  # Include created_at in run object
            "test_counts": test_counts,
            "failed_tests": []
        }

        if not should_process_task(status):
            agave_task = fetch_agave_task(task_id)
            run["failed_tests"] = process_failed_tests(task_id, agave_task)

        runs.append(run)

    # Calculate oldest start date per branch from all runs
    branch_start_dates = {}  # {branch: oldest_datetime}
    
    for run in runs:
        branch = run.get("branch")
        created_at = run.get("created_at")
        
        if branch and created_at:
            try:
                # Parse created_at time
                if isinstance(created_at, str):
                    # Handle ISO format
                    if 'T' in created_at:
                        dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                    else:
                        # Try other common formats
                        try:
                            dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            dt = datetime.strptime(created_at, "%Y-%m-%d")
                elif isinstance(created_at, dict) and "$date" in created_at:
                    # MongoDB date format
                    dt = datetime.fromtimestamp(created_at["$date"] / 1000)
                else:
                    continue
                
                # Track oldest date per branch
                if branch not in branch_start_dates or dt < branch_start_dates[branch]:
                    branch_start_dates[branch] = dt
            except (ValueError, TypeError) as e:
                logger.debug(f"Could not parse created_at for branch {branch}: {created_at}, error: {e}")
                continue
    
    # Format branch start dates as readable strings
    branch_start_dates_formatted = {}
    for branch, dt in branch_start_dates.items():
        branch_start_dates_formatted[branch] = dt.strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"[DEBUG] Oldest start date for branch '{branch}': {branch_start_dates_formatted[branch]}")
    
    # Calculate overall oldest start date
    oldest_start_date = None
    if branch_start_dates:
        oldest_start_date = min(branch_start_dates.values())
        oldest_start_date_str = oldest_start_date.strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"[DEBUG] Overall oldest start date: {oldest_start_date_str}")

    # Log final branch distribution after normalization
    if runs:
        final_branch_dist = {}
        for run in runs:
            br = run.get("branch", "None")
            final_branch_dist[br] = final_branch_dist.get(br, 0) + 1
        logger.info(f"[DEBUG] Final branch distribution after normalization: {final_branch_dist}")
        if "master" in final_branch_dist:
            logger.info(f"[DEBUG] Master branch tasks found: {final_branch_dist['master']} tasks")
        else:
            logger.warning(f"[DEBUG] No 'master' branch found in final distribution! Available branches: {list(final_branch_dist.keys())}")

    logger.info(
        f"[END] runs={len(runs)} | time={time.time() - start:.2f}s"
    )
    
    # Include metadata about missing task IDs if using task_ids mode
    response_data = {
        "tag": tag,
        "generated_at": datetime.utcnow().isoformat(),
        "total_runs": len(runs),
        "runs": runs,
        "branch_start_dates": branch_start_dates_formatted,  # Oldest start date per branch
        "oldest_start_date": oldest_start_date.strftime("%Y-%m-%d %H:%M:%S") if oldest_start_date else None
    }
    
    if requested_task_ids and not tag:
        found_task_ids = [task["_id"]["$oid"] for task in tasks]
        missing_task_ids = list(set(requested_task_ids) - set(found_task_ids))
        if missing_task_ids:
            response_data["missing_task_ids"] = missing_task_ids
            response_data["requested_count"] = len(requested_task_ids)
            response_data["found_count"] = len(found_task_ids)
            logger.warning(f"Some task IDs were not found: {len(missing_task_ids)} missing out of {len(requested_task_ids)} requested")

    return jsonify(response_data)


# ---------------------------------------------------
# Manual Tasks Endpoints
# ---------------------------------------------------
@app.route("/mcp/regression/manual-tasks", methods=["GET"])
@jwt_required
def get_manual_tasks():
    tag = request.args.get("tag")
    branch = request.args.get("branch")

    if not tag:
        return jsonify({"error": "tag is required"}), 400

    if not branch:
        return jsonify({"error": "branch is required"}), 400

    # Get manual tasks for the given tag and branch
    tag_store = manual_tasks_store.get(tag, {})
    task_ids = tag_store.get(branch, [])

    return jsonify({
        "tag": tag,
        "branch": branch,
        "manual_tasks": task_ids
    })


@app.route("/mcp/regression/manual-tasks", methods=["POST"])
@jwt_required
def add_manual_tasks():
    data = request.get_json()
    tag = data.get("tag")
    branch = data.get("branch")
    task_ids = data.get("task_ids", [])

    if not tag:
        return jsonify({"error": "tag is required"}), 400

    if not branch:
        return jsonify({"error": "branch is required"}), 400

    if not task_ids:
        return jsonify({"error": "task_ids is required"}), 400

    # Initialize storage if needed
    if tag not in manual_tasks_store:
        manual_tasks_store[tag] = {}

    if branch not in manual_tasks_store[tag]:
        manual_tasks_store[tag][branch] = []

    # Add new task IDs (avoid duplicates)
    existing = set(manual_tasks_store[tag][branch])
    for task_id in task_ids:
        if task_id not in existing:
            manual_tasks_store[tag][branch].append(task_id)

    return jsonify({
        "tag": tag,
        "branch": branch,
        "manual_tasks": manual_tasks_store[tag][branch]
    })


@app.route("/mcp/regression/manual-tasks", methods=["DELETE"])
@jwt_required
def remove_manual_task():
    tag = request.args.get("tag")
    branch = request.args.get("branch")
    task_id = request.args.get("task_id")

    if not tag:
        return jsonify({"error": "tag is required"}), 400

    if not branch:
        return jsonify({"error": "branch is required"}), 400

    if not task_id:
        return jsonify({"error": "task_id is required"}), 400

    # Remove task ID if it exists
    if tag in manual_tasks_store and branch in manual_tasks_store[tag]:
        if task_id in manual_tasks_store[tag][branch]:
            manual_tasks_store[tag][branch].remove(task_id)

    return jsonify({
        "tag": tag,
        "branch": branch,
        "manual_tasks": manual_tasks_store.get(tag, {}).get(branch, [])
    })


# ---------------------------------------------------
# Configuration Endpoints
# ---------------------------------------------------
@app.route("/mcp/regression/config", methods=["GET"])
@jwt_required
def get_regression_config():
    """Get regression dashboard configuration from JSON file"""
    try:
        config = load_regression_config()
        return jsonify(config)
    except Exception as e:
        logger.error(f"Error getting regression config: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/config", methods=["POST"])
@jwt_required
def save_regression_config_endpoint():
    """Save regression dashboard configuration to JSON file"""
    try:
        data = request.get_json()
        
        # Validate required fields
        input_mode = data.get("input_mode")
        if input_mode not in ["tag", "task_ids"]:
            return jsonify({"error": "input_mode must be 'tag' or 'task_ids'"}), 400
        
        default_tag = data.get("default_tag")
        added_tags = data.get("added_tags")
        if added_tags is None:
            added_tags = []
        if not isinstance(added_tags, list):
            added_tags = []
        
        config = {
            "input_mode": input_mode,
            "default_tag": default_tag if default_tag else None,
            "added_tags": [str(t).strip() for t in added_tags if t and str(t).strip()],
            "tag": (default_tag or "").strip() if input_mode == "tag" else "",
            "task_ids": data.get("task_ids", []) if input_mode == "task_ids" else []
        }
        
        # Validate based on input mode
        if input_mode == "tag":
            if config["default_tag"] and config["default_tag"] not in config["added_tags"]:
                return jsonify({"error": "default_tag must be in added_tags or null"}), 400
            config["task_ids"] = []
        elif input_mode == "task_ids":
            if not config["task_ids"] or len(config["task_ids"]) == 0:
                return jsonify({"error": "task_ids is required when input_mode is 'task_ids'"}), 400
            config["tag"] = ""
            config["default_tag"] = None
        
        save_regression_config(config)
        if input_mode == "task_ids":
            invalidate_triage_accuracy_cache(None)
        logger.info(f"Saved regression config: input_mode={input_mode}, default_tag={config.get('default_tag')}, added_tags={len(config.get('added_tags', []))}")
        return jsonify(config)
    except Exception as e:
        logger.error(f"Error saving regression config: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/config/tags", methods=["POST"])
@jwt_required
def add_config_tag():
    """Add a tag to added_tags. Validation is lenient - tag is added even if JITA returns empty."""
    try:
        data = request.get_json() or {}
        tag = (data.get("tag") or "").strip()
        if not tag:
            return jsonify({"error": "tag is required"}), 400
        
        config = load_regression_config()
        added = list(config.get("added_tags", []))
        if tag in added:
            return jsonify({"added_tags": added, "message": "Tag already in list"}), 200
        
        # Optional validation: if JITA returns tasks, tag is validated; otherwise still add (lenient)
        validated = False
        try:
            tasks = fetch_regression_tasks(tag=tag)
            validated = bool(tasks)
        except Exception as e:
            logger.warning(f"Tag validation skipped for '{tag}': {e}")
        
        added.append(tag)
        config["added_tags"] = added
        save_regression_config(config)
        logger.info(f"Added tag to config: {tag} (validated={validated})")
        return jsonify({"added_tags": added, "tag": tag, "validated": validated})
    except Exception as e:
        logger.error(f"Error adding tag: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/config/tags", methods=["DELETE"])
@jwt_required
def delete_config_tag():
    """Remove tag from added_tags and delete per-tag triage accuracy JSON."""
    try:
        tag = request.args.get("tag", "").strip()
        if not tag:
            return jsonify({"error": "tag query param is required"}), 400
        
        config = load_regression_config()
        added = list(config.get("added_tags", []))
        if tag not in added:
            return jsonify({"error": f"Tag '{tag}' not in added_tags"}), 404
        
        added = [t for t in added if t != tag]
        config["added_tags"] = added
        if config.get("default_tag") == tag:
            config["default_tag"] = None
            config["tag"] = ""
        save_regression_config(config)
        
        invalidate_triage_accuracy_cache(tag)
        logger.info(f"Deleted tag from config and triage JSON: {tag}")
        return jsonify({"added_tags": added})
    except Exception as e:
        logger.error(f"Error deleting tag: {e}")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# Fetch Branches from Tag Endpoint
# ---------------------------------------------------
@app.route("/mcp/regression/branches", methods=["GET"])
@jwt_required
def get_branches_from_tag():
    tag = request.args.get("tag")
    
    if not tag:
        return jsonify({"error": "tag is required"}), 400
    
    try:
        # Fetch tasks using the same API as regression_home
        tasks = fetch_regression_tasks(tag)
        
        # Extract unique branches
        branches = set()
        for task in tasks:
            branch = task.get("branch")
            if branch:
                branches.add(branch)
        
        return jsonify({
            "tag": tag,
            "branches": sorted(list(branches))
        })
    except Exception as e:
        logger.error(f"Error fetching branches: {e}")
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# Helper: Resolve Owner from Test Name
# ---------------------------------------------------
def resolve_owner(test_name):
    """Resolve owner from test name using prefix mapping"""
    for prefix, owner in owner_mapping.items():
        if test_name.startswith(prefix):
            return owner
    return "Unknown"


# ---------------------------------------------------
# Helper: Fetch Test Results (POST API – batch)
# NOTE: This function is deprecated. Use fetch_test_results_batch_with_pagination() instead.
# Kept for backward compatibility but redirects to pagination version.
# ---------------------------------------------------
def fetch_test_results_batch(task_ids):
    """Fetch test results for multiple tasks in batch (deprecated - uses pagination version)"""
    return fetch_test_results_batch_with_pagination(task_ids)


# ---------------------------------------------------
# Helper function to calculate QI impact for bulk issues
# ---------------------------------------------------
def calculate_bulk_issues_qi_impact(bulk_issues, test_data, tag=None):
    """
    Calculate QI impact for bulk issues using TCMS API.
    
    Args:
        bulk_issues: Dictionary mapping ticket -> list of testcase names
        test_data: List of test result data from API
        tag: Optional tag to extract milestone from
    
    Returns:
        Dictionary with:
        - bulk_issues_with_qi: Dict mapping ticket -> QI impact data
        - test_qi_map: Dict mapping testcase_name -> QI value
    """
    if not bulk_issues:
        return {"bulk_issues_with_qi": {}, "test_qi_map": {}}
    
    # Extract milestone from tag or use default
    milestone = "7.5.1"  # Default milestone
    if tag:
        # Try to extract milestone from tag (e.g., "cdp_master_full_reg" -> "master", "7.5.1" -> "7.5.1")
        milestone_match = re.search(r'(\d+\.\d+(?:\.\d+)?)', tag)
        if milestone_match:
            milestone = milestone_match.group(1)
        elif "master" in tag.lower():
            milestone = "master"
    
    # Collect all unique testcases from bulk issues for TCMS API calls
    all_bulk_testcases = set()
    for test_names in bulk_issues.values():
        all_bulk_testcases.update(test_names)
    
    # Fetch QI values from TCMS API for all testcases in bulk issues
    logger.info(f"Fetching QI from TCMS API for {len(all_bulk_testcases)} testcases in bulk issues (milestone: {milestone})")
    test_qi_map = {}
    
    # Use ThreadPoolExecutor to fetch QI values in parallel
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_testcase = {
            executor.submit(fetch_qi_from_tcms, testcase_name, milestone): testcase_name
            for testcase_name in all_bulk_testcases
        }
        
        for future in as_completed(future_to_testcase):
            testcase_name = future_to_testcase[future]
            try:
                qi_value = future.result()
                if qi_value is not None:
                    test_qi_map[testcase_name] = qi_value
                else:
                    # Fallback: use status-based QI if TCMS API doesn't return data
                    # Find the test in test_data to get status
                    status_qi = 0
                    for test in test_data:
                        if test.get("test", {}).get("name") == testcase_name:
                            status = test.get("status", "")
                            if status == "Succeeded":
                                status_qi = 100
                            elif status == "Warning":
                                status_qi = 50
                            else:
                                status_qi = 0
                            break
                    test_qi_map[testcase_name] = status_qi
            except Exception as e:
                logger.warning(f"Error fetching QI for {testcase_name}: {e}")
                # Fallback to 0 if error
                test_qi_map[testcase_name] = 0
    
    logger.info(f"Fetched QI values for {len(test_qi_map)} testcases from TCMS API")
    
    # Calculate QI impact for each bulk issue using generate_qi_impact logic
    bulk_issues_with_qi = {}
    # Use total unique test cases from all test data (not just failed)
    all_unique_tests = set()
    for test in test_data:
        test_name = test.get("test", {}).get("name", "")
        if test_name:
            all_unique_tests.add(test_name)
    total_test_cases = len(all_unique_tests) if all_unique_tests else 1  # Avoid division by zero
    
    for ticket, test_names in bulk_issues.items():
        # Get QI values for all testcases affected by this ticket
        qi_values = []
        testcase_qi_details = []  # Store individual testcase QI details
        for test_name in test_names:
            qi_value = test_qi_map.get(test_name, 0)
            qi_values.append(qi_value)
            testcase_qi_details.append({
                "testcase": test_name,
                "qi": qi_value
            })
        
        if qi_values:
            # Calculate average QI (matching generate_qi_impact logic)
            average_qi = sum(qi_values) / len(qi_values)
            nr_test_cases = len(test_names)
            
            # Calculate QI impact: (average_qi - 100) * nr_test_cases
            qi_impact = (average_qi - 100) * nr_test_cases
            
            # Calculate overall QI impact: 100 * (qi_impact / (100 * total_test_cases))
            if total_test_cases > 0:
                overall_qi_impact = 100 * (qi_impact / (100 * total_test_cases))
            else:
                overall_qi_impact = 0
            
            bulk_issues_with_qi[ticket] = {
                "testcases": test_names,
                "testcase_count": nr_test_cases,
                "average_qi": round(average_qi, 2),
                "qi_impact": round(qi_impact, 2),
                "overall_qi_impact": round(overall_qi_impact, 2),
                "testcase_qi_details": testcase_qi_details  # Include individual testcase QI details
            }
        else:
            # Fallback if no QI data
            bulk_issues_with_qi[ticket] = {
                "testcases": test_names,
                "testcase_count": len(test_names),
                "average_qi": 0,
                "qi_impact": 0,
                "overall_qi_impact": 0,
                "testcase_qi_details": []
            }
    
    return {
        "bulk_issues_with_qi": bulk_issues_with_qi,
        "test_qi_map": test_qi_map
    }


# ---------------------------------------------------
# Helper function to fetch QI from TCMS API
# ---------------------------------------------------
def fetch_qi_from_tcms(testcase_name, milestone="7.5.1"):
    """
    Fetch QI (operation_success_percentage) from TCMS API for a given testcase.
    
    Args:
        testcase_name: Name of the testcase (e.g., "cdp.counter.fio.test_fio_counters.CountersFIOTest.test_fio_end_to_end")
        milestone: Target milestone (default: "7.5.1")
    
    Returns:
        float: QI value (operation_success_percentage) or None if not found/error
    """
    try:
        # Construct payload based on the provided example
        # Use more flexible matching - try exact name match first, then regex
        payload = [{
            "$match": {
                "$and": [
                    {"target_milestone": milestone},
                    {"last_result": {"$elemMatch": {"pass_name": "overall"}}},
                    {"deleted": False},
                    {"test_case.metadata.tags": {"$nin": ["SYSTEST_LONGEVITY", "LIMITED_RUNS"]}},
                    {
                        "$or": [
                            {"test_case.name": testcase_name},  # Exact match first
                            {"test_case.name": {"$regex": testcase_name, "$options": "i"}}  # Case-insensitive regex
                        ]
                    },
                    {"test_case.deprecated": False}
                ]
            }
        }, {"$sort": {"name": 1}}, {"$skip": 0}, {"$limit": 50}]
        
        # Make POST request to TCMS API
        response = requests.post(
            f"{TCMS_BASE}/milestone_all_test_cases/aggregate",
            json=payload,
            headers={"Content-Type": "application/json"},
            verify=False,
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get("success") and data.get("data") and len(data["data"]) > 0:
                # Extract operation_success_percentage from published section
                testcase_data = data["data"][0]
                last_result = testcase_data.get("last_result", [])
                if last_result and len(last_result) > 0:
                    published = last_result[0].get("published", {})
                    if published:
                        operation_success_percentage = published.get("operation_success_percentage")
                        if operation_success_percentage is not None:
                            return float(operation_success_percentage)
            
            logger.warning(f"TCMS API: No QI data found for testcase: {testcase_name}")
            return None
        else:
            logger.warning(f"TCMS API error for {testcase_name}: HTTP {response.status_code}")
            return None
            
    except requests.exceptions.Timeout:
        logger.warning(f"TCMS API timeout for testcase: {testcase_name}")
        return None
    except Exception as e:
        logger.warning(f"Error fetching QI from TCMS for {testcase_name}: {e}")
        return None


# ---------------------------------------------------
# TCMS Tags Fetch Endpoint
# ---------------------------------------------------
def _tcms_aggregate_headers():
    h = {"Content-Type": "application/json", "Accept": "application/json"}
    token = (
        (os.getenv("TCMS_API_TOKEN") or os.getenv("TCMS_TOKEN") or os.getenv("TCMS_BEARER") or "")
        .strip()
    )
    if token:
        h["Authorization"] = f"Bearer {token}"
    return h


def _tcms_aggregate_post(payload, timeout=60):
    return requests.post(
        f"{TCMS_BASE}/milestone_all_test_cases/aggregate",
        json=payload,
        headers=_tcms_aggregate_headers(),
        verify=False,
        timeout=timeout,
    )


def _tcms_response_rows(data):
    """Normalize aggregate JSON body to a list of row dicts."""
    if not data or not isinstance(data, dict):
        return None
    rows = data.get("data")
    if rows is None:
        rows = data.get("result") or data.get("rows") or data.get("items")
    if not isinstance(rows, list):
        return None
    return rows


def _parse_grouped_tag_ids(rows):
    tags = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        tid = item.get("_id")
        if tid is None or isinstance(tid, (dict, list)):
            continue
        s = str(tid).strip()
        if s:
            tags.append(s)
    return tags


def _tags_from_projected_docs(rows, *paths):
    seen = set()
    out = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        for p in paths:
            t = item
            for part in p.split("."):
                if not isinstance(t, dict):
                    t = None
                    break
                t = t.get(part)
            if isinstance(t, list):
                for x in t:
                    if x and str(x).strip():
                        s = str(x).strip()
                        if s and s not in seen:
                            seen.add(s)
                            out.append(s)
            elif t and str(t).strip():
                s = str(t).strip()
                if s and s not in seen:
                    seen.add(s)
                    out.append(s)
    return sorted(out, key=str.lower)


def _fetch_tcms_distinct_tags_group(milestone):
    """$group distinct tag values (preferred, smallest payload)."""
    payload = [
        {
            "$match": {
                "$and": [
                    {"target_milestone": milestone},
                    {"deleted": False},
                    {"test_case.deprecated": False},
                    {"test_case.metadata.tags": {"$exists": True, "$ne": []}},
                ]
            }
        },
        {"$unwind": "$test_case.metadata.tags"},
        {"$group": {"_id": "$test_case.metadata.tags"}},
        {"$sort": {"_id": 1}},
        {"$limit": 2000},
    ]
    r = _tcms_aggregate_post(payload)
    if r.status_code != 200:
        return None, r
    j = r.json()
    rows = _tcms_response_rows(j)
    if not rows and j and j.get("data") is None and isinstance(j, dict) and "success" in j:
        logger.warning(f"TCMS tags group: unexpected body keys: {list(j.keys())} snippet={str(j)[:400]}")
    if not rows:
        return [], r
    return _parse_grouped_tag_ids(rows), r


def _fetch_tcms_tags_project_scan(milestone, limit=5000):
    """
    Fallback: return many documents and build distinct tags in Python.
    Handles different shapes / empty $group responses.
    """
    payload = [
        {
            "$match": {
                "target_milestone": milestone,
                "deleted": False,
            }
        },
        {
            "$project": {
                "m_tags": "$test_case.metadata.tags",
                "alt_tags": "$test_case.tags",
            }
        },
        {"$limit": int(limit)},
    ]
    r = _tcms_aggregate_post(payload, timeout=90)
    if r.status_code != 200:
        return None, r
    j = r.json()
    rows = _tcms_response_rows(j)
    if not rows:
        return [], r
    tags = _tags_from_projected_docs(rows, "m_tags", "alt_tags")
    return tags, r


@app.route("/mcp/regression/tcms/tags", methods=["GET"])
def fetch_tcms_tags():
    """
    Fetch available tags from TCMS API for a given milestone.
    Query params: milestone (default: env TCMS_MILESTONE or 7.5.1)
    On empty result, tries TCMS_MILESTONE_FALLBACKS (comma env) and project-scan fallback.
    Set TCMS_API_TOKEN if your TCMS read requires bearer auth.
    """
    try:
        default_ms = (os.getenv("TCMS_MILESTONE") or "7.5.1").strip()
        milestone = (request.args.get("milestone") or default_ms).strip()
        raw_fb = (os.getenv("TCMS_MILESTONE_FALLBACKS") or "7.5,7.5.1,7.3,7.3.1").strip()
        fallbacks = [m.strip() for m in raw_fb.split(",") if m.strip()]
        try_order = [milestone] + [m for m in fallbacks if m != milestone]

        last_resp = None
        used_ms = milestone
        for ms in try_order:
            used_ms = ms
            tags, last_resp = _fetch_tcms_distinct_tags_group(ms)
            if tags is None and last_resp is not None:
                try:
                    logger.error(
                        f"TCMS tags group failed milestone={ms} status={last_resp.status_code} "
                        f"body={last_resp.text[:500]}"
                    )
                except Exception:
                    pass
                continue
            if tags:
                logger.info(f"TCMS tags (group) milestone={ms} count={len(tags)}")
                return jsonify({"success": True, "tags": tags, "milestone": ms, "source": "group"})

        # project-scan for first milestone in try_order
        for ms in try_order:
            tags, last_resp = _fetch_tcms_tags_project_scan(ms)
            if tags is None and last_resp is not None and last_resp.status_code != 200:
                try:
                    logger.error(
                        f"TCMS tags scan failed milestone={ms} status={last_resp.status_code} "
                        f"body={last_resp.text[:500]}"
                    )
                except Exception:
                    pass
                continue
            if tags:
                logger.info(f"TCMS tags (scan) milestone={ms} count={len(tags)}")
                return jsonify(
                    {
                        "success": True,
                        "tags": tags,
                        "milestone": ms,
                        "source": "scan",
                    }
                )

        if last_resp is not None and last_resp.status_code != 200:
            return (
                jsonify(
                    {
                        "success": False,
                        "error": f"TCMS API HTTP {last_resp.status_code}",
                        "tags": [],
                    }
                ),
                502,
            )

        logger.warning(f"TCMS API returned no tags (tried: {', '.join(try_order[:6])})")
        return jsonify(
            {
                "success": True,
                "tags": [],
                "milestone": used_ms,
                "source": "empty",
                "hint": "Set TCMS_MILESTONE and TCMS_API_TOKEN; check network to tcms.eng.nutanix.com",
            }
        )

    except requests.exceptions.Timeout:
        logger.error("TCMS API timeout (tags)")
        return jsonify({"error": "TCMS API timeout", "success": False, "tags": []}), 504
    except Exception as e:
        logger.error(f"Error fetching TCMS tags: {e}", exc_info=True)
        return jsonify({"error": str(e), "success": False, "tags": []}), 500


@app.route("/mcp/regression/tcms/testcases", methods=["POST"])
def fetch_tcms_testcases_by_tags():
    """
    Fetch testcases from TCMS API filtered by tags.
    Body: {"tags": ["tag1", "tag2"], "milestone": "7.5.1"}
    Returns: list of testcase names
    """
    try:
        req_data = request.json or {}
        tags = req_data.get("tags", [])
        milestone = req_data.get("milestone", "7.5.1")
        
        if not tags or not isinstance(tags, list):
            return jsonify({"error": "tags array is required"}), 400
        
        # Build match query for tags (test must have ALL specified tags)
        payload = [{
            "$match": {
                "$and": [
                    {"target_milestone": milestone},
                    {"deleted": False},
                    {"test_case.deprecated": False},
                    {"test_case.metadata.tags": {"$all": tags}}
                ]
            }
        }, {
            "$project": {
                "name": "$test_case.name",
                "tags": "$test_case.metadata.tags",
                "description": "$test_case.description"
            }
        }, {
            "$sort": {"name": 1}
        }, {
            "$limit": 500
        }]
        
        response = requests.post(
            f"{TCMS_BASE}/milestone_all_test_cases/aggregate",
            json=payload,
            headers={"Content-Type": "application/json"},
            verify=False,
            timeout=30
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get("success") and data.get("data"):
                testcases = data["data"]
                logger.info(f"Fetched {len(testcases)} TCMS testcases for tags {tags}")
                return jsonify({
                    "success": True,
                    "testcases": testcases,
                    "count": len(testcases),
                    "milestone": milestone,
                    "tags": tags
                })
            else:
                return jsonify({
                    "success": True,
                    "testcases": [],
                    "count": 0,
                    "milestone": milestone,
                    "tags": tags
                })
        else:
            logger.error(f"TCMS API error: HTTP {response.status_code}")
            return jsonify({"error": f"TCMS API error: {response.status_code}"}), 502
            
    except requests.exceptions.Timeout:
        logger.error("TCMS API timeout")
        return jsonify({"error": "TCMS API timeout"}), 504
    except Exception as e:
        logger.error(f"Error fetching TCMS testcases: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# Triage Count Endpoint
# ---------------------------------------------------
@app.route("/mcp/regression/triage-count", methods=["GET"])
@jwt_required
def get_triage_count():
    start = time.time()
    tag = request.args.get("tag")
    task_ids_param = request.args.get("task_ids")  # Comma-separated task IDs
    
    # Parse task_ids if provided
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]
    
    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400
    
    if tag:
        logger.info(f"[START] Triage Count | tag={tag}")
    else:
        logger.info(f"[START] Triage Count | task_ids={len(task_ids)} tasks")
    
    try:
        # Reload owner mapping in case it was updated
        load_owner_mapping()

        try:
            tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
        except TimeoutError as e:
            logger.error(f"Triage count: JITA task list timeout: {e}")
            return jsonify({
                "error": str(e),
                "type": "jita_timeout",
                "tag": tag,
                "generated_at": datetime.utcnow().isoformat(),
                "triage_summary": {},
                "owner_ticket_map": {},
                "bulk_issues": {},
                "bulk_issues_with_qi": {},
                "pending_tests": 0,
                "total_tests_processed": 0,
            }), 504
        except ConnectionError as e:
            logger.error(f"Triage count: JITA connection error: {e}")
            return jsonify({
                "error": str(e),
                "type": "jita_connection_error",
                "tag": tag,
                "generated_at": datetime.utcnow().isoformat(),
                "triage_summary": {},
                "owner_ticket_map": {},
                "bulk_issues": {},
                "bulk_issues_with_qi": {},
                "pending_tests": 0,
                "total_tests_processed": 0,
            }), 503

        logger.info(f"Tasks count: {len(tasks)}")
        
        if not tasks:
            return jsonify({
                "tag": tag,
                "generated_at": datetime.utcnow().isoformat(),
                "triage_summary": {},
                "owner_ticket_map": {},
                "bulk_issues": {},
                "pending_tests": 0,
                "message": "No tasks found for this tag"
            })
        
        # Collect all task IDs
        task_ids = []
        for task in tasks:
            task_id = task["_id"]["$oid"]
            task_ids.append(task_id)
        
        # Validate tasks exist and fetch test results in batch
        logger.info(f"Fetching test results for {len(task_ids)} tasks")
        try:
            # Use longer timeout for triage count as it may process many tasks
            test_data = fetch_test_results_batch_with_pagination(task_ids, timeout=180)
            logger.info(f"Fetched {len(test_data)} merged test results")
        except requests.exceptions.Timeout as e:
            logger.error(f"Timeout error in triage count: {e}")
            return jsonify({
                "error": f"Request timed out while fetching test results. This may be due to a large number of tasks ({len(task_ids)}). Please try with fewer tasks or contact support.",
                "type": "timeout_error",
                "task_count": len(task_ids)
            }), 504
        except requests.exceptions.RequestException as e:
            logger.error(f"Request error in triage count: {e}")
            return jsonify({
                "error": f"Failed to fetch test results: {str(e)}",
                "type": "request_error"
            }), 500
        
        # Initialize data structures
        summary = defaultdict(lambda: {
            "Total Failed": 0,
            "Triaged": 0,
            "UnTriaged": 0
        })
        ticket_case_map = defaultdict(set)  # ticket → set of unique testcases
        owner_ticket_map = defaultdict(lambda: defaultdict(int))  # owner → ticket → count
        inprogress_tests = 0  # Count pending/running from merged test results
        
        # Deduplicate by test name to ensure each unique test is counted only once
        # This handles cases where merge might not fully deduplicate or same test appears with different statuses
        seen_tests = set()  # Track unique test names we've processed
        seen_pending_running = set()  # Track unique test names for pending/running
        
        # Process test data - matching the working script logic
        for test in test_data:
            status = test.get("status")
            test_name = test.get("test", {}).get("name", "")
            
            # Skip if test name is empty
            if not test_name:
                continue
            
            # Skip succeeded tests
            if status == "Succeeded":
                continue
            
            # Count pending/running tests as in-progress (from merged results)
            # Deduplicate by test name to avoid double counting
            if status == "Pending":
                if test_name not in seen_pending_running:
                    inprogress_tests += 1
                    seen_pending_running.add(test_name)
                continue
            if status == "Running":
                if test_name not in seen_pending_running:
                    inprogress_tests += 1
                    seen_pending_running.add(test_name)
                continue
            
            # Process failed tests only (excluding Succeeded, Pending, Running)
            # Deduplicate by test name to ensure each unique test is counted only once
            if test_name in seen_tests:
                continue  # Skip if we've already processed this test
            
            seen_tests.add(test_name)
            tickets = test.get("jira_tickets", [])
            owner = resolve_owner(test_name)
            
            # Summary stats
            summary[owner]["Total Failed"] += 1
            if tickets:
                summary[owner]["Triaged"] += 1
            else:
                summary[owner]["UnTriaged"] += 1
            
            # Update ticket-to-test map (using set to avoid duplicates)
            for ticket in tickets:
                ticket_case_map[ticket].add(test_name)
                owner_ticket_map[owner][ticket] += 1
        
        # Identify bulk issues (tickets with >5 testcases)
        # Convert sets to lists for JSON serialization
        bulk_issues = {ticket: list(tests) for ticket, tests in ticket_case_map.items() if len(tests) > 5}
        
        # Calculate QI impact for bulk issues only if requested (to speed up triage count)
        include_bulk_qi = request.args.get("include_bulk_qi", "false").lower() == "true"
        bulk_issues_with_qi = {}
        if include_bulk_qi and bulk_issues:
            logger.info("Calculating QI impact for bulk issues (this may take longer)...")
            qi_calculation_result = calculate_bulk_issues_qi_impact(bulk_issues, test_data, tag)
            bulk_issues_with_qi = qi_calculation_result["bulk_issues_with_qi"]
        else:
            logger.info("Skipping bulk issues QI calculation for faster triage count response")
        
        # Update bulk issues count - use total count of unique bulk tickets
        # (not per-owner, since the same ticket can be tagged on tests from multiple owners)
        total_bulk_issues_count = len(bulk_issues)
        
        # Convert defaultdict to regular dict for JSON serialization
        triage_summary = {k: dict(v) for k, v in summary.items()}
        owner_ticket_dict = {k: dict(v) for k, v in owner_ticket_map.items()}
        bulk_issues_dict = {k: v for k, v in bulk_issues.items()}
        
        logger.info(f"[END] Triage Count | time={time.time() - start:.2f}s")
        logger.info(f"Triage Summary: {triage_summary}")
        logger.info(f"Owner Ticket Map: {owner_ticket_dict}")
        logger.info(f"Bulk Issues: {bulk_issues_dict}")
        logger.info(f"Bulk Issues with QI: {bulk_issues_with_qi}")
        logger.info(f"Pending Tests: {inprogress_tests}")
        logger.info(f"Total Tests Processed: {len(test_data)}")
        
        return jsonify({
            "tag": tag or None,
            "task_ids": task_ids if task_ids else None,
            "generated_at": datetime.utcnow().isoformat(),
            "triage_summary": triage_summary,
            "owner_ticket_map": owner_ticket_dict,
            "bulk_issues": bulk_issues_dict,
            "bulk_issues_with_qi": bulk_issues_with_qi,
            "bulk_issues_count": total_bulk_issues_count,
            "pending_tests": inprogress_tests,
            "total_tests_processed": len(test_data)
        })
    except Exception as e:
        logger.error(f"Error in triage count: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# Triage Accuracy Analyzer Endpoint
# ---------------------------------------------------
def _config_matches_cached(cached, tag, task_ids):
    """Check if cached data matches current tag/task_ids config."""
    if not cached:
        return False
    cached_tag = (cached.get("tag") or "").strip()
    cached_task_ids = cached.get("task_ids") or []
    req_tag = (tag or "").strip()
    req_task_ids = sorted([str(t).strip() for t in (task_ids or []) if t and str(t).strip()])
    cached_task_ids_sorted = sorted([str(t).strip() for t in cached_task_ids if t])
    # Tag mode: match by tag only
    if req_tag:
        return req_tag == cached_tag
    # Task IDs mode: match by task_ids
    return req_task_ids == cached_task_ids_sorted


@app.route("/mcp/regression/triage-accuracy", methods=["GET"])
@app.route("/api/mcp/regression/triage-accuracy", methods=["GET"])
@jwt_required
def get_triage_accuracy():
    """Triage Accuracy Analyzer: fetch Failed+Warning testcases, compare Jira vs Triage Genie, store in JSON."""
    start = time.time()
    tag = request.args.get("tag")
    if tag is not None:
        tag = (tag or "").strip() or None
    task_ids_param = request.args.get("task_ids")
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]

    # Fall back to config if params missing
    if not tag and not task_ids:
        config = load_regression_config()
        if config.get("input_mode") == "tag":
            tag = config.get("default_tag") or config.get("tag", "") or ""
        if not tag and config.get("input_mode") == "task_ids" and config.get("task_ids"):
            task_ids = config.get("task_ids", [])

    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400

    if tag:
        logger.info(f"[START] Triage Accuracy | tag={tag}")
    else:
        logger.info(f"[START] Triage Accuracy | task_ids={len(task_ids)} tasks")

    try:
        load_owner_mapping()
        cache_tag = tag if tag else None
        reload = request.args.get("reload", "false").lower() == "true"
        if reload:
            invalidate_triage_accuracy_cache(cache_tag)
            cached = None
            logger.info("[Triage Accuracy] Cache invalidated, fetching fresh data")
        else:
            cached = load_triage_accuracy_data(cache_tag)
        if cached and _config_matches_cached(cached, tag, task_ids):
            logger.info("[Triage Accuracy] Using cached data")
            return jsonify(cached)

        tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
        if not tasks:
            result = {
                "generated_time": datetime.utcnow().isoformat(),
                "tag": tag or None,
                "task_ids": list(task_ids) if task_ids else [],
                "testcases": [],
                "triage_summary": {
                    "total_failed_warning_count": 0,
                    "triaged_count": 0,
                    "triage_genie_count": 0,
                    "total_triage_genie_count": 0,
                    "triage_completed_percent": 0,
                    "triage_genie_percent": 0,
                    "total_triage_genie_percent": 0,
                    "matched_count": 0,
                    "unmatched_count": 0,
                    "matched_percent": 0,
                    "unmatched_percent": 0,
                },
            }
            save_triage_accuracy_data(result, cache_tag)
            return jsonify(result)

        collected_task_ids = [t["_id"]["$oid"] for t in tasks]
        logger.info(f"Fetching test results for {len(collected_task_ids)} tasks")
        test_data = fetch_test_results_batch_with_pagination(collected_task_ids, timeout=180)

        # Filter Failed or Warning; deduplicate by test name
        failed_warning = [
            tr for tr in test_data
            if tr.get("status", "").lower() in ("failed", "failure", "warning", "warn")
        ]
        seen_tests = set()
        unique_results = []
        for tr in failed_warning:
            test_name = (tr.get("test") or {}).get("name", "") if isinstance(tr.get("test"), dict) else ""
            if not test_name or test_name in seen_tests:
                continue
            seen_tests.add(test_name)
            unique_results.append(tr)

        # Batch pre-fetch Triage Genie tickets via direct /api/tasks/{id} lookups
        _test_result_ids = []
        for tr in unique_results:
            _rid = tr.get("_id")
            if isinstance(_rid, dict) and "$oid" in _rid:
                _test_result_ids.append(_rid["$oid"])
            elif _rid:
                _test_result_ids.append(str(_rid))
        tg_ticket_map = build_triage_genie_ticket_map(_test_result_ids)

        def process_one(tr):
            try:
                test_field = tr.get("test", {})
                testcase_name = (test_field.get("name", "") if isinstance(test_field, dict) else
                                str(test_field) if test_field else "")
                status = tr.get("status", "Failed")
                jira_tickets = tr.get("jira_tickets", [])
                jira_ticket = (jira_tickets[0] if jira_tickets else "") or ""
                if isinstance(jira_ticket, dict):
                    jira_ticket = jira_ticket.get("$oid", "") or str(jira_ticket)
                jira_ticket = str(jira_ticket).strip() if jira_ticket else ""

                testcase_id = None
                if isinstance(tr.get("_id"), dict) and "$oid" in tr.get("_id", {}):
                    testcase_id = tr["_id"]["$oid"]
                else:
                    testcase_id = str(tr.get("_id", "")) if tr.get("_id") else ""

                triage_genie_ticket = ""
                tg = (
                    tg_ticket_map.get(testcase_id)
                    or tg_ticket_map.get(testcase_name)
                )
                if tg:
                    triage_genie_ticket = str(tg).strip()

                if jira_ticket and triage_genie_ticket:
                    match_status = "Matched" if jira_ticket.upper() == triage_genie_ticket.upper() else "Unmatched"
                else:
                    match_status = "N/A" if not jira_ticket and not triage_genie_ticket else ("" if not (jira_ticket and triage_genie_ticket) else "N/A")

                regression_owner = resolve_owner(testcase_name) if testcase_name else "Unknown"
                return {
                    "testcase_name": testcase_name,
                    "regression_owner": regression_owner,
                    "status": status,
                    "triage_genie_ticket": triage_genie_ticket,
                    "jira_ticket": jira_ticket,
                    "match_status": match_status,
                }
            except Exception as e:
                logger.warning(f"Error processing test result for triage accuracy: {e}")
                return None

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            raw = list(executor.map(process_one, unique_results))
            testcases = [tc for tc in raw if tc is not None]

        total = len(testcases)
        triaged_count = sum(1 for tc in testcases if tc.get("jira_ticket"))
        # Triage Genie % = among triaged (JITA tagged), how many have Triage Genie ticket
        triage_genie_count = sum(1 for tc in testcases if tc.get("jira_ticket") and tc.get("triage_genie_ticket"))
        # Total Triage Genie Tagged = among ALL failed/warning, how many have Triage Genie ticket
        total_triage_genie_count = sum(1 for tc in testcases if tc.get("triage_genie_ticket"))
        matched_count = sum(1 for tc in testcases if tc.get("match_status") == "Matched")
        unmatched_count = sum(1 for tc in testcases if tc.get("match_status") == "Unmatched")

        triage_completed_percent = round(100 * triaged_count / total, 1) if total else 0
        triage_genie_percent = round(100 * triage_genie_count / triaged_count, 1) if triaged_count else 0
        total_triage_genie_percent = round(100 * total_triage_genie_count / total, 1) if total else 0
        denom = matched_count + unmatched_count
        matched_percent = round(100 * matched_count / denom, 1) if denom else 0
        unmatched_percent = round(100 * unmatched_count / denom, 1) if denom else 0

        result = {
            "generated_time": datetime.utcnow().isoformat(),
            "tag": tag if tag else None,
            "task_ids": collected_task_ids,
            "testcases": testcases,
            "triage_summary": {
                "total_failed_warning_count": total,
                "triaged_count": triaged_count,
                "triage_genie_count": triage_genie_count,
                "total_triage_genie_count": total_triage_genie_count,
                "triage_completed_percent": triage_completed_percent,
                "triage_genie_percent": triage_genie_percent,
                "total_triage_genie_percent": total_triage_genie_percent,
                "matched_count": matched_count,
                "unmatched_count": unmatched_count,
                "matched_percent": matched_percent,
                "unmatched_percent": unmatched_percent,
            },
        }
        save_triage_accuracy_data(result, cache_tag)
        logger.info(f"[END] Triage Accuracy | testcases={len(testcases)} | time={time.time() - start:.2f}s")
        return jsonify(result)
    except Exception as e:
        logger.error(f"Error in triage accuracy: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/triage-accuracy/export-excel", methods=["GET"])
@app.route("/api/mcp/regression/triage-accuracy/export-excel", methods=["GET"])
@jwt_required
def export_triage_accuracy_excel():
    """Export triage accuracy data as Excel file."""
    try:
        tag = request.args.get("tag")
        if not tag:
            config = load_regression_config()
            if config.get("input_mode") == "tag":
                tag = config.get("default_tag") or config.get("tag") or ""
        cache_tag = tag if tag else None
        data = load_triage_accuracy_data(cache_tag)
        if not data or not isinstance(data, dict):
            return jsonify({"error": "No triage accuracy data available. Load Triage Accuracy Analyzer first."}), 404

        testcases = data.get("testcases", [])
        triage_summary = data.get("triage_summary", {})

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Sheet 1: Triage Analysis
            df_analysis = pd.DataFrame(testcases, columns=[
                "testcase_name", "regression_owner", "status",
                "triage_genie_ticket", "jira_ticket", "match_status"
            ])
            df_analysis.columns = [
                "Testcase Name", "Regression Owner", "Status",
                "Triage Genie Ticket", "Jira Ticket", "Matched/Unmatched"
            ]
            df_analysis.to_excel(writer, sheet_name="Triage Analysis", index=False)

            # Sheet 2: Triage Summary (Metric, Count, Percentage)
            tg_count = triage_summary.get("triage_genie_count", 0)
            matched = triage_summary.get("matched_count", 0)
            unmatched = triage_summary.get("unmatched_count", 0)
            summary_rows = [
                ("Triage Genie Ticket %(based on completed triaged)", tg_count, triage_summary.get("triage_genie_percent", 0)),
                ("Matched %", matched, triage_summary.get("matched_percent", 0)),
                ("Unmatched %", unmatched, triage_summary.get("unmatched_percent", 0)),
            ]
            df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Count", "Percentage"])
            df_summary.to_excel(writer, sheet_name="Triage Summary", index=False)

        output.seek(0)
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="triage_accuracy_report.xlsx",
        )
    except Exception as e:
        logger.error(f"Error exporting triage accuracy Excel: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# QI Summary Report Endpoint
# ---------------------------------------------------
@app.route("/mcp/regression/qi-summary", methods=["GET"])
@jwt_required
def get_qi_summary():
    start = time.time()
    tag = request.args.get("tag")
    task_ids_param = request.args.get("task_ids")  # Comma-separated task IDs
    
    # Parse task_ids if provided
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]
    
    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400
    
    if tag:
        logger.info(f"[START] QI Summary | tag={tag}")
    else:
        logger.info(f"[START] QI Summary | task_ids={len(task_ids)} tasks")
    
    try:
        # Fetch tasks for the tag or task IDs
        tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
        
        # Collect all task IDs from fetched tasks
        collected_task_ids = [task["_id"]["$oid"] for task in tasks]
        
        # Fetch test results using agave_test_results API for accurate counts
        logger.info(f"Fetching test results for {len(collected_task_ids)} tasks using agave_test_results API")
        test_results = []
        if collected_task_ids:
            try:
                test_results = fetch_test_results_batch_with_pagination(collected_task_ids)
                logger.info(f"Fetched {len(test_results)} test results")
            except Exception as e:
                logger.warning(f"Failed to fetch test results from agave_test_results API: {e}. Falling back to test_result_count.")
                test_results = []
        
        # Track whether we successfully fetched merged results
        merged_results_available = len(test_results) > 0
        
        # Group test results by task_id and branch
        test_results_by_task = defaultdict(list)
        task_branch_map = {}
        for task in tasks:
            task_id = task["_id"]["$oid"]
            branch = task.get("branch", "unknown")
            task_branch_map[task_id] = branch
        
        for test_result in test_results:
            agave_task_id = test_result.get("agave_task_id")
            if agave_task_id:
                # Handle both string and $oid format
                if isinstance(agave_task_id, dict) and "$oid" in agave_task_id:
                    task_id = agave_task_id["$oid"]
                else:
                    task_id = str(agave_task_id)
                if task_id in task_branch_map:
                    test_results_by_task[task_id].append(test_result)
        
        # Generate QI Summary Report
        summary = {
            "tag": tag or None,
            "task_ids": collected_task_ids if collected_task_ids else None,
            "generated_at": datetime.utcnow().isoformat(),
            "total_tasks": len(tasks),
            "status_summary": {
                "testing": 0,
                "completed": 0,
                "pending": 0,
                "failed": 0
            },
            "test_summary": {
                "total": 0,
                "succeeded": 0,
                "failed": 0,
                "pending": 0,
                "warning": 0,
                "running": 0,
                "skipped": 0,
                "killed": 0
            },
            "branch_summary": {}
        }
        
        for task in tasks:
            task_id = task["_id"]["$oid"]
            status = task.get("status", "").lower()
            branch = task.get("branch", "unknown")
            
            # Count test statuses from agave_test_results if available
            task_test_counts = {
                "total": 0,
                "Succeeded": 0,
                "Failed": 0,
                "Pending": 0,
                "Warning": 0,
                "Running": 0,
                "Skipped": 0,
                "Killed": 0
            }
            
            if task_id in test_results_by_task:
                # Count statuses from actual test results
                for test_result in test_results_by_task[task_id]:
                    test_status = test_result.get("status", "")
                    task_test_counts["total"] += 1
                    
                    # Normalize status names (handle case-insensitive matching)
                    status_lower = test_status.lower() if test_status else ""
                    
                    if status_lower == "succeeded" or status_lower == "success":
                        task_test_counts["Succeeded"] += 1
                    elif status_lower == "failed" or status_lower == "failure":
                        task_test_counts["Failed"] += 1
                    elif status_lower == "pending" or status_lower == "waiting":
                        task_test_counts["Pending"] += 1
                    elif status_lower == "warning" or status_lower == "warn":
                        task_test_counts["Warning"] += 1
                    elif status_lower == "running" or status_lower == "executing" or status_lower == "in_progress":
                        task_test_counts["Running"] += 1
                    elif status_lower == "skipped" or status_lower == "skip":
                        task_test_counts["Skipped"] += 1
                    elif status_lower == "killed" or status_lower == "terminated" or status_lower == "cancelled":
                        task_test_counts["Killed"] += 1
                    else:
                        # For unknown statuses, try to infer from common patterns
                        if any(x in status_lower for x in ["pending", "waiting", "queued"]):
                            task_test_counts["Pending"] += 1
                        elif any(x in status_lower for x in ["running", "executing", "in_progress", "active"]):
                            task_test_counts["Running"] += 1
                        elif any(x in status_lower for x in ["skipped", "skip"]):
                            task_test_counts["Skipped"] += 1
                        elif any(x in status_lower for x in ["killed", "terminated", "cancelled", "aborted"]):
                            task_test_counts["Killed"] += 1
                        else:
                            # Default to pending only if truly unknown
                            task_test_counts["Pending"] += 1
            elif not merged_results_available:
                # Fallback to test_result_count ONLY when the agave_test_results API
                # call failed entirely. When merge=True results were fetched successfully,
                # tasks not in the map have their tests already counted via other tasks'
                # merged results — using the stale test_result_count would double-count.
                tc = task.get("test_result_count", {})
                task_test_counts = {
                    "total": tc.get("Total", 0),
                    "Succeeded": tc.get("Succeeded", 0),
                    "Failed": tc.get("Failed", 0),
                    "Pending": tc.get("Pending", 0),
                    "Warning": tc.get("Warning", 0),
                    "Running": tc.get("Running", 0),
                    "Skipped": tc.get("Skipped", 0),
                    "Killed": tc.get("Killed", 0)
                }
            
            # Status summary
            if status == "testing":
                summary["status_summary"]["testing"] += 1
            elif status == "pending":
                summary["status_summary"]["pending"] += 1
            elif task_test_counts["Failed"] > 0:
                summary["status_summary"]["failed"] += 1
            else:
                summary["status_summary"]["completed"] += 1
            
            # Test summary (aggregate across all tasks)
            summary["test_summary"]["total"] += task_test_counts["total"]
            summary["test_summary"]["succeeded"] += task_test_counts["Succeeded"]
            summary["test_summary"]["failed"] += task_test_counts["Failed"]
            summary["test_summary"]["pending"] += task_test_counts["Pending"]
            summary["test_summary"]["warning"] += task_test_counts["Warning"]
            summary["test_summary"]["running"] += task_test_counts["Running"]
            summary["test_summary"]["skipped"] += task_test_counts["Skipped"]
            summary["test_summary"]["killed"] += task_test_counts["Killed"]
            
            # Branch summary
            if branch not in summary["branch_summary"]:
                summary["branch_summary"][branch] = {
                    "total_tasks": 0,
                    "total_tests": 0,
                    "failed_tests": 0
                }
            
            summary["branch_summary"][branch]["total_tasks"] += 1
            summary["branch_summary"][branch]["total_tests"] += task_test_counts["total"]
            summary["branch_summary"][branch]["failed_tests"] += task_test_counts["Failed"]
        
        logger.info(f"[END] QI Summary | time={time.time() - start:.2f}s")
        logger.info(f"Test Summary: {summary['test_summary']}")
        
        return jsonify(summary)
    except Exception as e:
        logger.error(f"Error in QI summary: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# TCMS Overall QI Endpoint
# ---------------------------------------------------
def _resolve_tcms_milestone(branch_name):
    """
    Convert a full branch name (as shown in the Run Summary table) to the
    short milestone name the TCMS API expects.

    Lookup order:
      1. Explicit entry in BRANCH_SHORT_NAME_MAP
      2. Regex extraction of a version pattern like X.Y or X.Y.Z
      3. Fall back to the original branch_name unchanged
    """
    if branch_name in BRANCH_SHORT_NAME_MAP:
        return BRANCH_SHORT_NAME_MAP[branch_name]

    version_match = re.search(r'(\d+\.\d+(?:\.\d+)?)', branch_name)
    if version_match:
        return version_match.group(1)

    return branch_name


@app.route("/mcp/regression/tcms-overall-qi", methods=["GET"])
@jwt_required
def get_tcms_overall_qi():
    """
    Fetch the aggregate QI (average_total_op_success_percentage) from the
    TCMS Summary API for a given team, branch, and time filter.

    Branch handling:
      * **master** — uses team-specific filters (additional_data.team,
        team test_sets regex, release_name exclusion, tag exclusions) and
        ``feat_type=regression``.
      * **release branches** (e.g. ganges-7.6-stable → milestone "7.6") —
        uses simpler filters (test_sets regex + deprecated flag only) and
        ``feat_type=all``.
    """
    start = time.time()
    team_name = request.args.get("team_name")
    branch_name = request.args.get("branch_name")
    time_filter = request.args.get("time_filter", "all")

    if not team_name or not branch_name:
        return jsonify({"error": "team_name and branch_name are required"}), 400

    milestone = _resolve_tcms_milestone(branch_name)
    is_master = branch_name.lower() in ("master", "main")

    logger.info(
        f"[START] TCMS Overall QI | team={team_name} branch={branch_name} "
        f"milestone={milestone} is_master={is_master} time_filter={time_filter}"
    )

    try:
        if is_master:
            # Master: team-specific filters, feat_type=regression
            filters = json.dumps({
                "$and": [
                    {"test_case.test_sets": {"$regex": f"test_sets/milestones/{milestone}/", "$options": "i"}},
                    {"release_name": {"$ne": milestone}},
                    {"test_case.metadata.tags": {"$nin": ["SYSTEST_LONGEVITY", "LIMITED_RUNS"]}},
                    {"additional_data.team": f"{milestone}/{team_name}"},
                    {"test_case.test_sets": {"$regex": f"test_sets/milestones/{milestone}/{team_name}/", "$options": "i"}},
                    {"test_case.deprecated": False},
                ]
            })
            feat_type = "regression"
        else:
            # Release branch: team-specific filters, feat_type=regression
            filters = json.dumps({
                "$and": [
                    {"additional_data.team": f"{milestone}/{team_name}"},
                    {"test_case.test_sets": {"$regex": f"test_sets/milestones/{milestone}/{team_name}/", "$options": "i"}},
                    {"test_case.deprecated": False},
                ]
            })
            feat_type = "regression"

        params = {
            "aggregation_field": "target_package_type",
            "time_filter": time_filter,
            "target_milestone": milestone,
            "feat_type": feat_type,
            "filters": filters,
        }

        url = f"{TCMS_SUMMARY_BASE}/milestone_all_test_cases/aggregate/metrics"
        response = requests.get(
            url,
            params=params,
            headers={"Content-Type": "application/json"},
            verify=False,
            timeout=30,
        )

        if response.status_code != 200:
            logger.error(f"TCMS Summary API returned {response.status_code}: {response.text[:500]}")
            return jsonify({"error": f"TCMS API error: {response.status_code}"}), 502

        data = response.json()
        if not data.get("success") or not data.get("data"):
            logger.warning("TCMS Summary API returned no data")
            return jsonify({
                "qi_value": None,
                "message": "No data returned from TCMS",
                "team_name": team_name,
                "branch_name": branch_name,
                "milestone": milestone,
                "time_filter": time_filter,
            })

        overall = data["data"][0]
        qi_value = overall.get("average_total_op_success_percentage")

        logger.info(
            f"[END] TCMS Overall QI | qi={qi_value} | time={time.time() - start:.2f}s"
        )

        return jsonify({
            "qi_value": qi_value,
            "team_name": team_name,
            "branch_name": branch_name,
            "milestone": milestone,
            "time_filter": time_filter,
            "total_tests": overall.get("total"),
            "run": overall.get("run"),
            "passed": overall.get("passed"),
            "failed": overall.get("failed"),
            "not_run": overall.get("not_run"),
            "blocked": overall.get("blocked"),
            "run_percentage": overall.get("run_percentage"),
            "overall_effectiveness": overall.get("overall_effectiveness"),
            "overall_stability": overall.get("overall_stability"),
            "total_triaged": overall.get("total_triaged"),
            "triage_percentage": overall.get("triage_percentage"),
            "total_product_issues": overall.get("total_product_issues"),
            "total_test_issues": overall.get("total_test_issues"),
            "total_other_issues": overall.get("total_other_issues"),
            "total_infra_issues": overall.get("total_infra_issues"),
            "total_framework_issues": overall.get("total_framework_issues"),
            "openBugs": overall.get("openBugs"),
            "unique_tickets": overall.get("unique_tickets", []),
            "execution_passed_percentage": overall.get("execution_passed_percentage"),
            "execution_failed_percentage": overall.get("execution_failed_percentage"),
        })

    except requests.exceptions.Timeout:
        logger.error("TCMS Summary API request timed out")
        return jsonify({"error": "TCMS API request timed out"}), 504
    except Exception as e:
        logger.error(f"Error fetching TCMS Overall QI: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------
# Team Config Endpoint
# ---------------------------------------------------
@app.route("/mcp/regression/team-config", methods=["GET"])
@jwt_required
def get_team_config():
    """Return the tag-to-team configuration and branch short-name mapping."""
    return jsonify({
        "team_config": TEAM_CONFIG,
        "branch_short_names": BRANCH_SHORT_NAME_MAP,
    })


# ======================================================
# Run Report - QI Analysis Endpoint
# ======================================================
def generate_qi_analysis(run_folder):
    """Generate QI analysis Excel file from CSV files in run folder"""
    try:
        # File paths
        tcms_csv_path = os.path.join(run_folder, 'tcms.csv')
        regression_owners_path = os.path.join(run_folder, 'regression_owners.csv')
        jita_csv_path = os.path.join(run_folder, 'jita.csv')
        tcms_bugs_csv_path = os.path.join(run_folder, 'tcms_bugs.csv')
        
        # Check if open-bugs.csv exists (alternative name)
        if not os.path.exists(tcms_bugs_csv_path):
            tcms_bugs_csv_path = os.path.join(run_folder, 'open-bugs.csv')
        
        # Validate required files exist
        required_files = {
            'tcms.csv': tcms_csv_path,
            'regression_owners.csv': regression_owners_path,
            'jita.csv': jita_csv_path,
            'tcms_bugs.csv or open-bugs.csv': tcms_bugs_csv_path
        }
        
        missing_files = [name for name, path in required_files.items() if not os.path.exists(path)]
        if missing_files:
            raise FileNotFoundError(f"Missing required files: {', '.join(missing_files)}")
        
        # Read CSV files
        logger.info("Reading CSV files...")
        df = pd.read_csv(tcms_csv_path)
        df_jita = pd.read_csv(jita_csv_path)
        df_o = pd.read_csv(regression_owners_path)
        tcms_bug_list = pd.read_csv(tcms_bugs_csv_path)
        
        # Process JITA data (remove Pending, handle duplicates)
        df_jita = df_jita[~df_jita.status.isin(['Pending'])]
        temp_df = df_jita[df_jita['start_time'] != "-"]
        if len(temp_df) > 0:
            temp_df.loc[:, 'start_time'] = pd.to_datetime(temp_df['start_time'], dayfirst=True)
            min_time = temp_df['start_time'].min()
            df_jita.loc[df_jita['start_time'] == '-', "start_time"] = min_time
            df_jita.loc[:, 'start_time'] = pd.to_datetime(df_jita['start_time'], dayfirst=True)
            df_jita = df_jita.sort_values(['start_time'], ascending=False).drop_duplicates(['name'])
        
        # Process TCMS data
        df['Test_Set'] = 'None'
        ll = []
        
        for index, row in df.iterrows():
            # Extract test set name
            if isinstance(row.get('Test Sets'), str):
                for testset in row['Test Sets'].split(','):
                    if re.search('test_sets/milestones/.*/cdp/.*/Regression_team_owned_lst', testset):
                        df.loc[index, 'Test_Set'] = testset.split('/')[-1]
                        break
            
            # Extract last passed ops
            if isinstance(row.get('Last Passed Ops'), str):
                parts = row['Last Passed Ops'].split('/')
                if len(parts) >= 2:
                    passed_ops = parts[0]
                    total_ops = parts[1].split('(')[0]
                    df.loc[index, 'last_passed_ops'] = passed_ops
                    df.loc[index, 'last_passed_total_ops'] = total_ops
            
            # Extract last run ops
            col_name = 'Last Run Ops'
            if isinstance(row.get(col_name), str):
                parts = row[col_name].split('/')
                if len(parts) >= 2:
                    passed_ops = parts[0]
                    total_ops = parts[1].split('(')[0]
                    df.loc[index, 'last_run_ops'] = int(passed_ops) if passed_ops.isdigit() else 0
                    df.loc[index, 'last_run_total_ops'] = int(total_ops) if total_ops.isdigit() else 0
                    if df.loc[index, 'last_run_total_ops'] > 0:
                        df.loc[index, 'last_run_qi'] = 100 * (df.loc[index, 'last_run_ops'] / df.loc[index, 'last_run_total_ops'])
                    else:
                        df.loc[index, 'last_run_qi'] = 0
                else:
                    df.loc[index, 'last_run_ops'] = 0
                    df.loc[index, 'last_run_total_ops'] = 0
                    df.loc[index, 'last_run_qi'] = 0
            else:
                df.loc[index, 'last_run_ops'] = 0
                df.loc[index, 'last_run_total_ops'] = 0
                df.loc[index, 'last_run_qi'] = 0
            
            # Extract test sets
            if isinstance(row.get('Test Sets'), str):
                test_sets = row['Test Sets'].split(',')
                for test_set in test_sets:
                    if re.search('test_sets/milestones/.*/cdp/.*/Regression_team_owned_lst/', test_set):
                        df.loc[index, 'Test_Set'] = test_set.split('/')[-1]
                        break
                    else:
                        df.loc[index, 'Test_Set'] = test_set.split('/')[-1]
            
            # Process open bugs
            if isinstance(row.get('Open Bugs'), str):
                bug_list = row['Open Bugs'].split(',')
                bugtype = tcms_bug_list[tcms_bug_list.Ticket.isin(bug_list)].groupby(['Type'])['Ticket'].count().to_dict()
                for key in bugtype.keys():
                    df.loc[index, f'{key}_bug_count'] = bugtype.get(key, 0)
                
                # Create bug-testcase mapping
                for bug in bug_list:
                    m = {
                        'bug_id': bug,
                        'test_case': row['Name'],
                        'last_run_qi': df.loc[index, 'last_run_qi'],
                        'last_run_ops': df.loc[index, 'last_run_ops'],
                        'last_run_total_ops': df.loc[index, 'last_run_total_ops'],
                        'Last Run Status': row.get('Last Run Status', '')
                    }
                    ll.append(m)
        
        df['last_run_ops'] = df['last_run_ops'].astype(int)
        df['last_run_total_ops'] = df['last_run_total_ops'].astype(int)
        
        df_bugs = pd.DataFrame(ll)
        
        # Extract test area
        testcasenames_split = df['Name'].str.split(pat='.', expand=True)
        t = testcasenames_split[[0, 1, 2, 3]].agg('.'.join, axis=1)
        df.insert(loc=1, column='Test Area', value=t)
        
        # Join with regression owners
        df = df.join(df_o.set_index('Test Area'), on='Test Area', how='left')
        
        nr_test_cases = df['Name'].count()
        
        def generate_qi_impact(df_data, colname, nr_test_cases):
            s1 = df_data.groupby([colname])['last_run_qi'].agg("mean").sort_values()
            s2 = df_data.groupby([colname])[colname].count()
            if colname == 'bug_id':
                s3 = df_bugs.groupby([colname, 'Last Run Status'])[[colname]].count().unstack()
                df_testarea = pd.concat([s1, s2, s3], axis=1)
            else:
                df_testarea = pd.concat([s1, s2], axis=1)
            cols = ['average_qi', 'nr_test_cases']
            cols.extend(df_testarea.columns[2:])
            df_testarea.columns = cols
            df_testarea['qi_impact'] = (df_testarea['average_qi'] - 100) * df_testarea['nr_test_cases']
            df_testarea['overall_qi_impact'] = 100 * (df_testarea['qi_impact'] / (100 * nr_test_cases))
            return df_testarea.sort_values(['overall_qi_impact'])
        
        df_testareas = generate_qi_impact(df, 'Test Area', nr_test_cases)
        df_bugid = generate_qi_impact(df_bugs, 'bug_id', nr_test_cases)
        
        # Process dates
        baseline_date = datetime.now() - timedelta(days=30)
        try:
            df.loc[:, 'Last Run Date'] = pd.to_datetime(df['Last Run Date'], format="%Y-%m-%d")
        except:
            df.loc[:, 'Last Run Date'] = pd.to_datetime(df['Last Run Date'])
        df = df.sort_values(['Last Run Date']).drop_duplicates(['Name'])
        
        df_tcms = df.copy()
        df_jita_tcms = df_jita.join(df.set_index('Name'), on='name', how='left')
        
        # Merge bugs with TCMS bug list
        tcms_bug_list['Name'] = tcms_bug_list['Ticket']
        del tcms_bug_list['Ticket']
        tcms_bug_qi = tcms_bug_list.join(df_bugid, on='Name', how='left')
        tcms_bug_qi['tcms_test_cases'] = tcms_bug_qi['nr_test_cases']
        del tcms_bug_qi['nr_test_cases']
        tcms_bug_qi = tcms_bug_qi.set_index(['Name'])
        
        # Generate output list for summary
        output_list = []
        output_list.append(f"Analysis ran on: {datetime.now()}")
        output_list.append(f"Total number of test cases: {df['Name'].count()}")
        output_list.append(f"Total number of test that passed in last run: {df[df['Last Run Status']=='succeeded']['Name'].count()}")
        output_list.append(f"Total number of test that failed in last run: {df[df['Last Run Status']=='failed']['Name'].count()}")
        output_list.append(f"Total number of test that warned in last run: {df[df['Last Run Status']=='warning']['Name'].count()}")
        output_list.append(f"Total number of test cases with bugs: {df[~df['Open Bugs'].isna()]['Name'].count()}")
        output_list.append(f"Last Run QI: {df['last_run_qi'].mean():.2f}")
        output_list.append(f"Total possible QI: {100*df['Name'].count():.0f}")
        output_list.append(f"Total QI of all test cases: {df[['last_run_qi']].sum().iloc[0]:.0f}")
        output_list.append(f"Total QI impacted due to bugs (overestimated): {df_bugid['qi_impact'].sum():.0f} in Percentage: {df_bugid['qi_impact'].sum()/(df['Name'].count()):.2f}%")
        output_list.append(f"Total number of bugs identified by the runs so far: {len(df_bugs['bug_id'].unique())}")
        output_list.append(f"Tests that never passed: {df[df['Last Passed Date'].isna()].count().iloc[0]}")
        output_list.append(f"Tests that never passed and have last_run_qi<50: {df[((df['Last Passed Date'].isna())&(df['last_run_qi']<50))].count().iloc[0]}")
        t = (100-(df[((df['Last Passed Date'].isna())&(df['last_run_qi']<50))]['last_run_qi'].sum()))*df[((df['Last Passed Date'].isna())&(df['last_run_qi']<50))].count().iloc[0]
        output_list.append(f"QI impact of tests that never passed and have last_run_qi<50: {t:.2f} in Percentage: {(t/df['Name'].count()):.2f}%")
        output_list.append(f"Failed tests with no open Bugs: {df_tcms[(df_tcms['Last Run Status'].isin(['failed','warning'])) & (df_tcms['Open Bugs'].isna())]['Name'].count()}")
        output_list.append(f"Failed tests that are not triaged (TCMS): {df[(df['Last Run Status'] != 'succeeded') & (df.get('Is Last Run Triaged', pd.Series([False]*len(df))) != True)]['Name'].count()}")
        
        # Generate Excel file
        filename = 'analysis_' + datetime.now().isoformat().replace(':', '_').replace('-', '_')
        xlsfilepath = os.path.join(run_folder, filename + '.xlsx')
        
        # Use openpyxl engine for writing
        with pd.ExcelWriter(xlsfilepath, engine='openpyxl') as writer:
            # Summary sheet
            pd.DataFrame(output_list).to_excel(writer, sheet_name='summary', startrow=0, startcol=0, index=False, header=False)
            
            # Bug QI Summary sheet
            tcms_bug_qi.groupby(['Type', 'Priority'])[['overall_qi_impact']].agg(['count', 'sum']).to_excel(
                writer, sheet_name='bug_qi_summary', startrow=0, startcol=0, index=True
            )
            
            # Bugs QI Analysis sheet
            tcms_bug_qi.sort_values(['overall_qi_impact']).to_excel(
                writer, sheet_name='bugs_qi_analysis', startrow=0, startcol=0, index=True
            )
        
        logger.info(f"Generated QI analysis file: {xlsfilepath}")
        return xlsfilepath
        
    except Exception as e:
        logger.error(f"Error generating QI analysis: {e}", exc_info=True)
        raise

@app.route("/mcp/regression/run-report/list-analysis-files", methods=["POST"])
@jwt_required
def list_analysis_files():
    """List all analysis_*.xlsx files in a given directory"""
    try:
        req_data = request.json
        folder_path = req_data.get("folder_path")
        
        if not folder_path:
            return jsonify({"error": "folder_path is required"}), 400
        
        if not os.path.isdir(folder_path):
            return jsonify({"error": f"Folder path does not exist: {folder_path}"}), 400
        
        # Find all files matching analysis_*.xlsx pattern
        pattern = os.path.join(folder_path, "analysis_*.xlsx")
        matching_files = glob.glob(pattern)
        
        # Sort by modification time (newest first)
        matching_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
        
        # Return just the filenames
        file_list = [os.path.basename(f) for f in matching_files]
        
        return jsonify({
            "success": True,
            "files": file_list,
            "folder_path": folder_path
        })
    except Exception as e:
        logger.error(f"Error listing analysis files: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

def read_existing_analysis_file(analysis_file_path):
    """Read and extract data from an existing analysis Excel file"""
    try:
        if not os.path.exists(analysis_file_path):
            raise FileNotFoundError(f"Analysis file not found: {analysis_file_path}")
        
        logger.info(f"Reading QI analysis from existing file: {analysis_file_path}")
        
        # Read Excel file sheets (same logic as in qi_analysis_from_folder)
        excel_data = {}
        
        # 1. Read summary sheet
        try:
            df_summary = pd.read_excel(analysis_file_path, sheet_name="summary", header=None)
            summary_text = "\n".join(df_summary[0].astype(str).tolist()) if len(df_summary.columns) > 0 else ""
            excel_data["summary"] = summary_text
        except Exception as e:
            logger.warning(f"Could not read summary sheet: {e}")
            excel_data["summary"] = ""
        
        # 2. Read bug_qi_summary sheet
        try:
            df_bug_qi_summary = pd.read_excel(analysis_file_path, sheet_name="bug_qi_summary", index_col=[0, 1], header=[0, 1])
            # Handle multi-level columns (from groupby with agg)
            bug_qi_summary_data = []
            type_totals = {}  # Track total impacting QI by type (test, product, framework, other)
            
            for (bug_type, priority), row in df_bug_qi_summary.iterrows():
                # Extract count and sum from multi-level columns
                count_val = 0
                sum_val = 0.0
                
                # Try to find count and sum columns
                for col in df_bug_qi_summary.columns:
                    if isinstance(col, tuple):
                        if 'overall_qi_impact' in str(col[0]).lower() or 'overall_qi_impact' in str(col):
                            if 'count' in str(col[1]).lower() or 'count' in str(col):
                                count_val = int(row[col]) if pd.notna(row[col]) else 0
                            elif 'sum' in str(col[1]).lower() or 'sum' in str(col):
                                sum_val = float(row[col]) if pd.notna(row[col]) else 0.0
                
                # Fallback: if columns are flattened
                if count_val == 0 and sum_val == 0.0:
                    for col in df_bug_qi_summary.columns:
                        col_str = str(col)
                        if 'count' in col_str.lower():
                            count_val = int(row[col]) if pd.notna(row[col]) else 0
                        elif 'sum' in col_str.lower():
                            sum_val = float(row[col]) if pd.notna(row[col]) else 0.0
                
                # Accumulate total by type
                type_str = str(bug_type).lower()
                if type_str not in type_totals:
                    type_totals[type_str] = 0.0
                type_totals[type_str] += sum_val
                
                bug_qi_summary_data.append({
                    "type": str(bug_type),
                    "priority": str(priority),
                    "testcase_count": count_val,
                    "impacting_qi": sum_val
                })
            
            # Calculate totals for test, product, framework, other
            type_summary = {
                "test": 0.0,
                "product": 0.0,
                "framework": 0.0,
                "other": 0.0
            }
            
            for type_key, total_val in type_totals.items():
                type_lower = type_key.lower()
                if 'test' in type_lower:
                    type_summary["test"] += total_val
                elif 'product' in type_lower:
                    type_summary["product"] += total_val
                elif 'framework' in type_lower:
                    type_summary["framework"] += total_val
                else:
                    type_summary["other"] += total_val
            
            excel_data["bug_qi_summary"] = bug_qi_summary_data
            excel_data["type_summary"] = type_summary
        except Exception as e:
            logger.warning(f"Could not read bug_qi_summary sheet: {e}")
            excel_data["bug_qi_summary"] = []
        
        # 3. Read bugs_qi_analysis sheet (Top QI Impacting bugs)
        try:
            df_bugs_qi_analysis = pd.read_excel(analysis_file_path, sheet_name="bugs_qi_analysis", index_col=0)
            # Sort by overall_qi_impact in ascending order (most negative/impactful first) and get top 30
            if "overall_qi_impact" in df_bugs_qi_analysis.columns:
                df_bugs_qi_analysis = df_bugs_qi_analysis.sort_values("overall_qi_impact", ascending=True).head(30)
            
            # Extract required columns
            top_bugs = []
            for bug_name, row in df_bugs_qi_analysis.iterrows():
                bug_data = {
                    "name": str(bug_name),
                    "type": str(row.get("Type", "")) if "Type" in row else "",
                    "priority": str(row.get("Priority", "")) if "Priority" in row else "",
                    "summary": str(row.get("Summary", "")) if "Summary" in row else "",
                    "assignee": str(row.get("Assignee", "")) if "Assignee" in row else "",
                    "impacted_tcs_latest_run": int(row.get("Impacted TCs (Latest Run)", 0)) if "Impacted TCs (Latest Run)" in row else 0,
                    "deferred": str(row.get("Deferred", "")) if "Deferred" in row else "",
                    "average_qi": float(row.get("average_qi", 0)) if "average_qi" in row else 0.0,
                    "overall_qi_impact": float(row.get("overall_qi_impact", 0)) if "overall_qi_impact" in row else 0.0
                }
                top_bugs.append(bug_data)
            
            excel_data["top_qi_impacting_bugs"] = top_bugs
        except Exception as e:
            logger.warning(f"Could not read bugs_qi_analysis sheet: {e}")
            excel_data["top_qi_impacting_bugs"] = []
        
        return excel_data
    except Exception as e:
        logger.error(f"Error reading existing analysis file: {e}", exc_info=True)
        raise

@app.route("/mcp/regression/run-report/qi-analysis", methods=["POST"])
@jwt_required
def qi_analysis_from_folder():
    """Generate QI analysis Excel file and extract data from it, or read from existing file"""
    try:
        req_data = request.json
        run_folder = req_data.get("run_folder")
        analysis_file_name = req_data.get("analysis_file")  # Optional: if provided, use existing file
        
        if not run_folder:
            return jsonify({"error": "run_folder path is required"}), 400
        
        if not os.path.isdir(run_folder):
            return jsonify({"error": f"Folder path does not exist: {run_folder}"}), 400
        
        # If analysis_file is provided, read from existing file
        if analysis_file_name:
            analysis_file_path = os.path.join(run_folder, analysis_file_name)
            if not os.path.exists(analysis_file_path):
                return jsonify({"error": f"Analysis file not found: {analysis_file_name}"}), 400
            
            excel_data = read_existing_analysis_file(analysis_file_path)
            
            return jsonify({
                "success": True,
                "analysis_file": analysis_file_name,
                "run_folder": run_folder,
                "data": excel_data
            })
        
        # Otherwise, generate the analysis Excel file
        logger.info(f"Generating QI analysis for folder: {run_folder}")
        analysis_file = generate_qi_analysis(run_folder)
        
        # Read the generated file using the shared function
        excel_data = read_existing_analysis_file(analysis_file)
        
        return jsonify({
            "success": True,
            "analysis_file": os.path.basename(analysis_file),
            "run_folder": run_folder,
            "data": excel_data
        })
        
    except Exception as e:
        logger.error(f"Error reading QI analysis: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ======================================================
# Run Report - Email Endpoints
# ======================================================
@app.route("/mcp/regression/run-report/preview-email", methods=["POST"])
@jwt_required
def preview_qi_bug_email():
    """Generate email preview for Top QI Impacting Bugs"""
    try:
        req_data = request.json
        bugs = req_data.get("bugs", [])
        branch_name = req_data.get("branch_name", "Unknown Branch")
        run_folder = req_data.get("run_folder", "")
        
        if not bugs or len(bugs) == 0:
            return jsonify({"error": "Bug data is required"}), 400
        
        # Collect unique assignees
        assignees = set()
        for bug in bugs:
            assignee = bug.get("assignee", "")
            if assignee:
                assignees.add(assignee)
        
        if not assignees:
            return jsonify({"error": "No assignees found in bug data"}), 400
        
        # Convert assignees to email format
        recipient_emails = []
        for assignee in assignees:
            if "@" in assignee:
                recipient_emails.append(assignee)
            else:
                recipient_emails.append(f"{assignee}@nutanix.com")
        
        # Create email subject
        subject = f"Top QI Impacting Bugs on {branch_name}"
        
        # Create email body with HTML table for all bugs
        bugs_table_rows = ""
        for idx, bug in enumerate(bugs, 1):
            bugs_table_rows += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{bug.get('name', 'N/A')}</td>
                    <td>{bug.get('type', 'N/A')}</td>
                    <td>{bug.get('priority', 'N/A')}</td>
                    <td>{bug.get('summary', 'N/A')}</td>
                    <td>{bug.get('assignee', 'N/A')}</td>
                    <td>{bug.get('impacted_tcs_latest_run', 0)}</td>
                    <td>{bug.get('deferred', 'N/A')}</td>
                    <td>{bug.get('average_qi', 0):.2f}</td>
                    <td>{bug.get('overall_qi_impact', 0):.2f}</td>
                </tr>
            """
        
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
                .summary {{ margin: 20px 0; padding: 15px; background-color: #f8f9fa; border-left: 4px solid #3498db; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; font-size: 12px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #3498db; color: white; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                tr:hover {{ background-color: #e8f4f8; }}
                .note {{ margin-top: 20px; padding: 15px; background-color: #fff3cd; border-left: 4px solid #ffc107; }}
            </style>
        </head>
        <body>
            <div class="summary">
                <h2>Top QI Impacting Bugs Notification</h2>
                <p>The following bugs have been identified as top QI impacting bugs on branch: <strong>{branch_name}</strong></p>
                <p><strong>Total Bugs:</strong> {len(bugs)}</p>
                <p><strong>Recipients:</strong> {', '.join(recipient_emails)}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Bug Name</th>
                        <th>Type</th>
                        <th>Priority</th>
                        <th>Summary</th>
                        <th>Assignee</th>
                        <th>Impacted TCs</th>
                        <th>Deferred</th>
                        <th>Average QI</th>
                        <th>Overall QI Impact</th>
                    </tr>
                </thead>
                <tbody>
                    {bugs_table_rows}
                </tbody>
            </table>
            
            <div class="note">
                <p><strong>Note:</strong> These bugs impact more than 4 test cases and have significant overall QI impact. Please review and take appropriate action.</p>
                <p><strong>Run Folder:</strong> {run_folder}</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Top QI Impacting Bugs Notification

The following bugs have been identified as top QI impacting bugs on branch: {branch_name}

Total Bugs: {len(bugs)}
Recipients: {', '.join(recipient_emails)}

Bug Details:
"""
        for idx, bug in enumerate(bugs, 1):
            text_body += f"""
{idx}. {bug.get('name', 'N/A')}
   - Type: {bug.get('type', 'N/A')}
   - Priority: {bug.get('priority', 'N/A')}
   - Summary: {bug.get('summary', 'N/A')}
   - Assignee: {bug.get('assignee', 'N/A')}
   - Impacted Test Cases: {bug.get('impacted_tcs_latest_run', 0)}
   - Deferred: {bug.get('deferred', 'N/A')}
   - Average QI: {bug.get('average_qi', 0):.2f}
   - Overall QI Impact: {bug.get('overall_qi_impact', 0):.2f}
"""
        
        text_body += f"""

Note: These bugs impact more than 4 test cases and have significant overall QI impact. Please review and take appropriate action.

Run Folder: {run_folder}
        """
        
        return jsonify({
            "success": True,
            "subject": subject,
            "html_body": html_body,
            "text_body": text_body,
            "recipients": recipient_emails,
            "bugs": bugs,
            "branch_name": branch_name,
            "run_folder": run_folder
        })
        
    except Exception as e:
        logger.error(f"Error in preview email: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-report/send-email", methods=["POST"])
@jwt_required
def send_qi_bug_email():
    """Send email for Top QI Impacting Bugs"""
    try:
        req_data = request.json
        bugs = req_data.get("bugs", [])
        branch_name = req_data.get("branch_name", "Unknown Branch")
        run_folder = req_data.get("run_folder", "")
        recipients = req_data.get("recipients", [])
        
        if not bugs or len(bugs) == 0:
            return jsonify({"error": "Bug data is required"}), 400
        
        if not recipients or len(recipients) == 0:
            return jsonify({"error": "Recipients are required"}), 400
        
        # Create email subject
        subject = f"Top QI Impacting Bugs on {branch_name}"
        
        # Create email body with HTML table for all bugs
        bugs_table_rows = ""
        for idx, bug in enumerate(bugs, 1):
            bugs_table_rows += f"""
                <tr>
                    <td>{idx}</td>
                    <td>{bug.get('name', 'N/A')}</td>
                    <td>{bug.get('type', 'N/A')}</td>
                    <td>{bug.get('priority', 'N/A')}</td>
                    <td>{bug.get('summary', 'N/A')}</td>
                    <td>{bug.get('assignee', 'N/A')}</td>
                    <td>{bug.get('impacted_tcs_latest_run', 0)}</td>
                    <td>{bug.get('deferred', 'N/A')}</td>
                    <td>{bug.get('average_qi', 0):.2f}</td>
                    <td>{bug.get('overall_qi_impact', 0):.2f}</td>
                </tr>
            """
        
        html_body = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; }}
                .summary {{ margin: 20px 0; padding: 15px; background-color: #f8f9fa; border-left: 4px solid #3498db; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; font-size: 12px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #3498db; color: white; font-weight: bold; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                tr:hover {{ background-color: #e8f4f8; }}
                .note {{ margin-top: 20px; padding: 15px; background-color: #fff3cd; border-left: 4px solid #ffc107; }}
            </style>
        </head>
        <body>
            <div class="summary">
                <h2>Top QI Impacting Bugs Notification</h2>
                <p>The following bugs have been identified as top QI impacting bugs on branch: <strong>{branch_name}</strong></p>
                <p><strong>Total Bugs:</strong> {len(bugs)}</p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>#</th>
                        <th>Bug Name</th>
                        <th>Type</th>
                        <th>Priority</th>
                        <th>Summary</th>
                        <th>Assignee</th>
                        <th>Impacted TCs</th>
                        <th>Deferred</th>
                        <th>Average QI</th>
                        <th>Overall QI Impact</th>
                    </tr>
                </thead>
                <tbody>
                    {bugs_table_rows}
                </tbody>
            </table>
            
            <div class="note">
                <p><strong>Note:</strong> These bugs impact more than 4 test cases and have significant overall QI impact. Please review and take appropriate action.</p>
                <p><strong>Run Folder:</strong> {run_folder}</p>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
Top QI Impacting Bugs Notification

The following bugs have been identified as top QI impacting bugs on branch: {branch_name}

Total Bugs: {len(bugs)}

Bug Details:
"""
        for idx, bug in enumerate(bugs, 1):
            text_body += f"""
{idx}. {bug.get('name', 'N/A')}
   - Type: {bug.get('type', 'N/A')}
   - Priority: {bug.get('priority', 'N/A')}
   - Summary: {bug.get('summary', 'N/A')}
   - Assignee: {bug.get('assignee', 'N/A')}
   - Impacted Test Cases: {bug.get('impacted_tcs_latest_run', 0)}
   - Deferred: {bug.get('deferred', 'N/A')}
   - Average QI: {bug.get('average_qi', 0):.2f}
   - Overall QI Impact: {bug.get('overall_qi_impact', 0):.2f}
"""
        
        text_body += f"""

Note: These bugs impact more than 4 test cases and have significant overall QI impact. Please review and take appropriate action.

Run Folder: {run_folder}
        """
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = os.getenv('SMTP_FROM_EMAIL', 'regression-dashboard@nutanix.com')
        msg['To'] = ', '.join(recipients)
        
        # Add both plain text and HTML versions
        part1 = MIMEText(text_body, 'plain')
        part2 = MIMEText(html_body, 'html')
        
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email using SMTP
        smtp_server = os.getenv('SMTP_SERVER', 'smtp.nutanix.com')
        smtp_port = int(os.getenv('SMTP_PORT', '587'))
        smtp_user = os.getenv('SMTP_USER', '')
        smtp_password = os.getenv('SMTP_PASSWORD', '')
        
        try:
            # For now, we'll just log the email instead of actually sending it
            # Uncomment the SMTP code below when SMTP credentials are configured
            logger.info(f"Email prepared for {', '.join(recipients)}:")
            logger.info(f"Subject: {subject}")
            logger.info(f"Number of bugs: {len(bugs)}")
            logger.info(f"Body length: {len(text_body)} chars")
            
            # Uncomment below to actually send email:
            # with smtplib.SMTP(smtp_server, smtp_port) as server:
            #     if smtp_user and smtp_password:
            #         server.starttls()
            #         server.login(smtp_user, smtp_password)
            #     server.send_message(msg)
            
            return jsonify({
                "success": True,
                "message": f"Email prepared for {len(recipients)} recipient(s)",
                "recipients": recipients,
                "note": "Email sending is currently logged. Configure SMTP settings to enable actual email sending."
            })
        except Exception as e:
            logger.error(f"Error sending email: {e}", exc_info=True)
            return jsonify({"error": f"Failed to send email: {str(e)}"}), 500
        
    except Exception as e:
        logger.error(f"Error in send email: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ======================================================
# Run Plan Endpoints
# ======================================================

# JITA Authentication
from base64 import b64decode

def safe_b64decode(s):
    """Safely decode base64 string, adding padding if needed"""
    # Add padding if needed (base64 strings must be multiple of 4)
    missing_padding = len(s) % 4
    if missing_padding:
        s += '=' * (4 - missing_padding)
    return b64decode(s).decode("utf-8")

# Service account credentials for batch operations (matching reference script)
JITA_SVC_USERNAME = safe_b64decode("c3ZjLmNkcC5yZWdyZXNzaW9u")
JITA_SVC_PASSWORD = safe_b64decode("Knh0WTFtNiYlVko0akZXZzJlZHY=")
JITA_SVC_AUTH = (JITA_SVC_USERNAME, JITA_SVC_PASSWORD)

# Named service accounts selectable per run plan
RUN_PLAN_SERVICE_ACCOUNTS = {
    "svc.teamchandra": (
        safe_b64decode("c3ZjLnRlYW1jaGFuZHJh"),
        safe_b64decode("KndIZTM5JWZieVcmblFFTEQ1dHU="),
    ),
    "svc.cdp.regression": (JITA_SVC_USERNAME, JITA_SVC_PASSWORD),
}

# User credentials for triggering (matching reference script)
JITA_USERNAME = safe_b64decode("c3VkaGFyc2hhbi5tdXNhbGk=")
JITA_PASSWORD = safe_b64decode("V29ya291dEAy")
JITA_AUTH = (JITA_USERNAME, JITA_PASSWORD)

# Helper function to update tester_tags for job profiles
def update_job_profiles_tester_tags(job_profile_ids, tag_name, action="add"):
    """
    Update tester_tags for multiple job profiles
    action: "add" to append tag, "remove" to remove tag
    """
    updated_count = 0
    failed_updates = []
    
    def update_single_job_tags(job_id):
        try:
            # Fetch existing profile
            get_url = f"{JITA_BASE}/job_profiles/{job_id}"
            get_resp = requests.get(get_url, headers={"Content-Type": "application/json"}, auth=JITA_SVC_AUTH, verify=False, timeout=30)
            
            if get_resp.status_code != 200:
                return {"job_id": job_id, "success": False, "error": f"Failed to fetch: {get_resp.status_code}"}
            
            existing_profile = get_resp.json().get("data", {})
            if not existing_profile:
                return {"job_id": job_id, "success": False, "error": "Empty profile data"}
            
            updated_profile = existing_profile.copy()
            tester_tags = existing_profile.get("tester_tags", [])
            
            if not isinstance(tester_tags, list):
                tester_tags = []
            
            if action == "add":
                # Add tag if not already present
                if tag_name not in tester_tags:
                    tester_tags.append(tag_name)
                    updated_profile["tester_tags"] = tester_tags
                else:
                    return {"job_id": job_id, "success": True, "message": "Tag already exists"}
            elif action == "remove":
                # Remove tag if present
                if tag_name in tester_tags:
                    tester_tags.remove(tag_name)
                    updated_profile["tester_tags"] = tester_tags
                else:
                    return {"job_id": job_id, "success": True, "message": "Tag not found"}
            
            # Ensure JSON serializable
            serializable_payload = {}
            for k, v in updated_profile.items():
                if isinstance(v, (set, tuple)):
                    serializable_payload[k] = list(v)
                elif v is Ellipsis:
                    serializable_payload[k] = None
                else:
                    serializable_payload[k] = v
            
            # PUT update
            put_resp = requests.put(
                f"{JITA_BASE}/job_profiles/{job_id}",
                headers={"Content-Type": "application/json"},
                json=serializable_payload,
                auth=JITA_SVC_AUTH,
                verify=False,
                timeout=30
            )
            
            if put_resp.status_code == 200:
                resp_json = put_resp.json()
                if resp_json.get("success", True):
                    return {"job_id": job_id, "success": True}
                else:
                    return {"job_id": job_id, "success": False, "error": resp_json.get("message", "Update failed")}
            else:
                error_msg = put_resp.text[:200] if put_resp.text else f"HTTP {put_resp.status_code}"
                return {"job_id": job_id, "success": False, "error": error_msg}
        except Exception as e:
            logger.error(f"Exception updating tester_tags for {job_id}: {e}", exc_info=True)
            return {"job_id": job_id, "success": False, "error": str(e)}
    
    # Parallel execution
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(update_single_job_tags, jp_id) for jp_id in job_profile_ids]
        for future in as_completed(futures):
            result = future.result()
            if result["success"]:
                updated_count += 1
            else:
                failed_updates.append(result)
    
    return updated_count, failed_updates

@app.route("/mcp/regression/run-plan/service-accounts", methods=["GET"])
@jwt_required
def list_service_accounts():
    """Return the names of available service accounts for run plan triggers."""
    return jsonify({"service_accounts": list(RUN_PLAN_SERVICE_ACCOUNTS.keys())})


@app.route("/mcp/regression/run-plan", methods=["GET"])
@jwt_required
def list_run_plans():
    """List all run plans"""
    try:
        data = load_run_plans()
        return jsonify({"run_plans": data.get("run_plans", [])})
    except Exception as e:
        logger.error(f"Error listing run plans: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan", methods=["POST"])
@jwt_required
def create_run_plan():
    """Create a new run plan"""
    try:
        req_data = request.json
        data = load_run_plans()
        
        # Validate tag name uniqueness
        tag_name = req_data.get("tag_name")
        for rp in data.get("run_plans", []):
            if rp.get("tag_name") == tag_name:
                return jsonify({"error": f"Tag name '{tag_name}' already exists"}), 400
        
        # Validate and filter job profiles
        job_profiles = req_data.get("job_profiles", [])
        # Filter out empty strings and invalid IDs
        job_profiles = [jp_id for jp_id in job_profiles if jp_id and isinstance(jp_id, str) and jp_id.strip()]
        
        if not job_profiles:
            return jsonify({"error": "At least one valid job profile is required"}), 400
        
        # Generate tag_name automatically if not provided
        tag_name = req_data.get("tag_name")
        if not tag_name:
            # Extract branch from name (e.g., CDP_Regression_Upgrade_master -> master)
            name_parts = req_data.get("name", "").split("_")
            branch = name_parts[-1] if name_parts else "master"
            timestamp = int(time.time() * 1000)
            tag_name = f"{branch}_{timestamp}"
        
        # Create new run plan
        new_id = str(int(time.time() * 1000))
        new_run_plan = {
            "id": new_id,
            "name": req_data.get("name"),
            "branch": req_data.get("branch", ""),
            "job_profiles": job_profiles,
            "tag_name": tag_name,
            "schedule_date": req_data.get("schedule_date"),
            "schedule_triggered": False,
            "service_account": req_data.get("service_account", ""),
            "created_at": datetime.now().isoformat(),
            "last_triggered": None
        }
        
        data["run_plans"].append(new_run_plan)
        save_run_plans(data)
        
        # Append tag_name to tester_tags for all job profiles
        if tag_name and job_profiles:
            logger.info(f"Updating tester_tags for {len(job_profiles)} job profile(s) with tag: {tag_name}")
            updated_count, failed = update_job_profiles_tester_tags(job_profiles, tag_name, action="add")
            logger.info(f"Updated tester_tags: {updated_count} succeeded, {len(failed)} failed")
        
        return jsonify({"success": True, "run_plan": new_run_plan}), 201
    except Exception as e:
        logger.error(f"Error creating run plan: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>", methods=["PUT"])
@jwt_required
def update_run_plan(run_plan_id):
    """Update an existing run plan"""
    try:
        req_data = request.json
        data = load_run_plans()
        
        # Find and update run plan
        for i, rp in enumerate(data.get("run_plans", [])):
            if rp.get("id") == run_plan_id:
                # Check if already triggered (restrict edits)
                if rp.get("last_triggered"):
                    # Only allow editing schedule_date, name, branch, service_account, and tag_name
                    if "schedule_date" in req_data:
                        rp["schedule_date"] = req_data["schedule_date"]
                        rp["schedule_triggered"] = False
                    if "name" in req_data:
                        rp["name"] = req_data["name"]
                    if "branch" in req_data:
                        rp["branch"] = req_data["branch"]
                    if "service_account" in req_data:
                        rp["service_account"] = req_data["service_account"]
                    if "tag_name" in req_data:
                        # Validate uniqueness
                        tag_name = req_data["tag_name"]
                        for other_rp in data.get("run_plans", []):
                            if other_rp.get("id") != run_plan_id and other_rp.get("tag_name") == tag_name:
                                return jsonify({"error": f"Tag name '{tag_name}' already exists"}), 400
                        rp["tag_name"] = tag_name
                else:
                    # Full edit allowed before first trigger
                    rp["name"] = req_data.get("name", rp.get("name"))
                    if "branch" in req_data:
                        rp["branch"] = req_data["branch"]
                    if "service_account" in req_data:
                        rp["service_account"] = req_data["service_account"]
                    
                    # Validate and filter job profiles if provided
                    if "job_profiles" in req_data:
                        new_job_profiles = req_data.get("job_profiles", [])
                        new_job_profiles = [jp_id for jp_id in new_job_profiles if jp_id and isinstance(jp_id, str) and jp_id.strip()]
                        if not new_job_profiles:
                            return jsonify({"error": "At least one valid job profile is required"}), 400
                        rp["job_profiles"] = new_job_profiles
                    
                    if "schedule_date" in req_data:
                        rp["schedule_date"] = req_data.get("schedule_date")
                        rp["schedule_triggered"] = False
                
                save_run_plans(data)
                
                # Update tester_tags if job_profiles were updated
                # Tag name remains unchanged (auto-generated on create)
                tag_name = rp.get("tag_name")
                if "job_profiles" in req_data and tag_name:
                    job_profiles = rp.get("job_profiles", [])
                    job_profiles = [jp_id for jp_id in job_profiles if jp_id and isinstance(jp_id, str) and jp_id.strip()]
                    
                    if job_profiles:
                        logger.info(f"Ensuring tag '{tag_name}' exists in tester_tags for {len(job_profiles)} job profile(s)")
                        updated_count, failed = update_job_profiles_tester_tags(job_profiles, tag_name, action="add")
                        logger.info(f"Updated tester_tags: {updated_count} succeeded, {len(failed)} failed")
                
                return jsonify({"success": True, "run_plan": rp})
        
        return jsonify({"error": "Run plan not found"}), 404
    except Exception as e:
        logger.error(f"Error updating run plan: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/search-job-profiles", methods=["POST"])
@jwt_required
def search_job_profiles():
    """Search job profiles by ID or pattern"""
    try:
        req_data = request.json
        search_type = req_data.get("search_type")  # 'id' or 'pattern'
        search_value = req_data.get("search_value")
        
        if not search_value:
            return jsonify({"error": "Search value is required"}), 400
        
        # Build raw_query based on search type
        if search_type == "id":
            # Comma-separated IDs
            ids = [id.strip() for id in search_value.split(",")]
            raw_query = {
                "_id": {"$in": [{"$oid": id} for id in ids]}
            }
        else:  # pattern
            raw_query = {
                "name": {
                    "$regex": f"^{search_value}",
                    "$options": "i"
                }
            }
        
        # Call JITA API
        # Note: JITA API expects raw_query as a URL-encoded JSON string in query params
        from urllib.parse import quote
        raw_query_str = quote(json.dumps(raw_query))
        params = {
            "raw_query": raw_query_str,
            "limit": 100
        }
        
        response = requests.get(
            f"{JITA_BASE}/job_profiles",
            params=params,
            auth=JITA_SVC_AUTH,
            verify=False,
            timeout=30
        )
        
        if response.status_code != 200:
            return jsonify({"error": f"JITA API error: {response.status_code}"}), 500
        
        result = response.json()
        job_profiles = result.get("data", [])
        
        return jsonify({"success": True, "job_profiles": job_profiles})
    except Exception as e:
        logger.error(f"Error searching job profiles: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>/trigger", methods=["POST"])
@jwt_required
def trigger_run_plan(run_plan_id):
    """Trigger a run plan using its configured service account, or the logged-in user's LDAP credentials."""
    try:
        logger.info(f"[START] Trigger Run Plan | run_plan_id={run_plan_id}")

        data = load_run_plans()

        # Find run plan first so we can check for a service account
        run_plan = None
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                run_plan = rp
                break

        if not run_plan:
            logger.error(f"Run plan not found: {run_plan_id}")
            return jsonify({"error": "Run plan not found"}), 404

        # Resolve credentials: prefer run-plan-level service account, fall back to LDAP
        svc_name = run_plan.get("service_account", "")
        if svc_name and svc_name in RUN_PLAN_SERVICE_ACCOUNTS:
            user_auth = RUN_PLAN_SERVICE_ACCOUNTS[svc_name]
            logger.info(f"Using service account '{svc_name}' for Jita trigger")
        else:
            current_username = g.current_user.get("sub", "")
            user_auth = _get_user_credentials(current_username)
            if not user_auth:
                logger.warning(f"No cached credentials for user '{current_username}' — asking for re-auth")
                return jsonify({
                    "error": "Session credentials expired. Please re-login to trigger runs.",
                    "code": "CREDENTIALS_EXPIRED"
                }), 401
            logger.info(f"Using LDAP credentials of '{current_username}' for Jita trigger")
        
        logger.info(f"Found run plan: {run_plan.get('name')} (ID: {run_plan_id})")
        
        job_profile_ids = run_plan.get("job_profiles", [])
        logger.info(f"Original job_profiles from run plan: {job_profile_ids}")
        
        # Filter out empty strings and invalid IDs
        original_count = len(job_profile_ids)
        job_profile_ids = [jp_id for jp_id in job_profile_ids if jp_id and isinstance(jp_id, str) and jp_id.strip()]
        filtered_count = len(job_profile_ids)
        
        logger.info(f"Filtered job profiles: {original_count} -> {filtered_count} valid IDs")
        
        if not job_profile_ids:
            error_msg = f"Run plan '{run_plan.get('name')}' has no valid job profiles. Original list: {run_plan.get('job_profiles')}. Please add at least one valid job profile to the run plan."
            logger.error(error_msg)
            return jsonify({
                "error": error_msg,
                "run_plan_name": run_plan.get("name"),
                "original_job_profiles": run_plan.get("job_profiles"),
                "filtered_count": filtered_count
            }), 400
        
        logger.info(f"Triggering {filtered_count} job profile(s): {job_profile_ids}")
        
        # Trigger job profiles in parallel
        task_ids = []
        failed_jobs = []
        
        def trigger_single_job(job_id):
            try:
                if not job_id or not isinstance(job_id, str) or not job_id.strip():
                    return {"job_id": job_id, "success": False, "error": "Invalid job profile ID"}
                
                url = f"{JITA_BASE}/job_profiles/{job_id}/trigger"
                payload = {}
                headers = {"Content-Type": "application/json"}
                
                # Update NOS commit if provided (use service account for updates)
                if run_plan.get("nos_commit"):
                    # Fetch existing profile first
                    get_url = f"{JITA_BASE}/job_profiles/{job_id}"
                    get_resp = requests.get(get_url, headers=headers, auth=JITA_SVC_AUTH, verify=False, timeout=30)
                    if get_resp.status_code == 200:
                        existing_profile = get_resp.json().get("data", {})
                        if isinstance(existing_profile, dict):
                            build_selection = existing_profile.get("build_selection", {})
                            build_selection["commit_id"] = run_plan.get("nos_commit")
                            build_selection["by_commit_id"] = True
                            
                            # Update profile
                            update_payload = existing_profile.copy()
                            update_payload["build_selection"] = build_selection
                            
                            # Ensure JSON serializable
                            serializable_payload = {}
                            for k, v in update_payload.items():
                                if isinstance(v, (set, tuple)):
                                    serializable_payload[k] = list(v)
                                elif v is Ellipsis:
                                    serializable_payload[k] = None
                                else:
                                    serializable_payload[k] = v
                            
                            update_resp = requests.put(
                                f"{JITA_BASE}/job_profiles/{job_id}",
                                headers=headers,
                                json=serializable_payload,
                                auth=JITA_SVC_AUTH,
                                verify=False,
                                timeout=30
                            )
                            if update_resp.status_code != 200:
                                logger.warning(f"Failed to update commit for {job_id}: {update_resp.text[:200]}")
                
                # Trigger using the logged-in user's LDAP credentials
                logger.info(f"Triggering Job Profile ID: {job_id} as user '{user_auth[0]}'")
                resp = requests.post(
                    url,
                    headers=headers,
                    auth=user_auth,
                    json=payload,
                    verify=False,
                    timeout=60
                )
                
                if resp.status_code == 200:
                    try:
                        res_data = resp.json()
                        if res_data.get("success") and "task_ids" in res_data:
                            # Extract task IDs (matching reference script pattern)
                            ids = [item["$oid"] if isinstance(item, dict) and "$oid" in item else item for item in res_data["task_ids"]]
                            logger.info(f"✅ Triggered: {job_id} → Task ID(s): {ids}")
                            return {
                                "job_id": job_id,
                                "task_ids": ids,
                                "success": True
                            }
                        else:
                            error_msg = res_data.get("message", "Trigger failed") if isinstance(res_data, dict) else str(res_data)
                            logger.error(f"❌ Trigger failed for {job_id}: {error_msg}")
                            return {"job_id": job_id, "success": False, "error": error_msg}
                    except Exception as e:
                        logger.error(f"❌ Response parse error for {job_id}: {e}")
                        return {"job_id": job_id, "success": False, "error": f"Response parse error: {str(e)}"}
                else:
                    error_msg = resp.text[:200] if resp.text else f"HTTP {resp.status_code}"
                    logger.error(f"❌ HTTP {resp.status_code} → Failed to trigger job profile {job_id}: {error_msg}")
                    return {"job_id": job_id, "success": False, "error": f"HTTP {resp.status_code}: {error_msg}"}
            except Exception as e:
                logger.error(f"Exception in trigger_single_job for {job_id}: {e}", exc_info=True)
                return {"job_id": job_id, "success": False, "error": str(e)}
        
        # Parallel execution
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(trigger_single_job, jp_id) for jp_id in job_profile_ids]
            for future in as_completed(futures):
                result = future.result()
                if result["success"]:
                    task_ids.extend(result["task_ids"])
                else:
                    failed_jobs.append(result)
        
        # Update run plan
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                rp["last_triggered"] = datetime.now().isoformat()
                break
        
        # Save history
        history_entry = {
            "id": str(int(time.time() * 1000)),
            "run_plan_id": run_plan_id,
            "triggered_at": datetime.now().isoformat(),
            "triggered_by": current_username,
            "task_ids": task_ids,
            "failed_jobs": failed_jobs,
            "status": "success" if not failed_jobs else "partial"
        }
        
        if "history" not in data:
            data["history"] = []
        data["history"].append(history_entry)
        
        save_run_plans(data)
        
        logger.info(f"[END] Trigger Run Plan | run_plan_id={run_plan_id} | task_ids={len(task_ids)} | failed={len(failed_jobs)}")
        
        return jsonify({
            "success": True,
            "triggered_by": current_username,
            "task_ids": task_ids,
            "failed_jobs": failed_jobs,
            "total_triggered": len(task_ids),
            "total_failed": len(failed_jobs)
        })
    except Exception as e:
        logger.error(f"Error triggering run plan {run_plan_id}: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>/batch-update", methods=["POST"])
@jwt_required
def batch_update_job_profiles(run_plan_id):
    """Batch update job profiles in a run plan"""
    try:
        req_data = request.json
        data = load_run_plans()
        
        # Find run plan
        run_plan = None
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                run_plan = rp
                break
        
        if not run_plan:
            return jsonify({"error": "Run plan not found"}), 404
        
        job_profile_ids = run_plan.get("job_profiles", [])
        # Filter out empty strings and invalid IDs
        job_profile_ids = [jp_id for jp_id in job_profile_ids if jp_id and isinstance(jp_id, str) and jp_id.strip()]
        
        if not job_profile_ids:
            return jsonify({"error": "No valid job profiles in run plan"}), 400
        
        # Handle new components array structure or legacy single component structure
        components = req_data.get("components", [])
        if not components:
            # Legacy support: single component update
            component = req_data.get("component")
            if component:
                components = [{
                    "component": component,
                    "branch": req_data.get("branch", ""),
                    "update_type": req_data.get("update_type", ""),
                    "build_type": req_data.get("build_type", ""),
                    "tag": req_data.get("tag", ""),
                    "commit_id": req_data.get("commit_id", ""),
                    "gbn": req_data.get("gbn", "")
                }]
        
        updated_count = 0
        failed_updates = []
        
        def update_single_job(job_id):
            try:
                # Fetch existing profile (use service account for fetching)
                get_url = f"{JITA_BASE}/job_profiles/{job_id}"
                get_resp = requests.get(get_url, headers={"Content-Type": "application/json"}, auth=JITA_SVC_AUTH, verify=False, timeout=30)
                
                if get_resp.status_code != 200:
                    return {"job_id": job_id, "success": False, "error": f"Failed to fetch: {get_resp.status_code}"}
                
                existing_profile = get_resp.json().get("data", {})
                if not existing_profile:
                    return {"job_id": job_id, "success": False, "error": "Empty profile data"}
                
                # Start with full existing profile (like reference script)
                updated_profile = existing_profile.copy()
                
                # Process each component update
                for comp_data in components:
                    component = comp_data.get("component")
                    branch = comp_data.get("branch", "")
                    update_type = comp_data.get("update_type", "")
                    build_type = comp_data.get("build_type", "")
                    tag = comp_data.get("tag", "")
                    commit_id = comp_data.get("commit_id", "")
                    gbn = comp_data.get("gbn", "")
                    
                    if component == "NOS_CLUSTER":
                        # Update git.branch (always update if branch is provided, or preserve existing)
                        git = existing_profile.get("git", {})
                        if not git:
                            git = {"repo": "main"}
                        if branch:
                            git["branch"] = branch
                        git["repo"] = "main"  # Always ensure repo is "main"
                        updated_profile["git"] = git
                        
                        # Update build_selection (always update if update_type or build_type is provided)
                        if update_type or build_type:
                            build_selection = existing_profile.get("build_selection", {})
                            if not build_selection:
                                build_selection = {}
                            
                            # Always set build_type if provided
                            if build_type:
                                build_selection["build_type"] = build_type
                            
                            if update_type == "tag":
                                # By tag - always set these flags (reference shows they're always set)
                                build_selection["commit_must_be_newer"] = False
                                build_selection["by_latest_smoked"] = True
                                # Remove commit-related fields if they exist
                                build_selection.pop("by_commit_id", None)
                                build_selection.pop("commit_id", None)
                                build_selection.pop("gbn", None)
                            elif update_type == "commit":
                                # By commit - always set by_commit_id flag
                                build_selection["by_commit_id"] = True
                                # Remove tag-related fields if they exist
                                build_selection.pop("commit_must_be_newer", None)
                                build_selection.pop("by_latest_smoked", None)
                                
                                # Set commit_id and gbn if provided
                                if commit_id:
                                    build_selection["commit_id"] = commit_id
                                if gbn:
                                    # GBN should be an integer
                                    try:
                                        build_selection["gbn"] = int(gbn) if isinstance(gbn, str) else gbn
                                    except (ValueError, TypeError):
                                        build_selection["gbn"] = gbn
                            
                            updated_profile["build_selection"] = build_selection
                    
                    elif component == "PRISM_CENTRAL":
                        # Update PRISM_CENTRAL (following reference script pattern)
                        # Initialize resource_manager_json structure
                        resource_manager_json = existing_profile.get("resource_manager_json", {})
                        if not resource_manager_json:
                            resource_manager_json = {}
                        
                        PRISM_CENTRAL = resource_manager_json.get("PRISM_CENTRAL", {})
                        if not PRISM_CENTRAL:
                            PRISM_CENTRAL = {}
                        
                        PC_BUILD = PRISM_CENTRAL.get("build", {})
                        if not PC_BUILD:
                            PC_BUILD = {}
                        
                        # Update branch and component (if branch is provided)
                        if branch:
                            PC_BUILD["branch"] = branch
                        PC_BUILD["component"] = "main"  # Always set to "main"
                        
                        # Update build selection based on update_type
                        if update_type == "tag":
                            if tag:
                                PC_BUILD["build_selection_option"] = tag
                        elif update_type == "commit":
                            # For PRISM_CENTRAL, by commit uses build_selection_option for commit_id
                            if commit_id:
                                PC_BUILD["build_selection_option"] = commit_id
                            if gbn:
                                # GBN should be an integer
                                try:
                                    PC_BUILD["gbn"] = int(gbn) if isinstance(gbn, str) else gbn
                                except (ValueError, TypeError):
                                    PC_BUILD["gbn"] = gbn
                        
                        # Update build_selection_build_type if build_type is provided
                        if build_type:
                            PC_BUILD["build_selection_build_type"] = build_type
                        
                        # Always update PRISM_CENTRAL structure if component is selected
                        # This ensures the structure is properly initialized even if fields are optional
                        PRISM_CENTRAL["build"] = PC_BUILD
                        resource_manager_json["PRISM_CENTRAL"] = PRISM_CENTRAL
                        updated_profile["resource_manager_json"] = resource_manager_json
                
                # Update test framework branch if provided
                if req_data.get("nutest_branch"):
                    updated_profile["nutest-py3-tests_branch"] = req_data.get("nutest_branch")
                
                # Update test framework metadata (patch URLs and branch)
                if req_data.get("nutest_branch") or req_data.get("patch_url") or req_data.get("framework_patch_url"):
                    test_framework_metadata = existing_profile.get("test_framework_metadata", {})
                    if not test_framework_metadata:
                        test_framework_metadata = {"framework": {}, "test": {}}
                    
                    # Get existing metadata or create new (preserve existing values)
                    test_metadata = test_framework_metadata.get("test", {})
                    if not test_metadata:
                        test_metadata = {}
                    else:
                        # Preserve existing test metadata (branch, commit, etc.)
                        test_metadata = test_metadata.copy()
                    
                    framework_metadata = test_framework_metadata.get("framework", {})
                    if not framework_metadata:
                        framework_metadata = {}
                    else:
                        # Preserve existing framework metadata (branch, commit, etc.)
                        framework_metadata = framework_metadata.copy()
                    
                    # Update branch in both test and framework if nutest_branch is provided
                    if req_data.get("nutest_branch"):
                        test_metadata["branch"] = req_data.get("nutest_branch")
                        framework_metadata["branch"] = req_data.get("nutest_branch")
                    else:
                        # Preserve existing branch if not updating
                        if "branch" not in framework_metadata:
                            # Get from existing if available
                            existing_framework = test_framework_metadata.get("framework", {})
                            if existing_framework and "branch" in existing_framework:
                                framework_metadata["branch"] = existing_framework["branch"]
                    
                    # Update test patch URL if provided
                    if req_data.get("patch_url"):
                        test_metadata["patch_url"] = req_data.get("patch_url")
                    
                    # Update framework patch URL if provided
                    framework_patch_url = req_data.get("framework_patch_url")
                    if framework_patch_url:
                        framework_metadata["patch_url"] = framework_patch_url
                        logger.info(f"Updating framework patch_url to: {framework_patch_url}")
                    
                    # Ensure commit is preserved (set to null if not present, or preserve existing)
                    if "commit" not in framework_metadata:
                        existing_framework = test_framework_metadata.get("framework", {})
                        if existing_framework and "commit" in existing_framework:
                            framework_metadata["commit"] = existing_framework["commit"]
                        else:
                            framework_metadata["commit"] = None
                    if "commit" not in test_metadata:
                        existing_test = test_framework_metadata.get("test", {})
                        if existing_test and "commit" in existing_test:
                            test_metadata["commit"] = existing_test["commit"]
                        else:
                            test_metadata["commit"] = None
                    
                    test_framework_metadata["test"] = test_metadata
                    test_framework_metadata["framework"] = framework_metadata
                    updated_profile["test_framework_metadata"] = test_framework_metadata
                    logger.info(f"Updated test_framework_metadata: framework.patch_url={framework_metadata.get('patch_url')}, framework.branch={framework_metadata.get('branch')}")
                
                # Update top-level patch_url to match framework_patch_url (JITA uses this field)
                if req_data.get("framework_patch_url"):
                    updated_profile["patch_url"] = req_data.get("framework_patch_url")
                
                # Update tester_tags if provided (optional batch update)
                if req_data.get("tester_tags_action"):  # "add" or "remove"
                    tester_tags_action = req_data.get("tester_tags_action")
                    tester_tag_value = req_data.get("tester_tag_value", "")
                    tester_tags_to_remove = req_data.get("tester_tags_to_remove", [])
                    
                    tester_tags = existing_profile.get("tester_tags", [])
                    if not isinstance(tester_tags, list):
                        tester_tags = []
                    
                    if tester_tags_action == "add" and tester_tag_value:
                        if tester_tag_value not in tester_tags:
                            tester_tags.append(tester_tag_value)
                            updated_profile["tester_tags"] = tester_tags
                    elif tester_tags_action == "remove":
                        if isinstance(tester_tags_to_remove, list) and tester_tags_to_remove:
                            original_len = len(tester_tags)
                            tester_tags = [t for t in tester_tags if t not in tester_tags_to_remove]
                            if len(tester_tags) != original_len:
                                updated_profile["tester_tags"] = tester_tags
                                logger.info(f"Removed {original_len - len(tester_tags)} tag(s) from job profile {job_id}")
                        elif tester_tag_value:
                            if tester_tag_value in tester_tags:
                                tester_tags.remove(tester_tag_value)
                                updated_profile["tester_tags"] = tester_tags
                
                # Overwrite run_tests_with_additional_tags if provided in the request
                if "run_tests_with_additional_tags" in req_data:
                    new_additional_tags = req_data["run_tests_with_additional_tags"]
                    if not isinstance(new_additional_tags, list):
                        new_additional_tags = []
                    updated_profile["run_tests_with_additional_tags"] = new_additional_tags
                    logger.info(f"Overwriting run_tests_with_additional_tags for {job_id}: {new_additional_tags}")
                
                # Ensure JSON serializable (following reference script pattern)
                serializable_payload = {}
                for k, v in updated_profile.items():
                    if isinstance(v, (set, tuple)):
                        serializable_payload[k] = list(v)
                    elif v is Ellipsis:
                        serializable_payload[k] = None
                    else:
                        serializable_payload[k] = v
                
                # PUT update (use service account for batch updates)
                put_resp = requests.put(
                    f"{JITA_BASE}/job_profiles/{job_id}",
                    headers={"Content-Type": "application/json"},
                    json=serializable_payload,
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
                
                if put_resp.status_code == 200:
                    resp_json = put_resp.json()
                    if resp_json.get("success", True):
                        logger.info(f"Successfully updated job profile {job_id}")
                        return {"job_id": job_id, "success": True}
                    else:
                        error_msg = resp_json.get("message", "Update failed")
                        logger.error(f"Failed to update job profile {job_id}: {error_msg}")
                        return {"job_id": job_id, "success": False, "error": error_msg}
                else:
                    error_msg = put_resp.text[:500] if put_resp.text else f"HTTP {put_resp.status_code}"
                    logger.error(f"Failed to update job profile {job_id}: {error_msg}")
                    return {"job_id": job_id, "success": False, "error": error_msg}
            except Exception as e:
                logger.error(f"Exception updating job profile {job_id}: {e}", exc_info=True)
                return {"job_id": job_id, "success": False, "error": str(e)}
        
        # Parallel execution
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(update_single_job, jp_id) for jp_id in job_profile_ids]
            for future in as_completed(futures):
                result = future.result()
                if result["success"]:
                    updated_count += 1
                else:
                    failed_updates.append(result)
        
        return jsonify({
            "success": True,
            "updated_count": updated_count,
            "failed_updates": failed_updates
        })
    except Exception as e:
        logger.error(f"Error batch updating job profiles: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>/history", methods=["GET"])
@jwt_required
def get_run_plan_history(run_plan_id):
    """Get history for a run plan"""
    try:
        data = load_run_plans()
        history = [
            entry for entry in data.get("history", [])
            if entry.get("run_plan_id") == run_plan_id
        ]
        # Sort by triggered_at descending
        history.sort(key=lambda x: x.get("triggered_at", ""), reverse=True)
        return jsonify({"history": history})
    except Exception as e:
        logger.error(f"Error fetching history: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/history/<history_id>/retry", methods=["POST"])
@jwt_required
def retry_history_entry(history_id):
    """Retry a history entry trigger"""
    try:
        data = load_run_plans()
        
        # Find history entry
        history_entry = None
        for entry in data.get("history", []):
            if entry.get("id") == history_id:
                history_entry = entry
                break
        
        if not history_entry:
            return jsonify({"error": "History entry not found"}), 404
        
        # Trigger the run plan again
        run_plan_id = history_entry.get("run_plan_id")
        return trigger_run_plan(run_plan_id)
    except Exception as e:
        logger.error(f"Error retrying history entry: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/history/<history_id>", methods=["DELETE"])
@jwt_required
def delete_history_entry(history_id):
    """Delete a history entry"""
    try:
        data = load_run_plans()
        
        # Remove history entry
        data["history"] = [
            entry for entry in data.get("history", [])
            if entry.get("id") != history_id
        ]
        
        save_run_plans(data)
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Error deleting history entry: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/history/<history_id>/kill", methods=["POST"])
@jwt_required
def kill_history_tasks(history_id):
    """Kill (abort) all JITA tasks associated with a history entry."""
    try:
        current_username = g.current_user.get("sub", "")
        user_auth = _get_user_credentials(current_username)
        if not user_auth:
            return jsonify({
                "error": "Session credentials expired. Please re-login to kill tasks.",
                "code": "CREDENTIALS_EXPIRED"
            }), 401

        data = load_run_plans()
        history_entry = None
        for entry in data.get("history", []):
            if entry.get("id") == history_id:
                history_entry = entry
                break

        if not history_entry:
            return jsonify({"error": "History entry not found"}), 404

        task_ids = history_entry.get("task_ids", [])
        if not task_ids:
            return jsonify({"error": "No task IDs to kill"}), 400

        killed = []
        failed = []

        def kill_single_task(tid):
            try:
                resp = requests.put(
                    f"{JITA_BASE}/tasks/{tid}",
                    headers={"Content-Type": "application/json"},
                    json={"status": "killed"},
                    auth=user_auth,
                    verify=False,
                    timeout=30
                )
                if resp.status_code == 200:
                    return {"task_id": tid, "success": True}
                return {"task_id": tid, "success": False, "error": f"HTTP {resp.status_code}: {resp.text[:200]}"}
            except Exception as exc:
                return {"task_id": tid, "success": False, "error": str(exc)}

        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(kill_single_task, tid) for tid in task_ids]
            for future in as_completed(futures):
                result = future.result()
                if result["success"]:
                    killed.append(result["task_id"])
                else:
                    failed.append(result)

        if len(killed) == len(task_ids):
            history_entry["status"] = "killed"
        elif killed:
            history_entry["status"] = "partially_killed"
        save_run_plans(data)

        logger.info(f"Kill tasks for history {history_id}: {len(killed)} killed, {len(failed)} failed (by {current_username})")
        return jsonify({
            "success": len(failed) == 0,
            "killed": killed,
            "failed": failed,
            "total_killed": len(killed),
            "total_failed": len(failed),
            "status": history_entry["status"]
        })
    except Exception as e:
        logger.error(f"Error killing tasks for history {history_id}: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/run-plan/<run_plan_id>/clone", methods=["POST"])
@jwt_required
def clone_run_plan(run_plan_id):
    """Clone a run plan with a new unique tag_name"""
    try:
        data = load_run_plans()
        
        # Find run plan to clone
        source_run_plan = None
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                source_run_plan = rp
                break
        
        if not source_run_plan:
            return jsonify({"error": "Run plan not found"}), 404
        
        # Generate new unique tag_name
        # Extract branch from name (e.g., CDP_Regression_Upgrade_master -> master)
        name_parts = source_run_plan.get("name", "").split("_")
        branch = name_parts[-1] if name_parts else "master"
        timestamp = int(time.time() * 1000)
        new_tag_name = f"{branch}_{timestamp}"
        
        # Check for uniqueness (should be unique due to timestamp, but double-check)
        for rp in data.get("run_plans", []):
            if rp.get("tag_name") == new_tag_name:
                # If somehow duplicate, add random suffix
                new_tag_name = f"{branch}_{timestamp}_{random.randint(1000, 9999)}"
                break
        
        # Create cloned run plan
        new_id = str(int(time.time() * 1000))
        original_name = source_run_plan.get("name", "")
        cloned_name = f"{original_name}_clone" if original_name else "cloned_run_plan"
        
        cloned_run_plan = {
            "id": new_id,
            "name": cloned_name,
            "branch": source_run_plan.get("branch", ""),
            "job_profiles": source_run_plan.get("job_profiles", []).copy(),
            "tag_name": new_tag_name,
            "schedule_date": source_run_plan.get("schedule_date"),
            "schedule_triggered": False,
            "service_account": source_run_plan.get("service_account", ""),
            "created_at": datetime.now().isoformat(),
            "last_triggered": None
        }
        
        data["run_plans"].append(cloned_run_plan)
        save_run_plans(data)
        
        # Append new tag_name to tester_tags for all job profiles
        job_profiles = cloned_run_plan.get("job_profiles", [])
        job_profiles = [jp_id for jp_id in job_profiles if jp_id and isinstance(jp_id, str) and jp_id.strip()]
        
        if new_tag_name and job_profiles:
            logger.info(f"Updating tester_tags for {len(job_profiles)} job profile(s) with new tag: {new_tag_name}")
            updated_count, failed = update_job_profiles_tester_tags(job_profiles, new_tag_name, action="add")
            logger.info(f"Updated tester_tags: {updated_count} succeeded, {len(failed)} failed")
        
        return jsonify({
            "success": True,
            "run_plan": cloned_run_plan,
            "message": f"Run plan cloned successfully with new tag: {new_tag_name}"
        })
        
    except Exception as e:
        logger.error(f"Error cloning run plan: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>", methods=["DELETE"])
@jwt_required
def delete_run_plan(run_plan_id):
    """Delete a run plan and all its associated history entries"""
    try:
        data = load_run_plans()
        
        # Find run plan
        run_plan = None
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                run_plan = rp
                break
        
        if not run_plan:
            return jsonify({"error": "Run plan not found"}), 404
        
        # Remove run plan from list
        data["run_plans"] = [
            rp for rp in data.get("run_plans", [])
            if rp.get("id") != run_plan_id
        ]
        
        # Remove all history entries associated with this run plan
        data["history"] = [
            entry for entry in data.get("history", [])
            if entry.get("run_plan_id") != run_plan_id
        ]
        
        save_run_plans(data)
        logger.info(f"Deleted run plan: {run_plan.get('name')} (ID: {run_plan_id})")
        
        return jsonify({"success": True, "message": f"Run plan '{run_plan.get('name')}' deleted successfully"})
    except Exception as e:
        logger.error(f"Error deleting run plan: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/<run_plan_id>/delete-tag", methods=["POST"])
@jwt_required
def delete_tag_from_job_profiles(run_plan_id):
    """Delete a tag from tester_tags of all job profiles in a run plan"""
    try:
        req_data = request.json
        tag_name = req_data.get("tag_name")
        
        if not tag_name:
            return jsonify({"error": "Tag name is required"}), 400
        
        data = load_run_plans()
        
        # Find run plan
        run_plan = None
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                run_plan = rp
                break
        
        if not run_plan:
            return jsonify({"error": "Run plan not found"}), 404
        
        job_profile_ids = run_plan.get("job_profiles", [])
        # Filter out empty strings and invalid IDs
        job_profile_ids = [jp_id for jp_id in job_profile_ids if jp_id and isinstance(jp_id, str) and jp_id.strip()]
        
        if not job_profile_ids:
            return jsonify({"error": "No valid job profiles in run plan"}), 400
        
        logger.info(f"Removing tag '{tag_name}' from tester_tags for {len(job_profile_ids)} job profile(s)")
        updated_count, failed = update_job_profiles_tester_tags(job_profile_ids, tag_name, action="remove")
        
        return jsonify({
            "success": True,
            "updated_count": updated_count,
            "failed_updates": failed,
            "tag_name": tag_name
        })
    except Exception as e:
        logger.error(f"Error deleting tag: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/run-plan/tags", methods=["GET"])
@jwt_required
def get_available_tags():
    """Get list of available tags from JITA"""
    try:
        # Use logged-in user's credentials, fall back to service account for read-only ops
        current_username = g.current_user.get("sub", "")
        auth_creds = _get_user_credentials(current_username) or JITA_SVC_AUTH

        # Fetch recent tasks to get unique tags
        params = {
            "limit": 1000,
            "only": "tester_tags"
        }
        
        response = session.get(
            f"{JITA_BASE}/tasks",
            params=params,
            auth=auth_creds,
            timeout=30
        )
        
        if response.status_code != 200:
            return jsonify({"error": f"JITA API error: {response.status_code}"}), 500
        
        result = response.json()
        tasks = result.get("data", [])
        
        # Extract unique tags
        tags_set = set()
        for task in tasks:
            task_tags = task.get("tester_tags", [])
            tags_set.update(task_tags)
        
        tags_list = sorted(list(tags_set))
        return jsonify({"tags": tags_list})
    except Exception as e:
        logger.error(f"Error fetching tags: {e}")
        return jsonify({"error": str(e)}), 500

# ======================================================
# Run Plan Bulk (Category) Actions
# ======================================================
@app.route("/mcp/regression/run-plan/bulk-trigger", methods=["POST"])
@jwt_required
def bulk_trigger_by_branch():
    """Trigger all run plans that belong to a given branch."""
    try:
        current_username = g.current_user.get("sub", "")
        user_auth = _get_user_credentials(current_username)
        if not user_auth:
            return jsonify({"error": "Session credentials expired. Please re-login.", "code": "CREDENTIALS_EXPIRED"}), 401

        branch = (request.json or {}).get("branch", "")
        if not branch:
            return jsonify({"error": "branch is required"}), 400

        data = load_run_plans()
        targets = [rp for rp in data.get("run_plans", []) if rp.get("branch") == branch]
        if not targets:
            return jsonify({"error": f"No run plans found for branch '{branch}'"}), 404

        results = []
        for rp in targets:
            rp_id = rp["id"]
            job_profile_ids = [jp for jp in rp.get("job_profiles", []) if jp and isinstance(jp, str) and jp.strip()]
            task_ids = []
            failed_jobs = []
            for jp_id in job_profile_ids:
                try:
                    resp = requests.post(
                        f"{JITA_BASE}/job_profiles/{jp_id}/trigger",
                        json={}, headers={"Content-Type": "application/json"},
                        auth=user_auth, verify=False, timeout=60
                    )
                    if resp.status_code == 200:
                        res_data = resp.json()
                        if res_data.get("success") and "task_ids" in res_data:
                            ids = [item["$oid"] if isinstance(item, dict) and "$oid" in item else item for item in res_data["task_ids"]]
                            task_ids.extend(ids)
                        else:
                            failed_jobs.append({"job_id": jp_id, "error": res_data.get("message", "unknown")})
                    else:
                        failed_jobs.append({"job_id": jp_id, "error": f"HTTP {resp.status_code}"})
                except Exception as exc:
                    failed_jobs.append({"job_id": jp_id, "error": str(exc)})

            rp["last_triggered"] = datetime.now().isoformat()
            if "history" not in data:
                data["history"] = []
            data["history"].append({
                "id": str(int(time.time() * 1000)),
                "run_plan_id": rp_id,
                "triggered_at": rp["last_triggered"],
                "triggered_by": current_username,
                "task_ids": task_ids,
                "failed_jobs": failed_jobs,
                "status": "success" if not failed_jobs else "partial"
            })
            results.append({"run_plan_id": rp_id, "name": rp.get("name"), "task_ids": task_ids, "failed": len(failed_jobs)})

        save_run_plans(data)
        return jsonify({"success": True, "branch": branch, "results": results})
    except Exception as e:
        logger.error(f"Error in bulk trigger: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/run-plan/bulk-schedule", methods=["POST"])
@jwt_required
def bulk_schedule_by_branch():
    """Set the same schedule_date on every run plan in a branch."""
    try:
        req = request.json or {}
        branch = req.get("branch", "")
        schedule_date = req.get("schedule_date", "")
        if not branch or not schedule_date:
            return jsonify({"error": "branch and schedule_date are required"}), 400

        data = load_run_plans()
        updated = 0
        for rp in data.get("run_plans", []):
            if rp.get("branch") == branch:
                rp["schedule_date"] = schedule_date
                rp["schedule_triggered"] = False
                updated += 1
        if updated == 0:
            return jsonify({"error": f"No run plans found for branch '{branch}'"}), 404

        save_run_plans(data)
        return jsonify({"success": True, "branch": branch, "updated": updated})
    except Exception as e:
        logger.error(f"Error in bulk schedule: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# Run Plan Calendar Endpoints
# ======================================================
@app.route("/mcp/regression/run-plan/calendar", methods=["GET"])
@jwt_required
def run_plan_calendar():
    """Return all triggered and scheduled events grouped by date for calendar view."""
    try:
        data = load_run_plans()
        run_plans = data.get("run_plans", [])
        history = data.get("history", [])

        rp_map = {rp["id"]: rp for rp in run_plans}
        events = []

        for entry in history:
            triggered_at = entry.get("triggered_at", "")
            date_str = triggered_at[:10] if len(triggered_at) >= 10 else ""
            rp = rp_map.get(entry.get("run_plan_id", ""))
            events.append({
                "type": "triggered",
                "date": date_str,
                "datetime": triggered_at,
                "run_plan_id": entry.get("run_plan_id"),
                "run_plan_name": rp.get("name") if rp else "Deleted Run Plan",
                "status": entry.get("status"),
                "triggered_by": entry.get("triggered_by", ""),
                "task_ids": entry.get("task_ids", []),
                "history_id": entry.get("id"),
            })

        for rp in run_plans:
            sched = rp.get("schedule_date")
            if not sched:
                continue
            date_str = sched[:10] if len(sched) >= 10 else ""
            events.append({
                "type": "scheduled",
                "date": date_str,
                "datetime": sched,
                "run_plan_id": rp["id"],
                "run_plan_name": rp.get("name", ""),
                "schedule_triggered": rp.get("schedule_triggered", False),
                "last_triggered": rp.get("last_triggered"),
            })

        return jsonify({"events": events, "run_plans": run_plans})
    except Exception as e:
        logger.error(f"Error building calendar data: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/run-plan/<run_plan_id>/schedule", methods=["PUT", "DELETE"])
@jwt_required
def schedule_run_plan(run_plan_id):
    """Set, update, or clear the schedule_date on an existing run plan."""
    try:
        data = load_run_plans()
        for rp in data.get("run_plans", []):
            if rp.get("id") == run_plan_id:
                if request.method == "DELETE":
                    rp["schedule_date"] = None
                    rp["schedule_triggered"] = False
                    save_run_plans(data)
                    return jsonify({"success": True, "run_plan": rp})

                req_data = request.json or {}
                schedule_date = req_data.get("schedule_date")
                if schedule_date is None:
                    rp["schedule_date"] = None
                    rp["schedule_triggered"] = False
                else:
                    if not schedule_date:
                        return jsonify({"error": "schedule_date is required (or send null to clear)"}), 400
                    rp["schedule_date"] = schedule_date
                    rp["schedule_triggered"] = False
                save_run_plans(data)
                return jsonify({"success": True, "run_plan": rp})
        return jsonify({"error": "Run plan not found"}), 404
    except Exception as e:
        logger.error(f"Error scheduling run plan: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# Triage Genie Endpoints
# ======================================================
@app.route("/mcp/regression/triage-genie/jobs", methods=["GET"])
@jwt_required
def list_triage_genie_jobs():
    """List all Triage Genie jobs - primarily from JSON file"""
    try:
        # Load stored jobs from JSON file (primary source)
        stored_data = load_triage_genie_jobs()
        stored_jobs = stored_data.get("jobs", [])
        
        # Optionally fetch from API to update status for existing jobs
        # But prioritize stored jobs
        try:
            page = request.args.get("page", "1")
            per_page = request.args.get("per_page", "10")
            run_status = request.args.get("run_status", "")
            show_all = request.args.get("show_all", "true")
            name_search = request.args.get("name_search", "")
            
            timestamp = int(time.time() * 1000)
            url = f"{TRIAGE_GENIE_BASE}/jobs?page={page}&per_page={per_page}&run_status={run_status}&show_all={show_all}&name_search={name_search}&_={timestamp}"
            
            tg_session = create_triage_genie_session()
            if tg_session:
                response = tg_session.get(url, verify=False, timeout=30)
            else:
                response = requests.get(url, verify=False, timeout=30)
            
            if response.status_code == 200:
                api_data = response.json()
                api_jobs = api_data.get("data", [])
                # Create a map of API jobs by ID for quick lookup
                api_jobs_map = {job.get("id"): job for job in api_jobs if job.get("id")}
                
                # Update stored jobs with latest status from API if available
                for stored_job in stored_jobs:
                    job_id = stored_job.get("id")
                    if job_id in api_jobs_map:
                        api_job = api_jobs_map[job_id]
                        # Update status fields but keep our stored data (name, jita_task_ids, etc.)
                        stored_job["run_status"] = api_job.get("run_status")
                        stored_job["triage_status"] = api_job.get("triage_status")
                        stored_job["last_check_time"] = api_job.get("last_check_time")
                        stored_job["last_check_status"] = api_job.get("last_check_status")
                        stored_job["last_check_triage_status"] = api_job.get("last_check_triage_status")
                        stored_job["last_check_review_status"] = api_job.get("last_check_review_status")
        except Exception as api_error:
            logger.warning(f"Failed to fetch from Triage Genie API: {api_error}, using stored jobs only")
        
        # Sort by ID descending (newest first)
        stored_jobs.sort(key=lambda x: x.get("id", 0), reverse=True)
        
        return jsonify({
            "success": True,
            "jobs": stored_jobs,
            "total": len(stored_jobs)
        })
            
    except Exception as e:
        logger.error(f"Error listing Triage Genie jobs: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

@app.route("/mcp/regression/triage-genie/jobs", methods=["POST"])
@jwt_required
def create_triage_genie_job():
    """Create a new Triage Genie job"""
    try:
        req_data = request.json
        name = req_data.get("name")
        jita_task_ids = req_data.get("jita_task_ids")  # Comma-separated string
        skip_review = req_data.get("skip_review", False)
        created_by = req_data.get("created_by", "")
        
        if not name:
            return jsonify({"error": "Name is required"}), 400
        
        if not jita_task_ids:
            return jsonify({"error": "JITA task IDs are required"}), 400
        
        # Build payload
        payload = {
            "name": name,
            "jita_task_ids": jita_task_ids,
            "skip_review": skip_review,
            "created_by": created_by
        }
        
        tg_session = create_triage_genie_session()
        if tg_session:
            response = tg_session.post(
                f"{TRIAGE_GENIE_BASE}/jobs",
                json=payload,
                verify=False,
                timeout=30,
            )
        else:
            response = requests.post(
                f"{TRIAGE_GENIE_BASE}/jobs",
                json=payload,
                verify=False,
                timeout=30,
            )
        
        if response.status_code in [200, 201]:
            data = response.json()
            # Handle different response structures
            if isinstance(data, dict):
                # Check if job is nested in response
                job_data = data.get("data") or data.get("job") or data
            else:
                job_data = {}
            
            # Always set name from request (ensure it's always present in stored data)
            # The name from the form should always be saved, even if API returns a different one
            job_data["name"] = name
            
            # Always set created_by from request
            if created_by:
                job_data["created_by"] = created_by
            
            # Always set jita_task_ids from request
            job_data["jita_task_ids"] = jita_task_ids
            
            # Convert jita_task_ids string to list and ensure jita_task_id_list is always set
            # Use API response if available, otherwise create from request
            if "jita_task_id_list" in job_data and job_data.get("jita_task_id_list"):
                # API already provided the list, use it
                pass
            else:
                # Create list from request jita_task_ids
                if isinstance(jita_task_ids, str):
                    job_data["jita_task_id_list"] = [tid.strip() for tid in jita_task_ids.split(",") if tid.strip()]
                elif isinstance(jita_task_ids, list):
                    job_data["jita_task_id_list"] = jita_task_ids
                else:
                    job_data["jita_task_id_list"] = []
            
            # Add created_at timestamp if not present
            if "create_time" not in job_data:
                job_data["create_time"] = datetime.now().isoformat()
            
            # Ensure skip_review is set
            job_data["skip_review"] = skip_review
            
            # Store job in JSON file
            stored_data = load_triage_genie_jobs()
            stored_jobs = stored_data.get("jobs", [])
            
            # Check if job already exists (by ID or name)
            job_id = job_data.get("id")
            job_name = job_data.get("name")
            
            # Remove existing job with same ID or name
            stored_jobs = [j for j in stored_jobs if j.get("id") != job_id and j.get("name") != job_name]
            
            # Add new job
            stored_jobs.append(job_data)
            stored_data["jobs"] = stored_jobs
            save_triage_genie_jobs(stored_data)
            
            logger.info(f"Triage Genie job created and stored: ID={job_id}, Name={job_name}")
            
            return jsonify({
                "success": True,
                "job": job_data
            })
        else:
            logger.error(f"Triage Genie API error: {response.status_code} - {response.text}")
            return jsonify({"error": f"Failed to create job: {response.status_code} - {response.text}"}), response.status_code
            
    except Exception as e:
        logger.error(f"Error creating Triage Genie job: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500

# ======================================================
# Failed Testcase Analysis Endpoint
# ======================================================

# Constants for Jira and Glean APIs
JIRA_BASE = "https://jira.nutanix.com/rest/api/2"
GLEAN_BASE = "https://nutanix-be.glean.com/api/v1"

def get_jira_headers():
    """Get Jira API headers with authentication"""
    jira_token = os.getenv("JIRA_TOKEN", "")
    return {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {jira_token}" if jira_token else ""
    }

def get_glean_headers():
    """Get Glean API headers with authentication"""
    glean_token = os.getenv("GLEAN_TOKEN", "")
    return {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {glean_token}" if glean_token else ""
    }

def fetch_jira_ticket(ticket_id):
    """Fetch Jira ticket details"""
    try:
        if not os.getenv("JIRA_TOKEN"):
            logger.warning("JIRA_TOKEN not set, skipping Jira API call")
            return None
        
        headers = get_jira_headers()
        resp = session.get(
            f"{JIRA_BASE}/issue/{ticket_id}",
            headers=headers,
            timeout=30
        )
        if resp.status_code == 200:
            return resp.json()
        else:
            logger.warning(f"Failed to fetch Jira ticket {ticket_id}: {resp.status_code}")
            return None
    except Exception as e:
        logger.warning(f"Error fetching Jira ticket {ticket_id}: {e}")
        return None

def search_glean(query_text):
    """Search Glean for similar issues"""
    try:
        if not os.getenv("GLEAN_TOKEN"):
            logger.warning("GLEAN_TOKEN not set, skipping Glean API call")
            return None
        
        headers = get_glean_headers()
        payload = {
            "query": query_text,
            "max_results": 5
        }
        resp = session.post(
            f"{GLEAN_BASE}/search",
            headers=headers,
            json=payload,
            timeout=30
        )
        if resp.status_code == 200:
            return resp.json()
        else:
            logger.warning(f"Failed to search Glean: {resp.status_code}")
            return None
    except Exception as e:
        logger.warning(f"Error searching Glean: {e}")
        return None


def search_glean_jira(query_text, max_results=10):
    """Search Glean specifically for JIRA tickets matching the query.

    Filters results to the ``jira`` datasource so that only ENG/JIRA issues
    are returned.  Falls back to the generic search if the JIRA-scoped call
    fails.
    """
    if not os.getenv("GLEAN_TOKEN"):
        logger.warning("GLEAN_TOKEN not set, skipping Glean JIRA search")
        return None
    headers = get_glean_headers()
    payload = {
        "query": query_text,
        "pageSize": max_results,
        "requestOptions": {
            "datasourcesFilter": {
                "datasources": ["jira"],
            },
        },
    }
    try:
        resp = session.post(
            f"{GLEAN_BASE}/search",
            headers=headers,
            json=payload,
            timeout=30,
        )
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f"Glean JIRA search returned {resp.status_code}, falling back to generic search")
    except Exception as e:
        logger.warning(f"Glean JIRA search failed: {e}, falling back")
    return search_glean(query_text)


def _extract_jira_tickets_from_glean(glean_data):
    """Walk Glean search results and extract JIRA ticket IDs + metadata.

    Returns a list of dicts:
        [{"ticket": "ENG-12345", "title": "...", "url": "...", "snippet": "..."}, ...]
    """
    import re as _re
    jira_pat = _re.compile(r'[A-Z][A-Z0-9]+-\d+')

    if not glean_data or not isinstance(glean_data, dict):
        return []

    raw_results = glean_data.get("results", [])
    if isinstance(raw_results, dict):
        raw_results = raw_results.get("results", [])
    if not isinstance(raw_results, list):
        return []

    seen_tickets = {}
    for r in raw_results:
        if not isinstance(r, dict):
            continue
        doc = r.get("document", r)
        title = doc.get("title", "")
        url = doc.get("url", "")
        # Collect snippet text from all available snippet objects
        snippet_text = ""
        snippets = r.get("snippets", [])
        if isinstance(snippets, list):
            parts = []
            for s in snippets:
                if isinstance(s, dict):
                    inner = s.get("snippet", s)
                    if isinstance(inner, dict):
                        parts.append(inner.get("text", ""))
                    elif isinstance(inner, str):
                        parts.append(inner)
                elif isinstance(s, str):
                    parts.append(s)
            snippet_text = " ".join(parts)[:500]

        combined_text = f"{title} {url} {snippet_text}"
        found_ids = jira_pat.findall(combined_text)

        # If the URL is a direct JIRA link, the ticket from the URL is the primary one
        if url and "jira" in url.lower():
            url_ids = jira_pat.findall(url)
            found_ids = url_ids + [t for t in found_ids if t not in url_ids]

        for tid in found_ids:
            if tid not in seen_tickets:
                seen_tickets[tid] = {
                    "ticket": tid,
                    "title": title[:200],
                    "url": url if tid in (jira_pat.findall(url) if url else []) else f"https://jira.nutanix.com/browse/{tid}",
                    "snippet": snippet_text[:300],
                }
    return list(seen_tickets.values())

def determine_failure_stage(test_result):
    """Determine failure stage from test result"""
    # Try to extract from exception_summary or exception
    exception_summary = test_result.get("exception_summary", "").lower()
    exception = test_result.get("exception", "").lower()
    combined = f"{exception_summary} {exception}"
    
    if any(keyword in combined for keyword in ["setup", "set up", "before", "precondition"]):
        return "Test Setup"
    elif any(keyword in combined for keyword in ["teardown", "tear down", "after", "cleanup"]):
        return "Teardown"
    elif any(keyword in combined for keyword in ["infra", "infrastructure", "connection", "timeout", "network"]):
        return "Infra"
    else:
        return "Test Body"

# Intermittent (mark for rerun) patterns: loaded from JSON; if exception_summary
# matches any pattern, mark for rerun as Yes.
INTERMITTENT_PATTERNS_JSON = os.path.join(os.path.dirname(__file__), "intermittent_patterns.json")

def _load_intermittent_patterns():
    """Load intermittent regex patterns from JSON file."""
    default_patterns = [
        r"Timedout executing command source.*cluster start in .* secs with error:",
        r"Timedout executing command source.*cluster.*create in .* secs with error",
        r"Couldn't get handle to SVM VM object for ip",
    ]
    try:
        if os.path.exists(INTERMITTENT_PATTERNS_JSON):
            with open(INTERMITTENT_PATTERNS_JSON, "r") as f:
                data = json.load(f)
            raw = data.get("intermittent_patterns", data) if isinstance(data, dict) else data
            if isinstance(raw, list) and raw:
                return [re.compile(p, re.IGNORECASE) for p in raw]
    except Exception as e:
        logger.warning(f"Could not load intermittent_patterns.json: {e}, using defaults")
    return [re.compile(p, re.IGNORECASE) for p in default_patterns]

SETUP_EXC_LIST = _load_intermittent_patterns()

def is_intermittent_rerun(exception_summary):
    """If exception_summary matches setup_exc_list patterns, mark for rerun as Yes."""
    if not exception_summary:
        return "No"
    text = (exception_summary or "").strip()
    for pattern in SETUP_EXC_LIST:
        if pattern.search(text):
            return "Yes"
    return "No"

def classify_failure(exception_summary, exception, jira_data, glean_data):
    """Classify failure as Test Issue or Product Issue"""
    exception_lower = (exception_summary or "").lower() + " " + (exception or "").lower()
    
    # Test Issue indicators
    test_issue_keywords = [
        "assertion", "assert", "expected", "actual", "python", "import error",
        "syntax error", "indentation", "nameerror", "attributeerror", "typeerror",
        "test framework", "pytest", "unittest", "dependency", "library", "module not found"
    ]
    
    # Product Issue indicators
    product_issue_keywords = [
        "api", "backend", "server", "500", "503", "timeout", "connection refused",
        "feature", "regression", "bug", "defect", "broken", "not working"
    ]
    
    test_issue_score = sum(1 for keyword in test_issue_keywords if keyword in exception_lower)
    product_issue_score = sum(1 for keyword in product_issue_keywords if keyword in exception_lower)
    
    # Check Jira ticket if available
    if jira_data:
        jira_summary = jira_data.get("fields", {}).get("summary", "").lower()
        jira_description = jira_data.get("fields", {}).get("description", "").lower()
        jira_combined = f"{jira_summary} {jira_description}"
        
        if "test" in jira_combined and "fix" in jira_combined:
            test_issue_score += 2
        if "product" in jira_combined or "regression" in jira_combined:
            product_issue_score += 2
    
    if test_issue_score > product_issue_score and test_issue_score > 0:
        return "Test Issue"
    elif product_issue_score > test_issue_score and product_issue_score > 0:
        return "Product Issue"
    else:
        return "Unknown / Needs Manual Review"

def validate_triage_genie_ticket(jira_ticket_id, exception_summary, exception):
    """Validate Triage Genie / Jira ticket relevance"""
    if not jira_ticket_id:
        return "Invalid"
    
    jira_data = fetch_jira_ticket(jira_ticket_id)
    if not jira_data:
        return "Invalid"
    
    fields = jira_data.get("fields", {})
    ticket_status = fields.get("status", {}).get("name", "")
    resolution = fields.get("resolution", {}).get("name", "") if fields.get("resolution") else None
    summary = fields.get("summary", "").lower()
    description = fields.get("description", "").lower()
    
    # Check if ticket is resolved/closed
    if resolution or ticket_status in ["Closed", "Resolved", "Done"]:
        return "Invalid"
    
    # Compare exception with ticket description
    exception_lower = (exception_summary or "").lower() + " " + (exception or "").lower()
    ticket_text = f"{summary} {description}"
    
    # Check for keyword overlap
    exception_words = set(exception_lower.split())
    ticket_words = set(ticket_text.split())
    common_words = exception_words.intersection(ticket_words)
    
    if len(common_words) >= 3:
        return "Valid"
    elif len(common_words) >= 1:
        return "Partial"
    else:
        return "Invalid"

def generate_ai_suggestion(issue_type, exception_summary, exception, jira_tickets, jira_data, glean_data):
    """Generate AI suggestion based on analysis"""
    suggestions = []
    
    if issue_type == "Test Issue":
        exception_lower = (exception_summary or "").lower() + " " + (exception or "").lower()
        
        if "python" in exception_lower or "import" in exception_lower:
            suggestions.append("Failure caused by Python dependency or import issue. Update test dependencies and verify Python environment.")
        elif "assertion" in exception_lower or "assert" in exception_lower:
            suggestions.append("Assertion failure detected. Review expected vs actual values. Update test assertions if product behavior has changed.")
        elif "syntax" in exception_lower or "indentation" in exception_lower:
            suggestions.append("Syntax error in test code. Review and fix test script syntax.")
        else:
            suggestions.append("Test logic issue identified. Review test implementation and update test code. Consider creating a Nugerrit CR for test fixes.")
        
        if jira_tickets:
            suggestions.append(f"Review Jira ticket(s): {', '.join(jira_tickets)}")
    
    elif issue_type == "Product Issue":
        if jira_tickets and jira_data:
            ticket_id = jira_tickets[0]
            suggestions.append(f"Failure aligns with known product issue in {ticket_id}. Test behavior is valid. Monitor Jira ticket for fix.")
        elif jira_tickets:
            suggestions.append(f"Product regression detected. Track via Jira ticket(s): {', '.join(jira_tickets)}. Test should remain as-is until product fix.")
        else:
            suggestions.append("Product-level failure identified. Test behavior is valid. Consider creating a Jira ticket to track this product issue.")
        
        if glean_data:
            suggestions.append("Similar issues found in Glean search. Review related documentation or known issues.")
    
    else:
        suggestions.append("Unable to definitively classify failure. Manual review recommended. Check test logs and Jira tickets for context.")
        if jira_tickets:
            suggestions.append(f"Review existing Jira ticket(s): {', '.join(jira_tickets)}")
    
    return " ".join(suggestions)

def fetch_detailed_test_result(testcase_id):
    """Fetch detailed test result including exception and failure stage"""
    try:
        resp = session.get(
            f"{PHX_BASE}/agave_test_results/{testcase_id}",
            timeout=30
        )
        resp.raise_for_status()
        return resp.json().get("data", {})
    except Exception as e:
        logger.warning(f"Error fetching detailed test result {testcase_id}: {e}")
        return {}

def convert_testcase_name_to_path(testcase_name):
    """
    Convert testcase name to directory path.
    Example: cdp.curator.goldsuite_iointegrity.test_goldsuite.CuratorGoldSuiteTest.test_gold___cluster_upgrade_with_disk_balancing
    -> cdp/curator/goldsuite_iointegrity/test_goldsuite/CuratorGoldSuiteTest/test_gold___cluster_upgrade_with_disk_balancing
    """
    if not testcase_name:
        return ""
    # Replace dots with slashes
    path = testcase_name.replace(".", "/")
    return path

def fetch_log_from_url(log_url, timeout=30):
    """
    Fetch log content from a URL.
    Returns the log content as string, or empty string if fetch fails.
    """
    try:
        resp = session.get(log_url, timeout=timeout, verify=False)
        if resp.status_code == 200:
            return resp.text
        else:
            logger.warning(f"Failed to fetch log from {log_url}: HTTP {resp.status_code}")
            return ""
    except Exception as e:
        logger.warning(f"Error fetching log from {log_url}: {e}")
        return ""

def fetch_testcase_logs(testcase_name, tester_log_url):
    """
    Fetch steps.log and nutest_test.log for a testcase.
    
    Args:
        testcase_name: Full testcase name (e.g., cdp.curator.goldsuite_iointegrity.test_goldsuite.CuratorGoldSuiteTest.test_gold___cluster_upgrade_with_disk_balancing)
        tester_log_url: Base URL for logs directory (e.g., http://10.40.234.216/logs/...)
    
    Returns:
        dict with 'steps_log' and 'nutest_test_log' keys, containing log content or empty strings
    """
    logs = {
        "steps_log": "",
        "nutest_test_log": ""
    }
    
    if not testcase_name or not tester_log_url:
        return logs
    
    # Convert testcase name to path
    testcase_path = convert_testcase_name_to_path(testcase_name)
    if not testcase_path:
        return logs
    
    # Construct log URLs
    # Ensure tester_log_url ends with /
    base_url = tester_log_url.rstrip("/") + "/"
    
    # Construct full paths
    steps_log_url = f"{base_url}{testcase_path}/steps.log"
    nutest_test_log_url = f"{base_url}{testcase_path}/nutest_test.log"
    
    logger.info(f"Fetching logs for {testcase_name}")
    logger.debug(f"Steps log URL: {steps_log_url}")
    logger.debug(f"Nutest test log URL: {nutest_test_log_url}")
    
    # Fetch logs
    logs["steps_log"] = fetch_log_from_url(steps_log_url)
    logs["nutest_test_log"] = fetch_log_from_url(nutest_test_log_url)
    
    return logs

def generate_ai_failure_summary(exception, exception_summary, tester_log_url, testcase_name, steps_log="", nutest_test_log=""):
    """
    Generate AI-based failure summary using the Nutanix AI endpoint.
    
    Args:
        exception: Full exception text
        exception_summary: Exception summary
        tester_log_url: Base URL for logs
        testcase_name: Testcase name
        steps_log: Content of steps.log
        nutest_test_log: Content of nutest_test.log
    
    Returns:
        tuple (success: bool, summary: str or error message)
    """
    try:
        # Build the prompt content
        log_content = ""
        if steps_log:
            log_content += f"=== steps.log ===\n{steps_log[:5000]}\n\n"  # Limit to 5000 chars per log
        if nutest_test_log:
            log_content += f"=== nutest_test.log ===\n{nutest_test_log[:5000]}\n\n"
        
        # Build user content
        user_content = (
            f"Testcase: {testcase_name}\n\n"
            f"Exception Summary: {exception_summary or 'N/A'}\n\n"
            f"Exception:\n```\n{exception or 'N/A'}\n```\n\n"
        )
        
        if log_content:
            user_content += f"Test Logs:\n{log_content}\n"
        
        user_content += (
            "Provide a concise failure summary:\n"
            "(1) Root cause in one sentence\n"
            "(2) Failing component or line if clear\n"
            "(3) Suggested fix or next step"
        )
        
        # System prompt
        system_prompt = (
            "You are a test failure analyst. Given a Python test failure with exception details and logs, "
            "provide a concise failure summary: (1) Root cause in one sentence, (2) failing component or line if clear, "
            "(3) suggested fix or next step. Be specific and actionable."
        )
        
        # Prepare payload
        payload = {
            "model": "hack-reason",
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content}
            ],
            "max_tokens": 1024,
            "stream": False
        }
        
        # Make request to AI endpoint
        url = f"{AI_BASE}/chat/completions"
        headers = {
            "Authorization": f"Bearer {AI_API_KEY}",
            "Accept": "application/json",
            "Content-Type": "application/json"
        }
        
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        
        with urllib.request.urlopen(req, context=SSL_CTX, timeout=60) as resp:
            if resp.getcode() != 200:
                return False, f"AI API returned HTTP {resp.getcode()}"
            
            response_data = json.loads(resp.read().decode())
            choices = response_data.get("choices", [])
            if not choices:
                return False, "AI returned no choices"
            
            content = (choices[0].get("message") or {}).get("content", "")
            return True, content.strip()
            
    except urllib.error.HTTPError as e:
        body = e.read().decode() if e.fp else ""
        try:
            err_json = json.loads(body)
        except Exception:
            err_json = {"raw": body}
        return False, f"AI API error HTTP {e.code}: {json.dumps(err_json)[:300]}"
    except Exception as e:
        logger.error(f"Error generating AI failure summary: {e}", exc_info=True)
        return False, f"Error generating AI summary: {str(e)}"

def create_triage_genie_session():
    """
    Create a requests.Session authenticated via login for Triage Genie API calls.

    Uses TRIAGE_GENIE_USERNAME / TRIAGE_GENIE_PASSWORD env vars if set,
    otherwise falls back to the service account (JITA_SVC_USERNAME / JITA_SVC_PASSWORD).
    """
    tg_user = TRIAGE_GENIE_USERNAME or JITA_SVC_USERNAME
    tg_pass = TRIAGE_GENIE_PASSWORD or JITA_SVC_PASSWORD

    if not tg_user or not tg_pass:
        logger.warning("Triage Genie credentials not configured (set TRIAGE_GENIE_USERNAME / TRIAGE_GENIE_PASSWORD)")
        return None

    session_triage = requests.Session()
    session_triage.verify = False
    try:
        login_response = session_triage.post(
            LOGIN_URL,
            data={"username": tg_user, "password": tg_pass},
            verify=False,
            timeout=15,
            allow_redirects=False,
        )
        # Successful login: 302 redirect to dashboard (not back to /login).
        # Failed login: 200 with login page re-rendered.
        if login_response.status_code == 302:
            location = login_response.headers.get("Location", "")
            if "/login" not in location:
                logger.info("Triage Genie session authenticated successfully")
                return session_triage

        logger.warning(
            "Triage Genie login failed (credentials rejected). "
            "Set TRIAGE_GENIE_USERNAME and TRIAGE_GENIE_PASSWORD env vars "
            "with valid Triage Genie / LDAP credentials."
        )
        return None
    except Exception as e:
        logger.warning(f"Triage Genie login error: {e}")
        return None


def build_triage_genie_ticket_map(test_result_ids):
    """
    Batch pre-fetch Triage Genie dup_ticket_id for each test result.

    Uses GET /api/tasks/{testcase_id} (the direct Triage Genie task endpoint).
    Authentication: session-based login via create_triage_genie_session().

    Args:
        test_result_ids: List of JITA test result _id values (testcase IDs)

    Returns:
        dict: Mapping of testcase_id -> dup_ticket_id
    """
    ticket_map = {}
    if not test_result_ids:
        return ticket_map

    unique_ids = list(set(tid for tid in test_result_ids if tid))
    logger.info(f"build_triage_genie_ticket_map: looking up {len(unique_ids)} testcase(s) via /api/tasks/{{id}}")

    tg_session = create_triage_genie_session()
    if not tg_session:
        logger.warning("build_triage_genie_ticket_map: no TG session — cannot fetch tickets")
        return ticket_map

    def _lookup_single(tc_id):
        """GET /api/tasks/{tc_id} and return (tc_id, dup_ticket_id) or (tc_id, None)."""
        try:
            resp = tg_session.get(
                f"{TRIAGE_GENIE_BASE}/tasks/{tc_id}",
                verify=False,
                timeout=15,
            )
            if resp.status_code == 200:
                ct = resp.headers.get("Content-Type", "")
                if "json" in ct or "application/json" in ct:
                    data = resp.json()
                    dup = data.get("dup_ticket_id")
                    if dup:
                        return (tc_id, str(dup).strip())
                else:
                    logger.debug(f"TG /tasks/{tc_id} returned HTML (auth may have expired)")
            elif resp.status_code == 404:
                logger.debug(f"TG /tasks/{tc_id}: not found (test may not be triaged yet)")
            else:
                logger.debug(f"TG /tasks/{tc_id} returned {resp.status_code}")
        except Exception as exc:
            logger.debug(f"TG /tasks/{tc_id} error: {exc}")
        return (tc_id, None)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(_lookup_single, tid): tid for tid in unique_ids}
        for future in as_completed(futures):
            tc_id, dup_ticket = future.result()
            if dup_ticket:
                ticket_map[tc_id] = dup_ticket

    logger.info(f"build_triage_genie_ticket_map: found {len(ticket_map)} ticket(s) out of {len(unique_ids)} lookups")
    return ticket_map


def fetch_triage_genie_ticket_id(testcase_id, triage_session=None):
    """
    Fetch Triage Genie dup_ticket_id for a single test result.

    Calls GET /api/tasks/{testcase_id} using a session-authenticated connection.
    If triage_session is not provided, creates one via create_triage_genie_session().

    Returns:
        str: Jira ticket ID (e.g., "ENG-904593") or None if not found
    """
    if not testcase_id:
        return None

    session = triage_session or create_triage_genie_session()
    if not session:
        return None

    try:
        resp = session.get(
            f"{TRIAGE_GENIE_BASE}/tasks/{testcase_id}",
            verify=False,
            timeout=15,
        )
        if resp.status_code == 200:
            ct = resp.headers.get("Content-Type", "")
            if "json" in ct:
                data = resp.json()
                dup = data.get("dup_ticket_id")
                if dup:
                    logger.info(f"Found Triage Genie ticket {dup} for testcase {testcase_id}")
                    return str(dup).strip()
            else:
                logger.debug(f"TG /tasks/{testcase_id} returned HTML (auth expired or invalid)")
        elif resp.status_code == 404:
            logger.debug(f"TG /tasks/{testcase_id}: not found")
        else:
            logger.debug(f"TG /tasks/{testcase_id} returned {resp.status_code}")
    except Exception as e:
        logger.debug(f"TG /tasks/{testcase_id} error: {e}")

    return None

@app.route("/mcp/regression/failed-analysis/analyze", methods=["GET"])
@jwt_required
def analyze_failed_testcases():
    """Analyze failed testcases with AI agent"""
    start = time.time()
    
    tag = request.args.get("tag")
    task_ids_param = request.args.get("task_ids")
    include_param = request.args.get("include", "")
    include_set = {x.strip().lower() for x in include_param.split(",") if x.strip()} if include_param else set()
    # Default: basic + exception_summary + intermittent when no include given
    if not include_set:
        include_set = {"basic", "exception_summary", "intermittent"}
    
    # Parse task_ids if provided
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]
    
    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400
    
    try:
        # Fetch regression tasks
        tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
        if not tasks:
            return jsonify({
                "success": True,
                "results": [],
                "message": "No tasks found for the given criteria"
            })
        
        # Collect all task IDs
        collected_task_ids = [task["_id"]["$oid"] for task in tasks]
        
        # Fetch test results - only failed ones
        logger.info(f"Fetching failed test results for {len(collected_task_ids)} tasks")
        all_test_results = []
        
        if collected_task_ids:
            try:
                # Fetch all test results first
                all_results = fetch_test_results_batch_with_pagination(collected_task_ids)
                # Filter for failed tests
                failed_results = [
                    tr for tr in all_results 
                    if tr.get("status", "").lower() in ("failed", "failure")
                ]
                logger.info(f"Found {len(failed_results)} failed testcases")
                
                # Batch pre-fetch Triage Genie tickets via direct /api/tasks/{id} lookups
                tg_ticket_map = {}
                if "triage_genie_ticket" in include_set:
                    _failed_ids = []
                    for _fr in failed_results:
                        _rid = _fr.get("_id")
                        if isinstance(_rid, dict) and "$oid" in _rid:
                            _failed_ids.append(_rid["$oid"])
                        elif _rid:
                            _failed_ids.append(str(_rid))
                    tg_ticket_map = build_triage_genie_ticket_map(_failed_ids)
                
                # Current branch for history API (from first task)
                current_branch = (tasks[0].get("branch") or "") if tasks else ""
                
                # Analyze each failed testcase
                analysis_results = []
                
                for test_result in failed_results:
                    testcase_id = None
                    if isinstance(test_result.get("_id"), dict) and "$oid" in test_result.get("_id", {}):
                        testcase_id = test_result["_id"]["$oid"]
                    else:
                        testcase_id = str(test_result.get("_id", ""))
                    
                    # Fetch detailed test result for exception and failure stage
                    detailed_result = {}
                    if testcase_id:
                        detailed_result = fetch_detailed_test_result(testcase_id)
                    
                    # Extract data
                    test_field = test_result.get("test", {})
                    if isinstance(test_field, dict):
                        testcase_name = test_field.get("name", "")
                    elif isinstance(test_field, str):
                        testcase_name = test_field
                    else:
                        # Try from detailed result
                        detailed_test = detailed_result.get("test", {})
                        if isinstance(detailed_test, dict):
                            testcase_name = detailed_test.get("name", "")
                        elif isinstance(detailed_test, str):
                            testcase_name = detailed_test
                        else:
                            testcase_name = ""
                    
                    status = test_result.get("status", "FAILED")
                    exception_summary = test_result.get("exception_summary") or detailed_result.get("exception_summary", "")
                    exception = detailed_result.get("exception", "")
                    jira_tickets = test_result.get("jira_tickets", [])
                    test_log_url = test_result.get("test_log_url") or detailed_result.get("test_log_url", "")
                    comments = test_result.get("comments") or detailed_result.get("comments") or ""
                    
                    # Determine failure stage
                    failure_stage = determine_failure_stage({**test_result, **detailed_result})
                    
                    # Intermittent (mark for rerun) from exception_summary
                    intermittent_rerun = is_intermittent_rerun(exception_summary) if "intermittent" in include_set else None
                    
                    # Optional heavy fields: Jira/Glean, issue type, suggestion, Triage Genie ticket, AI Summary
                    jira_data = None
                    glean_data = None
                    issue_type = None
                    suggestion = None
                    triage_genie_ticket_id = None
                    ai_summary = None
                    if "issue_type" in include_set or "suggestion" in include_set:
                        if jira_tickets:
                            jira_data = fetch_jira_ticket(jira_tickets[0])
                        if exception_summary:
                            glean_data = search_glean(exception_summary)
                    if "issue_type" in include_set:
                        issue_type = classify_failure(exception_summary, exception, jira_data, glean_data)
                    if "suggestion" in include_set:
                        suggestion = generate_ai_suggestion(
                            issue_type or classify_failure(exception_summary, exception, jira_data, glean_data),
                            exception_summary, exception, jira_tickets, jira_data, glean_data
                        )
                    if "triage_genie_ticket" in include_set:
                        triage_genie_ticket_id = (
                            tg_ticket_map.get(testcase_id)
                            or tg_ticket_map.get(testcase_name)
                        )
                    
                    # Generate AI Summary if requested
                    if "ai_summary" in include_set:
                        try:
                            # Fetch logs
                            logs = fetch_testcase_logs(testcase_name, test_log_url)
                            
                            # Generate AI summary
                            success, summary = generate_ai_failure_summary(
                                exception=exception,
                                exception_summary=exception_summary,
                                tester_log_url=test_log_url,
                                testcase_name=testcase_name,
                                steps_log=logs.get("steps_log", ""),
                                nutest_test_log=logs.get("nutest_test_log", "")
                            )
                            
                            if success:
                                ai_summary = summary
                            else:
                                ai_summary = f"Error generating summary: {summary}"
                                logger.warning(f"Failed to generate AI summary for {testcase_name}: {summary}")
                        except Exception as e:
                            logger.error(f"Error generating AI summary for {testcase_name}: {e}", exc_info=True)
                            ai_summary = f"Error: {str(e)}"
                    
                    # Resolve regression owner
                    regression_owner = resolve_owner(testcase_name) if testcase_name else "Unknown"
                    
                    # Resolve agave_task_id for retrigger support
                    raw_atid = test_result.get("agave_task_id")
                    if isinstance(raw_atid, dict) and "$oid" in raw_atid:
                        agave_task_id = raw_atid["$oid"]
                    elif raw_atid:
                        agave_task_id = str(raw_atid)
                    else:
                        agave_task_id = None

                    row = {
                        "testcase_id": testcase_id,
                        "testcase_name": testcase_name,
                        "agave_task_id": agave_task_id,
                        "status": status,
                        "failure_stage": failure_stage,
                        "jira_tickets": jira_tickets,
                        "regression_owner": regression_owner,
                        "test_log_url": test_log_url,
                        "exception_summary": exception_summary,
                        "comments": comments,
                        "exception": exception[:200] if exception else ""
                    }
                    if intermittent_rerun is not None:
                        row["intermittent_rerun"] = intermittent_rerun
                    if issue_type is not None:
                        row["issue_type"] = issue_type
                    if suggestion is not None:
                        row["suggestion_by_ai_agent"] = suggestion
                    if triage_genie_ticket_id is not None:
                        row["triage_genie_ticket_id"] = triage_genie_ticket_id
                    if ai_summary is not None:
                        row["ai_summary"] = ai_summary

                    analysis_results.append(row)
                
                logger.info(f"[END] Failed Analysis | results={len(analysis_results)} | time={time.time() - start:.2f}s")
                
                return jsonify({
                    "success": True,
                    "results": analysis_results,
                    "total_analyzed": len(analysis_results),
                    "current_branch": current_branch,
                    "tag": tag or None
                })
                
            except requests.exceptions.ConnectionError as e:
                logger.error(f"Connection error: {e}", exc_info=True)
                return jsonify({
                    "error": "Failed to connect to JITA API. Please check your network connection and ensure 'jita.eng.nutanix.com' is accessible.",
                    "details": str(e)
                }), 503
            except requests.exceptions.Timeout as e:
                logger.error(f"Timeout error: {e}", exc_info=True)
                return jsonify({
                    "error": "Request to JITA API timed out. Please try again.",
                    "details": str(e)
                }), 504
            except Exception as e:
                logger.error(f"Error fetching test results: {e}", exc_info=True)
                return jsonify({"error": f"Failed to fetch test results: {str(e)}"}), 500
        else:
            return jsonify({
                "success": True,
                "results": [],
                "message": "No task IDs found"
            })
            
    except ConnectionError as e:
        logger.error(f"Connection error: {e}", exc_info=True)
        return jsonify({
            "error": str(e),
            "type": "connection_error"
        }), 503
    except TimeoutError as e:
        logger.error(f"Timeout error: {e}", exc_info=True)
        return jsonify({
            "error": str(e),
            "type": "timeout_error"
        }), 504
    except Exception as e:
        logger.error(f"Error analyzing failed testcases: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _analyze_failed_testcases_stream(tag, task_ids, include_set, status_set=None):
    """Generator that yields SSE events: progress, start, then one 'row' per result, then 'done' or 'error'."""
    if status_set is None:
        status_set = {"failed", "failure"}
    try:
        yield json.dumps({"type": "progress", "phase": "preparing"})

        tasks = fetch_regression_tasks(tag=tag, task_ids=task_ids)
        if not tasks:
            yield json.dumps({"type": "start", "total": 0, "current_branch": "", "tag": tag})
            yield json.dumps({"type": "done"})
            return
        collected_task_ids = [task["_id"]["$oid"] for task in tasks]

        yield json.dumps({"type": "progress", "phase": "fetching_results"})

        all_results = fetch_test_results_batch_with_pagination(collected_task_ids)
        failed_results = [
            tr for tr in all_results
            if tr.get("status", "").lower() in status_set
        ]
        current_branch = (tasks[0].get("branch") or "") if tasks else ""

        # Batch pre-fetch Triage Genie tickets via direct /api/tasks/{id} lookups
        tg_ticket_map = {}
        if "triage_genie_ticket" in include_set:
            yield json.dumps({"type": "progress", "phase": "fetching_triage_genie"})
            _failed_ids = []
            for _fr in failed_results:
                _rid = _fr.get("_id")
                if isinstance(_rid, dict) and "$oid" in _rid:
                    _failed_ids.append(_rid["$oid"])
                elif _rid:
                    _failed_ids.append(str(_rid))
            tg_ticket_map = build_triage_genie_ticket_map(_failed_ids)

        yield json.dumps({
            "type": "start",
            "total": len(failed_results),
            "current_branch": current_branch,
            "tag": tag or None
        })

        for test_result in failed_results:
            testcase_id = None
            if isinstance(test_result.get("_id"), dict) and "$oid" in test_result.get("_id", {}):
                testcase_id = test_result["_id"]["$oid"]
            else:
                testcase_id = str(test_result.get("_id", ""))
            detailed_result = {}
            if testcase_id:
                detailed_result = fetch_detailed_test_result(testcase_id)
            test_field = test_result.get("test", {})
            if isinstance(test_field, dict):
                testcase_name = test_field.get("name", "")
            elif isinstance(test_field, str):
                testcase_name = test_field
            else:
                detailed_test = detailed_result.get("test", {})
                if isinstance(detailed_test, dict):
                    testcase_name = detailed_test.get("name", "")
                elif isinstance(detailed_test, str):
                    testcase_name = detailed_test
                else:
                    testcase_name = ""
            status = test_result.get("status", "FAILED")
            exception_summary = test_result.get("exception_summary") or detailed_result.get("exception_summary", "")
            exception = detailed_result.get("exception", "")
            jira_tickets = test_result.get("jira_tickets", [])
            test_log_url = test_result.get("test_log_url") or detailed_result.get("test_log_url", "")
            comments = test_result.get("comments") or detailed_result.get("comments") or ""
            failure_stage = determine_failure_stage({**test_result, **detailed_result})
            intermittent_rerun = is_intermittent_rerun(exception_summary) if "intermittent" in include_set else None
            jira_data = None
            glean_data = None
            issue_type = None
            suggestion = None
            triage_genie_ticket_id = None
            ai_summary = None
            if "issue_type" in include_set or "suggestion" in include_set:
                if jira_tickets:
                    jira_data = fetch_jira_ticket(jira_tickets[0])
                if exception_summary:
                    glean_data = search_glean(exception_summary)
            if "issue_type" in include_set:
                issue_type = classify_failure(exception_summary, exception, jira_data, glean_data)
            if "suggestion" in include_set:
                suggestion = generate_ai_suggestion(
                    issue_type or classify_failure(exception_summary, exception, jira_data, glean_data),
                    exception_summary, exception, jira_tickets, jira_data, glean_data
                )
            if "triage_genie_ticket" in include_set:
                triage_genie_ticket_id = (
                    tg_ticket_map.get(testcase_id)
                    or tg_ticket_map.get(testcase_name)
                )
            
            # Generate AI Summary if requested
            if "ai_summary" in include_set:
                try:
                    # Fetch logs
                    logs = fetch_testcase_logs(testcase_name, test_log_url)
                    
                    # Generate AI summary
                    success, summary = generate_ai_failure_summary(
                        exception=exception,
                        exception_summary=exception_summary,
                        tester_log_url=test_log_url,
                        testcase_name=testcase_name,
                        steps_log=logs.get("steps_log", ""),
                        nutest_test_log=logs.get("nutest_test_log", "")
                    )
                    
                    if success:
                        ai_summary = summary
                    else:
                        ai_summary = f"Error generating summary: {summary}"
                        logger.warning(f"Failed to generate AI summary for {testcase_name}: {summary}")
                except Exception as e:
                    logger.error(f"Error generating AI summary for {testcase_name}: {e}", exc_info=True)
                    ai_summary = f"Error: {str(e)}"
            
            regression_owner = resolve_owner(testcase_name) if testcase_name else "Unknown"

            raw_atid = test_result.get("agave_task_id")
            if isinstance(raw_atid, dict) and "$oid" in raw_atid:
                agave_task_id = raw_atid["$oid"]
            elif raw_atid:
                agave_task_id = str(raw_atid)
            else:
                agave_task_id = None

            row = {
                "testcase_id": testcase_id,
                "testcase_name": testcase_name,
                "agave_task_id": agave_task_id,
                "status": status,
                "failure_stage": failure_stage,
                "jira_tickets": jira_tickets,
                "regression_owner": regression_owner,
                "test_log_url": test_log_url,
                "exception_summary": exception_summary,
                "comments": comments,
                "exception": exception[:200] if exception else ""
            }
            if intermittent_rerun is not None:
                row["intermittent_rerun"] = intermittent_rerun
            if issue_type is not None:
                row["issue_type"] = issue_type
            if suggestion is not None:
                row["suggestion_by_ai_agent"] = suggestion
            if triage_genie_ticket_id is not None:
                row["triage_genie_ticket_id"] = triage_genie_ticket_id
            if ai_summary is not None:
                row["ai_summary"] = ai_summary

            yield json.dumps({"type": "row", "result": row})
        yield json.dumps({"type": "done"})
    except Exception as e:
        logger.error(f"Error in analyze stream: {e}", exc_info=True)
        yield json.dumps({"type": "error", "message": str(e)})


@app.route("/mcp/regression/failed-analysis/analyze-stream", methods=["GET"])
@jwt_required
def analyze_failed_testcases_stream():
    """Stream analysis results as Server-Sent Events so the UI can display rows as they load."""
    tag = request.args.get("tag")
    task_ids_param = request.args.get("task_ids")
    include_param = request.args.get("include", "")
    include_set = {x.strip().lower() for x in include_param.split(",") if x.strip()} if include_param else set()
    if not include_set:
        include_set = {"basic", "exception_summary", "intermittent"}
    task_ids = None
    if task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]
    if not tag and not task_ids:
        return jsonify({"error": "Either tag or task_ids is required"}), 400

    statuses_param = request.args.get("statuses", "")
    STATUS_ALIASES = {
        "failed": ("failed", "failure"),
        "skipped": ("skipped", "skip"),
        "warning": ("warning", "warn"),
        "killed": ("killed", "terminated", "cancelled"),
    }
    if statuses_param:
        requested = {s.strip().lower() for s in statuses_param.split(",") if s.strip()}
        status_set = set()
        for key in requested:
            if key in STATUS_ALIASES:
                status_set.update(STATUS_ALIASES[key])
            else:
                status_set.add(key)
    else:
        status_set = {"failed", "failure"}

    def gen():
        for chunk in _analyze_failed_testcases_stream(tag, task_ids, include_set, status_set):
            yield f"data: {chunk}\n\n"

    return Response(
        stream_with_context(gen()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.route("/mcp/regression/failed-analysis/update-triage", methods=["PUT"])
@jwt_required
def update_triage_comments():
    """Update comment and/or jira_tickets for an agave_test_result via JITA PUT API."""
    try:
        data = request.get_json() or {}
        test_id = data.get("test_id")
        comment = data.get("comment", "")
        jira_tickets = data.get("jira_tickets")
        if not test_id:
            return jsonify({"error": "test_id is required"}), 400

        current_username = g.current_user.get("sub", "")
        user_auth = _get_user_credentials(current_username)
        if not user_auth:
            return jsonify({
                "error": "Session credentials expired. Please re-login to update triage.",
                "code": "CREDENTIALS_EXPIRED"
            }), 401

        update_fields = {
            "comments": comment,
            "triaged_by": current_username
        }
        if jira_tickets is not None:
            if isinstance(jira_tickets, str):
                jira_tickets = [jira_tickets] if jira_tickets else []
            update_fields["jira_tickets"] = jira_tickets
        payload = {
            "query": {"_id": {"$in": [{"$oid": test_id}]}},
            "data": {"$set": update_fields},
            "multi": True
        }
        url = f"{JITA_BASE}/agave_test_results"
        resp = requests.put(
            url,
            headers={"Content-Type": "application/json"},
            json=payload,
            auth=user_auth,
            verify=False,
            timeout=30
        )
        if resp.status_code == 200:
            return jsonify({"success": True, "message": "Updated", "triaged_by": current_username})
        return jsonify({"error": resp.text or f"HTTP {resp.status_code}"}), resp.status_code if resp.status_code >= 400 else 500
    except Exception as e:
        logger.error(f"Error updating triage: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# RDM Failure Analysis for Skipped Testcases
# ======================================================

RDM_PATTERNS_FILE = os.path.join(os.path.dirname(__file__), "..", "data", "rdm_failure_patterns.json")

def _load_rdm_patterns():
    """Load RDM failure patterns from JSON file and compile regexes."""
    try:
        if os.path.exists(RDM_PATTERNS_FILE):
            with open(RDM_PATTERNS_FILE, "r") as f:
                data = json.load(f)
            patterns = data.get("patterns", [])
            compiled = []
            for p in patterns:
                try:
                    compiled.append({
                        "id": p["id"],
                        "regex": re.compile(p["regex"], re.IGNORECASE | re.DOTALL),
                        "comment_template": p["comment_template"],
                        "category": p.get("category", ""),
                        "description": p.get("description", ""),
                        "jira": p.get("jira", ""),
                    })
                except re.error as e:
                    logger.warning(f"Invalid regex in RDM pattern '{p.get('id')}': {e}")
            return compiled
    except Exception as e:
        logger.warning(f"Could not load rdm_failure_patterns.json: {e}")
    return []


RDM_PATTERNS = _load_rdm_patterns()


def _extract_node_names(message):
    """Extract node names from RDM failure messages.

    Combines results from multiple extraction strategies so multi-node
    failures (where different nodes appear in different formats) are
    all captured.  Returns deduplicated list preserving first-seen order.
    """
    seen = dict()
    for n in re.findall(r'([\w-]+):\s*Received\s+"fatal"\s+in\s+waiting\s+for\s+event', message):
        seen.setdefault(n, None)
    for n in re.findall(r"(?:Nodes?:\s*)([a-zA-Z][\w\-]*\d+[\-\d]*)", message):
        seen.setdefault(n, None)
    if seen:
        return list(seen)
    all_nodes = re.findall(r"\b([a-zA-Z]+\d{2,}-\d+)\b", message)
    return list(dict.fromkeys(all_nodes))


def fetch_jita_deployments(task_id):
    """Fetch JITA deployments for a given agave_task_id."""
    payload = {
        "raw_query": {"task_id": {"$oid": task_id}},
        "start": 0,
        "limit": 20,
    }
    try:
        resp = session.post(
            f"{JITA_BASE}/reports/deployments",
            json=payload,
            timeout=30,
        )
        if resp.status_code == 200:
            return resp.json().get("data", [])
        logger.warning(f"JITA deployments fetch failed for task {task_id}: HTTP {resp.status_code}")
    except Exception as e:
        logger.warning(f"JITA deployments fetch error for task {task_id}: {e}")
    return []


def get_rdm_failure_info(task_id):
    """
    Get RDM failure information for a task by inspecting its JITA deployments.
    Returns a dict with rdm_message, rdm_link, provision_request_id, category, etc.
    or None if no failed deployments found.
    """
    deployments = fetch_jita_deployments(task_id)
    if not deployments:
        return None

    failed_deployments = [d for d in deployments if d.get("status") == "failed"]
    if not failed_deployments:
        return None

    # Collect failure messages from status_transitions
    all_reasons = []
    rdm_ids = []
    for dep in failed_deployments:
        prov_id = ""
        raw_prov = dep.get("provision_request_id")
        if isinstance(raw_prov, dict) and "$oid" in raw_prov:
            prov_id = raw_prov["$oid"]
        elif raw_prov:
            prov_id = str(raw_prov)
        if prov_id:
            rdm_ids.append(prov_id)

        for transition in dep.get("status_transitions", []):
            if transition.get("status") == "failed":
                reason = transition.get("reason", "")
                if reason and "RDM" in reason:
                    all_reasons.append(reason)

    if not all_reasons and not rdm_ids:
        return None

    combined_message = all_reasons[0] if all_reasons else ""
    primary_rdm_id = rdm_ids[0] if rdm_ids else ""

    # Try fetching detailed RDM failure_analysis from the RDM API
    rdm_category = ""
    rdm_resolution = ""
    rdm_source = ""
    if primary_rdm_id:
        try:
            rdm_resp = session.get(
                f"https://rdm.eng.nutanix.com/api/v1/scheduled_deployments/{primary_rdm_id}?expand=created_by",
                timeout=15,
            )
            if rdm_resp.status_code == 200:
                rdm_data = rdm_resp.json().get("data", {})
                rdm_msg = rdm_data.get("message", "")
                if rdm_msg and not combined_message:
                    combined_message = rdm_msg
                fa = rdm_data.get("failure_analysis", {})
                rdm_category = fa.get("category", "")
                rdm_resolution = fa.get("resolution", "")
                em = fa.get("error_metadata", {}).get("RDM", {})
                rdm_source = em.get("source", "")
        except Exception as e:
            logger.warning(f"RDM API fetch failed for {primary_rdm_id}: {e}")

    return {
        "rdm_message": combined_message,
        "rdm_link": f"https://rdm.eng.nutanix.com/scheduled_deployments/{primary_rdm_id}" if primary_rdm_id else "",
        "provision_request_id": primary_rdm_id,
        "rdm_category": rdm_category,
        "rdm_resolution": rdm_resolution,
        "rdm_source": rdm_source,
        "failed_deployment_count": len(failed_deployments),
        "total_deployment_count": len(deployments),
    }


def match_rdm_pattern(rdm_message):
    """
    Match an RDM failure message against known patterns.
    Returns dict with matched pattern info and generated comment, or None.
    Hot-reloads patterns from disk so file edits take effect without restart.
    """
    global RDM_PATTERNS
    if not rdm_message:
        return None

    RDM_PATTERNS = _load_rdm_patterns()
    for pattern in RDM_PATTERNS:
        match = pattern["regex"].search(rdm_message)
        if match:
            template = pattern["comment_template"]
            comment = template

            if "{node_name}" in template:
                nodes = _extract_node_names(rdm_message)
                if nodes:
                    comment = ", ".join(
                        template.replace("{node_name}", n) for n in nodes
                    )
                else:
                    comment = "regx_rerun"

            return {
                "pattern_id": pattern["id"],
                "category": pattern["category"],
                "description": pattern["description"],
                "generated_comment": comment,
                "failed_nodes": nodes if "{node_name}" in template and nodes else [],
                "jira": pattern.get("jira", ""),
                "matched": True,
            }

    return None


@app.route("/mcp/regression/failed-analysis/rdm-analyze", methods=["POST"])
@jwt_required
def rdm_analyze_skipped():
    """
    Analyze RDM failure for skipped testcases.
    Accepts: { task_ids: [agave_task_id, ...] }
    Returns RDM failure info with pattern-matched comments for each task.
    """
    try:
        data = request.get_json() or {}
        task_ids = data.get("task_ids", [])
        if isinstance(task_ids, str):
            task_ids = [tid.strip() for tid in task_ids.split(",") if tid.strip()]
        if not task_ids:
            return jsonify({"error": "task_ids is required"}), 400

        results = []
        for task_id in task_ids:
            rdm_info = get_rdm_failure_info(task_id)
            if not rdm_info:
                results.append({
                    "agave_task_id": task_id,
                    "rdm_found": False,
                    "rdm_message": "",
                    "generated_comment": "",
                    "pattern_matched": False,
                })
                continue

            pattern_result = match_rdm_pattern(rdm_info["rdm_message"])
            results.append({
                "agave_task_id": task_id,
                "rdm_found": True,
                "rdm_message": rdm_info["rdm_message"][:500],
                "rdm_link": rdm_info["rdm_link"],
                "rdm_category": rdm_info["rdm_category"],
                "rdm_resolution": rdm_info["rdm_resolution"],
                "rdm_source": rdm_info["rdm_source"],
                "failed_deployments": rdm_info["failed_deployment_count"],
                "total_deployments": rdm_info["total_deployment_count"],
                "pattern_matched": bool(pattern_result),
                "generated_comment": pattern_result["generated_comment"] if pattern_result else "",
                "pattern_id": pattern_result["pattern_id"] if pattern_result else "",
                "pattern_category": pattern_result["category"] if pattern_result else "",
                "pattern_description": pattern_result["description"] if pattern_result else "",
                "pattern_jira": pattern_result.get("jira", "") if pattern_result else "",
                "failed_nodes": pattern_result.get("failed_nodes", []) if pattern_result else [],
            })

        return jsonify({"success": True, "results": results})
    except Exception as e:
        logger.error(f"RDM analyze error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/failed-analysis/rdm-analyze-ai", methods=["POST"])
@jwt_required
def rdm_analyze_ai():
    """
    AI-powered analysis for RDM failures that don't match known patterns.
    Uses Glean search + AI summary to identify existing bugs and provide analysis.
    """
    try:
        data = request.get_json() or {}
        task_id = data.get("task_id", "")
        rdm_message = data.get("rdm_message", "")
        testcase_name = data.get("testcase_name", "")

        if not rdm_message and not task_id:
            return jsonify({"error": "task_id or rdm_message is required"}), 400

        if not rdm_message and task_id:
            rdm_info = get_rdm_failure_info(task_id)
            if rdm_info:
                rdm_message = rdm_info.get("rdm_message", "")

        if not rdm_message:
            return jsonify({"error": "No RDM failure message found"}), 404

        # Search Glean for similar failures
        glean_results = None
        search_query = rdm_message[:200]
        if testcase_name:
            short_name = testcase_name.split(".")[-1] if "." in testcase_name else testcase_name
            search_query = f"{short_name} {search_query}"
        try:
            glean_results = search_glean(search_query)
        except Exception as e:
            logger.warning(f"Glean search failed for RDM analysis: {e}")

        # Build AI analysis
        jira_refs = []
        glean_summary = ""
        if glean_results and isinstance(glean_results, list):
            for gr in glean_results[:5]:
                if isinstance(gr, dict):
                    title = gr.get("title", "")
                    url = gr.get("url", "")
                    if title:
                        glean_summary += f"- {title}"
                        if url:
                            glean_summary += f" ({url})"
                        glean_summary += "\n"
                        # Extract JIRA ticket IDs
                        jira_matches = re.findall(r"[A-Z]+-\d+", title)
                        jira_refs.extend(jira_matches)
                        if url:
                            jira_url_matches = re.findall(r"[A-Z]+-\d+", url)
                            jira_refs.extend(jira_url_matches)

        jira_refs = list(dict.fromkeys(jira_refs))

        summary_parts = []
        summary_parts.append(f"RDM Deployment Failure Analysis")
        summary_parts.append(f"Error: {rdm_message[:300]}")
        if jira_refs:
            summary_parts.append(f"Related Jira tickets: {', '.join(jira_refs[:5])}")
        if glean_summary:
            summary_parts.append(f"Similar issues found:\n{glean_summary}")
        else:
            summary_parts.append("No similar issues found in Glean.")

        ai_summary = "\n".join(summary_parts)

        suggested_comment = "regx_rerun"
        if jira_refs:
            suggested_comment = f"regx_rerun ({jira_refs[0]})"

        return jsonify({
            "success": True,
            "ai_summary": ai_summary,
            "jira_refs": jira_refs,
            "suggested_comment": suggested_comment,
            "glean_results_count": len(glean_results) if glean_results else 0,
        })
    except Exception as e:
        logger.error(f"RDM AI analyze error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/failed-analysis/ai-summary-single", methods=["POST"])
@jwt_required
def ai_summary_single_testcase():
    """Generate AI summary for a single testcase on demand."""
    try:
        body = request.get_json(force=True) or {}
        testcase_name = body.get("testcase_name", "")
        exception_summary = body.get("exception_summary", "")
        exception = body.get("exception", "")
        test_log_url = body.get("test_log_url", "")

        if not testcase_name:
            return jsonify({"success": False, "error": "testcase_name is required"}), 400

        logs = fetch_testcase_logs(testcase_name, test_log_url)
        success, summary = generate_ai_failure_summary(
            exception=exception,
            exception_summary=exception_summary,
            tester_log_url=test_log_url,
            testcase_name=testcase_name,
            steps_log=logs.get("steps_log", ""),
            nutest_test_log=logs.get("nutest_test_log", ""),
        )

        if success:
            return jsonify({"success": True, "ai_summary": summary})
        else:
            return jsonify({"success": False, "error": summary}), 500
    except Exception as e:
        logger.error(f"AI summary single testcase error: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/mcp/regression/failed-analysis/glean-search-single", methods=["POST"])
@jwt_required
def glean_search_single_testcase():
    """Search Glean for matching ENG JIRA tickets and generate an AI failure
    analysis for a single testcase.

    Strategy:
      1. Run a JIRA-scoped Glean search with the exception summary (most specific).
      2. Run a second JIRA-scoped search with the testcase short name + key error
         terms (catches broader matches).
      3. Merge and de-duplicate found tickets, then enrich the top ones via the
         JIRA REST API (status, summary, resolution) so the user sees whether
         tickets are still open.
      4. Feed everything into the AI model for classification and recommendation.
    """
    try:
        body = request.get_json(force=True) or {}
        testcase_name = body.get("testcase_name", "")
        exception_summary = body.get("exception_summary", "")
        exception_text = body.get("exception", "")
        ai_summary = body.get("ai_summary", "")
        test_log_url = body.get("test_log_url", "")
        failure_stage = body.get("failure_stage", "")
        jira_tickets = body.get("jira_tickets", [])

        if not testcase_name:
            return jsonify({"success": False, "error": "testcase_name is required"}), 400

        short_name = testcase_name.split(".")[-1] if "." in testcase_name else testcase_name

        # ---- Multi-query Glean search strategy ----

        all_glean_tickets = {}   # ticket_id -> metadata dict
        all_glean_snippets = []
        queries_used = []

        # Query 1: exception summary (most specific error text)
        q1 = (exception_summary or "")[:250].strip()
        if q1:
            queries_used.append(q1)
            g1 = search_glean_jira(q1, max_results=10)
            for t in _extract_jira_tickets_from_glean(g1):
                all_glean_tickets.setdefault(t["ticket"], t)
            all_glean_snippets.extend(_build_glean_snippets(g1))

        # Query 2: testcase short name + condensed error keywords
        error_keywords = _extract_error_keywords(exception_summary, exception_text)
        q2 = f"{short_name} {error_keywords}"[:250].strip()
        if q2 and q2 != q1:
            queries_used.append(q2)
            g2 = search_glean_jira(q2, max_results=10)
            for t in _extract_jira_tickets_from_glean(g2):
                all_glean_tickets.setdefault(t["ticket"], t)
            all_glean_snippets.extend(_build_glean_snippets(g2))

        # Query 3: AI summary as search input (if available and different)
        if ai_summary:
            q3 = ai_summary[:250].strip()
            if q3 and q3 != q1 and q3 != q2:
                queries_used.append(q3)
                g3 = search_glean_jira(q3, max_results=5)
                for t in _extract_jira_tickets_from_glean(g3):
                    all_glean_tickets.setdefault(t["ticket"], t)
                all_glean_snippets.extend(_build_glean_snippets(g3))

        # De-dup snippets by URL
        seen_urls = set()
        deduped_snippets = []
        for s in all_glean_snippets:
            key = s.get("url") or s.get("title")
            if key and key not in seen_urls:
                seen_urls.add(key)
                deduped_snippets.append(s)

        # ---- Enrich top JIRA tickets via JIRA REST API ----

        glean_ticket_list = list(all_glean_tickets.values())
        enriched_tickets = []
        for t_info in glean_ticket_list[:8]:
            tid = t_info["ticket"]
            jdata = fetch_jira_ticket(tid)
            entry = {
                "ticket": tid,
                "url": t_info.get("url", f"https://jira.nutanix.com/browse/{tid}"),
                "glean_title": t_info.get("title", ""),
                "glean_snippet": t_info.get("snippet", ""),
            }
            if jdata:
                fields = jdata.get("fields", {})
                entry["jira_summary"] = fields.get("summary", "")
                entry["jira_status"] = (fields.get("status") or {}).get("name", "Unknown")
                resolution = fields.get("resolution")
                entry["jira_resolution"] = (resolution.get("name", "") if isinstance(resolution, dict) else "") if resolution else ""
                entry["jira_priority"] = (fields.get("priority") or {}).get("name", "")
                entry["jira_type"] = (fields.get("issuetype") or {}).get("name", "")
                entry["is_open"] = entry["jira_status"] not in ("Closed", "Resolved", "Done", "Won't Fix")
            else:
                entry["jira_summary"] = t_info.get("title", "")
                entry["jira_status"] = "Unknown"
                entry["jira_resolution"] = ""
                entry["is_open"] = True
            enriched_tickets.append(entry)

        # Sort: open tickets first, then by whether title/snippet shares keywords with exception
        enriched_tickets.sort(key=lambda e: (0 if e.get("is_open") else 1))

        # Flat ref list for backward compat
        glean_jira_refs = [t["ticket"] for t in enriched_tickets]

        # ---- Classify ----
        first_jira_data = None
        all_ticket_ids = list(jira_tickets or []) + glean_jira_refs
        if all_ticket_ids:
            first_jira_data = fetch_jira_ticket(all_ticket_ids[0])
        issue_type = classify_failure(exception_summary, exception_text, first_jira_data, None)

        # ---- AI analysis combining everything ----

        prompt_parts = [
            f"Testcase: {testcase_name}",
            f"Exception Summary: {exception_summary or 'N/A'}",
        ]
        if exception_text:
            prompt_parts.append(f"Traceback:\n```\n{exception_text[:3000]}\n```")
        if ai_summary:
            prompt_parts.append(f"AI Failure Summary:\n{ai_summary[:1500]}")
        if failure_stage:
            prompt_parts.append(f"Failure Stage: {failure_stage}")
        if test_log_url:
            prompt_parts.append(f"Log URL: {test_log_url}")

        if enriched_tickets:
            ticket_lines = []
            for t in enriched_tickets[:8]:
                status_str = t["jira_status"]
                if t.get("jira_resolution"):
                    status_str += f" ({t['jira_resolution']})"
                ticket_lines.append(
                    f"- {t['ticket']} [{status_str}] {t.get('jira_type', '')} — {t.get('jira_summary', t.get('glean_title', ''))}"
                )
            prompt_parts.append("Matching JIRA tickets found via Glean:\n" + "\n".join(ticket_lines))

        if deduped_snippets:
            snippet_text = "\n".join(
                f"- [{s['title']}]({s['url']}): {s['snippet']}" for s in deduped_snippets[:5]
            )
            prompt_parts.append(f"Glean search results:\n{snippet_text}")

        prompt_parts.append(
            "\nProvide a concise analysis:\n"
            "1. **Failure Classification**: Is this a Product Issue, Test Issue, or Infrastructure Issue? Why?\n"
            "2. **Matching JIRA Tickets**: Which of the above JIRA tickets most closely match this failure? "
            "Are they still open? Should the regression owner link this failure to one of them?\n"
            "3. **Recent Product Change Impact**: Could a recent product change have caused this failure? What evidence?\n"
            "4. **Recommended Action**: What should the regression owner do next? "
            "(e.g. link to existing ticket, create new ticket, rerun, mark as infra issue)\n"
            "Keep the response concise and actionable."
        )

        system_prompt = (
            "You are a senior regression test failure analyst at Nutanix. "
            "Given test failure details and JIRA tickets found via Glean search, "
            "classify the failure (Product Issue / Test Issue / Infrastructure Issue), "
            "identify the most closely matching JIRA tickets (especially open ENG tickets), "
            "assess whether recent product changes may have caused it, "
            "and recommend concrete next steps. Be specific and actionable."
        )

        ai_analysis = ""
        try:
            payload = {
                "model": "hack-reason",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": "\n\n".join(prompt_parts)},
                ],
                "max_tokens": 1500,
                "stream": False,
            }
            url = f"{AI_BASE}/chat/completions"
            headers = {
                "Authorization": f"Bearer {AI_API_KEY}",
                "Accept": "application/json",
                "Content-Type": "application/json",
            }
            data = json.dumps(payload).encode("utf-8")
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, context=SSL_CTX, timeout=90) as resp:
                if resp.getcode() == 200:
                    response_data = json.loads(resp.read().decode())
                    choices = response_data.get("choices", [])
                    if choices:
                        ai_analysis = (choices[0].get("message") or {}).get("content", "").strip()
        except Exception as ai_err:
            logger.warning(f"AI analysis in glean-search-single failed: {ai_err}")
            ai_analysis = f"AI analysis unavailable: {str(ai_err)[:200]}"

        return jsonify({
            "success": True,
            "issue_type": issue_type,
            "glean_jira_refs": glean_jira_refs[:10],
            "enriched_tickets": enriched_tickets[:8],
            "glean_snippets": deduped_snippets[:5],
            "ai_analysis": ai_analysis,
            "search_queries": queries_used,
        })
    except Exception as e:
        logger.error(f"Glean search single testcase error: {e}", exc_info=True)
        return jsonify({"success": False, "error": str(e)}), 500


def _extract_error_keywords(exception_summary, exception_text):
    """Pull the most useful error keywords from exception text for search."""
    import re as _re
    text = f"{exception_summary or ''} {(exception_text or '')[:500]}"
    # Extract class-style error names like NuTestError, TimeoutException, etc.
    error_names = _re.findall(r'[A-Z][a-zA-Z]*(?:Error|Exception|Failure|Fault)', text)
    # Extract key phrases: status codes, method names, short phrases
    status_codes = _re.findall(r'\b[45]\d{2}\b', text)
    keywords = list(dict.fromkeys(error_names + status_codes))
    return " ".join(keywords[:5])


def _build_glean_snippets(glean_data):
    """Extract snippet entries from a Glean search response."""
    if not glean_data or not isinstance(glean_data, dict):
        return []
    raw = glean_data.get("results", [])
    if isinstance(raw, dict):
        raw = raw.get("results", [])
    if not isinstance(raw, list):
        return []
    snippets = []
    for r in raw[:10]:
        if not isinstance(r, dict):
            continue
        doc = r.get("document", r)
        title = doc.get("title", "")
        url = doc.get("url", "")
        snippet_text = ""
        slist = r.get("snippets", [])
        if isinstance(slist, list):
            parts = []
            for s in slist:
                if isinstance(s, dict):
                    inner = s.get("snippet", s)
                    parts.append(inner.get("text", "") if isinstance(inner, dict) else str(inner))
                elif isinstance(s, str):
                    parts.append(s)
            snippet_text = " ".join(parts)[:500]
        if title or snippet_text:
            snippets.append({"title": title, "url": url, "snippet": snippet_text})
    return snippets


@app.route("/mcp/regression/failed-analysis/rdm-patterns", methods=["GET"])
@jwt_required
def get_rdm_patterns():
    """Return current RDM failure patterns."""
    try:
        if os.path.exists(RDM_PATTERNS_FILE):
            with open(RDM_PATTERNS_FILE, "r") as f:
                return jsonify(json.load(f))
        return jsonify({"patterns": []})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/failed-analysis/rdm-patterns", methods=["PUT"])
@jwt_required
def update_rdm_patterns():
    """Update RDM failure patterns. Expects { patterns: [...] }"""
    global RDM_PATTERNS
    try:
        data = request.get_json() or {}
        patterns = data.get("patterns")
        if patterns is None:
            return jsonify({"error": "patterns array is required"}), 400
        with open(RDM_PATTERNS_FILE, "w") as f:
            json.dump({"patterns": patterns}, f, indent=2)
        RDM_PATTERNS = _load_rdm_patterns()
        return jsonify({"success": True, "count": len(RDM_PATTERNS)})
    except Exception as e:
        logger.error(f"Error updating RDM patterns: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _fetch_test_result_for_task_and_name(task_id, test_name):
    """Fetch test result for a single task and test name. Returns one result row or None."""
    payload = {
        "raw_query": {"agave_task_id": {"$oid": task_id}},
        "only": "test,status,jira_tickets,comments",
        "start": 0,
        "limit": 500,
        "merge": False
    }
    try:
        resp = session.post(
            f"{JITA_BASE}/reports/agave_test_results",
            json=payload,
            timeout=60
        )
        if resp.status_code != 200:
            return None
        data = resp.json()
        results = data.get("data", [])
        for r in results:
            t = r.get("test") or {}
            name = t.get("name") if isinstance(t, dict) else (t if isinstance(t, str) else "")
            if name == test_name:
                return {
                    "status": (r.get("status") or "").lower(),
                    "jira_ticket": (r.get("jira_tickets") or [None])[0] if r.get("jira_tickets") else None,
                    "comment": r.get("comments") or None
                }
        return None
    except Exception as e:
        logger.warning(f"History fetch for task {task_id} test {test_name}: {e}")
        return None


@app.route("/mcp/regression/failed-analysis/history", methods=["GET"])
@jwt_required
def failed_analysis_history():
    """Return past 3 runs for a test on same or other branch. For each run: status, jira_ticket or comment."""
    test_name = request.args.get("test_name")
    branch = request.args.get("branch", "")
    same_branch = request.args.get("same_branch", "true").lower() in ("1", "true", "yes")
    tag = request.args.get("tag")
    if not test_name or not tag:
        return jsonify({"error": "test_name and tag are required"}), 400
    try:
        tasks = fetch_regression_tasks(tag=tag, task_ids=None)
        if not tasks:
            return jsonify({"runs": []})
        if same_branch:
            filtered = [t for t in tasks if (t.get("branch") or "") == branch]
        else:
            filtered = [t for t in tasks if (t.get("branch") or "") != branch]
        # Sort by _id descending (newest first), take 3
        filtered.sort(key=lambda t: t.get("_id") or {}, reverse=True)
        selected = filtered[:3]
        runs = []
        for task in selected:
            task_id = task.get("_id", {}).get("$oid") if isinstance(task.get("_id"), dict) else task.get("_id")
            if not task_id:
                continue
            row = _fetch_test_result_for_task_and_name(task_id, test_name)
            if row:
                runs.append({
                    "status": "passed" if row["status"] in ("passed", "succeeded", "success") else "failed",
                    "jira_ticket": row["jira_ticket"],
                    "comment": row["comment"]
                })
            else:
                runs.append({"status": "unknown", "jira_ticket": None, "comment": None})
        return jsonify({"runs": runs})
    except Exception as e:
        logger.error(f"Error fetching history: {e}", exc_info=True)
        return jsonify({"error": str(e), "runs": []}), 500


# ======================================================
# Failed Analysis — Saved Tags CRUD
# ======================================================

@app.route("/mcp/regression/failed-analysis/saved-tags", methods=["GET"])
@jwt_required
def list_saved_tags():
    """Return the list of saved tag names."""
    data = load_failed_analysis_tags()
    return jsonify({"tags": data.get("tags", [])})


@app.route("/mcp/regression/failed-analysis/saved-tags", methods=["POST"])
@jwt_required
def add_saved_tag():
    """Add a new tag to the saved list."""
    body = request.get_json() or {}
    tag_name = (body.get("tag") or "").strip()
    if not tag_name:
        return jsonify({"error": "tag is required"}), 400
    data = load_failed_analysis_tags()
    existing = [t["name"] if isinstance(t, dict) else t for t in data["tags"]]
    if tag_name in existing:
        return jsonify({"error": "Tag already exists"}), 409
    data["tags"].append({"name": tag_name, "added_at": datetime.utcnow().isoformat() + "Z"})
    save_failed_analysis_tags(data)
    return jsonify({"success": True, "tags": data["tags"]})


@app.route("/mcp/regression/failed-analysis/saved-tags/<path:tag_name>", methods=["DELETE"])
@jwt_required
def delete_saved_tag(tag_name):
    """Remove a tag and its cached analysis results."""
    data = load_failed_analysis_tags()
    original_len = len(data["tags"])
    data["tags"] = [
        t for t in data["tags"]
        if (t["name"] if isinstance(t, dict) else t) != tag_name
    ]
    if len(data["tags"]) == original_len:
        return jsonify({"error": "Tag not found"}), 404
    save_failed_analysis_tags(data)
    delete_failed_analysis_results(tag_name)
    return jsonify({"success": True, "tags": data["tags"]})


@app.route("/mcp/regression/failed-analysis/saved-tags/<path:tag_name>/results", methods=["GET"])
@jwt_required
def get_saved_tag_results(tag_name):
    """Return cached analysis results for a saved tag."""
    cached = load_failed_analysis_results(tag_name)
    if cached is None:
        return jsonify({"results": [], "current_branch": "", "saved_at": None, "cursor_ai": {}})
    return jsonify(cached)


@app.route("/mcp/regression/failed-analysis/saved-tags/<path:tag_name>/results", methods=["PUT"])
@jwt_required
def save_saved_tag_results(tag_name):
    """Save / overwrite analysis results for a saved tag."""
    body = request.get_json() or {}
    results = body.get("results", [])
    current_branch = body.get("current_branch", "")
    cursor_ai = body.get("cursor_ai", {}) or {}
    payload = {
        "tag": tag_name,
        "results": results,
        "current_branch": current_branch,
        "cursor_ai": cursor_ai,
        "saved_at": datetime.utcnow().isoformat() + "Z",
        "count": len(results),
    }
    save_failed_analysis_results(tag_name, payload)
    return jsonify({"success": True, "count": len(results), "saved_at": payload["saved_at"]})


# ======================================================
# Re-trigger failed testcases via Jita agave_tasks rerun API
# ======================================================

@app.route("/mcp/regression/failed-analysis/retrigger", methods=["POST"])
@jwt_required
def retrigger_failed_testcases():
    """
    Re-trigger selected failed testcases.
    Groups tests by their agave_task_id, fetches the original task to build default
    payload, applies user overrides, and POSTs to the Jita rerun API.

    Request body:
      tests: list of {testcase_id, testcase_name, agave_task_id}
      overrides (all optional):
        nos_commit, nos_branch, nos_gbn,
        pc_commit, pc_branch, pc_gbn,
        nutest_branch, nutest_commit,
        patch_url, resource_pool, username
    """
    try:
        data = request.get_json() or {}
        tests = data.get("tests", [])
        if not tests:
            return jsonify({"error": "No tests provided"}), 400

        overrides = data.get("overrides", {})

        current_username = g.current_user.get("sub", "")
        user_auth = _get_user_credentials(current_username)
        if not user_auth:
            return jsonify({
                "error": "Session credentials expired. Please re-login to retrigger.",
                "code": "CREDENTIALS_EXPIRED"
            }), 401

        # Group test names by agave_task_id (as {"name": ...} dicts for JITA)
        tests_by_task = {}
        for t in tests:
            atid = t.get("agave_task_id")
            if not atid:
                continue
            name = t.get("testcase_name", "")
            if name:
                tests_by_task.setdefault(atid, []).append({"name": name})

        if not tests_by_task:
            return jsonify({"error": "No valid agave_task_id found in selected tests"}), 400

        rerun_results = []

        for task_id, task_tests in tests_by_task.items():
            try:
                # Fetch original task details
                task_data = fetch_agave_task(task_id)
                if not task_data:
                    rerun_results.append({
                        "agave_task_id": task_id,
                        "success": False,
                        "error": "Could not fetch task details"
                    })
                    continue

                # Build rerun payload from original task, apply overrides
                payload = _build_rerun_payload(task_data, task_tests, overrides, current_username)

                # POST to rerun endpoint (use data=json.dumps per JITA API)
                rerun_url = f"{JITA_BASE}/agave_tasks/{task_id}/rerun"
                logger.info(f"[retrigger] POST {rerun_url} with {len(task_tests)} test(s): {task_tests}")
                logger.info(f"[retrigger] Payload keys: {list(payload.keys())}")
                resp = requests.post(
                    rerun_url,
                    data=json.dumps(payload),
                    auth=user_auth,
                    verify=False,
                    timeout=60
                )

                resp_text = resp.text[:500]
                logger.info(f"[retrigger] Response {resp.status_code} for task {task_id}: {resp_text}")

                try:
                    resp_data = resp.json()
                except Exception:
                    resp_data = {}

                if resp_data.get("success") is True:
                    new_id = resp_data.get("_id")
                    if isinstance(new_id, dict) and "$oid" in new_id:
                        new_id = new_id["$oid"]
                    rerun_results.append({
                        "agave_task_id": task_id,
                        "success": True,
                        "rerun_task_id": new_id,
                        "message": resp_data.get("message", ""),
                    })
                else:
                    error_msg = resp_data.get("message") or resp_data.get("error") or f"HTTP {resp.status_code}: {resp_text}"
                    logger.warning(f"[retrigger] Failed for task {task_id}: {error_msg}")
                    rerun_results.append({
                        "agave_task_id": task_id,
                        "success": False,
                        "error": error_msg
                    })
            except Exception as e:
                logger.error(f"[retrigger] Error rerunning task {task_id}: {e}", exc_info=True)
                rerun_results.append({
                    "agave_task_id": task_id,
                    "success": False,
                    "error": str(e)
                })

        succeeded = [r for r in rerun_results if r.get("success")]
        failed = [r for r in rerun_results if not r.get("success")]
        return jsonify({
            "success": len(failed) == 0,
            "total": len(rerun_results),
            "succeeded": len(succeeded),
            "failed": len(failed),
            "results": rerun_results,
        })

    except Exception as e:
        logger.error(f"[retrigger] Unhandled error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _build_rerun_payload(task_data, task_tests, overrides, username):
    """
    Build a minimal rerun payload matching the JITA /rerun API contract.
    task_tests is a list of {"name": "..."} dicts for the tests to rerun.
    The JITA rerun API inherits most settings from the original task;
    we only send tests, infra, and explicit user overrides.
    """
    # Infra from the original task's requested_hardware
    requested_hw = task_data.get("requested_hardware", {})
    infra = requested_hw.get("infra", [])
    if overrides.get("resource_pool"):
        infra = [{
            "kind": "ON_PREM",
            "type": "node_pool",
            "entries": [overrides["resource_pool"]]
        }]

    payload = {
        "tests": task_tests,
        "infra": infra,
    }

    # --- NOS overrides ---
    if overrides.get("nos_branch"):
        payload["branch"] = overrides["nos_branch"]
    if overrides.get("nos_commit"):
        payload["commit_id"] = overrides["nos_commit"]
    if overrides.get("nos_gbn"):
        try:
            payload["gbn"] = int(overrides["nos_gbn"])
            payload["image_gbn"] = int(overrides["nos_gbn"])
        except (ValueError, TypeError):
            pass
    if overrides.get("nos_tag"):
        payload["tag"] = overrides["nos_tag"]
    if overrides.get("nos_build_type"):
        payload["build_type"] = overrides["nos_build_type"]

    # --- Nutest overrides ---
    if overrides.get("nutest_branch"):
        payload["nutest-py3-tests_branch"] = overrides["nutest_branch"]
    if overrides.get("nutest_commit"):
        payload["nutest-py3-tests_commit"] = overrides["nutest_commit"]

    # --- PC overrides (resource_manager_json) ---
    pc_keys = ("pc_commit", "pc_branch", "pc_gbn", "pc_tag", "pc_build_type")
    if any(overrides.get(k) for k in pc_keys):
        rm_json = {}
        pc_build = {}
        if overrides.get("pc_commit"):
            pc_build["commit_id"] = overrides["pc_commit"]
        if overrides.get("pc_branch"):
            pc_build["branch"] = overrides["pc_branch"]
        if overrides.get("pc_gbn"):
            try:
                pc_build["gbn"] = int(overrides["pc_gbn"])
            except (ValueError, TypeError):
                pass
        if overrides.get("pc_tag"):
            pc_build["tag"] = overrides["pc_tag"]
        if overrides.get("pc_build_type"):
            pc_build["build_type"] = overrides["pc_build_type"]
        if pc_build:
            rm_json["PRISM_CENTRAL"] = {"build": pc_build}
        payload["resource_manager_json"] = rm_json

    # --- Patch URLs → tester_tags ---
    tester_tags = list(task_data.get("tester_tags", []))
    tags_changed = False
    if overrides.get("patch_url"):
        patch_url = overrides["patch_url"].strip()
        if patch_url:
            patch_tag = f"patch_url__{patch_url}"
            if patch_tag not in tester_tags:
                tester_tags.append(patch_tag)
                tags_changed = True
    if overrides.get("framework_patch_url"):
        fw_url = overrides["framework_patch_url"].strip()
        if fw_url:
            fw_tag = f"framework_patch_url__{fw_url}"
            if fw_tag not in tester_tags:
                tester_tags.append(fw_tag)
                tags_changed = True
    if tags_changed:
        payload["tester_tags"] = tester_tags

    return payload


# ======================================================
# Testcase Management
# ======================================================

TESTCASE_MGMT_BRANCHES = {
    "master": {"milestone": "master", "team_prefix": "master", "test_set_regex": "test_sets/milestones/master/"},
    "ganges-7.6-stable": {"milestone": "7.6", "team_prefix": "7.6", "test_set_regex": "test_sets/milestones/7.6/"},
    "ganges-7.5-stable": {"milestone": "7.5", "team_prefix": "7.5", "test_set_regex": "test_sets/milestones/7.5/"},
}

TESTCASE_MGMT_TEAMS = ["CDP", "AHV"]


def _resolve_branch_config(branch):
    """Resolve branch config from static map or dynamically from a version string.

    Accepts full branch names (e.g. 'ganges-7.7-stable') as well as bare
    version numbers (e.g. '7.7', '7.8').  For master and pre-configured
    branches the static TESTCASE_MGMT_BRANCHES entry is returned.  For
    anything else a release-style config is generated on-the-fly.
    """
    if branch in TESTCASE_MGMT_BRANCHES:
        return TESTCASE_MGMT_BRANCHES[branch]

    version = branch
    if branch.startswith("ganges-") and branch.endswith("-stable"):
        version = branch.replace("ganges-", "").replace("-stable", "")

    if re.match(r"^\d+\.\d+(\.\d+)?$", version):
        return {
            "milestone": version,
            "team_prefix": version,
            "test_set_regex": f"test_sets/milestones/{version}/",
        }

    return None


def _tcms_auth():
    return (TCMS_USER, TCMS_PASSWORD)


def _build_aggregate_payload(milestone, team_prefix, team, test_set_regex, skip, limit):
    """Build the MongoDB aggregation pipeline payload for the TCMS POST API.

    Master and release branches use different filter clauses:
    - Master includes release_name exclusion and metadata tag exclusions.
    - Release branches include tc_type filter for regression/smart_qual.
    Both include a team-specific test_sets regex.
    """
    is_master = milestone == "master"
    team_test_set_regex = f"{test_set_regex}{team}/"

    if is_master:
        match_clause = {"$and": [
            {
                "target_milestone": milestone,
                "last_result": {"$elemMatch": {"pass_name": "overall"}},
                "deleted": False,
            },
            {"test_case.test_sets": {"$regex": test_set_regex, "$options": "i"}},
            {"release_name": {"$ne": milestone}},
            {"test_case.metadata.tags": {"$nin": ["SYSTEST_LONGEVITY", "LIMITED_RUNS"]}},
            {"additional_data.team": f"{team_prefix}/{team}"},
            {"test_case.test_sets": {"$regex": team_test_set_regex, "$options": "i"}},
            {"test_case.deprecated": False},
        ]}
    else:
        match_clause = {"$and": [
            {
                "target_milestone": milestone,
                "last_result": {"$elemMatch": {"pass_name": "overall"}},
                "deleted": False,
                "tc_type": {"$in": ["regression", "smart_qual"]},
            },
            {"test_case.test_sets": {"$regex": test_set_regex, "$options": "i"}},
            {"additional_data.team": f"{team_prefix}/{team}"},
            {"test_case.test_sets": {"$regex": team_test_set_regex, "$options": "i"}},
            {"test_case.deprecated": False},
        ]}

    return json.dumps([
        {"$match": match_clause},
        {"$sort": {"name": 1}},
        {"$skip": skip},
        {"$limit": limit},
    ])


def _normalize_testcase(item):
    """Extract a flat dict from a raw TCMS milestone_all_test_cases record."""
    tc = item.get("test_case", {})
    meta = tc.get("metadata", {})
    ad = item.get("additional_data", {})
    score = item.get("test_score", {})

    last_result_list = item.get("last_result", [])
    last_status = ""
    is_triaged = False
    issue_type = ""
    last_run_tickets = []
    last_run_date = None
    last_passed_date = None
    if isinstance(last_result_list, list) and last_result_list:
        entry = last_result_list[0]
        run_info = entry.get("run", {})
        last_status = run_info.get("status", "")
        is_triaged = run_info.get("is_triaged", False)
        issue_type = run_info.get("issue_type", "")
        last_run_tickets = run_info.get("tickets", [])
        run_start = run_info.get("start_time", {})
        if isinstance(run_start, dict) and "$date" in run_start:
            last_run_date = datetime.utcfromtimestamp(run_start["$date"] / 1000).strftime("%Y-%m-%d %H:%M")
        succeeded_info = entry.get("succeeded", {})
        if isinstance(succeeded_info, dict) and succeeded_info:
            succ_start = succeeded_info.get("start_time", {})
            if isinstance(succ_start, dict) and "$date" in succ_start:
                last_passed_date = datetime.utcfromtimestamp(succ_start["$date"] / 1000).strftime("%Y-%m-%d %H:%M")

    published_qi = None
    published_success_ops = None
    published_total_ops = None
    if isinstance(last_result_list, list) and last_result_list:
        published_info = last_result_list[0].get("published", {})
        if isinstance(published_info, dict) and published_info:
            published_qi = published_info.get("operation_success_percentage")
            published_success_ops = published_info.get("successful_operations")
            published_total_ops = published_info.get("total_operations")

    ect = item.get("execution_cycle_time", {})
    automated_date_raw = ad.get("automated_date", {})
    automated_date = None
    if isinstance(automated_date_raw, dict) and "$date" in automated_date_raw:
        automated_date = datetime.utcfromtimestamp(automated_date_raw["$date"] / 1000).strftime("%Y-%m-%d")

    return {
        "oid": (item.get("_id", {}).get("$oid", "") if isinstance(item.get("_id"), dict) else ""),
        "name": item.get("name", ""),
        "path": tc.get("path", ""),
        "owners": tc.get("owners", []),
        "priority": meta.get("priority", ""),
        "summary": meta.get("summary", ""),
        "components": meta.get("components", []),
        "primary_component": meta.get("primary_component", ""),
        "services": meta.get("services", []),
        "tags": [],
        "metadata_tags": meta.get("tags", []),
        "test_sets": tc.get("test_sets", []),
        "team": ad.get("team", []),
        "target_service": item.get("target_service", ""),
        "target": item.get("target", ""),
        "framework": tc.get("framework", ""),
        "last_status": last_status,
        "last_run_date": last_run_date,
        "last_passed_date": last_passed_date,
        "is_triaged": is_triaged,
        "issue_type": issue_type,
        "last_run_tickets": last_run_tickets,
        "success_percentage": ad.get("success_percentage"),
        "avg_run_duration": ad.get("avg_run_duration"),
        "automated_date": automated_date,
        "one_month_mttr": ect.get("one_month_mttr"),
        "three_months_mttr": ect.get("three_months_mttr"),
        "published_qi": published_qi,
        "published_success_ops": published_success_ops,
        "published_total_ops": published_total_ops,
        "stability": score.get("stability"),
        "effectiveness": score.get("effectiveness"),
        "total_results": score.get("total_results"),
        "tickets": item.get("tickets", []),
        "resource_spec": tc.get("resource_spec", []),
    }


def _fetch_tags_for_testcases(testcases, branch_key):
    """Batch-fetch tags from the GET all_test_cases API using regex matching."""
    if not testcases:
        return testcases

    name_map = {tc["name"]: tc for tc in testcases}
    target_branch = branch_key

    batch_size = 50
    names = list(name_map.keys())

    def _fetch_batch(batch_names):
        for tc_name in batch_names:
            try:
                raw_query = json.dumps({
                    "$and": [
                        {
                            "target_service": "NutestPy3Tests",
                            "target_branch": target_branch,
                            "target_package_type": "tar",
                            "deleted": False,
                        },
                        {"test_case.name": tc_name},
                        {"test_case.deprecated": False},
                    ]
                })
                url = (
                    f"{TCMS_TESTDB_BASE}/all_test_cases"
                    f"?raw_query={urllib.parse.quote(raw_query)}&sort=name&limit=1"
                )
                resp = requests.get(url, auth=_tcms_auth(), verify=False, timeout=30)
                if resp.status_code == 200:
                    data = resp.json().get("data", [])
                    if data:
                        tags = data[0].get("additional_data", {}).get("tags", [])
                        if tc_name in name_map:
                            name_map[tc_name]["tags"] = tags
            except Exception as exc:
                logger.warning(f"Failed to fetch tags for {tc_name}: {exc}")

    with ThreadPoolExecutor(max_workers=10) as pool:
        for i in range(0, len(names), batch_size):
            batch = names[i:i + batch_size]
            pool.submit(_fetch_batch, batch)

    return testcases


def _tc_data_file(branch, team):
    """Return the path for a per-branch/team JSON file."""
    safe_name = f"testcase_management_{branch}_{team}.json".replace("/", "_")
    return os.path.join(TESTCASE_MGMT_DATA_DIR, safe_name)


def _load_tc_data(branch, team):
    fpath = _tc_data_file(branch, team)
    try:
        with open(fpath, "r") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_updated": None, "branch": branch, "team": team, "testcases": []}


def _save_tc_data(branch, team, data):
    os.makedirs(TESTCASE_MGMT_DATA_DIR, exist_ok=True)
    fpath = _tc_data_file(branch, team)
    with open(fpath, "w") as f:
        json.dump(data, f, indent=2, default=str)


@app.route("/mcp/regression/testcase-mgmt/fetch-data", methods=["GET"])
@jwt_required
def testcase_mgmt_fetch_data():
    """Fetch all test cases from TCMS for a given branch/team and persist to JSON."""
    branch = request.args.get("branch", "master")
    team = request.args.get("team", "CDP")
    page_limit = 500

    branch_cfg = _resolve_branch_config(branch)
    if not branch_cfg:
        return jsonify({"error": f"Unknown branch: {branch}. Use a known branch name or a version number like 7.7"}), 400

    milestone = branch_cfg["milestone"]
    team_prefix = branch_cfg["team_prefix"]
    test_set_regex = branch_cfg["test_set_regex"]

    all_testcases = []
    skip = 0

    try:
        while True:
            payload = _build_aggregate_payload(milestone, team_prefix, team, test_set_regex, skip, page_limit)
            resp = requests.post(
                f"{TCMS_BASE}/milestone_all_test_cases/aggregate",
                data=payload,
                auth=_tcms_auth(),
                verify=False,
                timeout=120,
                headers={"Content-Type": "application/json"},
            )
            if resp.status_code != 200:
                logger.error(f"TCMS aggregate API returned {resp.status_code}: {resp.text[:500]}")
                break

            batch = resp.json().get("data", [])
            if not batch:
                break

            for item in batch:
                all_testcases.append(_normalize_testcase(item))

            logger.info(f"Fetched {len(batch)} testcases (skip={skip}) for {branch}/{team}")
            if len(batch) < page_limit:
                break
            skip += page_limit

        logger.info(f"Total testcases fetched from aggregate API: {len(all_testcases)} for {branch}/{team}")

        _fetch_tags_for_testcases(all_testcases, branch)

        now = datetime.utcnow().isoformat() + "Z"
        data = {
            "last_updated": now,
            "branch": branch,
            "team": team,
            "testcases": all_testcases,
        }
        _save_tc_data(branch, team, data)

        return jsonify({
            "status": "ok",
            "branch": branch,
            "team": team,
            "count": len(all_testcases),
            "last_updated": now,
        })

    except Exception as e:
        logger.error(f"Error fetching testcase data: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/testcase-mgmt/testcases", methods=["GET"])
@jwt_required
def testcase_mgmt_get_testcases():
    """Return testcases from local JSON with optional filters."""
    branch = request.args.get("branch", "master")
    team = request.args.get("team", "CDP")
    tag_filter = request.args.get("tags", "")
    name_filter = request.args.get("name", "").lower()
    status_filter = request.args.get("status", "")

    data = _load_tc_data(branch, team)
    all_testcases = data.get("testcases", [])
    testcases = list(all_testcases)

    if tag_filter:
        filter_tags = [t.strip().lower() for t in tag_filter.split(",") if t.strip()]
        testcases = [
            tc for tc in testcases
            if any(ft in [t.lower() for t in tc.get("tags", [])] for ft in filter_tags)
        ]

    if name_filter:
        exact_match = request.args.get("exact_match", "false").lower() in ("1", "true", "yes")
        name_terms = [t.strip() for t in name_filter.split(",") if t.strip()]
        if name_terms:
            if exact_match:
                testcases = [tc for tc in testcases if tc.get("name", "").lower() in name_terms]
            else:
                testcases = [tc for tc in testcases if any(term in tc.get("name", "").lower() for term in name_terms)]

    if status_filter:
        testcases = [tc for tc in testcases if tc.get("last_status", "").lower() == status_filter.lower()]

    all_tags = set()
    for tc in all_testcases:
        for t in tc.get("tags", []):
            all_tags.add(t)

    return jsonify({
        "branch": branch,
        "team": team,
        "count": len(testcases),
        "total_count": len(all_testcases),
        "last_updated": data.get("last_updated"),
        "available_tags": sorted(all_tags),
        "testcases": testcases,
    })


@app.route("/mcp/regression/testcase-mgmt/tags/add", methods=["POST"])
@jwt_required
def testcase_mgmt_add_tags():
    """Add tags to selected test cases via TCMS write API."""
    body = request.get_json(force=True)
    testcase_oids = body.get("testcase_oids", [])
    tags_to_add = body.get("tags", [])
    branch = body.get("branch", "master")
    team = body.get("team", "CDP")

    if not testcase_oids or not tags_to_add:
        return jsonify({"error": "testcase_oids and tags are required"}), 400

    results = {"success": 0, "failed": 0, "errors": []}
    successful_oids = []

    for oid in testcase_oids:
        try:
            url = f"{TCMS_WRITE_BASE}/all_test_cases/tags/{oid}"
            resp = requests.post(
                url,
                auth=_tcms_auth(),
                data=json.dumps({"tags": tags_to_add}),
                headers={"Content-Type": "application/json"},
                verify=False,
                timeout=30,
            )
            if resp.status_code in (200, 201):
                results["success"] += 1
                successful_oids.append(oid)
                logger.info(f"Successfully added tags to testcase {oid}: {resp.json() if resp.text else 'OK'}")
            else:
                results["failed"] += 1
                error_detail = {"oid": oid, "status": resp.status_code, "response": resp.text[:200]}
                results["errors"].append(error_detail)
                logger.error(f"Failed to add tags to testcase {oid}: status={resp.status_code}, response={resp.text[:200]}")
        except Exception as exc:
            results["failed"] += 1
            results["errors"].append({"oid": oid, "error": str(exc)})
            logger.error(f"Exception adding tags to testcase {oid}: {exc}")

    data = _load_tc_data(branch, team)
    for tc in data.get("testcases", []):
        if tc.get("oid") in successful_oids:
            existing = tc.get("tags", [])
            for tag in tags_to_add:
                if tag not in existing:
                    existing.append(tag)
            tc["tags"] = existing
    data["last_updated"] = datetime.utcnow().isoformat() + "Z"
    _save_tc_data(branch, team, data)

    return jsonify(results)


@app.route("/mcp/regression/testcase-mgmt/tags/delete", methods=["POST"])
@jwt_required
def testcase_mgmt_delete_tags():
    """Delete tags from selected test cases via TCMS write API."""
    body = request.get_json(force=True)
    testcase_oids = body.get("testcase_oids", [])
    tags_to_delete = body.get("tags", [])
    branch = body.get("branch", "master")
    team = body.get("team", "CDP")

    if not testcase_oids or not tags_to_delete:
        return jsonify({"error": "testcase_oids and tags are required"}), 400

    results = {"success": 0, "failed": 0, "errors": []}
    successful_oids = []

    for oid in testcase_oids:
        try:
            url = f"{TCMS_WRITE_BASE}/all_test_cases/tags/{oid}"
            resp = requests.delete(
                url,
                auth=_tcms_auth(),
                data=json.dumps({"tags": tags_to_delete}),
                headers={"Content-Type": "application/json"},
                verify=False,
                timeout=30,
            )
            if resp.status_code in (200, 204):
                results["success"] += 1
                successful_oids.append(oid)
                logger.info(f"Successfully deleted tags from testcase {oid}: {resp.json() if resp.text else 'OK'}")
            else:
                results["failed"] += 1
                error_detail = {"oid": oid, "status": resp.status_code, "response": resp.text[:200]}
                results["errors"].append(error_detail)
                logger.error(f"Failed to delete tags from testcase {oid}: status={resp.status_code}, response={resp.text[:200]}")
        except Exception as exc:
            results["failed"] += 1
            results["errors"].append({"oid": oid, "error": str(exc)})
            logger.error(f"Exception deleting tags from testcase {oid}: {exc}")

    data = _load_tc_data(branch, team)
    for tc in data.get("testcases", []):
        if tc.get("oid") in successful_oids:
            tc["tags"] = [t for t in tc.get("tags", []) if t not in tags_to_delete]
    data["last_updated"] = datetime.utcnow().isoformat() + "Z"
    _save_tc_data(branch, team, data)

    return jsonify(results)


@app.route("/mcp/regression/testcase-mgmt/resource-spec/download", methods=["GET"])
@jwt_required
def testcase_mgmt_resource_spec_download():
    """Generate an Excel workbook grouping test cases by unique resource_spec."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    branch = request.args.get("branch", "master")
    team = request.args.get("team", "CDP")
    data = _load_tc_data(branch, team)
    testcases = data.get("testcases", [])

    if not testcases:
        return jsonify({"error": "No testcase data found. Reload from TCMS first."}), 404

    def _spec_fingerprint(spec_list):
        """Canonical string key for grouping (ignores resource 'name')."""
        cleaned = []
        for item in (spec_list or []):
            entry = {k: v for k, v in sorted(item.items()) if k != "name"}
            cleaned.append(json.dumps(entry, sort_keys=True))
        return "|".join(sorted(cleaned))

    def _format_resource(r):
        """Return a multi-line human-readable string for one resource item."""
        lines = []
        lines.append(f"name: {r.get('name', '—')}")
        lines.append(f"type: {r.get('type', '—')}")
        hw = r.get("hardware", {})
        if hw:
            lines.append(f"min_host_gb_ram: {hw.get('min_host_gb_ram', '—')}")
            lines.append(f"min_host_cpu_cores: {hw.get('min_host_cpu_cores', '—')}")
            lines.append(f"cluster_min_nodes: {hw.get('cluster_min_nodes', r.get('cluster_min_nodes', '—'))}")
        elif "cluster_min_nodes" in r:
            lines.append(f"cluster_min_nodes: {r['cluster_min_nodes']}")
        sc = r.get("scaleout", {})
        if sc:
            lines.append(f"scaleout.num_instances: {sc.get('num_instances', '—')}")
        deps = r.get("dependencies")
        if deps:
            lines.append(f"dependencies: {', '.join(deps)}")
        prov = r.get("provider")
        if isinstance(prov, dict):
            lines.append(f"provider.host: {prov.get('host', '—')}")
        can_run = r.get("can_run_on_provider")
        if can_run:
            lines.append(f"can_run_on_provider: {', '.join(can_run) if isinstance(can_run, list) else can_run}")
        for k, v in sorted(r.items()):
            if k not in ("name", "type", "hardware", "cluster_min_nodes", "scaleout",
                         "dependencies", "provider", "can_run_on_provider"):
                lines.append(f"{k}: {json.dumps(v) if isinstance(v, (dict, list)) else v}")
        return "\n".join(lines)

    def _format_spec_full(spec_list):
        """Return the full resource_spec as readable text (all resources)."""
        if not spec_list:
            return "None"
        blocks = []
        for idx, r in enumerate(spec_list, 1):
            blocks.append(f"[Resource {idx}]\n{_format_resource(r)}")
        return "\n\n".join(blocks)

    groups = defaultdict(list)
    for tc in testcases:
        key = _spec_fingerprint(tc.get("resource_spec", []))
        groups[key].append(tc)
    key_to_id = {}
    for idx, key in enumerate(sorted(groups.keys(), key=lambda k: -len(groups[k])), 1):
        key_to_id[key] = idx

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    grp_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))
    wrap = Alignment(wrap_text=True, vertical="top")

    def _write_header(ws, headers):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = wrap
            c.border = thin

    # ── Sheet 1: Testcases ──
    ws1 = wb.active
    ws1.title = "Testcases"
    _write_header(ws1, ["Testcase Name", "Resource Spec", "Group ID"])

    row = 2
    for tc in sorted(testcases, key=lambda t: key_to_id.get(_spec_fingerprint(t.get("resource_spec", [])), 0)):
        spec = tc.get("resource_spec", [])
        gid = key_to_id.get(_spec_fingerprint(spec), 0)
        ws1.cell(row=row, column=1, value=tc.get("name", "")).border = thin
        ws1.cell(row=row, column=1).alignment = wrap
        c_spec = ws1.cell(row=row, column=2, value=_format_spec_full(spec))
        c_spec.border = thin
        c_spec.alignment = wrap
        c_gid = ws1.cell(row=row, column=3, value=gid)
        c_gid.border = thin
        c_gid.alignment = Alignment(horizontal="center", vertical="top")
        row += 1

    ws1.column_dimensions["A"].width = 80
    ws1.column_dimensions["B"].width = 70
    ws1.column_dimensions["C"].width = 12

    # ── Sheet 2: Grouped Resource Specs ──
    ws2 = wb.create_sheet("Grouped Resource Specs")
    _write_header(ws2, ["Group ID", "Testcase Count", "Resource Spec", "Testcase Names"])

    sorted_groups = sorted(key_to_id.items(), key=lambda kv: kv[1])
    for key, gid in sorted_groups:
        tcs = groups[key]
        spec_list = tcs[0].get("resource_spec", [])
        r = gid + 1
        ws2.cell(row=r, column=1, value=gid).border = thin
        ws2.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws2.cell(row=r, column=2, value=len(tcs)).border = thin
        ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center", vertical="top")
        c_spec = ws2.cell(row=r, column=3, value=_format_spec_full(spec_list))
        c_spec.border = thin
        c_spec.alignment = wrap
        tc_names = "\n".join(t.get("name", "") for t in tcs)
        c_names = ws2.cell(row=r, column=4, value=tc_names)
        c_names.border = thin
        c_names.alignment = wrap

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 70
    ws2.column_dimensions["D"].width = 80

    # ── Sheet 3: Resource Spec Detail (flat table) ──
    ws3 = wb.create_sheet("Resource Detail")
    detail_headers = [
        "Group ID", "Resource #", "name", "type",
        "min_host_gb_ram", "min_host_cpu_cores", "cluster_min_nodes",
        "scaleout.num_instances", "dependencies", "provider.host",
        "Extra Parameters",
    ]
    _write_header(ws3, detail_headers)
    KNOWN_KEYS = {"name", "type", "hardware", "cluster_min_nodes", "scaleout",
                  "dependencies", "provider", "can_run_on_provider", "can_run_on_hardware"}

    dr = 2
    for key, gid in sorted_groups:
        tcs = groups[key]
        spec_list = tcs[0].get("resource_spec", [])
        for ri, res in enumerate(spec_list, 1):
            hw = res.get("hardware", {})
            extras = {k: v for k, v in res.items() if k not in KNOWN_KEYS}
            extra_str = "; ".join(f"{k}={json.dumps(v) if isinstance(v, (dict, list)) else v}"
                                  for k, v in sorted(extras.items())) if extras else ""
            vals = [
                gid,
                ri,
                res.get("name", ""),
                res.get("type", ""),
                hw.get("min_host_gb_ram", ""),
                hw.get("min_host_cpu_cores", ""),
                hw.get("cluster_min_nodes", res.get("cluster_min_nodes", "")),
                (res.get("scaleout") or {}).get("num_instances", ""),
                ", ".join(res.get("dependencies", [])) if res.get("dependencies") else "",
                (res.get("provider") or {}).get("host", "") if isinstance(res.get("provider"), dict) else "",
                extra_str,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws3.cell(row=dr, column=ci, value=v)
                c.border = thin
                c.alignment = wrap
            if ri == 1:
                for ci in range(1, len(vals) + 1):
                    ws3.cell(row=dr, column=ci).fill = grp_fill
            dr += 1

    ws3.column_dimensions["A"].width = 10
    ws3.column_dimensions["B"].width = 12
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 18
    ws3.column_dimensions["E"].width = 16
    ws3.column_dimensions["F"].width = 18
    ws3.column_dimensions["G"].width = 18
    ws3.column_dimensions["H"].width = 20
    ws3.column_dimensions["I"].width = 30
    ws3.column_dimensions["J"].width = 20
    ws3.column_dimensions["K"].width = 40

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"resource_spec_{branch}_{team}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


@app.route("/mcp/regression/testcase-mgmt/resolve-job-profiles", methods=["POST"])
@jwt_required
def testcase_mgmt_resolve_job_profiles():
    """Search JITA job profiles by prefix and cross-reference with testcase test_sets."""
    from urllib.parse import quote

    body = request.get_json(force=True)
    branch = body.get("branch", "master")
    team = body.get("team", "CDP")
    jp_prefix = body.get("jp_prefix", "").strip()
    tc_names = body.get("testcase_names", [])

    if not jp_prefix:
        return jsonify({"error": "jp_prefix is required"}), 400

    data = _load_tc_data(branch, team)
    all_tcs = data.get("testcases", [])

    if tc_names:
        name_set = set(tc_names)
        target_tcs = [tc for tc in all_tcs if tc.get("name") in name_set]
    else:
        target_tcs = all_tcs

    if not target_tcs:
        return jsonify({"error": "No matching testcases found"}), 404

    raw_query = json.dumps({"name": {"$regex": f"^{jp_prefix}", "$options": "i"}})
    try:
        resp = requests.get(
            f"{JITA_BASE}/job_profiles",
            params={"raw_query": quote(raw_query), "limit": 200},
            auth=JITA_SVC_AUTH,
            verify=False,
            timeout=30,
        )
        resp.raise_for_status()
        job_profiles = resp.json().get("data", [])
    except Exception as exc:
        return jsonify({"error": f"JITA search failed: {exc}"}), 502

    if not job_profiles:
        return jsonify({
            "matched": [],
            "unmatched_testcases": [tc.get("name", "") for tc in target_tcs],
            "total_matched_testcases": 0,
            "total_unmatched_testcases": len(target_tcs),
            "job_profiles_found": 0,
        })

    jp_map = {}
    for jp in job_profiles:
        jp_id = jp.get("_id", {}).get("$oid", "") if isinstance(jp.get("_id"), dict) else str(jp.get("_id", ""))
        jp_name = jp.get("name", "")
        jp_test_sets = set()
        for ts in jp.get("test_sets", jp.get("test_set", [])):
            if isinstance(ts, str):
                jp_test_sets.add(ts.lower())
            elif isinstance(ts, dict):
                ts_name = ts.get("name", ts.get("test_set", ""))
                if ts_name:
                    jp_test_sets.add(str(ts_name).lower())
        jp_map[jp_id] = {"name": jp_name, "test_sets": jp_test_sets, "testcases": []}

    matched_tc_names = set()
    for tc in target_tcs:
        tc_test_sets = {s.lower() for s in tc.get("test_sets", []) if isinstance(s, str)}
        for jp_id, jp_info in jp_map.items():
            if tc_test_sets & jp_info["test_sets"]:
                jp_info["testcases"].append(tc.get("name", ""))
                matched_tc_names.add(tc.get("name", ""))

    matched = []
    for jp_id, jp_info in jp_map.items():
        if jp_info["testcases"]:
            matched.append({
                "job_profile_id": jp_id,
                "job_profile_name": jp_info["name"],
                "testcase_count": len(jp_info["testcases"]),
                "testcases": jp_info["testcases"],
            })
    matched.sort(key=lambda x: -x["testcase_count"])

    unmatched = [tc.get("name", "") for tc in target_tcs if tc.get("name", "") not in matched_tc_names]

    return jsonify({
        "matched": matched,
        "unmatched_testcases": unmatched,
        "total_matched_testcases": len(matched_tc_names),
        "total_unmatched_testcases": len(unmatched),
        "job_profiles_found": len(job_profiles),
    })


@app.route("/mcp/regression/testcase-mgmt/branches", methods=["GET"])
@jwt_required
def testcase_mgmt_branches():
    """Return available branches and teams for the testcase management module.

    Also signals that the UI can send arbitrary release versions (e.g. '7.7')
    which the backend resolves dynamically.
    """
    return jsonify({
        "branches": list(TESTCASE_MGMT_BRANCHES.keys()),
        "teams": TESTCASE_MGMT_TEAMS,
        "custom_release_supported": True,
    })
# Dynamic Job Profile APIs
# ======================================================

@app.route("/mcp/regression/dynamic-jp/test-execution-history", methods=["POST"])
def dynamic_jp_test_execution_history():
    """Fetch detailed test execution history from JITA.

    Mirrors JITA /test_history: each row is one execution. When ``branch`` is set in
    the JSON body, results are restricted to that ``system_under_test.branch`` (query
    + case-insensitive post-filter). ``test_set`` and ``job_profile`` come from the
    **history row** (``test_set`` / ``test_set_name`` / ``AgaveTask`` and run
    ``label`` for JP), with task lookups only as fallback.
    """
    try:
        req_data = request.json or {}
        test_name = (req_data.get("test_name") or "").strip()
        page = int(req_data.get("page", 1))
        limit = int(req_data.get("limit", 50))
        sort_field = req_data.get("sort", "-start_time")
        branch_filter = (req_data.get("branch") or "").strip()

        if not test_name:
            return jsonify({"error": "test_name is required"}), 400

        raw_query = {"test.name": test_name}
        if branch_filter:
            raw_query["system_under_test.branch"] = branch_filter

        start = max(0, limit * (page - 1))
        raw_items = []
        total = 0

        logger.info(f"[test-exec-history] Querying: test.name={test_name}, page={page}")

        # Primary: GET agave_test_results (mirrors JITA frontend)
        try:
            params = {
                "start": start,
                "limit": limit,
                "sort": sort_field,
                "raw_query": json.dumps(raw_query),
            }
            resp = requests.get(
                f"{JITA_BASE}/agave_test_results",
                params=params,
                auth=JITA_SVC_AUTH,
                verify=False,
                timeout=90,
            )
            if resp.status_code == 200:
                jita_data = resp.json()
                raw_items = jita_data.get("data", [])
                total = jita_data.get("total", 0)
                logger.info(f"[test-exec-history] GET returned {total} total, {len(raw_items)} items")
        except requests.exceptions.Timeout:
            logger.warning("[test-exec-history] GET timed out, trying POST fallback")
        except Exception as e:
            logger.warning(f"[test-exec-history] GET failed: {e}")

        # Fallback: POST reports/agave_test_results
        if not raw_items:
            try:
                payload = {
                    "raw_query": raw_query,
                    "start": start,
                    "limit": limit,
                    "sort": sort_field,
                }
                resp2 = requests.post(
                    f"{JITA_BASE}/reports/agave_test_results",
                    json=payload,
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=90,
                )
                if resp2.status_code == 200:
                    data2 = resp2.json()
                    raw_items = data2.get("data", [])
                    total = data2.get("total", data2.get("metadata", {}).get("total", 0))
                    logger.info(f"[test-exec-history] POST returned {total} total, {len(raw_items)} items")
            except requests.exceptions.Timeout:
                logger.warning("[test-exec-history] POST also timed out")
            except Exception as e:
                logger.warning(f"[test-exec-history] POST fallback failed: {e}")

        jita_total_pre_branch = total

        def _sut_branch(item):
            return (item.get("system_under_test") or {}).get("branch") or ""

        def _branch_matches(item):
            if not branch_filter:
                return True
            return (_sut_branch(item) or "").strip().lower() == branch_filter.lower()

        # Match UI: use history rows for the selected branch (backup if JITA query is loose)
        if branch_filter:
            raw_items = [it for it in raw_items if _branch_matches(it)]
            total = len(raw_items)
            logger.info(
                f"[test-exec-history] After branch filter {branch_filter!r}: {len(raw_items)} rows "
                f"(pre-filter total from JITA was {jita_total_pre_branch})"
            )

        def _oid(val):
            if isinstance(val, dict) and "$oid" in val:
                return val["$oid"]
            return str(val) if val else None

        def _date(val):
            if isinstance(val, dict) and "$date" in val:
                return val["$date"]
            return val

        def _test_set_name_from_embedded_agave(agt):
            """Per-result test set from AgaveTask embed only."""
            if not isinstance(agt, dict):
                return ""
            s = (agt.get("test_set_name") or "").strip()
            if s:
                return s
            tso = agt.get("test_set")
            if isinstance(tso, dict):
                s = (tso.get("name") or "").strip()
                if s:
                    return s
            elif isinstance(tso, str) and tso.strip():
                return tso.strip()
            return ""

        def _test_set_from_history_row(item, agt):
            """JITA test history row: prefer top-level + AgaveTask (same as /test_history table)."""
            if isinstance(item, dict):
                s = (item.get("test_set_name") or "").strip()
                if s:
                    return s
                tso = item.get("test_set")
                if isinstance(tso, dict):
                    s = (tso.get("name") or "").strip()
                    if s:
                        return s
                elif isinstance(tso, str) and tso.strip():
                    return tso.strip()
            return _test_set_name_from_embedded_agave(agt)

        def _label_to_jp_display(label):
            """Run label is the JP line on JITA history, e.g. Some_JP_Name-(42)."""
            if not label or not isinstance(label, str):
                return ""
            return re.sub(r"-\(\d+\)$", "", label.strip())

        def _jp_name_from_history_row(item, agt):
            """JP for display: JITA /test_history uses run label; prefer that over job_profile_name."""
            if isinstance(agt, dict):
                lab = (agt.get("label") or "").strip()
                if lab:
                    d = _label_to_jp_display(lab)
                    if d:
                        return d
                    return lab
                jn = (agt.get("job_profile_name") or "").strip()
                if jn:
                    return jn
            if isinstance(item, dict):
                lab = (item.get("label") or "").strip()
                if lab:
                    d = _label_to_jp_display(lab)
                    if d:
                        return d
                    return lab
                jn = (item.get("job_profile_name") or "").strip()
                if jn:
                    return jn
            return ""

        def _test_names_from_ts_doc_tests_field(tests_field):
            """Normalize JITA test_sets.tests entries to full testcase name strings."""
            out = set()
            if not isinstance(tests_field, list):
                return out
            for entry in tests_field:
                if isinstance(entry, str) and entry.strip():
                    out.add(entry.strip())
                elif isinstance(entry, dict):
                    n = entry.get("name") or entry.get("test") or entry.get("path")
                    if isinstance(n, str) and n.strip():
                        out.add(n.strip())
            return out

        # Collect unique task IDs so we can look up test_set / job_profile.
        # Only the rows whose embedded history fields are missing a test set OR a JP
        # need the (slow) secondary task/JP/test_set lookups; rows that already carry
        # both are resolved directly, so we skip those network round-trips entirely.
        unique_task_ids = list({
            _oid(item.get("agave_task_id"))
            for item in raw_items
            if _oid(item.get("agave_task_id"))
            and (
                not _test_set_from_history_row(item, item.get("AgaveTask") or {})
                or not _jp_name_from_history_row(item, item.get("AgaveTask") or {})
            )
        })

        task_info = {}  # task_id -> {test_set_name, job_profile_name, branch}
        if unique_task_ids:
            try:
                tids_for_query = [{"$oid": tid} for tid in unique_task_ids[:100]]
                rq = json.dumps({"_id": {"$in": tids_for_query}})
                task_resp = requests.get(
                    f"{JITA_BASE}/tasks",
                    params={
                        "raw_query": rq,
                        "limit": len(tids_for_query),
                        "only": "_id,test_sets,label,branch,job_profile",
                    },
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=60,
                )
                if task_resp.status_code == 200:
                    # Collect JP IDs to resolve names in bulk
                    jp_id_map = {}  # jp_oid -> None (will fill with name)
                    task_items = task_resp.json().get("data", [])
                    for t in task_items:
                        jp_ref = t.get("job_profile")
                        jp_id = _oid(jp_ref) if jp_ref else None
                        if jp_id:
                            jp_id_map[jp_id] = None

                    # Bulk-fetch JP names
                    if jp_id_map:
                        try:
                            jp_ids_for_q = [{"$oid": jid} for jid in list(jp_id_map.keys())[:100]]
                            jp_rq = json.dumps({"_id": {"$in": jp_ids_for_q}})
                            jp_resp = requests.get(
                                f"{JITA_BASE}/job_profiles",
                                params={"raw_query": jp_rq, "limit": len(jp_ids_for_q), "only": "_id,name"},
                                auth=JITA_SVC_AUTH, verify=False, timeout=30,
                            )
                            if jp_resp.status_code == 200:
                                for jp_item in jp_resp.json().get("data", []):
                                    jid = _oid(jp_item.get("_id"))
                                    if jid:
                                        jp_id_map[jid] = jp_item.get("name", "")
                        except Exception as e:
                            logger.warning(f"[test-exec-history] Failed to bulk-fetch JP names: {e}")

                    for t in task_items:
                        tid = _oid(t.get("_id"))
                        if not tid:
                            continue
                        ts_list = t.get("test_sets") or []
                        ts_name = ts_list[0].get("name", "") if ts_list else ""
                        ts_refs = []
                        for el in ts_list:
                            if not isinstance(el, dict):
                                continue
                            rid = _oid(el.get("_id") or el)
                            nm = (el.get("name") or "").strip()
                            if rid or nm:
                                ts_refs.append({"id": rid, "name": nm})

                        # Get JP name: prefer resolved name, fall back to label parsing
                        jp_ref = t.get("job_profile")
                        jp_id = _oid(jp_ref) if jp_ref else None
                        jp_name = jp_id_map.get(jp_id, "") if jp_id else ""
                        if not jp_name:
                            label = t.get("label", "")
                            jp_name = re.sub(r"-\(\d+\)$", "", label) if label else ""

                        task_info[tid] = {
                            "test_set_name": ts_name,
                            "ts_refs": ts_refs,
                            "job_profile_name": jp_name,
                            "branch": t.get("branch", ""),
                        }
                    logger.info(f"[test-exec-history] Fetched info for {len(task_info)} tasks, {len(jp_id_map)} unique JPs")
            except Exception as e:
                logger.warning(f"[test-exec-history] Failed to fetch task details: {e}")

        # When a task has multiple test sets and a row has no embedded test set, match by testcase name.
        ts_id_to_testnames = {}
        ts_id_to_tsname = {}
        if task_info:
            need_ids = set()
            for _tid, tmeta in task_info.items():
                refs = tmeta.get("ts_refs") or []
                if len(refs) > 1:
                    for r in refs:
                        rid = r.get("id")
                        if rid:
                            need_ids.add(rid)
            if need_ids:
                from urllib.parse import quote

                for chunk in (
                    list(need_ids)[i : i + 40] for i in range(0, min(len(need_ids), 200), 40)
                ):
                    if not chunk:
                        break
                    try:
                        tq = json.dumps({"_id": {"$in": [{"$oid": x} for x in chunk]}})
                        tsr = requests.get(
                            f"{JITA_BASE}/test_sets",
                            params={
                                "raw_query": quote(tq),
                                "limit": len(chunk),
                                "only": "_id,tests,name",
                            },
                            auth=JITA_SVC_AUTH,
                            verify=False,
                            timeout=45,
                        )
                        if tsr.status_code == 200:
                            for d in tsr.json().get("data", []):
                                oid = _oid(d.get("_id"))
                                if oid:
                                    ts_id_to_testnames[oid] = _test_names_from_ts_doc_tests_field(
                                        d.get("tests")
                                    )
                                    nm = (d.get("name") or "").strip()
                                    if nm:
                                        ts_id_to_tsname[oid] = nm
                    except Exception as e:
                        logger.warning(f"[test-exec-history] Batch test_sets fetch failed: {e}")

        rows = []
        seen_pairs = set()
        unique_pairs = []
        for item in raw_items:
            sut = item.get("system_under_test") or {}
            agave_task = item.get("AgaveTask") or {}
            exec_id = _oid(item.get("agave_task_id"))
            ti = task_info.get(exec_id, {})
            # Test set: JITA history row first (matches /test_history), then multi-TS membership, then task.
            row_ts = _test_set_from_history_row(item, agave_task)
            if not row_ts:
                refs = ti.get("ts_refs") or []
                if len(refs) > 1 and test_name:
                    for r in refs:
                        rid = r.get("id")
                        if not rid:
                            continue
                        tnames = ts_id_to_testnames.get(rid)
                        if tnames and test_name in tnames:
                            row_ts = (r.get("name") or "").strip() or ts_id_to_tsname.get(rid, "")
                            if row_ts:
                                break
            ts = row_ts or ti.get("test_set_name", "")
            # JP: history label / row fields first (user expects test set + label as on JITA for that branch).
            jp = _jp_name_from_history_row(item, agave_task) or ti.get("job_profile_name", "")
            rows.append({
                "id": _oid(item.get("_id")),
                "execution_id": exec_id,
                "branch": sut.get("branch", "") or ti.get("branch", ""),
                "test_set": ts,
                "job_profile": jp,
                "date_started": _date(item.get("start_time")),
                "date_ended": _date(item.get("end_time")),
                "status": item.get("status", ""),
                "jira_tickets": item.get("jira_tickets") or [],
                "exception_summary": item.get("exception_summary") or "",
                "label": agave_task.get("label", ""),
            })
            pair_key = f"{ts}|||{jp}"
            if (ts or jp) and pair_key not in seen_pairs:
                seen_pairs.add(pair_key)
                unique_pairs.append({"test_set": ts, "job_profile": jp})

        return jsonify({
            "success": True,
            "data": rows,
            "unique_pairs": unique_pairs,
            "total": total,
            "page": page,
            "limit": limit,
        })
    except requests.exceptions.Timeout:
        logger.warning("[test-exec-history] Timeout querying JITA")
        return jsonify({"error": "JITA request timed out", "data": [], "total": 0})
    except Exception as e:
        logger.error(f"[test-exec-history] Error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/testcase-history", methods=["POST"])
def dynamic_jp_testcase_history():
    """Fetch testcase run history from JITA and return associated JPs and test sets."""
    try:
        req_data = request.json
        if not req_data:
            return jsonify({"error": "Request body is required (JSON)"}), 400

        testcase_names = req_data.get("testcase_names", [])
        branch = req_data.get("branch", "master")

        if not testcase_names or not isinstance(testcase_names, list):
            return jsonify({"error": "testcase_names must be a non-empty list"}), 400

        # Sanitize: limit to 20 testcases to avoid overloading
        testcase_names = [str(tc) for tc in testcase_names[:20]]

        results = []

        def extract_search_keyword(tc_name):
            """Extract a meaningful keyword from a fully qualified testcase name.
            e.g. 'cdp.stargate.storage_policy.api.test_storage_policy...' -> 'storage_policy'
            """
            parts = tc_name.replace(".", " ").replace("/", " ").split()
            # Pick the most specific non-generic part (skip cdp, stargate, test_, api, etc.)
            skip = {"cdp", "stargate", "test", "api", "tests", "module", "class", "self"}
            for part in parts:
                cleaned = re.sub(r"^test_", "", part)
                if cleaned and cleaned.lower() not in skip and len(cleaned) > 3:
                    return cleaned
            # Fallback: use the 3rd component if available
            dot_parts = tc_name.split(".")
            if len(dot_parts) >= 3:
                return dot_parts[2]
            return tc_name.split(".")[-1] if "." in tc_name else tc_name

        def _parse_jita_items(data, kind="jp"):
            """Parse JITA response data into a flat list of dicts."""
            items = []
            raw_list = data.get("data", []) if isinstance(data, dict) else []
            for item in raw_list:
                if not isinstance(item, dict):
                    continue
                item_id = item.get("_id")
                if isinstance(item_id, dict) and "$oid" in item_id:
                    item_id = item_id["$oid"]
                elif not isinstance(item_id, str):
                    item_id = str(item_id) if item_id else None
                if kind == "jp":
                    items.append({
                        "_id": item_id,
                        "name": item.get("name", ""),
                        "description": item.get("description", ""),
                    })
                else:
                    items.append({
                        "_id": item_id,
                        "path": item.get("path", "") or "",
                        "name": item.get("name", "") or "",
                        "test_args": item.get("test_args", "") or "",
                        "framework_args": item.get("framework_args", "") or "",
                    })
            return items

        def _search_jps(keyword, branch_val):
            """Search job_profiles by keyword + branch, fall back to keyword only."""
            from urllib.parse import quote
            jp_details = []
            try:
                jp_pattern = f".*{re.escape(keyword)}.*{re.escape(branch_val)}"
                raw_q = json.dumps({"name": {"$regex": jp_pattern, "$options": "i"}})
                resp = requests.get(
                    f"{JITA_BASE}/job_profiles",
                    params={"raw_query": quote(raw_q), "limit": 10, "only": "_id,name,description"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=45
                )
                if resp.status_code == 200:
                    jp_details = _parse_jita_items(resp.json(), "jp")
            except (requests.exceptions.RequestException, ValueError) as e:
                logger.warning(f"JP search failed for '{keyword}+{branch_val}': {e}")

            if not jp_details:
                try:
                    raw_q2 = json.dumps({"name": {"$regex": f".*{re.escape(keyword)}.*", "$options": "i"}})
                    resp2 = requests.get(
                        f"{JITA_BASE}/job_profiles",
                        params={"raw_query": quote(raw_q2), "limit": 10, "only": "_id,name,description"},
                        auth=JITA_SVC_AUTH, verify=False, timeout=45
                    )
                    if resp2.status_code == 200:
                        jp_details = _parse_jita_items(resp2.json(), "jp")
                except (requests.exceptions.RequestException, ValueError) as e:
                    logger.warning(f"JP fallback search failed for '{keyword}': {e}")
            return jp_details

        def _search_test_sets(keyword):
            """Search test_sets by keyword."""
            from urllib.parse import quote
            try:
                raw_q = json.dumps({"name": {"$regex": f".*{re.escape(keyword)}.*", "$options": "i"}})
                resp = requests.get(
                    f"{JITA_BASE}/test_sets",
                    params={"raw_query": quote(raw_q), "limit": 10, "only": "_id,name,path,test_args,framework_args"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=60
                )
                if resp.status_code == 200:
                    return _parse_jita_items(resp.json(), "ts")
            except (requests.exceptions.RequestException, ValueError) as e:
                logger.warning(f"Test set search failed for '{keyword}': {e}")
            return []

        def fetch_single_testcase(tc_name):
            tc_name = tc_name.strip() if isinstance(tc_name, str) else ""
            if not tc_name:
                return None
            try:
                keyword = extract_search_keyword(tc_name)
                logger.info(f"[dynamic-jp] Searching for testcase '{tc_name}' using keyword '{keyword}' on branch '{branch}'")

                # Run JP and TS searches in parallel for speed
                from concurrent.futures import ThreadPoolExecutor, as_completed
                jp_details = []
                ts_details = []
                with ThreadPoolExecutor(max_workers=2) as mini_pool:
                    jp_future = mini_pool.submit(_search_jps, keyword, branch)
                    ts_future = mini_pool.submit(_search_test_sets, keyword)
                    try:
                        jp_details = jp_future.result(timeout=90)
                    except Exception as e:
                        logger.warning(f"JP parallel search error: {e}")
                    try:
                        ts_details = ts_future.result(timeout=90)
                    except Exception as e:
                        logger.warning(f"TS parallel search error: {e}")

                return {
                    "testcase": tc_name,
                    "keyword": keyword,
                    "runs": [],
                    "job_profiles": jp_details,
                    "test_sets": ts_details,
                }
            except requests.exceptions.Timeout:
                logger.warning(f"Timeout fetching history for {tc_name}")
                return {"testcase": tc_name, "error": "Request timed out", "runs": [], "job_profiles": [], "test_sets": []}
            except requests.exceptions.ConnectionError:
                logger.warning(f"Connection error fetching history for {tc_name}")
                return {"testcase": tc_name, "error": "Connection error to JITA", "runs": [], "job_profiles": [], "test_sets": []}
            except Exception as e:
                logger.error(f"Error fetching history for {tc_name}: {e}")
                return {"testcase": tc_name, "error": str(e), "runs": [], "job_profiles": [], "test_sets": []}

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(fetch_single_testcase, tc): tc for tc in testcase_names}
            for future in as_completed(futures):
                try:
                    result = future.result(timeout=60)
                    if result:
                        results.append(result)
                except Exception as e:
                    logger.error(f"Future exception in testcase-history: {e}")

        return jsonify({"success": True, "results": results})
    except Exception as e:
        logger.error(f"Error in dynamic-jp testcase-history: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


def _dynamic_jp_ts_prefix_for_date(dyn_date_str=None):
    """Prefixes for auto-named temporary test JP/TS."""
    if (
        dyn_date_str
        and isinstance(dyn_date_str, str)
        and len(dyn_date_str.strip()) == 8
        and dyn_date_str.strip().isdigit()
    ):
        date_str = dyn_date_str.strip()
    else:
        date_str = datetime.now().strftime("%Y%m%d")
    jp_p = f"Temp_Test_Profile_{date_str}_JP_"
    ts_p = f"Temp_Test_Profile_{date_str}_TS_"
    return jp_p, ts_p


# Monotonic sequence per YYYYMMDD; increments on every /dynamic-jp/create attempt (success or fail after bump).
DYN_NAME_SEQ_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".dyn_name_sequence.json")
_dyn_name_seq_lock = threading.Lock()


def _dyn_name_seq_date_key(dyn_name_date):
    d = (dyn_name_date or "").strip()
    if d and len(d) == 8 and d.isdigit():
        return d
    return datetime.now().strftime("%Y%m%d")


def _load_dyn_name_seq() -> dict:
    try:
        with open(DYN_NAME_SEQ_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict):
                out = {}
                for k, v in data.items():
                    ks = str(k)
                    if len(ks) == 8 and ks.isdigit():
                        try:
                            out[ks] = int(v)
                        except (TypeError, ValueError):
                            pass
                return out
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        pass
    return {}


def _save_dyn_name_seq(data: dict) -> None:
    try:
        with open(DYN_NAME_SEQ_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=0)
    except OSError as e:
        logger.warning(f"Could not write {DYN_NAME_SEQ_FILE}: {e}")


def _peek_next_dyn_name_seq(dyn_name_date):
    """Next sequence number to suggest (1 + last used); does not modify store."""
    key = _dyn_name_seq_date_key(dyn_name_date)
    with _dyn_name_seq_lock:
        data = _load_dyn_name_seq()
        return int(data.get(key, 0)) + 1


def _bump_dyn_name_seq(dyn_name_date):
    """Increment and return the sequence to use for this /create call (one bump per create request)."""
    key = _dyn_name_seq_date_key(dyn_name_date)
    with _dyn_name_seq_lock:
        data = _load_dyn_name_seq()
        n = int(data.get(key, 0)) + 1
        data[key] = n
        _save_dyn_name_seq(data)
    logger.info(f"[dyn-seq] date={key} last_reserved={n}")
    return n


def _apply_reserved_seq_to_dyn_custom_name(name: str, date_key: str, seq: int) -> str:
    """Rewrite generated JP/TS sequence to use `seq` (per create attempt)."""
    if not name or not date_key:
        return name
    s = re.sub(
        rf"((?:User_Dyn|Temp_Test_Profile)_{re.escape(date_key)}_JP_)\d+",
        rf"\g<1>{seq}",
        name,
        count=1,
        flags=re.IGNORECASE,
    )
    s = re.sub(
        rf"((?:User_Dyn|Temp_Test_Profile)_{re.escape(date_key)}_TS_)\d+",
        rf"\g<1>{seq}",
        s,
        count=1,
        flags=re.IGNORECASE,
    )
    ddmm = f"{date_key[6:8]}{date_key[4:6]}" if len(date_key) == 8 and date_key.isdigit() else ""
    if ddmm:
        s = re.sub(
            rf"([A-Z]{{2}}_{re.escape(ddmm)}_P)\d+((?:_(?:JP|TS))?(?:_|$))",
            rf"\g<1>{seq}\g<2>",
            s,
            count=1,
            flags=re.IGNORECASE,
        )
    return s


# Optional JITA test set "framework options" the UI injects only when the user
# clicks "Add defaults". Source TS test_args / framework_args are always preserved;
# these are merged on top only on demand.
DYN_TESTSET_OPTIONAL_FRAMEWORK_OPTS = {
    "no_log_collection": False,
    "use_logbay": True,
    "log_level": "DEBUG",
}
DYN_TESTSET_RETAIN_FRAMEWORK_OPTS = {
    "skip_teardown": True,
    "skip_class_teardown": True,
}


def _framework_args_value_to_dict(val):
    """Normalize test set framework_args from JITA (JSON string or dict) to a dict."""
    if val is None:
        return {}
    if isinstance(val, dict):
        return dict(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return {}
        try:
            parsed = json.loads(s)
            return dict(parsed) if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, TypeError, ValueError):
            return {}
    return {}


def _test_args_value_to_dict(val):
    """Normalize test set test_args from JITA (JSON string/dict) to a dict."""
    if val is None:
        return {}
    if isinstance(val, dict):
        return dict(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return {}
        try:
            parsed = json.loads(s)
            return dict(parsed) if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, TypeError, ValueError):
            # Some payloads arrive as python-literal dict strings with single quotes.
            try:
                import ast
                parsed = ast.literal_eval(s)
                return dict(parsed) if isinstance(parsed, dict) else {}
            except (SyntaxError, ValueError, TypeError):
                return {}
    return {}


def _merge_dyn_testset_framework_args(existing, retain_setup_on_failure=False, custom_options=None):
    """Preserve the source framework_args; apply retain-only teardown skips and any custom options.

    Defaults are NOT auto-injected here. Basic log options are added only when the
    UI's "Add defaults" toggle passes them in via custom_options.
    """
    merged = _framework_args_value_to_dict(existing)
    if retain_setup_on_failure:
        merged.update(DYN_TESTSET_RETAIN_FRAMEWORK_OPTS)
    else:
        for key in DYN_TESTSET_RETAIN_FRAMEWORK_OPTS:
            merged.pop(key, None)
    # Apply custom framework options from UI (overrides existing, only if has actual keys).
    if custom_options and isinstance(custom_options, dict) and len(custom_options) > 0:
        merged.update(custom_options)
    return merged


def _apply_default_framework_options_to_test_set_payload(ts_payload, retain_setup_on_failure=False, custom_options=None):
    """Mutate a POST /test_sets body with framework options used by both API and JITA UI."""
    if not isinstance(ts_payload, dict):
        return ts_payload
    existing = ts_payload.get("framework_args")
    if existing in (None, ""):
        existing = ts_payload.get("frameworkArgs")
    if existing in (None, ""):
        existing = ts_payload.get("agave_options")
    merged = _merge_dyn_testset_framework_args(existing, retain_setup_on_failure, custom_options)
    fa_str = json.dumps(merged, separators=(",", ":"))
    ts_payload["framework_args"] = fa_str
    ts_payload["frameworkArgs"] = fa_str
    # JITA Edit Test Set UI renders "Framework Options" from agave_options.
    ts_payload["agave_options"] = merged
    return ts_payload


def _ensure_test_args_on_test_set_payload(ts_payload, default_value="{}", custom_args=None):
    """Mutate a POST /test_sets body so existing test args are kept and user args are appended.

    JITA exposes test args across several fields (``test_args`` / ``testArgs`` as JSON
    strings and ``args_map`` as a dict, which is what the JITA UI "Test Args" panel
    renders). A cloned/created test set can have the real data in only one of them
    (e.g. ``test_args`` == "{}" while ``args_map`` holds the source values). To avoid
    dropping any pre-existing args we union ALL of these fields first (first value seen
    for a key wins), then append the user-supplied args BELOW the existing ones,
    skipping any key that already exists (duplicate).
    """
    if not isinstance(ts_payload, dict):
        return ts_payload

    # 1) Collect existing args from every field JITA may use (preserve insertion order).
    existing = {}
    for field in ("test_args", "testArgs", "args_map"):
        for k, v in _test_args_value_to_dict(ts_payload.get(field)).items():
            if k not in existing:
                existing[k] = v

    if not existing:
        existing = _test_args_value_to_dict(default_value)

    # 2) Append user args below existing ones; skip duplicates (existing value wins).
    if custom_args and isinstance(custom_args, dict):
        for k, v in custom_args.items():
            if k in existing:
                logger.info(f"[test_args] Skipping duplicate key already present: {k!r}")
                continue
            existing[k] = v

    ta = json.dumps(existing, separators=(",", ":"))

    # Mirror to every field JITA might read so UI + API stay consistent.
    ts_payload["test_args"] = ta
    ts_payload["testArgs"] = ta
    # JITA UI "Test Args" panel renders from args_map (dict, not JSON string).
    ts_payload["args_map"] = dict(existing)
    return ts_payload


@app.route("/mcp/regression/dynamic-jp/check-existing", methods=["POST"])
def dynamic_jp_check_existing():
    """Search for existing dynamic JP/TS by name prefix; suggest next numeric suffix.

    Defaults to User_Dyn_<YYYYMMDD>_JP_ / User_Dyn_<YYYYMMDD>_TS_ (local server date, or
    optional dyn_name_date=YYYYMMDD). Clients may still pass jp_pattern / ts_pattern.
    """
    try:
        req_data = request.json or {}
        dyn_name_date = (req_data.get("dyn_name_date") or "").strip()
        if dyn_name_date and (len(dyn_name_date) != 8 or not dyn_name_date.isdigit()):
            return jsonify({"error": "dyn_name_date must be YYYYMMDD"}), 400

        def_jp, def_ts = _dynamic_jp_ts_prefix_for_date(dyn_name_date or None)
        jp_raw = (req_data.get("jp_pattern") or "").strip() or def_jp
        ts_raw = (req_data.get("ts_pattern") or "").strip() or def_ts
        jp_suffix = (req_data.get("jp_suffix") or "").strip().strip("_")
        ts_suffix = (req_data.get("ts_suffix") or "").strip().strip("_")

        # Sanitize patterns to avoid regex injection
        jp_pattern = re.escape(jp_raw)
        ts_pattern = re.escape(ts_raw)

        from urllib.parse import quote

        existing_jps = []
        try:
            jp_regex = (
                f"^{jp_pattern}\\d+_{re.escape(jp_suffix)}$"
                if jp_suffix
                else f"^{jp_pattern}"
            )
            raw_query = json.dumps({"name": {"$regex": jp_regex, "$options": "i"}})
            params = {"raw_query": quote(raw_query), "limit": 100}
            resp = requests.get(
                f"{JITA_BASE}/job_profiles",
                params=params,
                auth=JITA_SVC_AUTH,
                verify=False,
                timeout=8
            )
            if resp.status_code == 200:
                resp_data = resp.json()
                jp_list = resp_data.get("data", []) if isinstance(resp_data, dict) else []
                for jp in jp_list:
                    if not isinstance(jp, dict):
                        continue
                    jp_id = jp.get("_id")
                    if isinstance(jp_id, dict) and "$oid" in jp_id:
                        jp_id = jp_id["$oid"]
                    elif not isinstance(jp_id, str):
                        jp_id = str(jp_id) if jp_id else None
                    existing_jps.append({
                        "_id": jp_id,
                        "name": jp.get("name", ""),
                        "description": jp.get("description", ""),
                        "created_at": jp.get("created_at"),
                    })
            else:
                logger.warning(f"check-existing: JP search returned {resp.status_code}")
        except (requests.exceptions.RequestException, ValueError) as e:
            logger.warning(f"check-existing: Failed to fetch JPs: {e}")

        existing_ts = []
        try:
            ts_regex = (
                f"^{ts_pattern}\\d+_{re.escape(ts_suffix)}$"
                if ts_suffix
                else f"^{ts_pattern}"
            )
            raw_query_ts = json.dumps({"name": {"$regex": ts_regex, "$options": "i"}})
            params_ts = {"raw_query": quote(raw_query_ts), "limit": 100}
            resp_ts = requests.get(
                f"{JITA_BASE}/test_sets",
                params=params_ts,
                auth=JITA_SVC_AUTH,
                verify=False,
                timeout=8
            )
            if resp_ts.status_code == 200:
                resp_ts_data = resp_ts.json()
                ts_list = resp_ts_data.get("data", []) if isinstance(resp_ts_data, dict) else []
                for ts in ts_list:
                    if not isinstance(ts, dict):
                        continue
                    ts_id = ts.get("_id")
                    if isinstance(ts_id, dict) and "$oid" in ts_id:
                        ts_id = ts_id["$oid"]
                    elif not isinstance(ts_id, str):
                        ts_id = str(ts_id) if ts_id else None
                    existing_ts.append({
                        "_id": ts_id,
                        "name": ts.get("name", ""),
                        "description": ts.get("description", ""),
                    })
            else:
                logger.warning(f"check-existing: TS search returned {resp_ts.status_code}")
        except (requests.exceptions.RequestException, ValueError) as e:
            logger.warning(f"check-existing: Failed to fetch test sets: {e}")

        def _next_seq_from_prefixed_name(name, prefix, suffix=""):
            if not isinstance(name, str) or not isinstance(prefix, str):
                return None
            if not name.lower().startswith(prefix.lower()):
                return None
            rest = name[len(prefix):]
            m = re.match(r"(\d+)(?:_|$)", rest)
            if not m:
                return None
            if suffix:
                existing_suffix = rest[m.end():].strip("_")
                if existing_suffix.lower() != suffix.lower():
                    return None
            try:
                return int(m.group(1)) + 1
            except (ValueError, TypeError):
                return None

        next_jp_num = 1
        next_ts_num = 1
        for jp in existing_jps:
            name = jp.get("name", "")
            num = _next_seq_from_prefixed_name(name, jp_raw, jp_suffix)
            if num is not None:
                next_jp_num = max(next_jp_num, num)
        for ts in existing_ts:
            name = ts.get("name", "")
            num = _next_seq_from_prefixed_name(name, ts_raw, ts_suffix)
            if num is not None:
                next_ts_num = max(next_ts_num, num)

        date_key = dyn_name_date if (dyn_name_date and len(dyn_name_date) == 8 and dyn_name_date.isdigit()) else datetime.now().strftime("%Y%m%d")
        seq_peek = _peek_next_dyn_name_seq(date_key)
        # Generic names keep the global daily floor. Suffix-based names use an
        # independent sequence so new source JP/TS suffixes start at P1.
        if not jp_suffix and not ts_suffix:
            next_both = max(next_jp_num, next_ts_num, seq_peek)
            next_jp_num = next_both
            next_ts_num = next_both
        else:
            if not jp_suffix:
                next_jp_num = max(next_jp_num, seq_peek)
            if not ts_suffix:
                next_ts_num = max(next_ts_num, seq_peek)

        return jsonify({
            "success": True,
            "job_profiles": existing_jps,
            "test_sets": existing_ts,
            "next_jp_number": next_jp_num,
            "next_ts_number": next_ts_num,
            "jp_name_prefix": jp_raw,
            "ts_name_prefix": ts_raw,
            "jp_suffix": jp_suffix,
            "ts_suffix": ts_suffix,
        })
    except Exception as e:
        logger.error(f"Error in dynamic-jp check-existing: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/fetch-testset", methods=["POST"])
def dynamic_jp_fetch_testset():
    """Fetch a test set's details (test_args, framework_args) by ID or path."""
    try:
        req_data = request.json
        if not req_data:
            return jsonify({"error": "Request body is required (JSON)"}), 400

        testset_id = req_data.get("testset_id")
        testset_path = req_data.get("testset_path")

        if not testset_id and not testset_path:
            return jsonify({"error": "testset_id or testset_path is required"}), 400

        # Sanitize inputs
        if testset_id:
            testset_id = str(testset_id).strip()
        if testset_path:
            testset_path = str(testset_path).strip()

        try:
            if testset_id:
                resp = requests.get(
                    f"{JITA_BASE}/test_sets/{testset_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
            else:
                from urllib.parse import quote
                raw_query = json.dumps({"path": testset_path})
                params = {"raw_query": quote(raw_query), "limit": 1}
                resp = requests.get(
                    f"{JITA_BASE}/test_sets",
                    params=params,
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
        except requests.exceptions.Timeout:
            return jsonify({"error": "Request to JITA timed out"}), 504
        except requests.exceptions.ConnectionError:
            return jsonify({"error": "Could not connect to JITA API"}), 503

        if resp.status_code != 200:
            return jsonify({"error": f"JITA API error: {resp.status_code}"}), 500

        try:
            data = resp.json()
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid JSON response from JITA"}), 500

        ts_data = data.get("data", {}) if isinstance(data, dict) else {}
        if isinstance(ts_data, list):
            ts_data = ts_data[0] if ts_data else {}
        if not isinstance(ts_data, dict):
            ts_data = {}

        ts_id = ts_data.get("_id")
        if isinstance(ts_id, dict) and "$oid" in ts_id:
            ts_id = ts_id["$oid"]
        elif ts_id and not isinstance(ts_id, str):
            ts_id = str(ts_id)

        tests = ts_data.get("tests", [])
        if not isinstance(tests, list):
            tests = []

        return jsonify({
            "success": True,
            "test_set": {
                "_id": ts_id,
                "name": ts_data.get("name", "") or "",
                "path": ts_data.get("path", "") or "",
                "test_args": ts_data.get("test_args", "") or "",
                "framework_args": ts_data.get("framework_args", "") or "",
                "tests": tests,
                "description": ts_data.get("description", "") or "",
            }
        })
    except Exception as e:
        logger.error(f"Error in dynamic-jp fetch-testset: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


JARVIS_BASE = "https://jarvis.eng.nutanix.com/api/v1"


@app.route("/mcp/regression/dynamic-jp/resolve-names", methods=["POST"])
def dynamic_jp_resolve_names():
    """Resolve JP and/or test set names to their JITA IDs."""
    try:
        req_data = request.json
        if not req_data:
            return jsonify({"error": "Request body is required (JSON)"}), 400

        jp_name = (req_data.get("jp_name") or "").strip()
        ts_name = (req_data.get("ts_name") or "").strip()

        if not jp_name and not ts_name:
            return jsonify({"error": "At least one of jp_name or ts_name is required"}), 400

        from urllib.parse import quote

        def _oid(val):
            if isinstance(val, dict) and "$oid" in val:
                return val["$oid"]
            return str(val) if val else None

        result = {"jp": None, "ts": None}

        if jp_name:
            try:
                raw_q = json.dumps({"name": jp_name})
                resp = requests.get(
                    f"{JITA_BASE}/job_profiles",
                    params={"raw_query": quote(raw_q), "limit": 40, "only": "_id,name,description,tags,tester_tags"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                logger.info(f"[resolve-names] JP lookup for '{jp_name}': HTTP {resp.status_code}")
                if resp.status_code == 200:
                    data = resp.json()
                    items = data.get("data", []) if isinstance(data, dict) else []
                    want = jp_name.strip()
                    want_l = want.lower()
                    chosen = None
                    if items and isinstance(items, list):
                        for el in items:
                            if not isinstance(el, dict):
                                continue
                            nm = (el.get("name") or "").strip()
                            if nm == want or nm.lower() == want_l:
                                chosen = el
                                break
                    if chosen:
                        result["jp"] = {
                            "_id": _oid(chosen.get("_id")),
                            "name": chosen.get("name", ""),
                            "description": chosen.get("description", ""),
                            "tags": chosen.get("tags", []) or [],
                            "tester_tags": chosen.get("tester_tags", []) or [],
                        }
                    elif items:
                        logger.warning(
                            f"[resolve-names] JP search returned {len(items)} row(s) but none named {want!r}; not guessing."
                        )
                else:
                    logger.warning(f"[resolve-names] JP search returned {resp.status_code}: {resp.text[:200]}")
            except Exception as e:
                logger.warning(f"[resolve-names] Failed to resolve JP '{jp_name}': {e}")

        if ts_name:
            try:
                raw_q = json.dumps({"name": ts_name})
                resp = requests.get(
                    f"{JITA_BASE}/test_sets",
                    params={"raw_query": quote(raw_q), "limit": 40, "only": "_id,name,test_args,framework_args"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                logger.info(f"[resolve-names] TS lookup for '{ts_name}': HTTP {resp.status_code}")
                if resp.status_code == 200:
                    data = resp.json()
                    items = data.get("data", []) if isinstance(data, dict) else []
                    want = ts_name.strip()
                    want_l = want.lower()
                    chosen = None
                    if items and isinstance(items, list):
                        for el in items:
                            if not isinstance(el, dict):
                                continue
                            nm = (el.get("name") or "").strip()
                            if nm == want or nm.lower() == want_l:
                                chosen = el
                                break
                    if chosen:
                        ta = chosen.get("test_args") or chosen.get("testArgs") or ""
                        fa = chosen.get("framework_args") or chosen.get("frameworkArgs") or ""
                        result["ts"] = {
                            "_id": _oid(chosen.get("_id")),
                            "name": chosen.get("name", ""),
                            "test_args": str(ta).strip() if ta is not None else "",
                            "framework_args": str(fa).strip() if fa is not None else "",
                        }
                    elif items:
                        logger.warning(
                            f"[resolve-names] TS search returned {len(items)} row(s) but none named {want!r}; not guessing."
                        )
                else:
                    logger.warning(f"[resolve-names] TS search returned {resp.status_code}: {resp.text[:200]}")
            except Exception as e:
                logger.warning(f"[resolve-names] Failed to resolve TS '{ts_name}': {e}")

        return jsonify({"success": True, **result})
    except Exception as e:
        logger.error(f"Error in dynamic-jp resolve-names: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/search-node-pools", methods=["POST"])
def dynamic_jp_search_node_pools():
    """Search Jarvis node pools by name keyword."""
    try:
        req_data = request.json or {}
        query = (req_data.get("query") or "").strip()
        if len(query) < 2:
            return jsonify({"pools": []})

        tokens = re.split(r'[\s_\-]+', query)
        tokens = [t for t in tokens if t]

        primary = max(tokens, key=len) if tokens else query
        pattern = re.escape(primary)

        raw_q = json.dumps({"name": {"$regex": pattern, "$options": "i"}})
        resp = requests.get(
            f"{JARVIS_BASE}/pools",
            params={"raw_query": raw_q, "limit": 100},
            auth=JITA_SVC_AUTH, verify=False, timeout=15
        )
        pools = []
        if resp.status_code == 200:
            for item in resp.json().get("data", []):
                name = item.get("name") or ""
                if not name or name in pools:
                    continue
                name_lower = name.lower()
                if all(t.lower() in name_lower for t in tokens):
                    pools.append(name)
        return jsonify({"pools": pools})
    except requests.exceptions.Timeout:
        return jsonify({"error": "Timed out searching node pools", "pools": []}), 504
    except Exception as e:
        logger.error(f"Error searching node pools: {e}", exc_info=True)
        return jsonify({"error": str(e), "pools": []}), 500


@app.route("/mcp/regression/dynamic-jp/search-branches", methods=["POST"])
def dynamic_jp_search_branches():
    """Search JITA branches by name."""
    try:
        req_data = request.json or {}
        query = (req_data.get("query") or "").strip()
        if len(query) < 2:
            return jsonify({"branches": []})

        pattern = re.escape(query)
        raw_q = json.dumps({"name": {"$regex": pattern, "$options": "i"}})
        resp = requests.get(
            f"{JITA_BASE}/branches",
            params={"raw_query": raw_q, "limit": 20},
            auth=JITA_SVC_AUTH, verify=False, timeout=15
        )
        branches = []
        if resp.status_code == 200:
            for item in resp.json().get("data", []):
                name = item.get("name") or ""
                if name and name not in branches:
                    branches.append(name)
        # Sort so exact-prefix matches come first
        q_lower = query.lower()
        branches.sort(key=lambda b: (0 if b.lower().startswith(q_lower) else 1, b.lower()))
        return jsonify({"branches": branches})
    except requests.exceptions.Timeout:
        return jsonify({"error": "Timed out", "branches": []}), 504
    except Exception as e:
        logger.error(f"Error searching branches: {e}", exc_info=True)
        return jsonify({"error": str(e), "branches": []}), 500


@app.route("/mcp/regression/dynamic-jp/search-clusters", methods=["POST"])
def dynamic_jp_search_clusters():
    """Search JITA clusters by name or IP."""
    try:
        req_data = request.json or {}
        query = (req_data.get("query") or "").strip()
        if len(query) < 2:
            return jsonify({"clusters": []})

        pattern = re.escape(query)
        raw_q = json.dumps({"name": {"$regex": pattern, "$options": "i"}})
        resp = requests.get(
            f"{JITA_BASE}/clusters",
            params={"raw_query": raw_q, "limit": 20},
            auth=JITA_SVC_AUTH, verify=False, timeout=15
        )
        clusters = []
        seen = set()
        if resp.status_code == 200:
            for item in resp.json().get("data", []):
                name = item.get("name") or ""
                if not name or name in seen:
                    continue
                seen.add(name)
                clusters.append({
                    "name": name,
                    "status": item.get("status", ""),
                })
        return jsonify({"clusters": clusters})
    except requests.exceptions.Timeout:
        return jsonify({"error": "Timed out searching clusters", "clusters": []}), 504
    except Exception as e:
        logger.error(f"Error searching clusters: {e}", exc_info=True)
        return jsonify({"error": str(e), "clusters": []}), 500


# Maps the tool's resource_type values to the QMS "kind" used for coupon validation
# (mirrors JITA's getQmsKind: "2.0" -> nested_ahv_2, "1.0"/nested -> nested, else physical).
DYN_QMS_KIND_BY_RESOURCE_TYPE = {
    "nested_2.0": "nested_ahv_2",
    "nested_1.0": "nested",
    "physical": "physical",
}


@app.route("/mcp/regression/dynamic-jp/validate-coupon", methods=["POST"])
@jwt_required
def dynamic_jp_validate_coupon():
    """Validate a QMS coupon for the Global Pool provider (mirrors JITA's Validate button).

    POSTs to QMS ``/coupons/<coupon>/validate`` with the resource kind and the current
    user's email. Returns ``{valid, category, message}``.
    """
    try:
        current_user_email = None
        if hasattr(g, "current_user") and isinstance(g.current_user, dict):
            current_user_email = (g.current_user.get("email") or "").strip()

        req_data = request.json or {}
        coupon = (req_data.get("coupon") or "").strip()
        resource_type = (req_data.get("resource_type") or "nested_2.0").strip()
        if not coupon:
            return jsonify({"valid": False, "error": "No coupon specified"}), 400

        kind = DYN_QMS_KIND_BY_RESOURCE_TYPE.get(resource_type, "nested_ahv_2")

        username = current_user_email or (req_data.get("username") or "").strip()
        if username and "@" not in username:
            username = f"{username}@nutanix.com"
        if not username:
            return jsonify({"valid": False, "error": "Could not determine user email for validation"}), 400

        from urllib.parse import quote as _quote
        try:
            resp = requests.post(
                f"{QMS_BASE_URL}/coupons/{_quote(coupon, safe='')}/validate",
                json={"kinds": [kind], "username": username},
                verify=False,
                timeout=20,
            )
        except requests.exceptions.Timeout:
            return jsonify({"valid": False, "error": "QMS validation timed out"}), 504
        except requests.exceptions.RequestException as e:
            return jsonify({"valid": False, "error": f"Could not reach QMS: {e}"}), 502

        try:
            body = resp.json()
        except (ValueError, TypeError):
            body = {}

        if resp.status_code == 200:
            return jsonify({
                "valid": True,
                "category": body.get("category"),
                "message": "Coupon is valid",
            })
        msg = body.get("message") if isinstance(body, dict) else None
        return jsonify({"valid": False, "error": msg or "Invalid coupon"})
    except Exception as e:
        logger.error(f"Error validating coupon: {e}", exc_info=True)
        return jsonify({"valid": False, "error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/create", methods=["POST"])
@jwt_required
def dynamic_jp_create():
    """Create a dynamic JP and test set. Two modes:
    - create_fresh=True: brand-new JP+TS from scratch with the given testcases
    - create_fresh=False: clone from source_jp_id, optionally copying test_args from source_testset_id
    """
    try:
        # Get current user info from JWT
        current_user_email = None
        if hasattr(g, 'current_user') and isinstance(g.current_user, dict):
            current_user_email = g.current_user.get('email', '').strip()
        
        req_data = request.json
        if not req_data:
            return jsonify({"error": "Request body is required (JSON)"}), 400

        create_fresh = bool(req_data.get("create_fresh", False))
        source_jp_id = req_data.get("source_jp_id")
        source_testset_id = req_data.get("source_testset_id")
        nos_branch = req_data.get("nos_branch", "master") or "master"
        nos_tag = req_data.get("nos_tag", "Latest Smoke Passed") or "Latest Smoke Passed"
        nos_update_type = req_data.get("nos_update_type", "by_tag") or "by_tag"
        nos_commit_id = (req_data.get("nos_commit_id") or "").strip() or ""
        nos_gbn = (req_data.get("nos_gbn") or "").strip() or ""
        pc_branch = req_data.get("pc_branch", "master") or "master"
        pc_tag = req_data.get("pc_tag", "Latest Smoke Passed") or "Latest Smoke Passed"
        pc_update_type = req_data.get("pc_update_type", "by_tag") or "by_tag"
        pc_commit_id = (req_data.get("pc_commit_id") or "").strip() or ""
        nutest_branch = req_data.get("nutest_branch", "master") or "master"
        provider = req_data.get("provider", "global_pool") or "global_pool"
        resource_type = req_data.get("resource_type", "nested_2.0") or "nested_2.0"
        # Optional Global Pool coupon (validated client-side via /validate-coupon).
        # When provided, it is passed to QMS in the infra params instead of auto-allocating.
        global_pool_coupon = (req_data.get("coupon") or "").strip() or None
        raw_np = req_data.get("node_pool") or []
        if isinstance(raw_np, list):
            node_pools = [p.strip() for p in raw_np if isinstance(p, str) and p.strip()]
        else:
            node_pools = [raw_np.strip()] if isinstance(raw_np, str) and raw_np.strip() else []
        framework_patch_url = (req_data.get("framework_patch_url") or "").strip() or None
        test_patch_url = (req_data.get("test_patch_url") or "").strip() or None
        testcase_names = req_data.get("testcase_names", [])
        custom_jp_name = (req_data.get("custom_jp_name") or "").strip() or None
        custom_ts_name = (req_data.get("custom_ts_name") or "").strip() or None
        jp_tags = req_data.get("jp_tags") or []
        if isinstance(jp_tags, str):
            jp_tags = [t.strip() for t in jp_tags.split(",") if t.strip()]
        elif not isinstance(jp_tags, list):
            jp_tags = []

        reuse_source_ts = bool(req_data.get("reuse_source_ts", False))
        retain_setup_on_failure = bool(req_data.get("retain_setup_on_failure", False))
        use_latest_commit = bool(req_data.get("use_latest_commit", False))
        sync_to_tcms = bool(req_data.get("sync_to_tcms", False))
        custom_test_args = req_data.get("custom_test_args")
        custom_framework_options = req_data.get("custom_framework_options")
        include_optional_defaults = bool(req_data.get("include_optional_defaults", False))
        
        # Convert None to empty dict
        if not custom_test_args or not isinstance(custom_test_args, dict):
            custom_test_args = {}
        if not custom_framework_options or not isinstance(custom_framework_options, dict):
            custom_framework_options = {}
        if include_optional_defaults:
            custom_framework_options = {**DYN_TESTSET_OPTIONAL_FRAMEWORK_OPTS, **custom_framework_options}
        if reuse_source_ts and create_fresh:
            return jsonify({
                "error": "reuse_source_ts is only valid when cloning from an existing job profile (not fresh create).",
            }), 400

        if not create_fresh:
            if not source_jp_id:
                return jsonify({"error": "source_jp_id is required when not creating fresh"}), 400
            source_jp_id = str(source_jp_id).strip()
            if not source_jp_id:
                return jsonify({"error": "source_jp_id cannot be empty"}), 400

        if source_testset_id:
            source_testset_id = str(source_testset_id).strip()

        if not isinstance(testcase_names, list):
            testcase_names = []
        testcase_names = [str(tc).strip() for tc in testcase_names if tc]

        if create_fresh and not testcase_names:
            return jsonify({"error": "testcase_names is required when creating fresh"}), 400

        logger.info(f"[create] mode={'fresh' if create_fresh else 'clone'}, "
                     f"source_jp_id={source_jp_id}, source_testset_id={source_testset_id}, "
                     f"source_testset_name={(req_data.get('source_testset_name') or '').strip() or None}, "
                     f"custom_jp_name={custom_jp_name}, custom_ts_name={custom_ts_name}, "
                     f"#testcases={len(testcase_names)}, #tags={len(jp_tags)}, sync_to_tcms={sync_to_tcms}")

        from urllib.parse import quote

        # 1. Fetch source JP (only when cloning)
        source_jp = {}
        if not create_fresh:
            try:
                jp_resp = requests.get(
                    f"{JITA_BASE}/job_profiles/{source_jp_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
            except requests.exceptions.Timeout:
                return jsonify({"error": "Timed out fetching source job profile from JITA"}), 504
            except requests.exceptions.ConnectionError:
                return jsonify({"error": "Could not connect to JITA to fetch source job profile"}), 503

            if jp_resp.status_code != 200:
                return jsonify({"error": f"Failed to fetch source JP (HTTP {jp_resp.status_code}). Verify the JP ID is correct."}), 500

            try:
                source_jp = jp_resp.json().get("data", {})
            except (ValueError, TypeError):
                return jsonify({"error": "Invalid JSON from JITA when fetching source JP"}), 500

            if not source_jp or not isinstance(source_jp, dict):
                return jsonify({"error": f"Source JP '{source_jp_id}' returned empty data. It may not exist."}), 404

        def _test_set_ref_oid(ref):
            if isinstance(ref, dict):
                if "$oid" in ref:
                    return str(ref["$oid"]).strip() or None
                inner = ref.get("_id")
                if isinstance(inner, dict) and "$oid" in inner:
                    return str(inner["$oid"]).strip() or None
                if isinstance(inner, str) and inner.strip():
                    return inner.strip()
            if isinstance(ref, str) and ref.strip():
                return ref.strip()
            return None

        def _jp_test_set_refs_list(jp):
            """Refs attached to a job profile: `test_sets` (list) or legacy `test_set` (list or one dict)."""
            if not isinstance(jp, dict):
                return []
            raw = jp.get("test_sets")
            if raw is None or raw == []:
                raw = jp.get("test_set")
            if raw is None:
                return []
            if isinstance(raw, list):
                return raw
            return [raw]

        def _jit_ts_arg_strings(ts):
            """JITA may expose args as snake_case or camelCase; normalize to strings for POST payload."""
            if not isinstance(ts, dict):
                return "", ""

            def _coerce(val):
                if val is None:
                    return ""
                if isinstance(val, dict):
                    try:
                        return json.dumps(val, separators=(",", ":"))
                    except (TypeError, ValueError):
                        return ""
                if isinstance(val, list):
                    try:
                        return json.dumps(val, separators=(",", ":"))
                    except (TypeError, ValueError):
                        return ""
                s = str(val).strip()
                return s

            ta = _coerce(ts.get("test_args")) or _coerce(ts.get("testArgs")) or _coerce(ts.get("args_map"))
            fa = _coerce(ts.get("framework_args")) or _coerce(ts.get("frameworkArgs"))
            return ta, fa

        def _jit_pick_ts_dict_from_response(ts_json):
            """Normalize GET /test_sets/:id (or similar) JSON to one test set dict."""
            if not isinstance(ts_json, dict):
                return {}
            data = ts_json.get("data")
            if isinstance(data, list):
                for el in data:
                    if isinstance(el, dict):
                        return el
                return {}
            if isinstance(data, dict):
                d = data
                for wrap in ("test_set", "document", "result", "item"):
                    inner = d.get(wrap)
                    if isinstance(inner, dict) and any(
                        k in inner
                        for k in ("tests", "test_args", "testArgs", "framework_args", "frameworkArgs", "name", "_id")
                    ):
                        return inner
                return d
            return {}

        def _build_clone_test_set_post_payload(source_doc, new_name, test_entries, description):
            """POST /test_sets with same top-level shape as JITA GET, plus mirrored arg keys.

            Preserves the source test set's test_args and framework_args so the clone
            keeps everything that was already configured.
            """
            import copy

            strip_keys = {
                "_id", "id", "created_at", "updated_at", "created_by", "updated_by",
                "__v", "createdAt", "updatedAt", "path",
            }
            if not isinstance(source_doc, dict) or not source_doc:
                return {
                    "name": new_name, "tests": test_entries, "description": description,
                    "test_args": "", "framework_args": "",
                    "testArgs": "", "frameworkArgs": "",
                }
            payload = {k: copy.deepcopy(v) for k, v in source_doc.items() if k not in strip_keys}
            payload["name"] = new_name
            payload["tests"] = test_entries
            payload["description"] = description

            for snake, camel in (("test_args", "testArgs"), ("framework_args", "frameworkArgs")):
                if snake in payload and camel not in payload:
                    payload[camel] = payload[snake]
                elif camel in payload and snake not in payload:
                    payload[snake] = payload[camel]
            return payload

        def _set_tcms_sync_flags(jp_payload, enabled):
            """Set JITA/TCMS sync flags for the created JP."""
            if not isinstance(jp_payload, dict):
                return

            jp_payload["sync_to_tcms"] = bool(enabled)

            # JITA manage UI computes "Sync To TCMS" from tester_tags.includes("official").
            tester_tags = jp_payload.get("tester_tags")
            if isinstance(tester_tags, list):
                if enabled:
                    jp_payload["tester_tags"] = list(dict.fromkeys(tester_tags + ["official"]))
                else:
                    jp_payload["tester_tags"] = [t for t in tester_tags if t != "official"]
            elif enabled:
                jp_payload["tester_tags"] = ["official"]

            sut = jp_payload.get("system_under_test")
            if isinstance(sut, dict):
                sut["sync_to_tcms"] = bool(enabled)

            if enabled:
                jp_payload.setdefault("package_type", "tar")
                jp_payload.setdefault("test_service", "NutestPy3Tests")
            else:
                # Empty strings survive JITA PUT; missing keys get defaulted back.
                for k in ("service", "package_type", "test_service"):
                    jp_payload[k] = ""

        def _apply_retain_setup_on_failure(jp_payload):
            """Retain deployment after each test failure, excluding DataCorruptionError."""
            if not isinstance(jp_payload, dict):
                return

            existing = jp_payload.get("retain_resources_config")
            existing_criteria = existing.get("criteria") if isinstance(existing, dict) else {}
            existing_test_failure = (
                existing_criteria.get("TEST_FAILURE") if isinstance(existing_criteria, dict) else {}
            )
            existing_params = (
                existing_test_failure.get("params") if isinstance(existing_test_failure, dict) else {}
            )
            duration = existing_params.get("duration", 720) if isinstance(existing_params, dict) else 720
            states = existing_params.get("states_to_track") if isinstance(existing_params, dict) else None
            if not isinstance(states, list) or not states:
                states = ["Failed"]
            required_states = ["Failed", "Aborted", "Timeout", "InfraError", "Warning"]
            states = list(dict.fromkeys(states + [s for s in required_states if s not in states]))
            exceptions = existing_params.get("exceptions") if isinstance(existing_params, dict) else []
            if not isinstance(exceptions, list):
                exceptions = []
            exceptions = [e for e in exceptions if e != "DataCorruptionError"]

            criteria = dict(existing_criteria) if isinstance(existing_criteria, dict) else {}
            criteria["TEST_FAILURE"] = {
                "entity": "DEPLOYMENT",
                "type": "AFTER_EACH",
                "params": {
                    "duration": duration,
                    "exceptions": exceptions,
                    "states_to_track": states,
                },
            }
            jp_payload["retain_resources_config"] = {"criteria": criteria}

            for plugin in jp_payload.get("plugins") or []:
                if not isinstance(plugin, dict):
                    continue
                args = plugin.get("args")
                if isinstance(args, dict) and isinstance(args.get("exceptions"), list):
                    args["exceptions"] = [e for e in args["exceptions"] if e != "DataCorruptionError"]

        source_testset_name = (req_data.get("source_testset_name") or "").strip()

        # Template test set id: prefer exact name from UI (execution history), then explicit id, then JP's first TS.
        template_ts_id = None
        ts_name_resolved_id = None
        if not create_fresh and source_testset_name:
            try:
                raw_q = json.dumps({"name": source_testset_name})
                nm_resp = requests.get(
                    f"{JITA_BASE}/test_sets",
                    params={
                        "raw_query": quote(raw_q),
                        "limit": 40,
                        "only": "_id,name,test_args,framework_args,tests,description",
                    },
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30,
                )
                if nm_resp.status_code == 200:
                    items = nm_resp.json().get("data", []) if isinstance(nm_resp.json(), dict) else []
                    want = source_testset_name.strip()
                    want_l = want.lower()
                    hit = None
                    if isinstance(items, list):
                        for el in items:
                            if not isinstance(el, dict):
                                continue
                            hn = (el.get("name") or "").strip()
                            if hn == want or hn.lower() == want_l:
                                hit = el
                                break
                    if hit:
                        cand_id = _test_set_ref_oid(hit.get("_id"))
                        if cand_id:
                            ts_name_resolved_id = cand_id
                            logger.info(
                                f"[create] Template from source_testset_name={source_testset_name!r} -> id={ts_name_resolved_id}"
                            )
                    elif items:
                        logger.warning(
                            f"[create] source_testset_name={source_testset_name!r}: JITA returned {len(items)} "
                            "row(s) but none matched that exact name; not guessing an id from order."
                        )
            except (requests.exceptions.RequestException, ValueError, TypeError) as e:
                logger.warning(f"[create] source_testset_name lookup failed: {e}")

        if source_testset_id:
            template_ts_id = str(source_testset_id).strip() or None
        elif ts_name_resolved_id:
            template_ts_id = ts_name_resolved_id
        if not create_fresh and not template_ts_id and source_jp:
            refs = _jp_test_set_refs_list(source_jp)
            if refs:
                template_ts_id = _test_set_ref_oid(refs[0])
                if template_ts_id and not source_testset_id and not source_testset_name:
                    logger.info(
                        f"[create] No source_testset_id/name in request; using source JP's first test set "
                        f"as clone template: {template_ts_id}"
                    )

        # 2. Fetch template test set (non-fatal if it fails)
        source_ts = {}
        ts_fetch_warning = None
        if template_ts_id:
            try:
                ts_resp = requests.get(
                    f"{JITA_BASE}/test_sets/{template_ts_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
                if ts_resp.status_code == 200:
                    source_ts = _jit_pick_ts_dict_from_response(ts_resp.json())
                    if source_ts:
                        logger.info(
                            f"[create] Loaded template TS id={template_ts_id} keys_sample={list(source_ts.keys())[:25]}"
                        )
                else:
                    ts_fetch_warning = f"Could not fetch source test set (HTTP {ts_resp.status_code}). Proceeding without test_args copy."
                    logger.warning(ts_fetch_warning)
            except (requests.exceptions.RequestException, ValueError) as e:
                ts_fetch_warning = f"Error fetching source test set: {e}. Proceeding without test_args copy."
                logger.warning(ts_fetch_warning)

        linked_ts_id_for_reuse = None
        reuse_ts_display_name = None

        if reuse_source_ts and not create_fresh:
            tid = source_testset_id or ts_name_resolved_id
            if not tid and source_jp:
                refs = _jp_test_set_refs_list(source_jp)
                if refs:
                    tid = _test_set_ref_oid(refs[0])
            if not tid:
                return jsonify({
                    "error": "reuse_source_ts requires a selected source test set or a source job profile "
                             "that has at least one test set.",
                }), 400
            linked_ts_id_for_reuse = str(tid).strip()
            st_oid = None
            if source_ts:
                st_oid = source_ts.get("_id")
                if isinstance(st_oid, dict) and "$oid" in st_oid:
                    st_oid = str(st_oid["$oid"])
                elif st_oid is not None:
                    st_oid = str(st_oid)
            if source_ts and st_oid == linked_ts_id_for_reuse:
                reuse_ts_display_name = (source_ts.get("name") or "").strip() or linked_ts_id_for_reuse
            else:
                try:
                    tr = requests.get(
                        f"{JITA_BASE}/test_sets/{linked_ts_id_for_reuse}",
                        auth=JITA_SVC_AUTH,
                        verify=False,
                        timeout=30,
                    )
                    if tr.status_code == 200:
                        tj = tr.json()
                        td = _jit_pick_ts_dict_from_response(tj)
                        if isinstance(td, dict):
                            reuse_ts_display_name = (td.get("name") or "").strip() or linked_ts_id_for_reuse
                    if not reuse_ts_display_name:
                        reuse_ts_display_name = linked_ts_id_for_reuse
                except (requests.exceptions.RequestException, ValueError):
                    reuse_ts_display_name = linked_ts_id_for_reuse
            logger.info(f"[create] reuse_source_ts=True, linked_ts_id={linked_ts_id_for_reuse}, name={reuse_ts_display_name}")

        # 3. Sequential names: User_Dyn_<YYYYMMDD>_JP_N / User_Dyn_<YYYYMMDD>_TS_N
        dyn_name_date = (req_data.get("dyn_name_date") or "").strip()
        if dyn_name_date and (len(dyn_name_date) != 8 or not dyn_name_date.isdigit()):
            return jsonify({"error": "dyn_name_date must be YYYYMMDD"}), 400
        jp_prefix, ts_prefix = _dynamic_jp_ts_prefix_for_date(dyn_name_date or None)
        # Keep a monotonic fallback counter for generic names, but do not
        # rewrite user-visible custom names. Those may be suffix-specific
        # suggestions like SW_2205_P1_<source-name>.
        reserved_seq = _bump_dyn_name_seq(dyn_name_date or None)

        if custom_jp_name:
            new_jp_name = custom_jp_name
        else:
            new_jp_name = f"{jp_prefix}{reserved_seq}"

        if reuse_source_ts and linked_ts_id_for_reuse:
            new_ts_name = reuse_ts_display_name or linked_ts_id_for_reuse
        elif custom_ts_name:
            new_ts_name = custom_ts_name
        else:
            new_ts_name = f"{ts_prefix}{reserved_seq}"

        # 4. Pre-check: verify JP and TS names; pick alternatives if the requested name exists
        def _name_exists(entity_type, name):
            """Check if a JP or TS with this exact name already exists. Returns ID or None."""
            try:
                rq = json.dumps({"name": name})
                logger.info(f"[create] Pre-check {entity_type} name='{name}', raw_query={rq}")
                resp = requests.get(
                    f"{JITA_BASE}/{entity_type}",
                    params={"raw_query": rq, "limit": 1, "only": "_id,name"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=8,
                )
                logger.info(f"[create] Pre-check {entity_type} response: HTTP {resp.status_code}, "
                            f"body={resp.text[:300]}")
                if resp.status_code == 200:
                    items = resp.json().get("data", [])
                    if items and isinstance(items[0], dict):
                        matched_name = items[0].get("name", "")
                        if matched_name == name:
                            eid = items[0].get("_id")
                            if isinstance(eid, dict) and "$oid" in eid:
                                return eid["$oid"]
                            return str(eid) if eid else None
                        else:
                            logger.info(f"[create] Pre-check returned '{matched_name}' which doesn't exactly match '{name}', treating as no match")
            except Exception as e:
                logger.warning(f"[create] Pre-check for {entity_type} '{name}' failed: {e}")
            return None

        def _pick_unique_name(entity_type, base_name, kind="Job Profile"):
            """If base_name is free, return (base_name, None). Otherwise use base_2, base_3, ... in JITA."""
            if not base_name:
                return base_name, None
            if not _name_exists(entity_type, base_name):
                return base_name, None
            orig = base_name
            for n in range(2, 5000):
                candidate = f"{orig}_{n}"
                if not _name_exists(entity_type, candidate):
                    msg = f"{kind} name {orig!r} already exists in JITA; using {candidate!r} instead."
                    logger.info(f"[create] {msg}")
                    return candidate, msg
            fallback = f"{orig}_{int(time.time())}"
            return fallback, f"{kind} name {orig!r} exists; using time-based name {fallback!r}."

        jp_name_adjust_warn = None
        new_jp_name, _adj = _pick_unique_name("job_profiles", new_jp_name, "Job Profile")
        if _adj:
            jp_name_adjust_warn = _adj

        def _apply_custom_args_to_existing_ts(ts_id):
            """Merge the newly passed test/framework args into an already-existing TS via PUT.

            When a test set is reused (name collision), JITA keeps it as-is, so the
            args the user just entered are otherwise lost. Fetch the TS, merge args
            on top of whatever it has, and PUT it back. Returns a warning string or None.
            """
            if not (custom_test_args or custom_framework_options):
                return None
            try:
                gr = requests.get(
                    f"{JITA_BASE}/test_sets/{ts_id}",
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                if gr.status_code != 200:
                    return f"Reused test set {ts_id}: could not fetch to apply new args (HTTP {gr.status_code})."
                ts_doc = _jit_pick_ts_dict_from_response(gr.json())
                if not isinstance(ts_doc, dict):
                    return f"Reused test set {ts_id}: unexpected shape, new args not applied."
                _apply_default_framework_options_to_test_set_payload(
                    ts_doc, retain_setup_on_failure, custom_framework_options
                )
                _ensure_test_args_on_test_set_payload(ts_doc, "{}", custom_test_args)
                pr = requests.put(
                    f"{JITA_BASE}/test_sets/{ts_id}",
                    json=ts_doc, auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                logger.info(
                    f"[create] Applied custom args to reused TS {ts_id}: PUT status={pr.status_code}, "
                    f"test_args_keys={list(_test_args_value_to_dict(ts_doc.get('test_args')).keys())}"
                )
                if pr.status_code not in (200, 201):
                    return f"Reused test set {ts_id}: failed to apply new args (HTTP {pr.status_code})."
            except (requests.exceptions.RequestException, ValueError) as e:
                logger.warning(f"[create] Error applying custom args to reused TS {ts_id}: {e}")
                return f"Reused test set {ts_id}: error applying new args ({e})."
            return None

        existing_ts_id = None
        if not (reuse_source_ts and linked_ts_id_for_reuse):
            existing_ts_id = _name_exists("test_sets", new_ts_name)

        ts_reused = False
        ts_create_warning = None

        if reuse_source_ts and linked_ts_id_for_reuse:
            created_ts_id = linked_ts_id_for_reuse
            ts_reused = True
            if testcase_names:
                ts_create_warning = (
                    "reuse_source_ts: cloned job profile uses the existing test set unchanged; "
                    "testcase names in the request were not written to JITA."
                )
            logger.info(f"[create] Linked JP to existing test set id={created_ts_id} (reuse_source_ts)")
        elif existing_ts_id:
            created_ts_id = existing_ts_id
            ts_reused = True
            ts_create_warning = f"Test set '{new_ts_name}' already exists (ID: {existing_ts_id}). Reusing it."
            logger.info(f"[create] TS '{new_ts_name}' already exists, reusing ID {existing_ts_id}")
            # Apply the newly passed args to the reused TS so they aren't lost.
            _reuse_args_warn = _apply_custom_args_to_existing_ts(created_ts_id)
            if _reuse_args_warn:
                ts_create_warning = f"{ts_create_warning} {_reuse_args_warn}"
        else:
            # Build test entries
            if testcase_names:
                row_tmpl = {}
                if isinstance(source_ts, dict):
                    src_tests = source_ts.get("tests") or []
                    if src_tests and isinstance(src_tests[0], dict):
                        row_tmpl = {
                            k: v
                            for k, v in src_tests[0].items()
                            if k not in ("_id", "name") and not str(k).startswith("__")
                        }
                test_entries = []
                for tc in testcase_names:
                    row = dict(row_tmpl)
                    row["name"] = tc
                    # Always honor user-selected test branch for newly added test rows.
                    row["branch"] = nutest_branch
                    row.setdefault("framework_version", "nutest-py3-tests")
                    row.setdefault("framework", "nutest-py3-tests")
                    row.setdefault("package_type", "tar")
                    row.setdefault("service", "NutestPy3Tests")
                    test_entries.append(row)
            else:
                test_entries = source_ts.get("tests", []) or []

            if not create_fresh and source_ts:
                ts_label = source_ts.get("name") or source_testset_id or template_ts_id or "unknown"
                desc = f"Dynamic test set cloned from {ts_label}"
                new_ts_payload = _build_clone_test_set_post_payload(source_ts, new_ts_name, test_entries, desc)
            else:
                _ta, _fa = _jit_ts_arg_strings(source_ts) if source_ts else ("", "")
                new_ts_payload = {
                    "name": new_ts_name,
                    "tests": test_entries,
                    "description": f"Dynamic test set with {len(test_entries)} testcase(s)",
                    "test_args": _ta,
                    "framework_args": _fa,
                    "testArgs": _ta,
                    "frameworkArgs": _fa,
                }
            _apply_default_framework_options_to_test_set_payload(
                new_ts_payload, retain_setup_on_failure, custom_framework_options
            )
            _ensure_test_args_on_test_set_payload(new_ts_payload, "{}", custom_test_args)
            logger.info(f"[create] TS POST payload: name={new_ts_name}, #tests={len(test_entries)}, "
                        f"test_args_keys={list(_test_args_value_to_dict(new_ts_payload.get('test_args')).keys())}, "
                        f"fw_args_keys={list(_framework_args_value_to_dict(new_ts_payload.get('framework_args')).keys())}")

            created_ts_id = None
            try:
                ts_create_resp = requests.post(
                    f"{JITA_BASE}/test_sets",
                    json=new_ts_payload,
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                ts_resp_json = ts_create_resp.json() if ts_create_resp.content else {}
                if ts_resp_json.get("success"):
                    created_ts_id = str(ts_resp_json["id"]) if ts_resp_json.get("id") else None
                    logger.info(f"Created test set: {new_ts_name} (ID: {created_ts_id})")
                    if created_ts_id:
                        try:
                            ts_get_resp = requests.get(
                                f"{JITA_BASE}/test_sets/{created_ts_id}",
                                auth=JITA_SVC_AUTH,
                                verify=False,
                                timeout=30,
                            )
                            if ts_get_resp.status_code == 200:
                                created_ts_doc = _jit_pick_ts_dict_from_response(ts_get_resp.json())
                                if isinstance(created_ts_doc, dict):
                                    _apply_default_framework_options_to_test_set_payload(
                                        created_ts_doc,
                                        retain_setup_on_failure,
                                        custom_framework_options,
                                    )
                                    _ensure_test_args_on_test_set_payload(
                                        created_ts_doc,
                                        "{}",
                                        custom_test_args,
                                    )
                                    ts_put_resp = requests.put(
                                        f"{JITA_BASE}/test_sets/{created_ts_id}",
                                        json=created_ts_doc,
                                        auth=JITA_SVC_AUTH,
                                        verify=False,
                                        timeout=30,
                                    )
                                    logger.info(
                                        f"[create] TS framework options PUT: ts_id={created_ts_id}, "
                                        f"status={ts_put_resp.status_code}"
                                    )
                        except (requests.exceptions.RequestException, ValueError) as e:
                            logger.warning(
                                f"[create] Could not verify/update test set framework options for {created_ts_id}: {e}"
                            )
                else:
                    msg = ts_resp_json.get("message", f"HTTP {ts_create_resp.status_code}")
                    ts_create_warning = f"Test set creation failed: {msg}"
                    logger.warning(f"Failed to create test set: {msg}")
                    return jsonify({"error": f"Failed to create test set '{new_ts_name}': {msg}"}), 500
            except (requests.exceptions.RequestException, ValueError) as e:
                logger.warning(f"Error creating test set: {e}")
                return jsonify({"error": f"Error creating test set: {e}"}), 500

        # 5. Build infra based on provider selection
        def _build_infra(prov, res_type, np_list):
            if prov == "global_pool":
                # Use the user-supplied coupon when present (validated via QMS); otherwise
                # let Jita/QMS auto-allocate by category.
                params = {"category": "general"}
                if global_pool_coupon:
                    params["coupon"] = global_pool_coupon
                if res_type == "physical":
                    return [{"type": "physical", "kind": "PRIVATE_CLOUD", "params": params}]
                # nested_2.0 → "nested", nested_1.0 → "nested_1"
                nested_type = "nested" if res_type in ("nested_2.0", "nested") else "nested_1"
                return [{"type": nested_type, "kind": "PRIVATE_CLOUD", "params": params}]
            elif prov == "node_pool":
                entries = np_list if np_list else ["unknown"]
                return [{"kind": "ON_PREM", "type": "node_pool", "entries": entries}]
            elif prov == "static":
                entries = np_list if np_list else ["unknown"]
                return [{"kind": "ON_PREM", "type": "cluster", "entries": entries}]
            return [{"type": "nested", "kind": "PRIVATE_CLOUD", "params": {"category": "general"}}]

        infra = _build_infra(provider, resource_type, node_pools)

        def _default_scheduling_options():
            return {
                "optimize_scheduling": True,
                "force_imaging": False,
                "task_priority": 10,
                "skip_resource_spec_match": False,
                "upgrade": False,
                "deployment_distribution_algorithm": "test_time_based",
                "retry_imaging": 1,
                "check_image_compatibility": True,
            }

        def _default_requested_hardware(res_type):
            hardware = {
                "hypervisor": "kvm",
                "hypervisor_version": "branch_symlink",
                "imaging_options": {
                    "enable_large_partitions": True,
                    "min_vcpus": 8,
                    "redundancy_factor": "default",
                    "additional_parameters": [
                        {
                            "key": "svm_rescue_args",
                            "value": "--large_partitions_enabled --boot_size=41940992 --home_nutanix_size=167770112",
                        }
                    ],
                },
            }
            if res_type in ("nested_2.0", "nested"):
                hardware["nested_params"] = {"version": "2.0", "is_nested": True}
            elif res_type == "nested_1.0":
                hardware["nested_params"] = {"version": "1.0", "is_nested": True}
            return hardware

        # Build new JP payload
        if create_fresh:
            import copy
            template_jp = None
            template_jp_id = os.getenv("DYNAMIC_JP_TEMPLATE_ID", "6a1e59718e79ce932625e7c0")
            try:
                template_resp = requests.get(
                    f"{JITA_BASE}/job_profiles/{template_jp_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30,
                )
                if template_resp.status_code == 200:
                    template_jp = template_resp.json().get("data", {})
                    if isinstance(template_jp, dict):
                        logger.info(f"[create] Fresh mode — loaded template JP {template_jp_id}")
                    else:
                        template_jp = None
                else:
                    logger.warning(
                        f"[create] Fresh mode — template JP fetch failed: HTTP {template_resp.status_code}"
                    )
            except Exception as exc:
                logger.warning(f"[create] Fresh mode — template JP fetch error: {exc}")

            if template_jp:
                new_jp_payload = copy.deepcopy(template_jp)
                for field in [
                    "_id",
                    "created_at",
                    "updated_at",
                    "created_by",
                    "created_by_user",
                    "__v",
                    "v",
                    "last_triggered",
                    "scheduled_jobs",
                    "auto_schedule_cron",
                ]:
                    new_jp_payload.pop(field, None)
                
                # Don't copy template's coupon - let QMS allocate a new one for this JP
                # Coupons are JP-specific and managed by Jita's QMS
                merged_infra = infra
                
                new_jp_payload.update({
                    "name": new_jp_name,
                    "description": f"Dynamic JP with {len(testcase_names)} testcase(s)",
                    "test_sets": [],
                    "git": {},
                    "build_selection": {},
                    "resource_manager_json": {},
                    "infra": merged_infra,
                    "requested_hardware": _default_requested_hardware(resource_type),
                    "services": ["NOS"],
                    "service": "AOS",
                    "test_framework": "nutest-py3-tests",
                    "nutest-py3-tests_branch": nutest_branch,
                    "scheduling_options": _default_scheduling_options(),
                    "private": True,
                    "user_groups": ["cdp_reg_jarvis"],
                    "auto_schedule_cron": False,
                    "allow_resource_sharing": False,
                    "allow_resource_sharing_across_tasks": False,
                    "advanced_options": {"run_tests_with_tags": False},
                    "skip_bad_tests": True,
                    "run_tests_with_priorities": [],
                    "run_tests_with_additional_tags": [],
                    "sdk_installation_options": {},
                    "demo_mode": False,
                    "image_build_type": "None",
                    "system_under_test": {"product": "aos", "component": "main", "branch": nos_branch},
                })
                new_jp_payload.pop("test_set", None)
            else:
                new_jp_payload = {
                    "name": new_jp_name,
                    "description": f"Dynamic JP with {len(testcase_names)} testcase(s)",
                    "test_sets": [],
                    "git": {},
                    "build_selection": {},
                    "resource_manager_json": {},
                    "infra": infra,
                    "requested_hardware": _default_requested_hardware(resource_type),
                    "services": ["NOS"],
                    "service": "AOS",
                    "test_framework": "nutest-py3-tests",
                    "nutest-py3-tests_branch": nutest_branch,
                    "scheduling_options": _default_scheduling_options(),
                    "private": True,
                    "user_groups": ["cdp_reg_jarvis"],
                    "auto_schedule_cron": False,
                    "allow_resource_sharing": False,
                    "allow_resource_sharing_across_tasks": False,
                    "advanced_options": {"run_tests_with_tags": False},
                    "skip_bad_tests": True,
                    "run_tests_with_priorities": [],
                    "run_tests_with_additional_tags": [],
                    "sdk_installation_options": {},
                    "demo_mode": False,
                    "image_build_type": "None",
                    "system_under_test": {"product": "aos", "component": "main", "branch": nos_branch},
                }
        else:
            import copy
            new_jp_payload = copy.deepcopy(source_jp)
            for field in ["_id", "created_at", "updated_at", "created_by", "__v"]:
                new_jp_payload.pop(field, None)
            new_jp_payload["name"] = new_jp_name
            new_jp_payload["description"] = f"Dynamic JP cloned from {source_jp.get('name', source_jp_id)}"
            # Legacy `test_set` on the source doc can override `test_sets` on POST; clear before we set links.
            new_jp_payload.pop("test_set", None)
            new_jp_payload["test_sets"] = []
            logger.info(f"[create] Source JP keys: {list(source_jp.keys())}")
            logger.info(f"[create] Source JP infra: {source_jp.get('infra')}")
            # Log email-related fields from source JP to understand structure
            for key in source_jp.keys():
                if 'email' in key.lower() or 'mail' in key.lower() or key in ['emails', 'scheduling_options', 'user_groups']:
                    logger.info(f"[create] Source JP email field '{key}': {source_jp.get(key)}")

        # Link to new test set if created
        if created_ts_id:
            new_jp_payload["test_sets"] = [{"$oid": created_ts_id}]
            new_jp_payload.pop("test_set", None)

        if create_fresh:
            # Fresh mode: apply all config from the UI
            git = new_jp_payload.get("git") or {}
            if not isinstance(git, dict):
                git = {}
            git["branch"] = nos_branch
            git["repo"] = "main"
            new_jp_payload["git"] = git

            nos_build_type = "opt" if nos_branch.strip().lower() == "master" else "release"
            pc_build_type = "opt" if pc_branch.strip().lower() == "master" else "release"

            # NOS build_selection based on update_type
            if nos_update_type == "by_commit":
                new_jp_payload["build_selection"] = {
                    "by_commit_id": True,
                    "commit_must_be_newer": False,
                    "build_type": nos_build_type,
                }
                if nos_commit_id:
                    new_jp_payload["build_selection"]["commit_id"] = nos_commit_id
                if nos_gbn:
                    try:
                        new_jp_payload["build_selection"]["gbn"] = int(nos_gbn) if isinstance(nos_gbn, str) else nos_gbn
                    except (ValueError, TypeError):
                        new_jp_payload["build_selection"]["gbn"] = nos_gbn
                # Remove tag-related fields
                new_jp_payload["build_selection"].pop("by_latest_smoked", None)
            else:
                # by_tag (default)
                new_jp_payload["build_selection"] = {
                    "by_latest_smoked": nos_tag == "Latest Smoke Passed",
                    "commit_must_be_newer": False,
                    "build_type": nos_build_type,
                }
                # Remove commit-related fields
                new_jp_payload["build_selection"].pop("by_commit_id", None)
                new_jp_payload["build_selection"].pop("commit_id", None)
                new_jp_payload["build_selection"].pop("gbn", None)

            resource_manager_json = new_jp_payload.get("resource_manager_json") or {}
            if not isinstance(resource_manager_json, dict):
                resource_manager_json = {}
            if "NOS_CLUSTER" not in resource_manager_json:
                resource_manager_json["NOS_CLUSTER"] = {}
            
            # PC configuration based on update_type
            pc_build = {
                "branch": pc_branch,
                "build_selection_build_type": pc_build_type,
            }
            if pc_update_type == "by_commit":
                if pc_commit_id:
                    pc_build["build_selection_option"] = pc_commit_id
            else:
                # by_tag (default)
                pc_build["build_selection_option"] = pc_tag
            
            resource_manager_json["PRISM_CENTRAL"] = {
                "build": pc_build
            }
            new_jp_payload["resource_manager_json"] = resource_manager_json

            test_framework_metadata = new_jp_payload.get("test_framework_metadata") or {}
            if not isinstance(test_framework_metadata, dict):
                test_framework_metadata = {}
            test_meta = test_framework_metadata.get("test") or {}
            if not isinstance(test_meta, dict):
                test_meta = {}
            test_meta["branch"] = nutest_branch
            test_meta["commit"] = None
            if test_patch_url:
                test_meta["patch_url"] = test_patch_url
            else:
                test_meta.pop("patch_url", None)
            framework_meta = test_framework_metadata.get("framework") or {}
            if not isinstance(framework_meta, dict):
                framework_meta = {}
            framework_meta["branch"] = nutest_branch
            framework_meta["commit"] = None
            if framework_patch_url:
                framework_meta["patch_url"] = framework_patch_url
            else:
                framework_meta.pop("patch_url", None)
            test_framework_metadata["test"] = test_meta
            test_framework_metadata["framework"] = framework_meta
            new_jp_payload["test_framework_metadata"] = test_framework_metadata
            new_jp_payload["test_framework"] = "nutest-py3-tests"
            new_jp_payload["nutest-py3-tests_branch"] = nutest_branch
            if framework_patch_url:
                new_jp_payload["patch_url"] = framework_patch_url
            else:
                new_jp_payload.pop("patch_url", None)
            new_jp_payload.pop("nutest_branch", None)

            new_jp_payload["infra"] = infra
        else:
            # Clone mode: preserve source JP's config (infra, git, build, etc.)
            # Only update name, description, test_sets (already done above)
            logger.info(f"[create] Clone mode — preserving source JP config "
                        f"(infra={new_jp_payload.get('infra', 'N/A')[:80] if isinstance(new_jp_payload.get('infra'), str) else 'present'})")
            # Always apply the user-selected test framework branch (Test Options page),
            # and layer patch URLs on top when provided. Previously the branch was only
            # set when a patch URL existed, so cloning without a patch left it unchanged.
            tmeta = new_jp_payload.get("test_framework_metadata")
            tmeta = dict(tmeta) if isinstance(tmeta, dict) else {}
            test_m = tmeta.get("test")
            test_m = dict(test_m) if isinstance(test_m, dict) else {}
            fw_m = tmeta.get("framework")
            fw_m = dict(fw_m) if isinstance(fw_m, dict) else {}
            test_m["branch"] = nutest_branch
            test_m["commit"] = None
            fw_m["branch"] = nutest_branch
            fw_m["commit"] = None
            if test_patch_url:
                test_m["patch_url"] = test_patch_url
            else:
                test_m.pop("patch_url", None)
            if framework_patch_url:
                fw_m["patch_url"] = framework_patch_url
            else:
                fw_m.pop("patch_url", None)
            tmeta["test"] = test_m
            tmeta["framework"] = fw_m
            new_jp_payload["test_framework_metadata"] = tmeta
            new_jp_payload["test_framework"] = "nutest-py3-tests"
            new_jp_payload["nutest-py3-tests_branch"] = nutest_branch
            if framework_patch_url:
                new_jp_payload["patch_url"] = framework_patch_url
            else:
                new_jp_payload.pop("patch_url", None)
            logger.info(f"[create] Clone mode — set test framework branch={nutest_branch}, "
                        f"patches(test={bool(test_patch_url)}, fw={bool(framework_patch_url)})")

            # Use Latest Commit: override build selection to Latest Smoke Passed with optimal build type
            if use_latest_commit:
                logger.info(f"[create] Clone mode — applying use_latest_commit configuration")
                
                # Extract actual branches from source JP (not from request)
                source_git = new_jp_payload.get("git") or {}
                actual_nos_branch = source_git.get("branch", "master") if isinstance(source_git, dict) else "master"
                
                # For PC branch, check resource_manager_json
                resource_manager_json = new_jp_payload.get("resource_manager_json") or {}
                if isinstance(resource_manager_json, dict):
                    pc_config = resource_manager_json.get("PRISM_CENTRAL") or {}
                    if isinstance(pc_config, dict):
                        pc_build = pc_config.get("build") or {}
                        if isinstance(pc_build, dict):
                            actual_pc_branch = pc_build.get("branch", actual_nos_branch)
                        else:
                            actual_pc_branch = actual_nos_branch
                    else:
                        actual_pc_branch = actual_nos_branch
                else:
                    actual_pc_branch = actual_nos_branch
                    resource_manager_json = {}
                
                # Determine build types based on actual branches from source JP
                nos_build_type = "opt" if actual_nos_branch.strip().lower() == "master" else "release"
                pc_build_type = "opt" if actual_pc_branch.strip().lower() == "master" else "release"
                
                logger.info(f"[create] Extracted branches from source JP: nos={actual_nos_branch}, pc={actual_pc_branch}")
                
                # Update git config (keep source branch)
                git = new_jp_payload.get("git") or {}
                if not isinstance(git, dict):
                    git = {}
                git["branch"] = actual_nos_branch
                git["repo"] = "main"
                new_jp_payload["git"] = git
                
                # Update build selection for NOS
                new_jp_payload["build_selection"] = {
                    "by_latest_smoked": True,  # Always true for Latest Smoke Passed
                    "commit_must_be_newer": False,
                    "build_type": nos_build_type,
                }
                # Ensure conflicting fields are removed
                new_jp_payload["build_selection"].pop("by_commit_id", None)
                new_jp_payload["build_selection"].pop("commit_id", None)
                new_jp_payload["build_selection"].pop("gbn", None)
                
                # Update resource_manager_json for PC
                if "NOS_CLUSTER" not in resource_manager_json:
                    resource_manager_json["NOS_CLUSTER"] = {}
                resource_manager_json["PRISM_CENTRAL"] = {
                    "build": {
                        "branch": actual_pc_branch,
                        "build_selection_build_type": pc_build_type,
                        "build_selection_option": "Latest Smoke Passed",
                    }
                }
                new_jp_payload["resource_manager_json"] = resource_manager_json
                
                logger.info(f"[create] Clone mode — set build_selection: nos_build_type={nos_build_type}, "
                           f"pc_build_type={pc_build_type}, by_latest_smoked=True, nos_branch={actual_nos_branch}, pc_branch={actual_pc_branch}")

        if retain_setup_on_failure:
            _apply_retain_setup_on_failure(new_jp_payload)

        _set_tcms_sync_flags(new_jp_payload, sync_to_tcms)

        def _force_email_on_and_clear_tag_filters(jp):
            """Turn 'Send Email Reports' ON (logged-in user as recipient) and disable
            'Run Tests With Tags' / clear any inherited TCMS tag filters. Mutates jp.

            JITA's "Send Email Reports" checkbox is NOT a boolean field — the UI derives
            its state from whether an ``EmailPlugin`` exists in ``plugins.post_run``
            (verified from the JITA frontend bundle). So we add that plugin entry.
            Recipients live in the ``emails`` list (``email_ids`` kept in sync).
            """
            if not isinstance(jp, dict):
                return
            # Add the EmailPlugin to plugins.post_run (idempotent) — this is what the
            # "Send Email Reports" toggle actually reflects.
            plugins = jp.get("plugins")
            if not isinstance(plugins, dict):
                plugins = {}
            post_run = plugins.get("post_run")
            if not isinstance(post_run, list):
                post_run = []
            if not any(isinstance(p, dict) and p.get("name") == "EmailPlugin" for p in post_run):
                post_run.append({
                    "name": "EmailPlugin",
                    "args": {},
                    "description": "Sends mail to the recipients.",
                    "stage": "post_run",
                    "metadata": {"kind": "task"},
                })
            plugins["post_run"] = post_run
            if not isinstance(plugins.get("pre_run"), list):
                plugins["pre_run"] = []
            jp["plugins"] = plugins
            # Legacy boolean flags (harmless; some API paths still read them).
            jp["send_email_on_completion"] = True
            jp["send_emails"] = True
            if current_user_email:
                recipients = [current_user_email]
                jp["emails"] = recipients
                jp["email_ids"] = recipients
                jp["private"] = True
                jp["user_groups"] = ["cdp_reg_jarvis"]
            # Disable "Run Tests With Tags" and drop inherited TCMS tag(s) (e.g. "unstable").
            adv = jp.get("advanced_options")
            if not isinstance(adv, dict):
                adv = {}
            adv["run_tests_with_tags"] = False
            adv["tags"] = []
            jp["advanced_options"] = adv
            jp["run_tests_with_additional_tags"] = []

        _force_email_on_and_clear_tag_filters(new_jp_payload)
        logger.info(
            f"[create] Email ON (recipient={current_user_email}); "
            f"run_tests_with_tags disabled and TCMS tag filters cleared"
        )

        # Tags will be applied via a separate PUT after creation (same
        # approach as Run Plan) because JITA's POST ignores tag fields.

        # Deep sanitize: ensure JSON serializable (handle ObjectId, sets, bytes, etc.)
        def sanitize_value(val):
            if isinstance(val, dict):
                return {k: sanitize_value(v) for k, v in val.items()}
            elif isinstance(val, list):
                return [sanitize_value(item) for item in val]
            elif isinstance(val, (set, tuple)):
                return [sanitize_value(item) for item in val]
            elif isinstance(val, bytes):
                return val.decode("utf-8", errors="replace")
            elif val is Ellipsis:
                return None
            else:
                return val

        serializable_payload = sanitize_value(new_jp_payload)
        logger.info(f"[create] Final JP payload — name: {serializable_payload.get('name')}, "
                    f"tags: {serializable_payload.get('tags', 'MISSING')}, "
                    f"build_selection: {serializable_payload.get('build_selection', 'MISSING')}, "
                    f"adv_opts_keys: {list((serializable_payload.get('advanced_options') or {}).keys())}, "
                    f"adv_tags: {(serializable_payload.get('advanced_options') or {}).get('tags', 'MISSING')}, "
                    f"test_sets: {serializable_payload.get('test_sets', 'MISSING')}")

        # 6. POST new JP
        try:
            jp_create_resp = requests.post(
                f"{JITA_BASE}/job_profiles",
                json=serializable_payload,
                auth=JITA_SVC_AUTH,
                verify=False,
                timeout=30
            )
        except requests.exceptions.Timeout:
            note = f" Note: Test set '{new_ts_name}' (ID: {created_ts_id}) was already created." if created_ts_id and not ts_reused else ""
            return jsonify({"error": f"Timed out creating job profile on JITA.{note}"}), 504
        except requests.exceptions.ConnectionError:
            note = f" Note: Test set '{new_ts_name}' (ID: {created_ts_id}) was already created." if created_ts_id and not ts_reused else ""
            return jsonify({"error": f"Could not connect to JITA to create job profile.{note}"}), 503

        try:
            jp_resp_json = jp_create_resp.json()
        except (ValueError, TypeError):
            jp_resp_json = {}

        if not jp_resp_json.get("success"):
            error_msg = jp_resp_json.get("message", f"HTTP {jp_create_resp.status_code}")
            logger.error(f"Failed to create JP: {error_msg}")
            note = f" Note: Test set '{new_ts_name}' (ID: {created_ts_id}) was already created." if created_ts_id and not ts_reused else ""
            return jsonify({
                "error": f"Failed to create Job Profile: {error_msg}.{note}",
            }), 500

        created_jp_id = jp_resp_json.get("id")
        if created_jp_id:
            created_jp_id = str(created_jp_id)

        logger.info(f"Created JP: {new_jp_name} (ID: {created_jp_id})")

        # JITA can fill TCMS fields back to defaults during POST. Unless the
        # user explicitly enabled Sync TCMS, re-fetch and clear them via PUT.
        tcms_cleanup_warning = None
        if not sync_to_tcms and created_jp_id:
            try:
                get_resp = requests.get(
                    f"{JITA_BASE}/job_profiles/{created_jp_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30,
                )
                if get_resp.status_code == 200:
                    jp_data = get_resp.json().get("data", {})
                    if isinstance(jp_data, dict):
                        _set_tcms_sync_flags(jp_data, False)
                        # Log infra from Jita's perspective after JP creation
                        infra_after = jp_data.get('infra', [])
                        coupon_allocated = None
                        if isinstance(infra_after, list) and len(infra_after) > 0:
                            params = infra_after[0].get('params', {})
                            coupon_allocated = params.get('coupon')
                        logger.info(
                            f"[create] Jita GET after POST: "
                            f"infra={infra_after}, "
                            f"coupon={coupon_allocated}, "
                            f"emails={jp_data.get('emails')}, "
                            f"send_emails={jp_data.get('send_emails')}, "
                            f"private={jp_data.get('private')}, "
                            f"user_groups={jp_data.get('user_groups')}"
                        )
                        # JITA can reset these on POST, so re-apply them on the PUT.
                        jp_data["auto_schedule_cron"] = False
                        jp_data["allow_resource_sharing"] = False
                        jp_data["allow_resource_sharing_across_tasks"] = False
                        jp_data["skip_bad_tests"] = True
                        jp_data["run_tests_with_priorities"] = jp_data.get("run_tests_with_priorities") or []
                        jp_data["sdk_installation_options"] = jp_data.get("sdk_installation_options") or {}
                        jp_data["demo_mode"] = False
                        jp_data["image_build_type"] = jp_data.get("image_build_type") or "None"
                        if create_fresh:
                            jp_data["requested_hardware"] = _default_requested_hardware(resource_type)
                        # Force email ON + clear inherited tag filters.
                        _force_email_on_and_clear_tag_filters(jp_data)
                        logger.info(
                            f"[create] Re-applied email + tag filters "
                            f"(send_emails={jp_data.get('send_emails')}, emails={jp_data.get('emails')})"
                        )
                        put_resp = requests.put(
                            f"{JITA_BASE}/job_profiles/{created_jp_id}",
                            json=sanitize_value(jp_data),
                            auth=JITA_SVC_AUTH,
                            verify=False,
                            timeout=30,
                        )
                        if put_resp.status_code == 200:
                            logger.info(
                                f"[create] TCMS fields cleared via PUT for JP {created_jp_id}"
                            )
                            # Log final state after PUT
                            try:
                                final_get = requests.get(
                                    f"{JITA_BASE}/job_profiles/{created_jp_id}",
                                    auth=JITA_SVC_AUTH,
                                    verify=False,
                                    timeout=15,
                                )
                                if final_get.status_code == 200:
                                    final_data = final_get.json().get("data", {})
                                    final_infra = final_data.get('infra', [])
                                    final_coupon = None
                                    if isinstance(final_infra, list) and len(final_infra) > 0:
                                        final_coupon = final_infra[0].get('params', {}).get('coupon')
                                    final_tfm = final_data.get('test_framework_metadata', {}) or {}
                                    final_test_branch = (final_tfm.get('test') or {}).get('branch')
                                    logger.info(
                                        f"[create] Final state after PUT: "
                                        f"emails={final_data.get('emails')}, "
                                        f"send_email_on_completion={final_data.get('send_email_on_completion')}, "
                                        f"send_emails={final_data.get('send_emails')}, "
                                        f"test_branch={final_test_branch}, "
                                        f"nutest_branch={final_data.get('nutest-py3-tests_branch')}, "
                                        f"private={final_data.get('private')}, "
                                        f"user_groups={final_data.get('user_groups')}, "
                                        f"coupon={final_coupon}"
                                    )
                            except Exception as e:
                                logger.warning(f"[create] Could not GET final state: {e}")
                        else:
                            tcms_cleanup_warning = (
                                f"JP created but TCMS sync cleanup failed (HTTP {put_resp.status_code})"
                            )
                            logger.warning(
                                f"[create] TCMS cleanup PUT failed: HTTP {put_resp.status_code} — {put_resp.text[:300]}"
                            )
                    else:
                        tcms_cleanup_warning = "JP created but TCMS sync cleanup could not re-fetch JP data"
                else:
                    tcms_cleanup_warning = (
                        f"JP created but TCMS sync cleanup re-fetch failed (HTTP {get_resp.status_code})"
                    )
                    logger.warning(
                        f"[create] TCMS cleanup GET failed: HTTP {get_resp.status_code} — {get_resp.text[:300]}"
                    )
            except Exception as e:
                tcms_cleanup_warning = f"JP created but TCMS sync cleanup failed: {e}"
                logger.warning(f"[create] TCMS cleanup error: {e}")

        # 7. Apply tags via PUT (same approach as Run Plan — JITA ignores
        # tag fields on POST but accepts them on PUT via tester_tags)
        tag_warning = None
        if (jp_tags or sync_to_tcms) and created_jp_id:
            effective_tags = list(dict.fromkeys(jp_tags + (["official"] if sync_to_tcms else [])))
            if not sync_to_tcms:
                effective_tags = [t for t in effective_tags if t != "official"]
            logger.info(f"[create] Applying tags {effective_tags} to JP {created_jp_id} via PUT (tester_tags)")
            try:
                get_resp = requests.get(
                    f"{JITA_BASE}/job_profiles/{created_jp_id}",
                    auth=JITA_SVC_AUTH,                     verify=False, timeout=30,
                )
                if get_resp.status_code == 200:
                    jp_data = get_resp.json().get("data", {})
                    if isinstance(jp_data, dict):
                        tester_tags = jp_data.get("tester_tags", [])
                        if not isinstance(tester_tags, list):
                            tester_tags = []
                        merged = list(dict.fromkeys(tester_tags + effective_tags))
                        if not sync_to_tcms:
                            merged = [t for t in merged if t != "official"]
                        jp_data["tester_tags"] = merged
                        _set_tcms_sync_flags(jp_data, sync_to_tcms)
                        # Keep email ON and tag filters cleared on this PUT too
                        # (covers the sync_to_tcms path where the block above is skipped).
                        _force_email_on_and_clear_tag_filters(jp_data)

                        put_payload = {}
                        for k, v in jp_data.items():
                            if isinstance(v, (set, tuple)):
                                put_payload[k] = list(v)
                            elif v is Ellipsis:
                                put_payload[k] = None
                            else:
                                put_payload[k] = v

                        put_resp = requests.put(
                            f"{JITA_BASE}/job_profiles/{created_jp_id}",
                            json=put_payload,
                            auth=JITA_SVC_AUTH,                             verify=False, timeout=30,
                        )
                        if put_resp.status_code == 200:
                            logger.info(f"[create] Tags applied successfully: tester_tags={merged}")
                        else:
                            tag_warning = f"JP created but tags could not be applied (HTTP {put_resp.status_code})"
                            logger.warning(f"[create] PUT tags failed: HTTP {put_resp.status_code} — {put_resp.text[:200]}")
                    else:
                        tag_warning = "JP created but could not re-fetch it to apply tags"
                else:
                    tag_warning = f"JP created but re-fetch for tags failed (HTTP {get_resp.status_code})"
                    logger.warning(f"[create] GET for tag update failed: HTTP {get_resp.status_code}")
            except Exception as e:
                tag_warning = f"JP created but tags could not be applied: {e}"
                logger.warning(f"[create] Tag update error: {e}")

        warnings = []
        if jp_name_adjust_warn:
            warnings.append(jp_name_adjust_warn)
        if ts_fetch_warning:
            warnings.append(ts_fetch_warning)
        if ts_create_warning:
            warnings.append(ts_create_warning)
        if tag_warning:
            warnings.append(tag_warning)
        if tcms_cleanup_warning:
            warnings.append(tcms_cleanup_warning)

        ts_msg = ""
        if created_ts_id:
            ts_msg = f" (reused existing {new_ts_name})" if ts_reused else f" and {new_ts_name}"

        return jsonify({
            "success": True,
            "reuse_source_ts": reuse_source_ts,
            "job_profile": {
                "_id": created_jp_id,
                "name": new_jp_name,
                "ui_url": _jita_browser_entity_url(JITA_WEB_JOB_PROFILE_URL, created_jp_id, new_jp_name),
            },
            "test_set": {
                "_id": created_ts_id,
                "name": new_ts_name,
                "reused": ts_reused,
                "ui_url": _jita_browser_entity_url(JITA_WEB_TEST_SET_URL, created_ts_id, new_ts_name),
            } if created_ts_id else None,
            "message": f"Created {new_jp_name}{ts_msg}",
            "warnings": warnings if warnings else None,
        })
    except Exception as e:
        logger.error(f"Error in dynamic-jp create: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/update", methods=["POST"])
def dynamic_jp_update():
    """Update an existing dynamic job profile or test set."""
    try:
        req_data = request.json
        if not req_data:
            return jsonify({"error": "Request body is required (JSON)"}), 400

        jp_id = req_data.get("jp_id")
        ts_id = req_data.get("ts_id")
        updates = req_data.get("updates", {})

        if not jp_id and not ts_id:
            return jsonify({"error": "jp_id or ts_id is required"}), 400

        if not isinstance(updates, dict):
            return jsonify({"error": "updates must be a JSON object"}), 400

        if jp_id:
            jp_id = str(jp_id).strip()
        if ts_id:
            ts_id = str(ts_id).strip()

        results = {}

        if jp_id and updates.get("jp_updates"):
            jp_updates = updates["jp_updates"]
            if not isinstance(jp_updates, dict):
                return jsonify({"error": "jp_updates must be a JSON object"}), 400

            try:
                get_resp = requests.get(
                    f"{JITA_BASE}/job_profiles/{jp_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
            except requests.exceptions.RequestException as e:
                return jsonify({"error": f"Failed to connect to JITA to fetch JP {jp_id}: {e}"}), 503

            if get_resp.status_code != 200:
                return jsonify({"error": f"Failed to fetch JP {jp_id} (HTTP {get_resp.status_code})"}), 500

            try:
                existing = get_resp.json().get("data", {})
            except (ValueError, TypeError):
                return jsonify({"error": "Invalid JSON from JITA when fetching JP"}), 500

            if not isinstance(existing, dict):
                existing = {}
            existing.update(jp_updates)

            def sanitize(val):
                if isinstance(val, dict):
                    return {k: sanitize(v) for k, v in val.items()}
                elif isinstance(val, list):
                    return [sanitize(item) for item in val]
                elif isinstance(val, (set, tuple)):
                    return [sanitize(item) for item in val]
                elif isinstance(val, bytes):
                    return val.decode("utf-8", errors="replace")
                elif val is Ellipsis:
                    return None
                return val

            serializable = sanitize(existing)

            try:
                put_resp = requests.put(
                    f"{JITA_BASE}/job_profiles/{jp_id}",
                    json=serializable,
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
                results["jp"] = {
                    "success": put_resp.status_code == 200,
                    "status_code": put_resp.status_code,
                    "message": "Updated" if put_resp.status_code == 200 else f"JITA returned {put_resp.status_code}",
                }
            except requests.exceptions.RequestException as e:
                results["jp"] = {"success": False, "error": str(e)}

        if ts_id and updates.get("ts_updates"):
            ts_updates = updates["ts_updates"]
            if not isinstance(ts_updates, dict):
                return jsonify({"error": "ts_updates must be a JSON object"}), 400

            try:
                get_resp = requests.get(
                    f"{JITA_BASE}/test_sets/{ts_id}",
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
            except requests.exceptions.RequestException as e:
                return jsonify({"error": f"Failed to connect to JITA to fetch test set {ts_id}: {e}"}), 503

            if get_resp.status_code != 200:
                return jsonify({"error": f"Failed to fetch test set {ts_id} (HTTP {get_resp.status_code})"}), 500

            try:
                existing_ts = get_resp.json().get("data", {})
            except (ValueError, TypeError):
                return jsonify({"error": "Invalid JSON from JITA when fetching test set"}), 500

            if not isinstance(existing_ts, dict):
                existing_ts = {}
            existing_ts.update(ts_updates)

            try:
                put_resp = requests.put(
                    f"{JITA_BASE}/test_sets/{ts_id}",
                    json=existing_ts,
                    auth=JITA_SVC_AUTH,
                    verify=False,
                    timeout=30
                )
                results["ts"] = {
                    "success": put_resp.status_code == 200,
                    "status_code": put_resp.status_code,
                    "message": "Updated" if put_resp.status_code == 200 else f"JITA returned {put_resp.status_code}",
                }
            except requests.exceptions.RequestException as e:
                results["ts"] = {"success": False, "error": str(e)}

        return jsonify({"success": True, "results": results})
    except Exception as e:
        logger.error(f"Error in dynamic-jp update: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/optional-defaults", methods=["GET"])
def dynamic_jp_optional_defaults():
    """Return the optional framework-option defaults the UI 'Add defaults' button injects."""
    return jsonify({
        "framework_options": DYN_TESTSET_OPTIONAL_FRAMEWORK_OPTS,
    })


def _oid_bounds_for_local_date(date_str):
    """Return (start_oid, end_oid) hex ObjectIds spanning a local calendar date.

    MongoDB ObjectIds embed the creation time (UTC seconds) in their leading 4
    bytes, so an ``_id`` range filters strictly by creation timestamp. ``date_str``
    is ``YYYY-MM-DD`` in the server's local timezone (same as the user's). Returns
    ``None`` if the date can't be parsed.
    """
    try:
        day_start = datetime.strptime(date_str.strip(), "%Y-%m-%d")
    except (ValueError, AttributeError):
        return None
    start_ts = int(day_start.timestamp())
    end_ts = int((day_start + timedelta(days=1)).timestamp())
    start_oid = f"{start_ts:08x}" + "0" * 16
    end_oid = f"{end_ts:08x}" + "0" * 16
    return start_oid, end_oid


@app.route("/mcp/regression/dynamic-jp/search", methods=["POST"])
def dynamic_jp_search():
    """Search Job Profiles and Test Sets by name and/or creation date.

    Either ``query`` (>= 2 chars name substring) or ``date`` (YYYY-MM-DD) is
    required; both may be combined. When ``date`` is supplied, results are limited
    to entities **created by this tool** (JITA service account) on that calendar
    date, using an ObjectId creation-timestamp range.
    """
    try:
        req_data = request.json or {}
        query = (req_data.get("query") or "").strip()
        date_str = (req_data.get("date") or "").strip()

        if len(query) < 2 and not date_str:
            return jsonify({"error": "Provide a search term (>= 2 chars) or pick a date"}), 400

        date_cond = None
        if date_str:
            bounds = _oid_bounds_for_local_date(date_str)
            if not bounds:
                return jsonify({"error": "date must be YYYY-MM-DD"}), 400
            start_oid, end_oid = bounds
            date_cond = {"_id": {"$gte": {"$oid": start_oid}, "$lt": {"$oid": end_oid}}}

        name_cond = (
            {"name": {"$regex": re.escape(query), "$options": "i"}}
            if len(query) >= 2
            else None
        )

        # "Created by this tool" marker: every dynamic JP/TS is tagged with the
        # cdp_reg_jarvis user group (JITA's created_by is an opaque ObjectId and is
        # unreliable). Only applied when a date is selected so plain name search keeps
        # its broad behaviour.
        tool_cond = {"user_groups": {"$in": ["cdp_reg_jarvis"]}}
        jp_tool_cond = tool_cond
        ts_tool_cond = tool_cond

        def _build_raw_query(tool_cond):
            conds = []
            if name_cond:
                conds.append(name_cond)
            if date_cond:
                conds.append(date_cond)
                conds.append(tool_cond)
            if not conds:
                return None
            return json.dumps({"$and": conds} if len(conds) > 1 else conds[0])

        limit = min(int(req_data.get("limit", 20)), 50)
        result = {"job_profiles": [], "test_sets": []}

        def _extract_id(item):
            eid = item.get("_id")
            if isinstance(eid, dict) and "$oid" in eid:
                return eid["$oid"]
            return str(eid) if eid else None

        jp_raw_q = _build_raw_query(jp_tool_cond)
        if jp_raw_q:
            try:
                jp_resp = requests.get(
                    f"{JITA_BASE}/job_profiles",
                    params={"raw_query": jp_raw_q, "limit": limit, "sort": "-_id",
                            "only": "_id,name,description,tags,created_at"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=20,
                )
                if jp_resp.status_code == 200:
                    for item in (jp_resp.json().get("data", []) or []):
                        if not isinstance(item, dict):
                            continue
                        result["job_profiles"].append({
                            "_id": _extract_id(item),
                            "name": item.get("name", ""),
                            "description": item.get("description", ""),
                            "created_at": item.get("created_at"),
                        })
            except Exception as e:
                logger.warning(f"[search] JP search failed: {e}")

        ts_raw_q = _build_raw_query(ts_tool_cond)
        if ts_raw_q:
            try:
                ts_resp = requests.get(
                    f"{JITA_BASE}/test_sets",
                    params={"raw_query": ts_raw_q, "limit": limit, "sort": "-_id",
                            "only": "_id,name,description,created_at"},
                    auth=JITA_SVC_AUTH, verify=False, timeout=20,
                )
                if ts_resp.status_code == 200:
                    for item in (ts_resp.json().get("data", []) or []):
                        if not isinstance(item, dict):
                            continue
                        result["test_sets"].append({
                            "_id": _extract_id(item),
                            "name": item.get("name", ""),
                            "description": item.get("description", ""),
                            "created_at": item.get("created_at"),
                        })
            except Exception as e:
                logger.warning(f"[search] TS search failed: {e}")

        return jsonify({"success": True, **result})
    except Exception as e:
        logger.error(f"Error in dynamic-jp search: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/dynamic-jp/delete", methods=["POST"])
def dynamic_jp_delete():
    """Delete one or more Job Profiles and/or Test Sets by ID."""
    try:
        req_data = request.json or {}
        jp_ids = req_data.get("jp_ids", [])
        ts_ids = req_data.get("ts_ids", [])

        if not jp_ids and not ts_ids:
            return jsonify({"error": "At least one of jp_ids or ts_ids is required"}), 400

        if not isinstance(jp_ids, list):
            jp_ids = [jp_ids]
        if not isinstance(ts_ids, list):
            ts_ids = [ts_ids]

        jp_ids = [str(i).strip() for i in jp_ids if i]
        ts_ids = [str(i).strip() for i in ts_ids if i]

        results = {"job_profiles": [], "test_sets": []}

        for jp_id in jp_ids:
            try:
                resp = requests.delete(
                    f"{JITA_BASE}/job_profiles/{jp_id}",
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                success = resp.status_code in (200, 204)
                msg = "Deleted" if success else f"JITA returned HTTP {resp.status_code}"
                if not success:
                    try:
                        msg = resp.json().get("message", msg)
                    except Exception:
                        pass
                results["job_profiles"].append({
                    "_id": jp_id,
                    "success": success,
                    "message": msg,
                })
                logger.info(f"[delete] JP {jp_id}: {'OK' if success else 'FAILED'} ({msg})")
            except Exception as e:
                results["job_profiles"].append({
                    "_id": jp_id,
                    "success": False,
                    "message": str(e),
                })
                logger.warning(f"[delete] JP {jp_id} error: {e}")

        for ts_id in ts_ids:
            try:
                resp = requests.delete(
                    f"{JITA_BASE}/test_sets/{ts_id}",
                    auth=JITA_SVC_AUTH, verify=False, timeout=30,
                )
                success = resp.status_code in (200, 204)
                msg = "Deleted" if success else f"JITA returned HTTP {resp.status_code}"
                if not success:
                    try:
                        msg = resp.json().get("message", msg)
                    except Exception:
                        pass
                results["test_sets"].append({
                    "_id": ts_id,
                    "success": success,
                    "message": msg,
                })
                logger.info(f"[delete] TS {ts_id}: {'OK' if success else 'FAILED'} ({msg})")
            except Exception as e:
                results["test_sets"].append({
                    "_id": ts_id,
                    "success": False,
                    "message": str(e),
                })
                logger.warning(f"[delete] TS {ts_id} error: {e}")

        all_ok = all(r["success"] for r in results["job_profiles"] + results["test_sets"])
        return jsonify({"success": all_ok, "results": results})
    except Exception as e:
        logger.error(f"Error in dynamic-jp delete: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/debug/inspect-jp/<jp_id>", methods=["GET"])
def debug_inspect_jp(jp_id):
    """Debug endpoint to inspect JP structure from JITA"""
    try:
        jp_resp = requests.get(
            f"{JITA_BASE}/job_profiles/{jp_id}",
            auth=JITA_SVC_AUTH,
            verify=False,
            timeout=30
        )
        if jp_resp.status_code != 200:
            return jsonify({"error": f"Failed to fetch JP (HTTP {jp_resp.status_code})"}), jp_resp.status_code
        
        jp_data = jp_resp.json().get("data", {})
        
        # Extract email-related fields
        email_fields = {}
        for key in jp_data.keys():
            if 'email' in key.lower() or 'mail' in key.lower() or key in ['emails', 'user_groups', 'scheduling_options']:
                email_fields[key] = jp_data.get(key)
        
        return jsonify({
            "jp_id": jp_id,
            "jp_name": jp_data.get("name"),
            "all_keys": list(jp_data.keys()),
            "email_related_fields": email_fields
        })
    except Exception as e:
        logger.error(f"Error inspecting JP: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# AI Analysis Endpoints
# ======================================================

def _call_ai_chat(system_prompt, user_content, max_tokens=2048):
    """Helper to call the Nutanix AI chat endpoint."""
    payload = {
        "model": "hack-reason",
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_content}
        ],
        "max_tokens": max_tokens,
        "stream": False
    }
    url = f"{AI_BASE}/chat/completions"
    headers = {
        "Authorization": f"Bearer {AI_API_KEY}",
        "Accept": "application/json",
        "Content-Type": "application/json"
    }
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=90) as resp:
        if resp.getcode() != 200:
            raise Exception(f"AI API returned HTTP {resp.getcode()}")
        response_data = json.loads(resp.read().decode())
        choices = response_data.get("choices", [])
        if not choices:
            raise Exception("AI returned no choices")
        content = (choices[0].get("message") or {}).get("content", "")
        return content.strip()


def _build_bulk_issues_fallback_analysis(issue_rows, tag, total_tests):
    """Generate deterministic markdown when external AI endpoint is unavailable."""
    table_header = (
        "| Bug ID | Affected Feature | Testcases Impacted | QI Impact | Risk Level | Impact | Recommended Action |\n"
        "| --- | --- | ---: | ---: | --- | --- | --- |"
    )
    table_rows = []
    for row in issue_rows:
        ticket = row["ticket"]
        feature = row["feature"] or "Unknown"
        testcase_count = row["testcase_count"]
        qi_impact = row["qi_impact"]
        risk_level = row["risk_level"]
        impact = (
            f"Potentially impacts {testcase_count} testcases in {feature}."
            if feature != "Unknown"
            else f"Potentially impacts {testcase_count} testcases across multiple areas."
        )
        recommended_action = (
            "Validate ownership, confirm reproducibility, and prioritize based on QI impact."
        )
        table_rows.append(
            f"| {ticket} | {feature} | {testcase_count} | {qi_impact} | {risk_level} | {impact} | {recommended_action} |"
        )

    critical_count = sum(1 for row in issue_rows if row["risk_level"] == "Critical")
    high_count = sum(1 for row in issue_rows if row["risk_level"] == "High")
    medium_count = sum(1 for row in issue_rows if row["risk_level"] == "Medium")
    risk_score = min(100, critical_count * 30 + high_count * 20 + medium_count * 10 + 15)
    readiness = (
        "Not Ready"
        if critical_count > 0
        else "At Risk"
        if high_count > 0
        else "Proceed with Caution"
        if medium_count > 0
        else "Ready with Monitoring"
    )

    top_three = sorted(
        issue_rows,
        key=lambda r: (
            {"Critical": 4, "High": 3, "Medium": 2, "Low": 1}.get(r["risk_level"], 0),
            abs(float(r["qi_impact"])),
        ),
        reverse=True,
    )[:3]

    priorities = "\n".join(
        f"- {row['ticket']} ({row['risk_level']}, QI impact {row['qi_impact']}%)"
        for row in top_three
    ) or "- No priority tickets identified"

    return (
        "_AI endpoint unreachable. Showing deterministic fallback analysis._\n\n"
        f"**Regression Tag:** `{tag or 'N/A'}`  \n"
        f"**Total Tests in Run:** `{total_tests}`  \n"
        f"**Bulk Issues:** `{len(issue_rows)}`\n\n"
        f"{table_header}\n"
        + "\n".join(table_rows)
        + "\n\n"
        + f"### Overall Summary\n"
        + f"- Overall Risk Score: **{risk_score}/100**\n"
        + f"- Release Readiness: **{readiness}**\n"
        + f"- Top 3 Priorities:\n{priorities}"
    )


@app.route("/mcp/regression/ai-analysis/bulk-issues", methods=["POST"])
@app.route("/api/mcp/regression/ai-analysis/bulk-issues", methods=["POST"])
@jwt_required
def ai_analyze_bulk_issues():
    """AI analysis for bulk issue JIRA tickets - analyzes patterns and provides risk assessment."""
    try:
        body = request.get_json(force=True) or {}
        bulk_issues = body.get("bulk_issues", {})
        bulk_issues_with_qi = body.get("bulk_issues_with_qi", {})
        tag = body.get("tag", "")
        total_tests = body.get("total_tests_processed", 0)

        if not bulk_issues:
            return jsonify({"error": "No bulk issues provided"}), 400

        issue_detail_blocks = []
        issue_rows = []
        for ticket, tests in bulk_issues.items():
            qi_data = bulk_issues_with_qi.get(ticket, {})
            qi_impact = qi_data.get("overall_qi_impact", "N/A")
            count = qi_data.get("testcase_count", len(tests)) if qi_data else len(tests)
            qi_numeric = None
            if qi_impact != "N/A":
                try:
                    qi_numeric = float(qi_impact)
                except Exception:
                    qi_numeric = None

            if qi_numeric is None:
                risk_level = "Low"
            elif qi_numeric <= -5:
                risk_level = "Critical"
            elif qi_numeric <= -2:
                risk_level = "High"
            elif qi_numeric <= -1:
                risk_level = "Medium"
            else:
                risk_level = "Low"

            tc_names = tests if isinstance(tests, list) else []
            # Extract affected feature from test name prefixes
            feature_prefixes = set()
            for name in tc_names[:20]:
                parts = name.split(".")
                if len(parts) >= 3:
                    feature_prefixes.add(".".join(parts[:3]))
                elif len(parts) >= 2:
                    feature_prefixes.add(".".join(parts[:2]))
            affected_features = ", ".join(sorted(feature_prefixes)[:3])
            tc_sample = tc_names[:10]
            tc_display = "\n".join(f"    - {name}" for name in tc_sample)
            if len(tc_names) > 10:
                tc_display += f"\n    ... and {len(tc_names) - 10} more"
            primary_feature = sorted(feature_prefixes)[0] if feature_prefixes else "Unknown"
            issue_rows.append(
                {
                    "ticket": ticket,
                    "feature": primary_feature,
                    "testcase_count": count,
                    "qi_impact": qi_numeric if qi_numeric is not None else 0.0,
                    "risk_level": risk_level,
                }
            )
            issue_detail_blocks.append(
                f"Ticket: {ticket}\n"
                f"  Testcases Impacted: {count}\n"
                f"  QI Impact: {qi_impact}{'%' if qi_impact != 'N/A' else ''}\n"
                f"  Affected Feature(s): {affected_features or 'Unknown'}\n"
                f"  Affected Testcases:\n{tc_display}"
            )

        user_content = (
            f"Regression Tag: {tag}\n"
            f"Total Tests in Run: {total_tests}\n"
            f"Number of Bulk Issues (tickets affecting >5 testcases): {len(bulk_issues)}\n\n"
            f"=== Bulk Issue Details ===\n\n" + "\n\n".join(issue_detail_blocks) + "\n\n"
            "IMPORTANT: Use EXACTLY the ticket IDs and testcase counts from the data above. "
            "Do NOT change or reorder the tickets. Analyse the actual testcase names to determine patterns.\n\n"
            "For EACH ticket listed above (in the same order), provide:\n"
            "1. Bug ID — the JIRA ticket ID exactly as given\n"
            "2. Affected Feature — derived from testcase name prefix (e.g., cdp.stargate.storage_policy)\n"
            "3. Risk Level (Critical/High/Medium/Low) — based on QI impact: <=-5% Critical, <=-2% High, <=-1% Medium, else Low\n"
            "4. Impact — one sentence describing what is broken\n"
            "5. Recommended Action — specific next step (if this looks like a test issue, note it can be fixed; "
            "if product issue, note the affected component)\n\n"
            "Then provide an overall summary:\n"
            "- Overall Risk Score (0-100, where 100 is highest risk)\n"
            "- Release Readiness assessment\n"
            "- Top 3 priorities to address\n\n"
            "Format using markdown. Use a table with columns: "
            "Bug ID | Affected Feature | Testcases Impacted | QI Impact | Risk Level | Impact | Recommended Action\n"
            "The Bug ID, Testcases Impacted, and QI Impact columns MUST match the input data exactly."
        )

        system_prompt = (
            "You are a regression testing expert analyzing bulk failure patterns in a CDP (Continuous Data Protection) "
            "software QA pipeline at Nutanix. "
            "Bulk issues are JIRA tickets that affect more than 5 testcases, indicating systemic problems. "
            "QI Impact is the Quality Index impact - negative values mean the bug is reducing overall quality. "
            "Risk Level thresholds: QI Impact <= -5% is Critical, <= -2% is High, <= -1% is Medium, > -1% is Low. "
            "You MUST use the exact ticket IDs and counts provided — never invent or reorder them. "
            "Infer the Affected Feature from testcase name prefixes (e.g., cdp.stargate.checksum_regions → CDP Stargate Checksum). "
            "For Recommended Action, distinguish between test issues (can be fixed in test code) and product issues "
            "(require product team attention). "
            "Provide actionable, concise analysis. Use markdown tables for structured data."
        )

        try:
            analysis = _call_ai_chat(system_prompt, user_content, max_tokens=3000)
        except urllib.error.URLError as err:
            logger.warning(
                f"AI bulk issues endpoint unreachable ({err.reason}); returning fallback analysis."
            )
            analysis = _build_bulk_issues_fallback_analysis(issue_rows, tag, total_tests)
        return jsonify({"success": True, "analysis": analysis})

    except Exception as e:
        logger.error(f"Error in AI bulk issues analysis: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/ai-analysis/deep-triage", methods=["POST"])
@app.route("/api/mcp/regression/ai-analysis/deep-triage", methods=["POST"])
@jwt_required
def ai_deep_triage():
    """
    Prepare context for deep triage of a single bug using the triage-cdp-test-failure skill.
    Returns structured context (log URL, test name, ticket, feature area) that the frontend
    uses to invoke the Cursor SDK agent with the skill.
    """
    try:
        body = request.get_json(force=True) or {}
        ticket_id = body.get("ticket_id", "")
        test_name = body.get("test_name", "")
        log_url = body.get("log_url", "")
        task_id = body.get("task_id", "")

        if not ticket_id and not test_name:
            return jsonify({"error": "ticket_id or test_name is required"}), 400

        # Extract feature area from test name
        feature_area = ""
        if test_name:
            parts = test_name.split(".")
            if len(parts) >= 3:
                feature_area = ".".join(parts[:3])
            elif len(parts) >= 2:
                feature_area = ".".join(parts[:2])

        # Build Jita log directory URL if task_id and test_name are provided but log_url is not
        if not log_url and task_id and test_name:
            # Convert test name to path format for log URL construction
            test_path = "/".join(test_name.split("."))
            log_url = f"http://10.40.234.216/logs/{task_id}/{test_path}/"

        # Construct the skill invocation command
        skill_command = f"/triage-cdp-test-failure"
        if ticket_id:
            skill_command += f" {ticket_id}"
        if log_url:
            skill_command += f" {log_url}"

        # Build context for the Cursor agent
        context = {
            "ticket_id": ticket_id,
            "test_name": test_name,
            "feature_area": feature_area,
            "log_url": log_url,
            "task_id": task_id,
            "skill_command": skill_command,
            "skill_prompt": (
                f"Triage the following CDP test failure:\n"
                f"- JIRA Ticket: {ticket_id}\n"
                f"- Test Name: {test_name}\n"
                f"- Feature Area: {feature_area}\n"
                f"- Jita Log Directory: {log_url}\n\n"
                f"Use the triage-cdp-test-failure skill to:\n"
                f"1. Fetch and analyse the test logs\n"
                f"2. Search the test source code in Sourcegraph (nugerrit.ntnxdpro.com/nutest-py3-tests)\n"
                f"3. Determine if this is a test issue or product issue\n"
                f"4. If test issue, propose a fix and offer to create a CR\n"
                f"5. Provide structured triage result"
            ),
        }

        return jsonify({"success": True, "context": context})

    except Exception as e:
        logger.error(f"Error in deep triage: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/jira-ticket-details", methods=["POST"])
@app.route("/api/mcp/regression/jira-ticket-details", methods=["POST"])
@jwt_required
def get_jira_ticket_details():
    """Fetch JIRA status and issue type for a list of ticket IDs."""
    try:
        body = request.get_json(force=True) or {}
        ticket_ids = body.get("ticket_ids", [])
        if not ticket_ids:
            return jsonify({"error": "No ticket IDs provided"}), 400

        results = {}
        for ticket_id in ticket_ids[:50]:
            jira_data = fetch_jira_ticket(ticket_id)
            if jira_data:
                fields = jira_data.get("fields", {})
                results[ticket_id] = {
                    "status": fields.get("status", {}).get("name", "Unknown"),
                    "issue_type": fields.get("issuetype", {}).get("name", "Unknown")
                }
            else:
                results[ticket_id] = {"status": "N/A", "issue_type": "N/A"}

        return jsonify({"success": True, "details": results})
    except Exception as e:
        logger.error(f"Error fetching JIRA ticket details: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/ai-analysis/owner-tickets", methods=["POST"])
@app.route("/api/mcp/regression/ai-analysis/owner-tickets", methods=["POST"])
@jwt_required
def ai_analyze_owner_tickets():
    """AI analysis for owner-wise JIRA ticket breakdown with JIRA data fetch."""
    try:
        body = request.get_json(force=True) or {}
        owner_ticket_map = body.get("owner_ticket_map", {})
        triage_summary = body.get("triage_summary", {})
        tag = body.get("tag", "")
        total_tests = body.get("total_tests_processed", 0)

        if not owner_ticket_map:
            return jsonify({"error": "No owner ticket data provided"}), 400

        # Collect all unique tickets and fetch JIRA data
        all_tickets = set()
        owner_for_ticket = {}
        ticket_test_count = {}
        for owner, tickets in owner_ticket_map.items():
            for ticket, count in tickets.items():
                all_tickets.add(ticket)
                owner_for_ticket.setdefault(ticket, []).append(owner)
                ticket_test_count[ticket] = ticket_test_count.get(ticket, 0) + count

        # Fetch JIRA data for each ticket
        jira_details = {}
        for ticket_id in all_tickets:
            jira_data = fetch_jira_ticket(ticket_id)
            if jira_data:
                fields = jira_data.get("fields", {})
                status = fields.get("status", {}).get("name", "Unknown")
                issue_type = fields.get("issuetype", {}).get("name", "Unknown")
                summary = fields.get("summary", "")
                description = (fields.get("description") or "")[:500]
                priority = fields.get("priority", {}).get("name", "Unknown") if fields.get("priority") else "Unknown"
                assignee = fields.get("assignee", {}).get("displayName", "Unassigned") if fields.get("assignee") else "Unassigned"
                # Get latest comments
                comments_data = fields.get("comment", {}).get("comments", [])
                latest_comments = ""
                for c in comments_data[-3:]:
                    comment_body = (c.get("body") or "")[:200]
                    latest_comments += f"  [{c.get('author', {}).get('displayName', 'Unknown')}]: {comment_body}\n"
                jira_details[ticket_id] = {
                    "status": status,
                    "issue_type": issue_type,
                    "summary": summary,
                    "description": description,
                    "priority": priority,
                    "assignee": assignee,
                    "comments": latest_comments.strip()
                }
            else:
                jira_details[ticket_id] = {
                    "status": "N/A",
                    "issue_type": "N/A",
                    "summary": "Unable to fetch",
                    "description": "",
                    "priority": "N/A",
                    "assignee": "N/A",
                    "comments": ""
                }

        # Build ticket detail lines for AI
        ticket_detail_lines = []
        for ticket_id in sorted(all_tickets):
            jd = jira_details[ticket_id]
            owners = ", ".join(owner_for_ticket.get(ticket_id, []))
            tc_count = ticket_test_count.get(ticket_id, 0)
            detail = (
                f"Ticket: {ticket_id}\n"
                f"  Owners: {owners}\n"
                f"  Testcases Impacted: {tc_count}\n"
                f"  JIRA Status: {jd['status']}\n"
                f"  JIRA Issue Type: {jd['issue_type']}\n"
                f"  Priority: {jd['priority']}\n"
                f"  Assignee: {jd['assignee']}\n"
                f"  Summary: {jd['summary']}\n"
            )
            if jd['description']:
                detail += f"  Description: {jd['description'][:300]}\n"
            if jd['comments']:
                detail += f"  Latest Comments:\n{jd['comments']}\n"
            ticket_detail_lines.append(detail)

        user_content = (
            f"Regression Tag: {tag}\n"
            f"Total Tests in Run: {total_tests}\n"
            f"Total Owners: {len(owner_ticket_map)}\n"
            f"Total Unique Tickets: {len(all_tickets)}\n\n"
            f"=== Ticket Details with JIRA Data ===\n\n"
            + "\n".join(ticket_detail_lines) + "\n\n"
            "Based on the JIRA data (status, description, comments), create a comprehensive analysis table:\n\n"
            "Create a markdown table with these columns:\n"
            "| Ticket ID | Owner | Status | Issue Type | Root Cause Category | Impact | AI Recommended Action |\n\n"
            "For Root Cause Category, classify as one of: Product | Test | Infra | Framework\n"
            "- Product: Bug in product code causing test failure\n"
            "- Test: Test code issue (flaky, outdated assertions, bad test data)\n"
            "- Infra: Infrastructure problem (environment, network, resources)\n"
            "- Framework: Test framework issue (automation framework, libraries)\n\n"
            "Base your classification on the JIRA description, comments, and issue type.\n"
            "For Impact, one sentence on what this ticket affects.\n"
            "For AI Recommended Action, provide a specific action based on the ticket status:\n"
            "- If Open/To Do: suggest priority and who should work on it\n"
            "- If In Progress: assess if it's on track\n"
            "- If Resolved/Closed: suggest verifying the fix and closing test gaps\n\n"
            "After the table, provide:\n"
            "1. **Summary** — Overall triage health assessment\n"
            "2. **Top 3 Priority Actions** — Most impactful actions to take now\n\n"
            "Use the EXACT ticket IDs, owner names, and status from the data."
        )

        system_prompt = (
            "You are a QA team lead analyzing JIRA tickets linked to regression test failures. "
            "You have access to real JIRA data including status, description, and comments. "
            "Classify each ticket's root cause category (Product/Test/Infra/Framework) based on the JIRA content. "
            "Provide specific, actionable recommendations based on ticket status and context. "
            "Use markdown tables. Use the exact ticket IDs and data provided — never invent data."
        )

        analysis = _call_ai_chat(system_prompt, user_content)

        # Return both analysis and jira_details for the frontend table
        return jsonify({
            "success": True,
            "analysis": analysis,
            "jira_details": jira_details
        })

    except Exception as e:
        logger.error(f"Error in AI owner tickets analysis: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/ai-analysis/testcase-summary", methods=["POST"])
@app.route("/api/mcp/regression/ai-analysis/testcase-summary", methods=["POST"])
@jwt_required
def ai_analyze_testcases():
    """AI analysis for ALL filtered testcase data across all pages - overall health, patterns, and recommendations."""
    try:
        body = request.get_json(force=True) or {}
        stats = body.get("stats", {})
        failed_testcases = body.get("failed_testcases", [])
        branch = body.get("branch", "")
        team = body.get("team", "")

        total = stats.get("total", 0)
        succeeded = stats.get("succeeded", 0)
        failed = stats.get("failed", 0)
        avg_stability = stats.get("avg_stability", "N/A")
        avg_success = stats.get("avg_success", "N/A")
        avg_qi = stats.get("avg_qi", "N/A")

        issue_type_breakdown = stats.get("issue_type_breakdown", {})
        component_breakdown = stats.get("component_breakdown", {})
        unique_tickets = stats.get("unique_tickets", [])

        if not issue_type_breakdown:
            for tc in failed_testcases:
                it = tc.get("issue_type", "Unknown")
                issue_type_breakdown[it] = issue_type_breakdown.get(it, 0) + 1

        if not unique_tickets:
            ticket_set = set()
            for tc in failed_testcases:
                for t in tc.get("last_run_tickets", []):
                    ticket_set.add(t)
            unique_tickets = list(ticket_set)

        failed_details = []
        for tc in failed_testcases[:50]:
            failed_details.append(
                f"  - {tc.get('name', 'Unknown')}: issue_type={tc.get('issue_type', 'N/A')}, "
                f"stability={tc.get('stability', 'N/A')}%, component={tc.get('primary_component', 'N/A')}, "
                f"tickets={tc.get('last_run_tickets', [])}"
            )

        component_lines = "\n".join(
            f"  {k}: {v} failed testcases" for k, v in sorted(component_breakdown.items(), key=lambda x: -x[1])[:15]
        ) if component_breakdown else "  No component data available"

        user_content = (
            f"Branch: {branch}\nTeam: {team}\n\n"
            f"=== Complete Test Suite Statistics (ALL filtered pages) ===\n"
            f"Total Testcases (all pages): {total}\n"
            f"Succeeded: {succeeded}\n"
            f"Failed: {failed}\n"
            f"Pass Rate: {(succeeded/total*100) if total else 0:.1f}%\n"
            f"Avg Stability: {avg_stability}%\n"
            f"Avg Success Rate: {avg_success}%\n"
            f"Avg Quality Index: {avg_qi}%\n\n"
            f"=== Issue Type Breakdown (ALL {failed} failed testcases) ===\n"
            + "\n".join(f"  {k}: {v}" for k, v in issue_type_breakdown.items()) + "\n\n"
            f"=== Component Breakdown (failed testcases by component) ===\n"
            + component_lines + "\n\n"
            f"=== Unique JIRA Tickets ({len(unique_tickets)}) ===\n"
            + ", ".join(unique_tickets[:30]) + "\n\n"
            f"=== Failed Testcase Samples (top 50 of {failed} total) ===\n"
            + "\n".join(failed_details) + "\n\n"
            "Provide a comprehensive analysis of the ENTIRE filtered dataset using markdown formatting with tables:\n"
            "1. **Overall Health Assessment** - Use a table with Metric/Value/Interpretation columns\n"
            "2. **Failure Pattern Analysis** - Common patterns across ALL failed testcases, systemic issues\n"
            "3. **Issue Type Summary** - Use a table with Issue Type/Count/Percentage/Comment columns\n"
            "4. **Component Risk Areas** - Table of components with highest failure concentration\n"
            "5. **Ticket Analysis** - Assessment of JIRA tickets linked to failed tests in a table\n"
            "6. **Recommendations** - Top 5 actionable items to improve test health\n"
            "7. **Risk Score** - Overall risk score 0-100 (100=highest risk)\n"
        )

        system_prompt = (
            "You are a QA analytics expert analyzing testcase data for a regression test suite. "
            "The data represents ALL filtered testcases across all pages (not just the first page). "
            "Provide actionable insights with specific recommendations. "
            "Focus on identifying systemic issues, high-risk areas, and concrete steps to improve. "
            "Use markdown formatting with tables (using | syntax) for structured data. "
            "Be concise but thorough."
        )

        analysis = _call_ai_chat(system_prompt, user_content)
        return jsonify({"success": True, "analysis": analysis})

    except Exception as e:
        logger.error(f"Error in AI testcase analysis: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/ai-analysis/run-plan-risk", methods=["POST"])
@app.route("/api/mcp/regression/ai-analysis/run-plan-risk", methods=["POST"])
@jwt_required
def ai_run_plan_risk():
    """AI risk analysis for a run plan based on its history and job profiles."""
    try:
        body = request.get_json(force=True) or {}
        run_plan_name = body.get("name", "")
        job_profile_count = body.get("job_profile_count", 0)
        history = body.get("history", [])
        tag_name = body.get("tag_name", "")

        history_summary = []
        success_count = 0
        fail_count = 0
        for entry in history[:10]:
            status = entry.get("status", "unknown")
            if status.lower() in ("success", "completed", "succeeded"):
                success_count += 1
            else:
                fail_count += 1
            history_summary.append(
                f"  - Date: {entry.get('triggered_at', 'N/A')}, Status: {status}, "
                f"Tasks: {len(entry.get('task_ids', []))}"
            )

        total_runs = success_count + fail_count
        success_rate = (success_count / total_runs * 100) if total_runs else 0

        user_content = (
            f"Run Plan: {run_plan_name}\n"
            f"Tag Name: {tag_name}\n"
            f"Job Profiles: {job_profile_count}\n\n"
            f"=== Run History (last {len(history[:10])} runs) ===\n"
            f"Success Rate: {success_rate:.0f}% ({success_count}/{total_runs})\n"
            + "\n".join(history_summary) + "\n\n"
            "Provide using markdown formatting:\n"
            "1. **Risk Score** (0-100, where 100 is highest risk)\n"
            "2. **Risk Level** (Low/Medium/High/Critical)\n"
            "3. **Confidence** (High/Medium/Low based on data available)\n"
            "4. **Key Factors** - What contributes to the risk\n"
            "5. **Recommendation** - One sentence actionable advice\n\n"
            "Keep the response concise (under 200 words). Use a summary table."
        )

        system_prompt = (
            "You are a regression testing risk analyst. Given run plan history, estimate "
            "the risk of the next scheduled run failing. Consider success rate trends, "
            "task counts, and patterns. Provide a numeric risk score and brief assessment. "
            "Use markdown formatting with tables. Be direct and actionable."
        )

        analysis = _call_ai_chat(system_prompt, user_content, max_tokens=512)

        risk_score_match = re.search(r"Risk Score[:\s]*(\d+)", analysis)
        risk_score = int(risk_score_match.group(1)) if risk_score_match else (
            90 if fail_count > success_count else
            60 if success_rate < 70 else
            30 if success_rate < 90 else 15
        )

        risk_level_match = re.search(r"Risk Level[:\s]*(Critical|High|Medium|Low)", analysis, re.IGNORECASE)
        risk_level = risk_level_match.group(1) if risk_level_match else (
            "Critical" if risk_score >= 80 else
            "High" if risk_score >= 60 else
            "Medium" if risk_score >= 40 else "Low"
        )

        return jsonify({
            "success": True,
            "risk_score": risk_score,
            "risk_level": risk_level,
            "analysis": analysis,
            "success_rate": round(success_rate, 1),
            "total_runs": total_runs
        })

    except Exception as e:
        logger.error(f"Error in AI run plan risk analysis: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# Cursor AI Bridge — Deep testcase analysis via Cursor SDK
# ======================================================
CURSOR_BRIDGE_URL = os.getenv("CURSOR_BRIDGE_URL", "http://localhost:5002")

# In-memory store for async Cursor AI analysis jobs
_cursor_ai_jobs = {}
_cursor_ai_jobs_lock = threading.Lock()


@app.route("/mcp/regression/cursor-ai/analyze-testcase", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/analyze-testcase", methods=["POST"])
@jwt_required
def cursor_ai_analyze_testcase():
    """Trigger deep AI analysis for a single failed testcase via the Cursor bridge."""
    try:
        body = request.get_json(force=True) or {}
        testcase_name = body.get("testcase_name", "")
        if not testcase_name:
            return jsonify({"error": "testcase_name is required"}), 400

        exception_summary = body.get("exception_summary", "")
        exception = body.get("exception", "")
        test_log_url = body.get("test_log_url", "")
        jira_tickets = body.get("jira_tickets", [])
        failure_stage = body.get("failure_stage", "")

        # Fetch logs if not provided in the request
        steps_log = body.get("steps_log", "")
        nutest_test_log = body.get("nutest_test_log", "")
        if not steps_log and not nutest_test_log and test_log_url:
            logs = fetch_testcase_logs(testcase_name, test_log_url)
            steps_log = logs.get("steps_log", "")
            nutest_test_log = logs.get("nutest_test_log", "")

        payload = {
            "testcase_name": testcase_name,
            "exception_summary": exception_summary,
            "exception": exception,
            "steps_log": steps_log[:5000],
            "nutest_test_log": nutest_test_log[:5000],
            "test_log_url": test_log_url,
            "jira_tickets": jira_tickets,
            "failure_stage": failure_stage,
        }

        resp = requests.post(
            f"{CURSOR_BRIDGE_URL}/analyze-testcase",
            json=payload,
            timeout=600,
        )
        if resp.status_code != 200:
            error_msg = resp.json().get("error", resp.text) if resp.headers.get("content-type", "").startswith("application/json") else resp.text
            return jsonify({"error": f"Bridge error: {error_msg}"}), 502

        data = resp.json()
        return jsonify({
            "success": True,
            "session_id": data.get("session_id", ""),
            "analysis": data.get("analysis", {}),
        })

    except requests.exceptions.ConnectionError:
        logger.error("Cursor bridge is not reachable at %s", CURSOR_BRIDGE_URL)
        return jsonify({"error": "Cursor AI bridge is not running. Start it with: cd cursor-bridge && npm start"}), 503
    except requests.exceptions.Timeout:
        return jsonify({"error": "Analysis timed out (>600s). Try batch mode for complex testcases."}), 504
    except Exception as e:
        logger.error(f"Error in cursor-ai analyze-testcase: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/analyze-batch", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/analyze-batch", methods=["POST"])
@jwt_required
def cursor_ai_analyze_batch():
    """Start async batch analysis for multiple selected failed testcases."""
    try:
        body = request.get_json(force=True) or {}
        testcases = body.get("testcases", [])
        if not testcases:
            return jsonify({"error": "testcases array is required"}), 400

        # Enrich each testcase with logs if missing
        enriched = []
        for tc in testcases:
            name = tc.get("testcase_name", "")
            log_url = tc.get("test_log_url", "")
            if name and log_url and not tc.get("steps_log") and not tc.get("nutest_test_log"):
                logs = fetch_testcase_logs(name, log_url)
                tc["steps_log"] = logs.get("steps_log", "")[:5000]
                tc["nutest_test_log"] = logs.get("nutest_test_log", "")[:5000]
            enriched.append(tc)

        resp = requests.post(
            f"{CURSOR_BRIDGE_URL}/analyze-batch",
            json={"testcases": enriched},
            timeout=30,
        )
        if resp.status_code != 200:
            error_msg = resp.json().get("error", resp.text) if resp.headers.get("content-type", "").startswith("application/json") else resp.text
            return jsonify({"error": f"Bridge error: {error_msg}"}), 502

        data = resp.json()
        job_id = data.get("job_id", "")

        # Track the job locally
        with _cursor_ai_jobs_lock:
            _cursor_ai_jobs[job_id] = {
                "status": "running",
                "total": len(enriched),
                "completed": 0,
                "created_at": datetime.now().isoformat(),
            }

        return jsonify({"success": True, "job_id": job_id, "total": len(enriched)})

    except requests.exceptions.ConnectionError:
        return jsonify({"error": "Cursor AI bridge is not running. Start it with: cd cursor-bridge && npm start"}), 503
    except Exception as e:
        logger.error(f"Error in cursor-ai analyze-batch: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/status/<job_id>", methods=["GET"])
@app.route("/api/mcp/regression/cursor-ai/status/<job_id>", methods=["GET"])
@jwt_required
def cursor_ai_job_status(job_id):
    """Poll the status of an async batch analysis job from the Cursor bridge."""
    try:
        resp = requests.get(
            f"{CURSOR_BRIDGE_URL}/status/{job_id}",
            timeout=10,
        )
        if resp.status_code == 404:
            return jsonify({"error": "Job not found"}), 404
        if resp.status_code != 200:
            return jsonify({"error": "Bridge error"}), 502

        data = resp.json()

        # Update local tracker
        with _cursor_ai_jobs_lock:
            if job_id in _cursor_ai_jobs:
                _cursor_ai_jobs[job_id]["status"] = data.get("status", "unknown")
                _cursor_ai_jobs[job_id]["completed"] = data.get("completed", 0)

        return jsonify(data)

    except requests.exceptions.ConnectionError:
        return jsonify({"error": "Cursor AI bridge is not reachable"}), 503
    except Exception as e:
        logger.error(f"Error polling cursor-ai status: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/result", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/result", methods=["POST"])
def cursor_ai_result_callback():
    """Callback endpoint for the bridge to push results (async mode).
    No JWT required since this is a server-to-server call from the bridge."""
    try:
        body = request.get_json(force=True) or {}
        job_id = body.get("job_id", "")
        testcase_id = body.get("testcase_id", "")
        analysis = body.get("analysis", {})

        if not job_id or not testcase_id:
            return jsonify({"error": "job_id and testcase_id are required"}), 400

        with _cursor_ai_jobs_lock:
            job = _cursor_ai_jobs.get(job_id)
            if job:
                if "results" not in job:
                    job["results"] = {}
                job["results"][testcase_id] = analysis
                job["completed"] = len(job["results"])
                if job["completed"] >= job.get("total", 0):
                    job["status"] = "done"

        logger.info(f"[cursor-ai] Received result for job={job_id}, testcase={testcase_id}")
        return jsonify({"success": True})

    except Exception as e:
        logger.error(f"Error in cursor-ai result callback: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/follow-up", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/follow-up", methods=["POST"])
@jwt_required
def cursor_ai_follow_up():
    """Send a follow-up question to an existing Cursor AI analysis session.
    The agent retains full conversation context from the initial triage."""
    try:
        body = request.get_json(force=True) or {}
        session_id = body.get("session_id", "")
        question = body.get("question", "")
        recovery_context = body.get("recovery_context", {}) or {}
        mode = (body.get("mode", "agent") or "agent").strip().lower()
        if mode not in {"ask", "agent", "plan"}:
            mode = "agent"

        if not session_id:
            return jsonify({"error": "session_id is required"}), 400
        if not question:
            return jsonify({"error": "question is required"}), 400

        resp = requests.post(
            f"{CURSOR_BRIDGE_URL}/follow-up",
            json={
                "session_id": session_id,
                "question": question,
                "mode": mode,
                "recovery_context": recovery_context,
            },
            timeout=600,
        )
        if resp.status_code == 404:
            return jsonify({"error": "Session expired or not found. Run a new analysis first."}), 404
        if resp.status_code != 200:
            error_msg = resp.json().get("error", resp.text) if resp.headers.get("content-type", "").startswith("application/json") else resp.text
            return jsonify({"error": f"Bridge error: {error_msg}"}), 502

        data = resp.json()
        return jsonify({
            "success": True,
            "session_id": session_id,
            "analysis": data.get("analysis", {}),
        })

    except requests.exceptions.ConnectionError:
        return jsonify({"error": "Cursor AI bridge is not reachable"}), 503
    except requests.exceptions.Timeout:
        return jsonify({"error": "Follow-up timed out (>600s)."}), 504
    except Exception as e:
        logger.error(f"Error in cursor-ai follow-up: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


# ======================================================
# Cursor AI Interactive Chat
# ======================================================

MCP_SERVER_CONFIGS = {
    "regx-data": {"url": "http://localhost:5003", "description": "RegX regression data"},
    "atlassian": {"url": "https://panacea-dev.eng.nutanix.com/mcp/atlassian", "description": "Jira & Confluence"},
    "gw-sourcegraph": {"url": "https://panacea-dev.eng.nutanix.com/mcp/sourcegraph", "description": "Code search"},
    "gw-jita": {"url": "https://panacea-dev.eng.nutanix.com/mcp/jita", "description": "JITA log access"},
    "gw-diamond": {"url": "https://panacea-dev.eng.nutanix.com/mcp/diamond", "description": "Diamond storage"},
    "gw-glean": {"url": "https://panacea-dev.eng.nutanix.com/mcp/glean", "description": "Internal knowledge search"},
    "gw-supportgpt": {"url": "https://panacea-dev.eng.nutanix.com/mcp/supportgpt", "description": "Support knowledge base"},
    "gw-nurag": {"url": "https://panacea-dev.eng.nutanix.com/mcp/nurag", "description": "Advanced RAG"},
    "gw-slack": {"url": "https://panacea-dev.eng.nutanix.com/mcp/slack", "description": "Slack integration"},
    "gw-panacea": {"url": "https://panacea-dev.eng.nutanix.com/mcp/panacea", "description": "Automated RCA"},
    "gw-live-debug": {"url": "https://panacea-dev.eng.nutanix.com/mcp/live-debug", "description": "Live debugging"},
    "auto-handoff": {"url": "http://10.40.224.6:9001/sse", "description": "Auto handoff"},
}

SOURCEGRAPH_SKILL_REPO = "nugerrit.ntnxdpro.com/nutest-py3-tests"
SOURCEGRAPH_WEB_BASE = "https://sourcegraph.ntnxdpro.com"
SYNCABLE_SOURCEGRAPH_SKILLS = {
    "triage-rdm-deployment-failure": ".cursor/skills/triage-rdm-deployment-failure/",
    "triage-cdp-test-failure": ".cursor/skills/triage-cdp-test-failure/",
    "glean-search": ".cursor/skills/glean-search/",
    "gerrit-comment-resolver": ".cursor/skills/gerrit-comment-resolver/",
}

MODE_SYSTEM_PROMPTS = {
    "agent": (
        "You are Cursor AI in Agent mode — a powerful implementation assistant with full access to MCP tools. "
        "You can search code, read logs, query knowledge bases, interact with Jira/Confluence, and perform "
        "automated triage. Provide detailed, actionable responses. When appropriate, explain which tools or "
        "data sources you would use to accomplish the task."
    ),
    "plan": (
        "You are Cursor AI in Plan mode — a read-only collaborative planning assistant. "
        "Help the user design approaches, explore trade-offs, and create step-by-step plans before implementation. "
        "Do not perform actions, only propose strategies with clear reasoning."
    ),
    "debug": (
        "You are Cursor AI in Debug mode — a systematic troubleshooting assistant. "
        "Help investigate bugs, failures, and unexpected behavior. Ask clarifying questions, "
        "suggest diagnostic steps, and methodically narrow down root causes using available MCP tools and logs."
    ),
    "ask": (
        "You are Cursor AI in Ask mode — a knowledge assistant for exploring code and answering questions. "
        "Provide clear, concise explanations. Reference specific files, functions, and architecture when relevant. "
        "Do not make changes, only inform."
    ),
}

# Store for MCP tool call results during chat (per-request)
_mcp_call_cache = {}
_mcp_sessions = {}
_mcp_rpc_counter = 0


def _next_mcp_rpc_id():
    global _mcp_rpc_counter
    _mcp_rpc_counter += 1
    return _mcp_rpc_counter


def _parse_mcp_http_response(resp):
    """Parse MCP JSON or SSE-like HTTP response into a JSON object."""
    text = resp.text or ""
    content_type = (resp.headers.get("Content-Type") or "").lower()

    # Standard JSON-RPC body
    if "application/json" in content_type:
        return resp.json()

    # Streamable transport commonly returns SSE framed data.
    data_payloads = []
    for line in text.splitlines():
        if line.startswith("data: "):
            payload = line[len("data: "):].strip()
            if payload:
                data_payloads.append(payload)
    if not data_payloads:
        raise RuntimeError("MCP response did not include JSON payload")
    return json.loads(data_payloads[-1])


def _normalize_mcp_tool_result(mcp_json):
    """Unwrap MCP tool result shape into plain dict."""
    if not isinstance(mcp_json, dict):
        return {"error": f"Unexpected MCP response type: {type(mcp_json).__name__}"}
    if mcp_json.get("error"):
        err = mcp_json["error"]
        return {"error": err.get("message", str(err)) if isinstance(err, dict) else str(err)}

    result = mcp_json.get("result", {})
    if isinstance(result, dict) and "content" in result and isinstance(result["content"], list):
        # Gateway format: result.content[0].text holds JSON-serialized tool output.
        text_chunks = []
        for item in result["content"]:
            if isinstance(item, dict) and item.get("type") == "text":
                text_chunks.append(item.get("text", ""))
        joined_text = "\n".join(chunk for chunk in text_chunks if chunk).strip()
        if joined_text:
            try:
                parsed = json.loads(joined_text)
                if isinstance(parsed, dict):
                    return parsed
                return {"result": parsed}
            except Exception:
                return {"result_text": joined_text}
    if isinstance(result, dict):
        return result
    return {"result": result}


def _call_mcp_tool(server_id, tool_name, arguments, timeout=30, _retry_on_auth=True):
    """Call an MCP tool and return normalized result."""
    config = MCP_SERVER_CONFIGS.get(server_id)
    if not config:
        return {"error": f"Unknown MCP server: {server_id}"}
    url = config["url"]
    try:
        session_id = _mcp_sessions.get(server_id)
        if not session_id:
            init_payload = {
                "jsonrpc": "2.0",
                "id": _next_mcp_rpc_id(),
                "method": "initialize",
                "params": {
                    "protocolVersion": "2024-11-05",
                    "capabilities": {},
                    "clientInfo": {"name": "regx-backend", "version": "1.0"},
                },
            }
            init_resp = requests.post(url, json=init_payload, timeout=timeout, verify=False)
            if init_resp.status_code != 200:
                return {"error": f"MCP initialize failed ({init_resp.status_code})"}
            session_id = init_resp.headers.get("mcp-session-id")
            if not session_id:
                return {"error": "MCP initialize did not return session ID"}
            _mcp_sessions[server_id] = session_id

        call_payload = {
            "jsonrpc": "2.0",
            "id": _next_mcp_rpc_id(),
            "method": "tools/call",
            "params": {
                "name": tool_name,
                "arguments": arguments or {},
            },
        }
        call_resp = requests.post(
            url,
            json=call_payload,
            headers={"mcp-session-id": session_id},
            timeout=timeout,
            verify=False,
        )

        # Session might expire; retry once with fresh initialize.
        if call_resp.status_code in (400, 401) and _retry_on_auth:
            _mcp_sessions.pop(server_id, None)
            return _call_mcp_tool(server_id, tool_name, arguments, timeout=timeout, _retry_on_auth=False)

        if call_resp.status_code != 200:
            return {"error": f"MCP call failed ({call_resp.status_code})"}

        parsed_resp = _parse_mcp_http_response(call_resp)
        return _normalize_mcp_tool_result(parsed_resp)
    except Exception as e:
        return {"error": f"MCP call failed: {str(e)}"}


def _strip_sourcegraph_line_prefixes(content):
    """Remove Sourcegraph read_file line-number prefixes."""
    cleaned = []
    for line in (content or "").splitlines():
        cleaned.append(re.sub(r"^\s*\d+:\s?", "", line))
    return "\n".join(cleaned).strip() + "\n"


def _list_sourcegraph_files_recursive(repo, root_path):
    """List all files under a Sourcegraph directory using MCP list_files."""
    normalized_root = root_path.rstrip("/")
    pending_dirs = [normalized_root]
    seen_dirs = set()
    files = []

    while pending_dirs:
        current_dir = pending_dirs.pop()
        if current_dir in seen_dirs:
            continue
        seen_dirs.add(current_dir)

        list_resp = _call_mcp_tool(
            "gw-sourcegraph",
            "sourcegraph__list_files",
            {"repo": repo, "path": current_dir},
            timeout=45,
        )
        if list_resp.get("error"):
            raise RuntimeError(list_resp["error"])

        entries = list_resp.get("files", [])
        if not isinstance(entries, list):
            raise RuntimeError(f"Invalid list_files response for {current_dir}")

        for entry in entries:
            entry_path = (entry or {}).get("path")
            if not entry_path:
                continue
            if (entry or {}).get("isDirectory"):
                pending_dirs.append(entry_path.rstrip("/"))
            else:
                files.append(entry_path)

    # Keep stable order for easier debugging and deterministic results.
    files.sort()
    return files


def _list_sourcegraph_files_recursive_http(repo, root_path):
    """Fallback: list Sourcegraph files by crawling tree pages."""
    normalized_root = root_path.rstrip("/")
    pending_dirs = [normalized_root]
    seen_dirs = set()
    files = []

    while pending_dirs:
        current_dir = pending_dirs.pop()
        if current_dir in seen_dirs:
            continue
        seen_dirs.add(current_dir)

        tree_url = f"{SOURCEGRAPH_WEB_BASE}/{repo}/-/tree/{current_dir}"
        resp = requests.get(tree_url, timeout=45, verify=False)
        if resp.status_code != 200:
            raise RuntimeError(f"Sourcegraph tree fetch failed ({resp.status_code}) for {current_dir}")

        html = resp.text
        tree_pattern = rf'/{re.escape(repo)}/-/tree/([^"?#]+)'
        blob_pattern = rf'/{re.escape(repo)}/-/blob/([^"?#]+)'

        for match in re.findall(tree_pattern, html):
            path = urllib.parse.unquote(match).rstrip("/")
            if path.startswith(normalized_root) and path not in seen_dirs:
                pending_dirs.append(path)

        for match in re.findall(blob_pattern, html):
            path = urllib.parse.unquote(match).rstrip("/")
            if path.startswith(normalized_root):
                files.append(path)

    files = sorted(set(files))
    return files


def _read_sourcegraph_file_http(repo, file_path):
    """Fallback: read Sourcegraph file via raw endpoint."""
    raw_url = f"{SOURCEGRAPH_WEB_BASE}/{repo}/-/raw/{file_path}"
    resp = requests.get(raw_url, timeout=45, verify=False)
    if resp.status_code != 200:
        raise RuntimeError(f"Sourcegraph raw fetch failed ({resp.status_code}) for {file_path}")
    return resp.text


@app.route("/mcp/regression/cursor-ai/sync-skills", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/sync-skills", methods=["POST"])
@jwt_required
def cursor_ai_sync_skills():
    """Sync selected Sourcegraph skills into local .cursor/skills."""
    try:
        body = request.get_json(force=True) or {}
        requested_ids = body.get("skill_ids") or list(SYNCABLE_SOURCEGRAPH_SKILLS.keys())
        if not isinstance(requested_ids, list):
            return jsonify({"error": "skill_ids must be a list"}), 400

        invalid_ids = [skill_id for skill_id in requested_ids if skill_id not in SYNCABLE_SOURCEGRAPH_SKILLS]
        if invalid_ids:
            return jsonify({
                "error": "Unsupported skill IDs requested",
                "invalid_skill_ids": invalid_ids,
                "allowed_skill_ids": list(SYNCABLE_SOURCEGRAPH_SKILLS.keys()),
            }), 400

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        skills_root = os.path.join(project_root, ".cursor", "skills")

        results = []
        success_count = 0
        for skill_id in requested_ids:
            source_dir = SYNCABLE_SOURCEGRAPH_SKILLS[skill_id].rstrip("/")
            target_dir = os.path.join(skills_root, skill_id)
            try:
                normalized_root = os.path.normpath(skills_root)
                source_files = []
                list_mode = "mcp"
                try:
                    source_files = _list_sourcegraph_files_recursive(SOURCEGRAPH_SKILL_REPO, source_dir)
                except Exception as mcp_list_err:
                    logger.warning(f"[cursor-ai-sync-skills] MCP list failed for {skill_id}, falling back to HTTP tree crawl: {mcp_list_err}")
                    source_files = _list_sourcegraph_files_recursive_http(SOURCEGRAPH_SKILL_REPO, source_dir)
                    list_mode = "http"
                if not source_files:
                    raise RuntimeError(f"No files found in Sourcegraph directory: {source_dir}")

                synced_files = []
                for source_file in source_files:
                    cleaned_content = ""
                    if list_mode == "mcp":
                        read_resp = _call_mcp_tool(
                            "gw-sourcegraph",
                            "sourcegraph__read_file",
                            {"repo": SOURCEGRAPH_SKILL_REPO, "path": source_file},
                            timeout=45,
                        )
                        if read_resp.get("error"):
                            logger.warning(f"[cursor-ai-sync-skills] MCP read failed for {source_file}, falling back to HTTP raw: {read_resp['error']}")
                            raw_content = _read_sourcegraph_file_http(SOURCEGRAPH_SKILL_REPO, source_file)
                            cleaned_content = raw_content if raw_content.endswith("\n") else f"{raw_content}\n"
                        else:
                            raw_content = read_resp.get("content", "")
                            cleaned_content = _strip_sourcegraph_line_prefixes(raw_content)
                    else:
                        raw_content = _read_sourcegraph_file_http(SOURCEGRAPH_SKILL_REPO, source_file)
                        cleaned_content = raw_content if raw_content.endswith("\n") else f"{raw_content}\n"

                    if not cleaned_content.strip():
                        raise RuntimeError(f"{source_file}: fetched file content is empty")

                    relative_source_path = os.path.relpath(source_file, source_dir)
                    target_file = os.path.normpath(os.path.join(target_dir, relative_source_path))
                    if not target_file.startswith(normalized_root + os.sep):
                        raise RuntimeError("Unsafe destination path")

                    os.makedirs(os.path.dirname(target_file), exist_ok=True)
                    with open(target_file, "w") as f:
                        f.write(cleaned_content)
                    synced_files.append(target_file)

                success_count += 1
                results.append({
                    "skill_id": skill_id,
                    "source_path": source_dir,
                    "source_fetch_mode": list_mode,
                    "target_dir": os.path.normpath(target_dir),
                    "synced_files": synced_files,
                    "synced_file_count": len(synced_files),
                    "success": True,
                })
            except Exception as sync_err:
                logger.error(f"[cursor-ai-sync-skills] Failed syncing {skill_id}: {sync_err}")
                results.append({
                    "skill_id": skill_id,
                    "source_path": source_dir,
                    "target_dir": os.path.normpath(target_dir),
                    "success": False,
                    "error": str(sync_err),
                })

        return jsonify({
            "success": success_count == len(requested_ids),
            "results": results,
            "summary": {
                "requested_count": len(requested_ids),
                "success_count": success_count,
                "failed_count": len(requested_ids) - success_count,
            },
        })
    except Exception as e:
        logger.error(f"Error syncing skills: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/chat", methods=["POST"])
@app.route("/api/mcp/regression/cursor-ai/chat", methods=["POST"])
@jwt_required
def cursor_ai_chat():
    """Interactive Cursor AI chat endpoint supporting multiple modes and MCP context."""
    try:
        body = request.get_json(force=True) or {}
        messages = body.get("messages", [])
        mode = body.get("mode", "agent")
        model = body.get("model", "claude-sonnet-4.6-high")
        mcp_servers = body.get("mcp_servers", [])

        if not messages:
            return jsonify({"error": "messages are required"}), 400

        if mode not in MODE_SYSTEM_PROMPTS:
            return jsonify({"error": f"Invalid mode: {mode}. Use: agent, plan, debug, ask"}), 400

        system_prompt = MODE_SYSTEM_PROMPTS[mode]

        # Add MCP context to system prompt
        if mcp_servers:
            active_tools = []
            for sid in mcp_servers:
                cfg = MCP_SERVER_CONFIGS.get(sid)
                if cfg:
                    active_tools.append(f"- {sid}: {cfg['description']}")
            if active_tools:
                system_prompt += (
                    "\n\nYou have access to the following MCP tools/servers:\n"
                    + "\n".join(active_tools)
                    + "\n\nWhen answering, reference which tools you would use and provide specific, "
                    "data-driven insights where possible."
                )

        # Build chat messages in OpenAI-compatible format
        chat_messages = [{"role": "system", "content": system_prompt}]
        for msg in messages:
            role = msg.get("role", "user")
            if role in ("user", "assistant"):
                chat_messages.append({"role": role, "content": msg.get("content", "")})

        # Map model names
        ai_model = model
        if model in ("claude-sonnet-4.6-high", "claude-sonnet-4.6"):
            ai_model = "hack-reason"

        payload = {
            "model": ai_model,
            "messages": chat_messages,
            "max_tokens": 4096,
            "stream": False,
        }

        url = f"{AI_BASE}/chat/completions"
        headers = {
            "Authorization": f"Bearer {AI_API_KEY}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }

        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(url, data=data, headers=headers, method="POST")
        with urllib.request.urlopen(req, context=SSL_CTX, timeout=120) as resp:
            if resp.getcode() != 200:
                return jsonify({"error": f"AI API returned HTTP {resp.getcode()}"}), 502

            response_data = json.loads(resp.read().decode())
            choices = response_data.get("choices", [])
            if not choices:
                return jsonify({"error": "AI returned no choices"}), 502

            content = (choices[0].get("message") or {}).get("content", "")

            tools_used = []
            if mcp_servers:
                for sid in mcp_servers[:5]:
                    cfg = MCP_SERVER_CONFIGS.get(sid)
                    if cfg and cfg["description"].lower() in content.lower():
                        tools_used.append(sid)

            return jsonify({
                "success": True,
                "reply": content.strip(),
                "mode": mode,
                "model": model,
                "tools_used": tools_used,
            })

    except (urllib.error.HTTPError, urllib.error.URLError) as e:
        # Handle both HTTP errors (401, 403, 500, etc) and connection errors
        if isinstance(e, urllib.error.HTTPError):
            error_body = e.read().decode() if e.fp else ""
            logger.error(f"[cursor-ai-chat] HTTP error: {e.code} - {error_body[:500]}")
            logger.info(f"[cursor-ai-chat] Nutanix AI returned HTTP {e.code}, falling back to Cursor Bridge")
        else:
            logger.error(f"[cursor-ai-chat] Nutanix AI unreachable ({e.reason}), falling back to Cursor Bridge")
        
        # Fallback: route through the Cursor Bridge when the Nutanix AI endpoint fails
        try:
            last_user_msg = ""
            for msg in reversed(messages):
                if msg.get("role") == "user":
                    last_user_msg = msg.get("content", "")
                    break
            if not last_user_msg:
                return jsonify({"error": "Cannot reach AI API endpoint and no user message for fallback"}), 503

            bridge_resp = requests.post(
                f"{CURSOR_BRIDGE_URL}/analyze-testcase",
                json={
                    "testcase_name": "cursor-ai-chat-fallback",
                    "exception_summary": last_user_msg,
                    "exception": "",
                    "steps_log": "",
                    "nutest_test_log": "",
                    "test_log_url": "",
                    "jira_tickets": [],
                    "failure_stage": "chat",
                },
                timeout=600,
            )
            if bridge_resp.status_code == 200:
                bridge_data = bridge_resp.json()
                analysis = bridge_data.get("analysis", {})
                reply = analysis.get("follow_up_answer") or analysis.get("root_cause") or json.dumps(analysis, indent=2)
                return jsonify({
                    "success": True,
                    "reply": reply,
                    "mode": mode,
                    "model": "cursor-bridge-fallback",
                    "tools_used": [],
                })
            else:
                return jsonify({"error": "Cannot reach AI API endpoint and Cursor Bridge also failed"}), 503
        except Exception as fallback_err:
            logger.error(f"[cursor-ai-chat] Cursor Bridge fallback also failed: {fallback_err}")
            return jsonify({"error": "Cannot reach AI API endpoint. Cursor Bridge fallback also unavailable."}), 503
    except Exception as e:
        logger.error(f"Error in cursor-ai chat: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/mcp/regression/cursor-ai/mcp-servers", methods=["GET"])
@app.route("/api/mcp/regression/cursor-ai/mcp-servers", methods=["GET"])
@jwt_required
def cursor_ai_list_mcp_servers():
    """List available MCP servers and their status."""
    servers = []
    for sid, cfg in MCP_SERVER_CONFIGS.items():
        servers.append({
            "id": sid,
            "url": cfg["url"],
            "description": cfg["description"],
        })
    return jsonify({"success": True, "servers": servers})




# ======================================================
# Handover / Onboarding & Deprecation (ported feature)
# ======================================================
# ======================================================
# Handover / JITA Analysis - Parse URL
# ======================================================
def parse_jita_url(url):
    """Extract task_ids from JITA results URL or direct API URL."""
    if not url or not url.strip():
        return []
    from urllib.parse import urlparse, parse_qs
    url = url.strip()
    parsed = urlparse(url)
    if "/agave_tasks/" in parsed.path:
        task_id = parsed.path.rstrip("/").split("/")[-1]
        if task_id and len(task_id) >= 20:
            return [task_id]
    qs = parse_qs(parsed.query)
    task_ids_param = qs.get("task_ids", [])
    if not task_ids_param:
        return []
    raw = task_ids_param[0] if isinstance(task_ids_param[0], str) else ",".join(task_ids_param)
    return [tid.strip() for tid in raw.split(",") if tid.strip()]


# ======================================================
# Jira Helper Functions (for Handover)
# ======================================================
def _categorize_bug_type_from_issuetype(issuetype):
    """Categorize bug type from Jira issuetype name."""
    if not issuetype:
        return None
    issuetype_lower = issuetype.lower()
    if "test bug" in issuetype_lower or "testbed" in issuetype_lower:
        return "Test Bug"
    elif "environment" in issuetype_lower:
        return "Environment"
    elif "flaky" in issuetype_lower:
        return "Flaky"
    elif "product bug" in issuetype_lower or ("product" in issuetype_lower and "bug" in issuetype_lower):
        return "Product Bug"
    elif issuetype_lower == "bug" or "bug" in issuetype_lower:
        return "Product Bug"
    return None


def _fetch_ticket_issuetype(ticket):
    """Fetch Jira ticket issuetype using the app's Jira integration
    (JIRA_BASE + get_jira_headers() + shared session). Returns (issuetype, error)."""
    if not ticket or not ticket.strip():
        return None, None
    ticket = ticket.strip().upper()
    if "-" in ticket and ticket.split("-")[0].isalpha():
        ticket = ticket.split("-")[0].upper() + "-" + (ticket.split("-", 1)[1] or "")

    if not os.getenv("JIRA_TOKEN"):
        return None, "Jira not configured"

    url = "%s/issue/%s" % (JIRA_BASE, ticket)

    try:
        resp = session.get(url, headers=get_jira_headers(), timeout=10)
        if resp.status_code == 401:
            return None, "Authentication required"
        if resp.status_code == 404:
            return None, "Ticket not found"
        if resp.status_code == 403:
            return None, "Forbidden"
        resp.raise_for_status()
        data = resp.json()
        issuetype = (data.get("fields") or {}).get("issuetype", {})
        name = (issuetype.get("name") or "").strip()
        return name, None
    except requests.RequestException as e:
        err_msg = str(e) or "Failed to fetch Jira issue."
        if "resolve" in err_msg.lower() or "nodename" in err_msg.lower():
            err_msg = "Cannot reach Jira server. Check network or VPN."
        elif "Connection" in err_msg and ("refused" in err_msg.lower() or "timeout" in err_msg.lower()):
            err_msg = "Cannot connect to Jira. Check network/VPN or try again later."
        return None, err_msg
    except Exception:
        return None, "Error"


def _get_task_id_from_run(run):
    """Extract task_id string from a test run's agave_task_id."""
    aid = run.get("agave_task_id")
    if not aid:
        return None
    if isinstance(aid, dict) and "$oid" in aid:
        return aid["$oid"]
    return str(aid)


def _aggregate_jita_test_cases(test_data, min_passes_for_success=1, auto_categorize_bug_types=False, latest_2_task_ids=None):
    """Aggregate test results by test_name. Returns (test_cases_list, total_executions, total_passed, all_tests_passed, summary, tickets_set).
    If latest_2_task_ids is provided, only consider runs from those 2 most recent tasks (consecutive latest runs).
    A test passes only if it passed in BOTH of the 2 latest runs (consecutive passes)."""
    by_name = defaultdict(list)
    tickets_set = set()
    tickets_by_test = defaultdict(set)
    for test in test_data:
        test_name = test.get("test", {}).get("name", "")
        if not test_name:
            continue
        # Filter to only 2 latest runs if specified
        if latest_2_task_ids:
            tid = _get_task_id_from_run(test)
            if tid not in latest_2_task_ids:
                continue
        by_name[test_name].append(test)
        test_tickets = test.get("jira_tickets") or []
        for t in test_tickets:
            if t and t.strip():
                tickets_set.add(t.strip())
                tickets_by_test[test_name].add(t.strip())

    # When using latest_2_task_ids, total_executions/total_passed reflect only those 2 runs
    filtered_data = test_data
    if latest_2_task_ids:
        filtered_data = [t for t in test_data if _get_task_id_from_run(t) in latest_2_task_ids]
    total_executions = len(filtered_data)
    total_passed = sum(1 for t in filtered_data if t.get("status") == "Succeeded")
    test_cases = []
    summary = {"total": 0, "succeeded": 0, "failed": 0, "warning": 0, "pending": 0, "running": 0, "skipped": 0}
    all_succeeded = True

    for test_name, runs in by_name.items():
        total_count = len(runs)
        passed_count = sum(1 for r in runs if r.get("status") == "Succeeded")
        # When using latest_2_task_ids: require BOTH runs to pass (consecutive). Otherwise use min_passes_for_success.
        required_passes = 2 if latest_2_task_ids else min_passes_for_success
        derived_status = "Succeeded" if passed_count >= required_passes else "Failed"
        if derived_status != "Succeeded":
            all_succeeded = False
        jira_tickets = []
        seen = set()
        for r in runs:
            raw_tickets = r.get("jira_tickets") or []
            for t in raw_tickets:
                if t and t.strip():
                    ticket_clean = t.strip()
                    if ticket_clean not in seen:
                        seen.add(ticket_clean)
                        jira_tickets.append(ticket_clean)

        bug_types = set()
        if auto_categorize_bug_types and jira_tickets and os.getenv("JIRA_TOKEN"):
            for ticket in jira_tickets:
                if not ticket or not ticket.strip():
                    continue
                ticket_clean = ticket.strip().upper()
                if "-" in ticket_clean:
                    parts = ticket_clean.split("-", 1)
                    if len(parts) == 2 and parts[0].isalpha() and parts[1]:
                        ticket_clean = f"{parts[0].upper()}-{parts[1]}"
                    else:
                        continue
                else:
                    continue

                issuetype, error = _fetch_ticket_issuetype(ticket_clean)
                if issuetype:
                    bug_type = _categorize_bug_type_from_issuetype(issuetype)
                    if bug_type:
                        bug_types.add(bug_type)
                if len(jira_tickets) > 1:
                    time.sleep(0.1)

        bug_type_str = ", ".join(sorted(bug_types)) if bug_types else None

        exception_summary = None
        test_log_url = None
        failure_analysis = None
        for r in runs:
            if r.get("exception_summary") and not exception_summary:
                exception_summary = r.get("exception_summary")
            if r.get("test_log_url") and not test_log_url:
                test_log_url = r.get("test_log_url")
            if r.get("failure_analysis") and not failure_analysis:
                failure_analysis = r.get("failure_analysis")
            if exception_summary and test_log_url and failure_analysis:
                break

        test_case_obj = {
            "test_name": test_name,
            "status": derived_status,
            "total_count": total_count,
            "passed_count": passed_count,
            "jira_tickets": jira_tickets,
            "exception_summary": exception_summary,
            "test_log_url": test_log_url,
            "failure_analysis": failure_analysis,
            "bug_type": bug_type_str,
        }
        test_cases.append(test_case_obj)
        summary["total"] += 1
        if derived_status == "Succeeded":
            summary["succeeded"] += 1
        else:
            summary["failed"] += 1

    all_tests_passed = all_succeeded and (summary["failed"] == 0)
    return test_cases, total_executions, total_passed, all_tests_passed, summary, list(tickets_set)


@app.route("/mcp/regression/jita-analysis", methods=["GET", "POST"])
@jwt_required
def jita_analysis():
    """Fetch by tag (same as Triage), or by JITA URL(s) / task_ids. Returns test cases with aggregated pass/fail."""
    import re
    if request.method == "OPTIONS":
        return jsonify({}), 200

    start = time.time()
    url = ""
    urls_param = []
    task_ids_param = ""
    tag_param = ""
    input_param = ""
    min_passes = 1

    if request.method == "POST" and request.is_json:
        data = request.get_json() or {}
        url = (data.get("url") or "").strip()
        urls_param = data.get("urls") or []
        if isinstance(urls_param, str):
            urls_param = [s.strip() for s in urls_param.split("\n") if s.strip()]
        else:
            urls_param = [(u or "").strip() for u in urls_param if (u or "").strip()]
        task_ids_param = (data.get("task_ids") or "").strip()
        tag_param = (data.get("tag") or "").strip()
        input_param = (data.get("input") or "").strip()
        min_passes = int(data.get("min_passes_for_success") or 1)
    else:
        url = request.args.get("url", "").strip()
        urls_param = request.args.getlist("urls") or []
        urls_param = [u.strip() for u in urls_param if u.strip()]
        task_ids_param = request.args.get("task_ids", "").strip()
        tag_param = request.args.get("tag", "").strip()
        input_param = request.args.get("input", "").strip()
        try:
            min_passes = int(request.args.get("min_passes_for_success") or 1)
        except Exception:
            min_passes = 1

    task_ids = []
    if tag_param:
        logger.info(f"[START] JITA Analysis (by tag) | tag={tag_param}")
        try:
            tasks = fetch_regression_tasks(tag_param)
            task_ids = [t["_id"]["$oid"] for t in tasks if t.get("_id", {}).get("$oid")]
            logger.info(f"JITA Analysis | tag yielded {len(task_ids)} task_ids")
        except Exception as e:
            logger.error(f"JITA Analysis by tag failed: {e}")
            return jsonify({"error": str(e)}), 500

    if not task_ids and urls_param:
        for u in urls_param:
            if u.startswith("http://") or u.startswith("https://"):
                task_ids.extend(parse_jita_url(u))
            else:
                potential_ids = re.split(r'[,\s\n]+', u)
                task_ids.extend([tid.strip() for tid in potential_ids if tid.strip() and re.match(r'^[a-f0-9]{20,}$', tid.strip())])
        task_ids = list(dict.fromkeys(task_ids))

    if not task_ids and input_param:
        if input_param.startswith("http://") or input_param.startswith("https://"):
            task_ids = parse_jita_url(input_param)
        else:
            potential_ids = re.split(r'[,\s\n]+', input_param)
            task_ids = [tid.strip() for tid in potential_ids if tid.strip() and re.match(r'^[a-f0-9]{20,}$', tid.strip())]

    if not task_ids and url:
        task_ids = parse_jita_url(url)

    if not task_ids and task_ids_param:
        task_ids = [tid.strip() for tid in task_ids_param.split(",") if tid.strip()]

    if not task_ids:
        return jsonify({
            "error": "Provide JITA URL(s) (e.g. https://jita.../results?task_ids=...) or task ID(s) (24-char hex, comma/space separated)"
        }), 400

    logger.info(f"[START] JITA Analysis | task_ids={len(task_ids)} | min_passes={min_passes}")
    try:
        try:
            test_data = fetch_test_results_batch_with_pagination(task_ids, timeout=180, merge=False)
            logger.info(f"JITA Analysis | fetched {len(test_data)} test results")
        except Exception as fetch_err:
            err_msg = str(fetch_err)
            if "timeout" in err_msg.lower() or "connection" in err_msg.lower() or "network" in err_msg.lower() or "resolve" in err_msg.lower():
                return jsonify({
                    "error": f"Network error while fetching JITA data: {err_msg}. Please check your VPN connection and ensure JITA server is accessible.",
                    "task_ids": task_ids,
                    "generated_at": datetime.utcnow().isoformat()
                }), 500
            raise

        task_metadata = []
        task_id_to_ts = {}
        for tid in task_ids:
            try:
                agave = fetch_agave_task(tid)
                tfm = agave.get("test_framework_metadata", {}) or {}
                test_branch = tfm.get("test", {}).get("branch") if tfm.get("test") else None
                framework_branch = tfm.get("framework", {}).get("branch") if tfm.get("framework") else None
                raw_ts = agave.get("updated_at") or agave.get("created_at") or ""
                # Normalize to comparable string: API may return dict e.g. {"$date": "..."} or nested
                if isinstance(raw_ts, dict):
                    val = raw_ts.get("$date") or raw_ts.get("$numberLong") or ""
                    while isinstance(val, dict):
                        val = val.get("$date") or val.get("$numberLong") or ""
                    updated_at = str(val) if val and not isinstance(val, dict) else ""
                else:
                    updated_at = str(raw_ts) if raw_ts else ""
                task_id_to_ts[tid] = str(updated_at)  # ensure always string
                if len(task_metadata) < 20:
                    task_metadata.append({
                        "task_id": tid,
                        "status": agave.get("status"),
                        "label": agave.get("label"),
                        "branch": test_branch or framework_branch,
                        "test_result_count": agave.get("test_result_count", {}),
                        "emails": agave.get("emails", []),
                        "test_framework_metadata": tfm,
                        "container_details": agave.get("container_details"),
                        "updated_at": updated_at,
                    })
            except Exception as e:
                logger.warning(f"Could not fetch agave_tasks for {tid}: {e}")
                task_id_to_ts[tid] = ""

        # Sort tasks by date (newest first), take 2 latest for "consecutive passes" logic
        def _ts_key(tid):
            v = task_id_to_ts.get(tid, "")
            return str(v) if not isinstance(v, str) else v  # always comparable string
        sorted_task_ids = sorted(task_ids, key=_ts_key, reverse=True)
        latest_2_task_ids = set(sorted_task_ids[:2]) if len(sorted_task_ids) >= 2 else None
        if latest_2_task_ids:
            logger.info(f"JITA Analysis | Using 2 latest runs only (consecutive): {list(latest_2_task_ids)}")

        test_cases, total_executions, total_passed, all_tests_passed, summary, tickets_set = _aggregate_jita_test_cases(
            test_data, min_passes_for_success=min_passes, auto_categorize_bug_types=True, latest_2_task_ids=latest_2_task_ids
        )
        logger.info(f"[END] JITA Analysis | all_passed={all_tests_passed} | time={time.time() - start:.2f}s")
        return jsonify({
            "task_ids": task_ids,
            "tag": tag_param or None,
            "jita_url": url or None,
            "urls": urls_param if urls_param else None,
            "task_metadata": task_metadata,
            "all_tests_passed": all_tests_passed,
            "summary": summary,
            "test_cases": test_cases,
            "assigned_tickets": tickets_set,
            "total_executions": total_executions,
            "total_passed": total_passed,
            "min_passes_for_success": min_passes,
            "generated_at": datetime.utcnow().isoformat(),
        })
    except Exception as e:
        logger.error(f"Error in JITA analysis: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return jsonify({"error": str(e)}), 500


# ======================================================
# Handover Records Support
# ======================================================
# Use abspath so the path is correct regardless of cwd when backend is started
_BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
_PROJECT_ROOT = os.path.dirname(_BACKEND_DIR)
HANDOVER_RECORDS_PATH = (os.getenv("HANDOVER_RECORDS_PATH") or "").strip() or os.path.join(_PROJECT_ROOT, "data", "handover_records.json")

def _load_handover_records():
    """Load handover records from JSON file"""
    try:
        parent = os.path.dirname(HANDOVER_RECORDS_PATH)
        if not os.path.exists(parent):
            os.makedirs(parent, exist_ok=True)
        if os.path.exists(HANDOVER_RECORDS_PATH):
            with open(HANDOVER_RECORDS_PATH, "r") as f:
                data = json.load(f)
                return data.get("records", [])
    except Exception as e:
        logger.warning(f"Could not load handover records: {e}")
    return []


def _save_handover_records(records):
    """Save handover records to JSON file"""
    try:
        parent = os.path.dirname(HANDOVER_RECORDS_PATH)
        os.makedirs(parent, exist_ok=True)
        with open(HANDOVER_RECORDS_PATH, "w") as f:
            json.dump({"records": records, "updated_at": datetime.utcnow().isoformat()}, f, indent=2)
    except Exception as e:
        logger.error(f"Could not save handover records: {e}")


def _add_handover_records(test_names, tickets, by_whom, branch="", lst_file="", test_bug_types=None, test_bug_tickets=None, handover_tickets=None):
    """Add handover records for test names"""
    records = _load_handover_records()
    now = datetime.utcnow().isoformat()
    ticket_list = tickets if isinstance(tickets, list) else []
    test_bug_types = test_bug_types if isinstance(test_bug_types, dict) else {}
    test_bug_tickets = test_bug_tickets if isinstance(test_bug_tickets, dict) else {}
    handover_tickets_list = handover_tickets if isinstance(handover_tickets, list) else []
    for test_name in test_names:
        if not (test_name and test_name.strip()):
            continue
        test_name = test_name.strip()
        tix = ticket_list
        if isinstance(tickets, dict) and test_name in tickets:
            tix = tickets[test_name] if isinstance(tickets[test_name], list) else [tickets[test_name]]
        bug_tix = []
        if isinstance(test_bug_tickets, dict) and test_name in test_bug_tickets:
            bug_tix = test_bug_tickets[test_name] if isinstance(test_bug_tickets[test_name], list) else [test_bug_tickets[test_name]]
        handover_tix = handover_tickets_list.copy()
        rec = {
            "test_name": test_name,
            "tickets": tix if isinstance(tix, list) else [tix] if tix else [],
            "bug_tickets": bug_tix,
            "handover_tickets": handover_tix,
            "handover_date": now,
            "by_whom": by_whom or "unknown",
            "branch": branch,
            "lst_file": lst_file,
        }
        if test_name in test_bug_types and test_bug_types[test_name]:
            rec["bug_type"] = test_bug_types[test_name]
        records.append(rec)
    _save_handover_records(records)
    logger.info(f"[HANDOVER-RECORD] Saved {len(test_names)} record(s), by {by_whom}")


def _delete_handover_record(test_name, handover_date, lst_file):
    """Delete a handover record"""
    records = _load_handover_records()
    lst = (lst_file or "").strip()
    date_str = (handover_date or "").strip()
    name = (test_name or "").strip()
    for i, r in enumerate(records):
        if (r.get("test_name") or "").strip() == name and (r.get("handover_date") or "").strip() == date_str and (r.get("lst_file") or "").strip() == lst:
            records.pop(i)
            _save_handover_records(records)
            logger.info(f"[HANDOVER-RECORD] Deleted record: {name} @ {date_str}")
            return True
    return False


@app.route("/mcp/regression/handover-record", methods=["POST"])
@jwt_required
def handover_record():
    """Create a handover record"""
    data = request.get_json() or {}
    test_names = data.get("test_names", [])
    tickets = data.get("tickets", [])
    test_tickets = data.get("test_tickets", {})
    test_bug_types = data.get("test_bug_types") or {}
    test_bug_tickets = data.get("test_bug_tickets", {})
    handover_tickets = data.get("handover_tickets", [])
    by_whom = (data.get("by_whom") or data.get("user_email") or data.get("user_name") or "").strip()
    branch = (data.get("branch") or "").strip()
    lst_file = (data.get("lst_file") or "").strip()
    if not test_names:
        return jsonify({"error": "test_names is required"}), 400
    _add_handover_records(test_names, test_tickets if test_tickets else tickets, by_whom or "unknown", branch, lst_file, test_bug_types=test_bug_types, test_bug_tickets=test_bug_tickets, handover_tickets=handover_tickets)
    return jsonify({"success": True, "message": "Recorded handover for %s test(s)." % len(test_names), "generated_at": datetime.utcnow().isoformat()})


@app.route("/mcp/regression/handover-record-delete", methods=["POST"])
@jwt_required
def handover_record_delete():
    """Delete a handover record"""
    data = request.get_json() or {}
    test_name = (data.get("test_name") or "").strip()
    handover_date = (data.get("handover_date") or "").strip()
    lst_file = (data.get("lst_file") or "").strip()
    if not test_name or not handover_date:
        return jsonify({"error": "test_name and handover_date are required"}), 400
    removed = _delete_handover_record(test_name, handover_date, lst_file)
    return jsonify({"success": removed, "message": "Record deleted." if removed else "No matching record found.", "generated_at": datetime.utcnow().isoformat()})


@app.route("/mcp/regression/validate-jira-ticket", methods=["GET", "POST"])
@jwt_required
def validate_jira_ticket():
    """Validate if a Jira ticket exists and return its issuetype."""
    if request.method == "OPTIONS":
        return jsonify({}), 200

    ticket = ""
    if request.method == "POST" and request.is_json:
        ticket = (request.get_json() or {}).get("ticket", "").strip().upper()
    else:
        ticket = (request.args.get("ticket") or "").strip().upper()

    if not ticket:
        return jsonify({"valid": False, "ticket": "", "issuetype": None, "error": "ticket is required"}), 400

    if "-" in ticket and ticket.split("-")[0].isalpha():
        ticket = ticket.split("-")[0].upper() + "-" + (ticket.split("-", 1)[1] or "")

    if not os.getenv("JIRA_TOKEN"):
        return jsonify({
            "valid": None, "ticket": ticket, "issuetype": None, "skipped": True,
            "message": "Jira validation is not configured. Set the JIRA_TOKEN environment variable (same one the dashboard uses) and restart the backend."
        }), 200

    url = "%s/issue/%s" % (JIRA_BASE, ticket)

    try:
        resp = session.get(url, headers=get_jira_headers(), timeout=10)
        if resp.status_code == 401:
            return jsonify({
                "valid": False, "ticket": ticket, "issuetype": None,
                "error": "Authentication required. Check the JIRA_TOKEN environment variable."
            }), 200
        if resp.status_code == 404:
            return jsonify({"valid": False, "ticket": ticket, "issuetype": None, "error": "Ticket not found"}), 200
        if resp.status_code == 403:
            return jsonify({"valid": False, "ticket": ticket, "issuetype": None, "error": "Forbidden"}), 200
        resp.raise_for_status()
        data = resp.json()
        issuetype = (data.get("fields") or {}).get("issuetype", {})
        name = (issuetype.get("name") or "").strip()
        bug_type = _categorize_bug_type_from_issuetype(name)
        is_valid = bug_type == "Product Bug"
        return jsonify({
            "valid": is_valid, "ticket": ticket, "issuetype": name, "bug_type": bug_type,
            "generated_at": datetime.utcnow().isoformat()
        })
    except requests.RequestException as e:
        err_msg = str(e)
        if "resolve" in err_msg.lower() or "nodename" in err_msg.lower():
            return jsonify({
                "valid": False, "ticket": ticket, "issuetype": None,
                "error": "Cannot reach Jira server. Check network or VPN."
            }), 200
        return jsonify({"valid": False, "ticket": ticket, "issuetype": None, "error": err_msg}), 200
    except Exception as e:
        return jsonify({"valid": False, "ticket": ticket, "issuetype": None, "error": str(e)}), 200


# ======================================================
# Handover - validate-lst, create-lst-cr, check-lst-testcases, search-lst-file, deprecate-lst-cr, deprecation-search
# ======================================================
def validate_with_sourcegraph(repo_name, branch, file_path):
    """Validate branch and file exist in repo via Sourcegraph."""
    results = {"branch_valid": None, "file_valid": None, "branch_error": None, "file_error": None, "file_suggestions": [], "auth_required": False}
    sg_token = (os.getenv("SOURCEGRAPH_TOKEN", "") or "").strip()
    sg_url = (os.getenv("SOURCEGRAPH_URL") or "https://sourcegraph.ntnxdpro.com").strip().rstrip("/")
    sg_graphql_path = (os.getenv("SOURCEGRAPH_GRAPHQL_PATH") or "/api/graphql").strip() or "/api/graphql"
    if not sg_token:
        results["auth_required"] = True
        return results
    base_url = sg_url
    if not base_url:
        results["branch_error"] = "SOURCEGRAPH_URL not set"
        return results
    api_urls_to_try = []
    sg_graphql_url = (os.getenv("SOURCEGRAPH_GRAPHQL_URL") or "").strip().rstrip("/")
    if sg_graphql_url:
        api_urls_to_try.append(sg_graphql_url)
    base = (base_url or "").strip().rstrip("/")
    graphql_path = sg_graphql_path
    if not graphql_path.startswith("/"):
        graphql_path = "/" + graphql_path
    if base:
        api_url = base + graphql_path
        if api_url not in api_urls_to_try:
            api_urls_to_try.append(api_url)
    if base:
        for p in ["/api/graphql", "/.api/graphql", "/graphql"]:
            full = base + p
            if full not in api_urls_to_try:
                api_urls_to_try.append(full)
    auth_headers = [
        {"Content-Type": "application/json", "Authorization": "token %s" % sg_token},
        {"Content-Type": "application/json", "Authorization": "Bearer %s" % sg_token},
    ]
    def run_search(search_query, url=None):
        urls = [url] if url else api_urls_to_try
        for u in urls:
            if not u:
                continue
            for headers in auth_headers:
                try:
                    resp = requests.post(u, json={"query": "query Search($q: String!) { search(query: $q) { results { matchCount resultCount results { ... on FileMatch { file { path } } } } } }", "variables": {"q": search_query}}, headers=headers, timeout=20, verify=False)
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("errors"):
                            continue
                        sr = data.get("data", {}).get("search", {}).get("results")
                        if sr is not None:
                            match_count = (sr or {}).get("matchCount") or (sr or {}).get("resultCount") or len((sr or {}).get("results", []))
                            return (match_count or 0), None
                except Exception:
                    continue
        return None, "Sourcegraph not configured or unreachable"
    try:
        branch_query = "repo:%s rev:%s count:1" % (repo_name, branch)
        match_count, err = run_search(branch_query)
        if err:
            results["branch_valid"] = None
            results["branch_error"] = err
        else:
            results["branch_valid"] = match_count > 0
            if not results["branch_valid"]:
                results["branch_error"] = "Branch '%s' not found" % branch
    except Exception as e:
        results["branch_valid"] = None
        results["branch_error"] = str(e)
    try:
        file_query = "repo:%s rev:%s file:%s type:path count:10" % (repo_name, branch, file_path)
        match_count, err = run_search(file_query)
        if err:
            results["file_valid"] = None
            results["file_error"] = err
        else:
            results["file_valid"] = match_count > 0
            if not results["file_valid"]:
                results["file_error"] = "File '%s' not found" % file_path
    except Exception as e:
        results["file_valid"] = None
        results["file_error"] = str(e)
    return results


def fetch_file_content_via_sourcegraph(repo_name, rev, file_path):
    """Fetch file content from Sourcegraph. Returns (content, None) or (None, error_message)."""
    sg_token = (os.getenv("SOURCEGRAPH_TOKEN", "") or "").strip()
    if not sg_token:
        return None, "SOURCEGRAPH_TOKEN environment variable not set. Set it on the backend or paste LST file content below."
    base_url = (os.getenv("SOURCEGRAPH_URL") or "https://sourcegraph.ntnxdpro.com").strip().rstrip("/")
    if not base_url:
        return None, "SOURCEGRAPH_URL not set"
    api_urls = [base_url + p for p in ["/api/graphql", "/.api/graphql", "/graphql"]]
    auth_headers = [
        {"Content-Type": "application/json", "Authorization": "token %s" % sg_token},
        {"Content-Type": "application/json", "Authorization": "Bearer %s" % sg_token},
    ]
    for query_name, query in [
        ("blob", "query RepoFile($repo: String!, $rev: String!, $path: String!) { repository(name: $repo) { commit(rev: $rev) { blob(path: $path) { content } } } }"),
        ("file", "query RepoFile($repo: String!, $rev: String!, $path: String!) { repository(name: $repo) { commit(rev: $rev) { file(path: $path) { content } } } }"),
    ]:
        payload = {"query": query, "variables": {"repo": repo_name, "rev": rev, "path": file_path}}
        for api_url in api_urls:
            for headers in auth_headers:
                try:
                    resp = requests.post(api_url, json=payload, headers=headers, timeout=20, verify=False)
                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("errors"):
                            continue
                        repo_data = (data.get("data") or {}).get("repository") or {}
                        commit_data = repo_data.get("commit") or {}
                        blob_or_file = commit_data.get("blob") or commit_data.get("file")
                        if blob_or_file is not None and "content" in blob_or_file:
                            return (blob_or_file.get("content") or ""), None
                except Exception:
                    continue
    return None, "Could not fetch file from Sourcegraph. Set the SOURCEGRAPH_TOKEN environment variable or paste LST content below."


def _check_testnames_in_lst_content(lst_content, test_names):
    """Check which test names are present in LST file content. Returns (present_list, not_present_list)."""
    if not lst_content or not isinstance(lst_content, str):
        return [], test_names
    existing = set()
    lines = lst_content.splitlines()
    for test_name in test_names:
        test_name = test_name.strip()
        if not test_name:
            continue
        for line in lines:
            if test_name in line:
                escaped_name = re.escape(test_name)
                pattern = r'(^|[\s,\[\]"\'])' + escaped_name + r'([\s,\[\]"\']|$)'
                if re.search(pattern, line):
                    existing.add(test_name)
    already_present = [t for t in test_names if t.strip() in existing]
    not_present = [t for t in test_names if t.strip() not in existing]
    return already_present, not_present


@app.route("/mcp/regression/validate-lst", methods=["POST"])
@jwt_required
def validate_lst():
    """Validate LST file path and branch via Sourcegraph."""
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json() or {}
    branch = (data.get("branch") or "master").strip()
    lst_file = (data.get("lst_file") or "").strip()
    repo_name = (data.get("repo_name") or os.getenv("SOURCEGRAPH_FIRST_REPO", "nugerrit.ntnxdpro.com/nutest-py3-tests")).strip()
    if not lst_file:
        return jsonify({"error": "lst_file is required", "branch_valid": None, "file_valid": None}), 400
    validation_results = validate_with_sourcegraph(repo_name, branch, lst_file)
    sg_url = (os.getenv("SOURCEGRAPH_URL") or "https://sourcegraph.ntnxdpro.com").strip().rstrip("/")
    sourcegraph_url = f"{sg_url}/{repo_name}@{branch}/-/blob/{lst_file.replace(' ', '+')}" if sg_url else None
    return jsonify({
        "branch": branch, "lst_file": lst_file, "repo_name": repo_name,
        "branch_valid": validation_results.get("branch_valid"),
        "file_valid": validation_results.get("file_valid"),
        "branch_error": validation_results.get("branch_error"),
        "file_error": validation_results.get("file_error"),
        "file_suggestions": validation_results.get("file_suggestions", []),
        "sourcegraph_url": sourcegraph_url,
        "message": validation_results.get("message", ""),
        "generated_at": datetime.utcnow().isoformat()
    })


@app.route("/mcp/regression/search-reviewers", methods=["GET"])
@jwt_required
def search_reviewers():
    """Search for reviewers by name (uses Gerrit accounts suggest API). Returns name, email, username."""
    q = (request.args.get("q") or "").strip()
    if not q or len(q) < 2:
        return jsonify({"results": [], "message": "Type at least 2 characters to search"})
    gerrit_url = (os.getenv("GERRIT_URL") or "https://nugerrit.ntnxdpro.com").strip().rstrip("/")
    gerrit_auth = (os.getenv("GERRIT_TOKEN") or "").strip()
    if not gerrit_auth:
        username = (os.getenv("GERRIT_USERNAME") or "").strip()
        password = (os.getenv("GERRIT_HTTP_PASSWORD") or "").strip()
        if username and password:
            gerrit_auth = "%s:%s" % (username, password)
    if not gerrit_auth:
        gerrit_auth = (os.getenv("GERRIT_HTTP_PASSWORD") or "").strip()
    prefix = "/a" if gerrit_auth else ""
    url = "%s%s/accounts/?suggest&q=%s&n=15" % (gerrit_url, prefix, requests.utils.quote(q))
    headers = {"Accept": "application/json"}
    if gerrit_auth:
        import base64
        if ":" in gerrit_auth:
            user, pw = gerrit_auth.split(":", 1)
        else:
            user, pw = "anonymous", gerrit_auth
        headers["Authorization"] = "Basic " + base64.b64encode(("%s:%s" % (user, pw)).encode()).decode()
    try:
        resp = requests.get(url, headers=headers, timeout=10, verify=False)
        if resp.status_code == 401:
            return jsonify({"results": [], "error": "Gerrit reviewer search requires authentication. Set GERRIT_TOKEN (username:http_password) or GERRIT_USERNAME + GERRIT_HTTP_PASSWORD environment variables. You can still add reviewers by typing an email and pressing Enter."})
        if resp.status_code != 200:
            return jsonify({"results": [], "error": "Could not fetch reviewers from Gerrit (status %s). Add reviewers manually by typing email and pressing Enter." % resp.status_code})
        text = resp.text
        if text.startswith(")]}'"):
            text = text[5:]
        data = json.loads(text)
        accounts = data if isinstance(data, list) else (data.get("accounts") if isinstance(data, dict) else [])
        if not isinstance(accounts, list):
            accounts = []
        results = []
        for acc in accounts:
            if not isinstance(acc, dict):
                continue
            name = acc.get("name") or acc.get("display_name") or ""
            email = acc.get("email") or acc.get("preferred_email") or ""
            if not email and isinstance(acc.get("emails"), list):
                for e in acc["emails"]:
                    if isinstance(e, dict) and e.get("preferred"):
                        email = e.get("email") or e.get("value") or ""
                        break
                if not email and acc["emails"]:
                    email = acc["emails"][0].get("email") or acc["emails"][0].get("value") or "" if isinstance(acc["emails"][0], dict) else ""
            username = acc.get("username") or ""
            if email or name or username:
                results.append({"name": name, "email": email, "username": username})
        return jsonify({"results": results, "generated_at": datetime.utcnow().isoformat()})
    except requests.exceptions.Timeout:
        return jsonify({"results": [], "error": "Gerrit request timed out. Add reviewers manually by typing email and pressing Enter."})
    except requests.exceptions.ConnectionError as e:
        err = str(e)
        if "resolve" in err.lower() or "nodename" in err.lower():
            return jsonify({"results": [], "error": "Cannot reach Gerrit (check VPN/network). Add reviewers manually by typing email and pressing Enter."})
        return jsonify({"results": [], "error": "Cannot connect to Gerrit. Add reviewers manually by typing email and pressing Enter."})
    except Exception as e:
        logger.warning("search_reviewers failed: %s", e)
        return jsonify({"results": [], "error": "Could not fetch reviewers: %s. Add reviewers manually by typing email and pressing Enter." % str(e)})


@app.route("/mcp/regression/gerrit-connectivity", methods=["GET"])
@jwt_required
def gerrit_connectivity():
    """Check if we can reach Gerrit and authenticate (for CR creation)."""
    if request.method == "OPTIONS":
        return jsonify({}), 200
    gerrit_url = (os.getenv("GERRIT_URL") or "https://nugerrit.ntnxdpro.com").strip().rstrip("/")
    auth = _get_gerrit_auth()
    if not auth:
        return jsonify({
            "ok": False,
            "message": "Gerrit credentials not configured",
            "details": "Set GERRIT_USERNAME and GERRIT_HTTP_PASSWORD environment variables (only needed for reviewer search)",
        })
    url = "%s/a/accounts/self" % gerrit_url
    headers = {"Accept": "application/json"}
    import base64
    user, pw = auth
    headers["Authorization"] = "Basic " + base64.b64encode(("%s:%s" % (user, pw)).encode()).decode()
    try:
        resp = requests.get(url, headers=headers, timeout=10, verify=False)
        if resp.status_code == 200:
            text = resp.text
            if text.startswith(")]}'"):
                text = text[5:]
            data = json.loads(text) if text else {}
            username = data.get("username") or data.get("name") or user
            return jsonify({
                "ok": True,
                "message": "Connected to Gerrit",
                "details": "Authenticated as %s. Ready to create CR." % username,
            })
        if resp.status_code == 401:
            return jsonify({
                "ok": False,
                "message": "Gerrit authentication failed",
                "details": "Invalid credentials. Check GERRIT_USERNAME and GERRIT_HTTP_PASSWORD.",
            })
        return jsonify({
            "ok": False,
            "message": "Gerrit returned status %s" % resp.status_code,
            "details": (resp.text or "")[:200],
        })
    except requests.exceptions.Timeout:
        return jsonify({
            "ok": False,
            "message": "Gerrit request timed out",
            "details": "Check VPN/network and try again.",
        })
    except requests.exceptions.ConnectionError as e:
        err = str(e)
        if "resolve" in err.lower() or "nodename" in err.lower():
            return jsonify({
                "ok": False,
                "message": "Cannot reach Gerrit",
                "details": "Check VPN/network. DNS resolution failed.",
            })
        return jsonify({
            "ok": False,
            "message": "Gerrit connection failed",
            "details": str(e)[:200],
        })
    except Exception as e:
        logger.exception("gerrit_connectivity failed")
        return jsonify({
            "ok": False,
            "message": "Unexpected error",
            "details": str(e)[:200],
        })


def _build_gerrit_push_ref(branch, reviewers):
    """Build Gerrit push ref with optional reviewers (e.g. refs/for/master%r=user@x.com)."""
    ref = "refs/for/%s" % (branch or "master")
    if reviewers and isinstance(reviewers, list):
        reviewers = [r.strip() for r in reviewers if r and str(r).strip()]
        if reviewers:
            ref += "%" + ",".join("r=" + r for r in reviewers)
    return ref


def _get_gerrit_auth():
    """Get Gerrit auth as (username, password) or None if not configured."""
    gerrit_token = (os.getenv("GERRIT_TOKEN") or "").strip()
    if gerrit_token and ":" in gerrit_token:
        parts = gerrit_token.split(":", 1)
        return (parts[0], parts[1])
    username = (os.getenv("GERRIT_USERNAME") or "").strip()
    password = (os.getenv("GERRIT_HTTP_PASSWORD") or "").strip()
    if username and password:
        return (username, password)
    return None


@app.route("/mcp/regression/create-lst-cr", methods=["POST"])
@jwt_required
def create_lst_cr():
    """Return the manual git steps to add tests to an LST file and push for review.

    Manual-only by design: the backend never pushes to Gerrit itself, so no Gerrit
    credentials are needed on the server. The reviewers are baked into the push ref.
    """
    if request.method == "OPTIONS":
        return jsonify({}), 200
    logger.info("[CREATE-LST-CR] POST received")
    data = request.get_json() or {}
    branch = (data.get("branch") or "master").strip()
    lst_file = (data.get("lst_file") or "").strip()
    test_names = data.get("test_names", [])
    logger.info("[CREATE-LST-CR] Received: branch=%r, lst_file=%r, test_names_count=%s, reviewers=%s",
                branch, lst_file, len(test_names) if isinstance(test_names, list) else "?", data.get("reviewers"))
    reviewers = data.get("reviewers") or []
    if not test_names:
        return jsonify({"error": "test_names is required"}), 400
    if not lst_file:
        return jsonify({"error": "lst_file is required"}), 400
    test_names = [str(t).strip() for t in test_names if (t or "").strip()]
    if not test_names:
        return jsonify({"error": "test_names is required"}), 400
    push_ref = _build_gerrit_push_ref(branch, reviewers)
    instructions = {
        "message": "Add the following %s test(s) to the LST file, then push for review." % len(test_names),
        "branch": branch, "lst_file": lst_file, "test_names": test_names,
        "manual_steps": [
            "1. Clone the repository and checkout branch '%s'" % branch,
            "2. Open the LST file: %s" % lst_file,
            "3. Add the following %s test name(s) to the file:" % len(test_names),
            "   " + "\n   ".join(test_names[:10]) + ("..." if len(test_names) > 10 else ""),
            "4. Commit with message: 'Add %s test(s) to LST'" % len(test_names),
            "5. Push for review: git push origin HEAD:%s" % push_ref,
        ],
    }
    return jsonify({
        "manual": True,
        "instructions": instructions,
        "message": "Follow the manual steps below to push the change for review.",
        "generated_at": datetime.utcnow().isoformat(),
    })


@app.route("/mcp/regression/check-lst-testcases", methods=["POST"])
@jwt_required
def check_lst_testcases():
    """Check which testcases are present in LST file."""
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json() or {}
    branch = (data.get("branch") or "master").strip()
    lst_file = (data.get("lst_file") or "").strip().replace("(pasted content)", "").strip()
    test_names = data.get("test_names") or []
    lst_file_content = data.get("lst_file_content")
    repo_name = (data.get("repo_name") or os.getenv("SOURCEGRAPH_FIRST_REPO", "nugerrit.ntnxdpro.com/nutest-py3-tests")).strip()
    if not lst_file and not (lst_file_content and isinstance(lst_file_content, str)):
        return jsonify({"error": "Enter LST file path or paste LST file content"}), 400
    test_names = [str(t).strip() for t in test_names if (t or "").strip()]
    if not test_names:
        return jsonify({"error": "test_names is required (at least one)"}), 400
    if lst_file_content:
        already_present, not_present = _check_testnames_in_lst_content(lst_file_content, test_names)
        return jsonify({"branch": branch, "lst_file": lst_file or "(pasted content)", "test_names": test_names, "present": already_present, "not_present": not_present, "generated_at": datetime.utcnow().isoformat()})
    if not lst_file:
        return jsonify({"error": "Enter LST file path or paste LST file content"}), 400
    content, err = fetch_file_content_via_sourcegraph(repo_name, branch, lst_file)
    if err:
        return jsonify({"error": err, "test_names": test_names, "present": [], "not_present": test_names, "generated_at": datetime.utcnow().isoformat()}), 200
    already_present, not_present = _check_testnames_in_lst_content(content, test_names)
    return jsonify({"branch": branch, "lst_file": lst_file, "test_names": test_names, "present": already_present, "not_present": not_present, "generated_at": datetime.utcnow().isoformat()})


def search_sourcegraph_for_test(repo_name, test_name, rev="master", lst_files_only=False):
    """Search Sourcegraph for files containing the test name."""
    test_name = str(test_name).strip() if test_name else ""
    if not test_name:
        return []
    token = (os.getenv("SOURCEGRAPH_TOKEN") or "").strip()
    if not token:
        return []
    base_url = (os.getenv("SOURCEGRAPH_URL") or "https://sourcegraph.ntnxdpro.com").strip().rstrip("/")
    # Sourcegraph's GraphQL endpoint is "/.api/graphql" (with the dot). Try the
    # configured path first (if any), then the standard candidates - same as
    # validate_with_sourcegraph / fetch_file_content_via_sourcegraph.
    configured_path = (os.getenv("SOURCEGRAPH_GRAPHQL_PATH") or "").strip()
    if configured_path and not configured_path.startswith("/"):
        configured_path = "/" + configured_path
    api_urls = []
    for p in [configured_path, "/.api/graphql", "/api/graphql", "/graphql"]:
        if not p:
            continue
        u = base_url + p
        if u not in api_urls:
            api_urls.append(u)
    auth_headers = [
        {"Content-Type": "application/json", "Authorization": "token %s" % token},
        {"Content-Type": "application/json", "Authorization": "Bearer %s" % token},
    ]
    escaped = test_name.replace("\\", "\\\\").replace('"', '\\"')
    search_query = f'repo:{repo_name} rev:{rev} "{escaped}" count:50'
    payload = {"query": "query Search($q: String!) { search(query: $q) { results { matchCount results { ... on FileMatch { file { path } } } } } }", "variables": {"q": search_query}}
    for api_url in api_urls:
        for headers in auth_headers:
            try:
                resp = requests.post(api_url, json=payload, headers=headers, timeout=20, verify=False)
                if resp.status_code != 200:
                    continue
                data = resp.json()
                if data.get("errors"):
                    continue
                search_node = (data.get("data") or {}).get("search")
                if not search_node:
                    continue
                results_node = search_node.get("results") or search_node.get("matches")
                if isinstance(results_node, dict):
                    hits = results_node.get("results", results_node.get("matches", [])) or []
                elif isinstance(results_node, list):
                    hits = results_node
                else:
                    continue
                out = [{"path": h["file"]["path"], "repo": repo_name, "rev": rev} for h in hits if isinstance(h, dict) and (h.get("file") or {}).get("path")]
                if lst_files_only:
                    out = [x for x in out if (x.get("path") or "").lower().endswith(".lst")]
                return out
            except Exception as e:
                logger.warning("[SOURCEGRAPH] Search failed (%s): %s", api_url, e)
                continue
    return []


@app.route("/mcp/regression/search-lst-file", methods=["POST"])
@jwt_required
def search_lst_file():
    """Search for LST files containing a test name via Sourcegraph."""
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json() or {}
    test_name = (data.get("test_name") or "").strip()
    if not test_name:
        return jsonify({"error": "test_name is required", "lst_files": []}), 400
    repo_name = os.getenv("SOURCEGRAPH_FIRST_REPO", "nugerrit.ntnxdpro.com/nutest-py3-tests")
    rev = (data.get("rev") or "master").strip()
    lst_files = search_sourcegraph_for_test(repo_name, test_name, rev=rev, lst_files_only=True)
    token = (os.getenv("SOURCEGRAPH_TOKEN") or "").strip()
    if not lst_files and not token:
        return jsonify({"test_name": test_name, "lst_files": [], "error": "Sourcegraph integration not configured. Set the SOURCEGRAPH_TOKEN environment variable.", "generated_at": datetime.utcnow().isoformat()})
    return jsonify({"test_name": test_name, "lst_files": [f["path"] for f in lst_files], "generated_at": datetime.utcnow().isoformat()})


@app.route("/mcp/regression/deprecate-lst-cr", methods=["POST"])
@jwt_required
def deprecate_lst_cr():
    """Return the manual git steps to remove tests from an LST file and push for review.

    Manual-only by design: no Gerrit credentials are needed on the server.
    """
    if request.method == "OPTIONS":
        return jsonify({}), 200
    data = request.get_json() or {}
    branch = (data.get("branch") or "master").strip()
    lst_file = (data.get("lst_file") or "").strip()
    test_names = data.get("test_names") or []
    reviewers = data.get("reviewers") or []
    if not test_names:
        return jsonify({"error": "test_names is required"}), 400
    if not lst_file:
        return jsonify({"error": "lst_file is required"}), 400
    test_names = [str(t).strip() for t in test_names if (t or "").strip()]
    if not test_names:
        return jsonify({"error": "test_names is required"}), 400
    push_ref = _build_gerrit_push_ref(branch, reviewers)
    instructions = {
        "message": "Remove the following %s test(s) from the LST file, then push for review." % len(test_names),
        "branch": branch, "lst_file": lst_file, "test_names": test_names,
        "manual_steps": [
            "1. Clone the repository and checkout branch '%s'" % branch,
            "2. Open the LST file: %s" % lst_file,
            "3. Remove the following %s test name(s) from the file:" % len(test_names),
            "   " + "\n   ".join(test_names[:10]) + ("..." if len(test_names) > 10 else ""),
            "4. Commit with message: 'Deprecated %s test(s) from LST'" % len(test_names),
            "5. Push for review: git push origin HEAD:%s" % push_ref,
        ],
    }
    return jsonify({
        "manual": True,
        "instructions": instructions,
        "message": "Follow the manual steps below to push the change for review.",
        "generated_at": datetime.utcnow().isoformat(),
    })


@app.route("/mcp/regression/deprecation-search", methods=["GET", "POST"])
@jwt_required
def deprecation_search():
    """Search handover records for deprecation. Includes Sourcegraph LST file hints for each query."""
    # Support both GET (params) and POST (body) to avoid URL length/encoding issues with long test names
    parts = request.args.getlist("q")
    if not parts:
        q0 = (request.args.get("q") or request.args.get("test_name") or "").strip()
        parts = [q0] if q0 else []
    if not parts and request.method == "POST":
        data = request.get_json(silent=True) or {}
        q_from_body = data.get("q") or data.get("queries") or data.get("test_names")
        if isinstance(q_from_body, list):
            parts = [str(x).strip() for x in q_from_body if str(x).strip()]
        elif q_from_body:
            parts = [str(q_from_body).strip()]
    queries = []
    for p in parts:
        if not p:
            continue
        s = str(p).strip()
        if not s:
            continue
        # Split on whitespace, comma, or newline to support "test_a test_b", "test_a,test_b", etc.
        queries.extend([t.strip() for t in re.split(r"[\s,\n\r]+", s) if t.strip()])
    seen_q = set()
    uniq_queries = []
    for q in queries:
        k = q.lower()
        if k not in seen_q:
            seen_q.add(k)
            uniq_queries.append(q)
    queries = uniq_queries
    if not queries:
        return jsonify({
            "results": [], "count": 0, "message": "Provide query parameter q (one or more test names; repeat q= for each, or comma/newline in a single q).",
            "sourcegraph_first_repo": [], "sourcegraph_other_repos": [],
            "sourcegraph_first_repo_name": "nugerrit.ntnxdpro.com/nutest-py3-tests",
            "generated_at": datetime.utcnow().isoformat()
        })
    records = _load_handover_records()
    q_lowers = [qt.lower() for qt in queries]
    matches = [r for r in records if any(qt in (r.get("test_name") or "").lower() for qt in q_lowers)]
    seen_key = set()
    unique_matches = []
    for r in matches:
        key = ((r.get("test_name") or "").strip(), (r.get("handover_date") or "").strip(), (r.get("lst_file") or "").strip())
        if key not in seen_key:
            seen_key.add(key)
            unique_matches.append(r)
    unique_matches.sort(key=lambda r: r.get("handover_date") or "", reverse=True)

    # Sourcegraph: search for LST files containing each query (first repo only for now)
    sourcegraph_first_repo = []
    sourcegraph_other_repos = []
    repo_name = os.getenv("SOURCEGRAPH_FIRST_REPO", "nugerrit.ntnxdpro.com/nutest-py3-tests")
    token = (os.getenv("SOURCEGRAPH_TOKEN") or "").strip()
    if token and queries:
        seen_paths = set()
        for test_name in queries[:10]:  # Limit to first 10 queries
            lst_files = search_sourcegraph_for_test(repo_name, test_name, rev="master", lst_files_only=True)
            for f in lst_files:
                path = f.get("path", "")
                if path and path not in seen_paths:
                    seen_paths.add(path)
                    sourcegraph_first_repo.append({"path": path, "test_name": test_name})
        sourcegraph_other_repos = []  # Can be extended with SOURCEGRAPH_OTHER_REPOS

    return jsonify({
        "q": " ".join(queries), "queries": queries, "results": unique_matches, "count": len(unique_matches),
        "sourcegraph_first_repo": sourcegraph_first_repo,
        "sourcegraph_other_repos": sourcegraph_other_repos,
        "sourcegraph_first_repo_name": repo_name,
        "generated_at": datetime.utcnow().isoformat()
    })






# ======================================================
# App Runner
# ======================================================
if __name__ == "__main__":
    scheduler_thread = threading.Thread(target=_run_plan_scheduler_loop, daemon=True)
    scheduler_thread.start()
    logger.info(f"[scheduler] Started — checking every {SCHEDULER_INTERVAL_SECONDS}s for due scheduled run plans")

    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT", "5001"))
    debug = os.getenv("FLASK_DEBUG", "false").lower() in ("true", "1", "yes")
    app.run(host=host, port=port, debug=debug, use_reloader=False)
